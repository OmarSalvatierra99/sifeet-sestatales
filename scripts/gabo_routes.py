import csv
from datetime import datetime
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import re
import sqlite3
import sys
import tempfile
import unicodedata

from flask import jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from backup_utils import prune_backup_dir
from scripts import financiamiento as financiamiento_service
from scripts.parsers import parse_cedula, parse_solventacion


def split_periodo_tokens(raw_value: str) -> list[str]:
    raw = (raw_value or "").strip()
    if not raw:
        return []
    tokens = re.split(r"(?:\r?\n|;|\|)+", raw)
    return [" ".join((token or "").strip().split()) for token in tokens if (token or "").strip()]


def normalize_periodo_key(ejercicio: str, periodo: str, *, label: str, strict: bool = True) -> str:
    clean = " ".join((periodo or "").strip().split())
    if not clean:
        return ""
    fecha_inicio, fecha_fin = parse_periodo_cedula(ejercicio, clean)
    if not fecha_inicio or not fecha_fin:
        if strict:
            raise ValueError(
                f"Titulares: {label} debe usar formato '01 de enero al 31 de diciembre'."
            )
        return clean.lower()
    return f"{fecha_inicio}|{fecha_fin}"


def normalize_solventacion_periodo_key(ejercicio: str, periodo: str) -> str:
    clean = " ".join((periodo or "").strip().split())
    if not clean:
        return ""
    months_es_to_num = {
        "enero": "01",
        "febrero": "02",
        "marzo": "03",
        "abril": "04",
        "mayo": "05",
        "junio": "06",
        "julio": "07",
        "agosto": "08",
        "septiembre": "09",
        "octubre": "10",
        "noviembre": "11",
        "diciembre": "12",
    }
    canonical = normalize_periodo_key(
        ejercicio,
        clean,
        label="periodo de solventación",
        strict=False,
    )
    if "|" in canonical:
        return canonical

    match = re.match(
        r"^\s*(\d{1,2})\s+([a-zA-ZáéíóúÁÉÍÓÚñÑ.]+)\s*[-–]\s*(\d{1,2})\s+([a-zA-ZáéíóúÁÉÍÓÚñÑ.]+)\s*$",
        clean,
    )
    if not match:
        return canonical

    month_aliases = {
        "ene": "enero",
        "feb": "febrero",
        "mar": "marzo",
        "abr": "abril",
        "may": "mayo",
        "jun": "junio",
        "jul": "julio",
        "ago": "agosto",
        "sep": "septiembre",
        "sept": "septiembre",
        "oct": "octubre",
        "nov": "noviembre",
        "dic": "diciembre",
    }

    start_day = int(match.group(1))
    start_month_key = normalize_text_key(match.group(2)).rstrip(".")
    end_day = int(match.group(3))
    end_month_key = normalize_text_key(match.group(4)).rstrip(".")
    start_month_key = month_aliases.get(start_month_key, start_month_key)
    end_month_key = month_aliases.get(end_month_key, end_month_key)

    start_month = months_es_to_num.get(start_month_key)
    end_month = months_es_to_num.get(end_month_key)
    if not start_month or not end_month:
        return canonical

    try:
        year = int(str(ejercicio).strip())
        start_date = datetime(year, int(start_month), start_day)
        end_date = datetime(year, int(end_month), end_day)
    except ValueError:
        return canonical
    return f"{start_date.strftime('%Y-%m-%d')}|{end_date.strftime('%Y-%m-%d')}"


def parse_cedula_periodos(
    ejercicio: str,
    raw_value: str,
    *,
    strict: bool = True,
) -> tuple[list[str], list[str]]:
    tokens = split_periodo_tokens(raw_value)
    periodos: list[str] = []
    keys: list[str] = []
    seen = set()
    for periodo in tokens:
        key = normalize_periodo_key(
            ejercicio,
            periodo,
            label="cada partición de cédula",
            strict=strict,
        )
        if key in seen:
            continue
        seen.add(key)
        periodos.append(periodo)
        keys.append(key)
    return periodos, keys


def register_gabo_routes(app, deps):
    globals().update(deps)
    TITULAR_EJERCICIO_FIJO = "2025"
    OBSERVACION_ESTADOS_VALIDOS = {"Emitido", "Pendiente", "Solventado"}
    GABO_READONLY_EJERCICIOS = {"2023", "2024"}
    DB_SNAPSHOT_KEEP_COUNT = 30
    MASS_UPLOAD_EJERCICIO_FIJO = "2025"
    MASS_UPLOAD_PREVIEW_LIMIT = 80
    MASS_UPLOAD_REQUIRED_HEADERS = {
        "ente_nombre_csv": "NOMBRE DEL ENTE",
        "asf": "ASF",
        "periodo": "PERIODO",
        "tipo_auditoria": "TIPO DE AUDITORÍA",
        "fuente_financiamiento": "FUENTE DE FINANCIAMIENTO",
        "tipo_anexo": "ANEXO",
        "numero_observacion": "NO. OBSERVACIÓN",
        "concepto_irregularidad": "CONCEPTO DE IRREGULARIDAD",
        "monto_observado": "MONTO OBSERVADO",
        "monto_solventado": "MONTO SOLVENTADO",
        "monto_pendiente": "MONTO PENDIENTE",
        "estatus": "ESTATUS",
    }
    MASS_UPLOAD_HEADER_ALIASES = {
        "nombre_del_ente": "ente_nombre_csv",
        "nombre_de_ente": "ente_nombre_csv",
        "nombre_ente": "ente_nombre_csv",
        "ente": "ente_nombre_csv",
        "asf": "asf",
        "periodo": "periodo",
        "tipo_de_auditoria": "tipo_auditoria",
        "tipo_auditoria": "tipo_auditoria",
        "fuente_de_financiamiento": "fuente_financiamiento",
        "fuente_financiamiento": "fuente_financiamiento",
        "fuente": "fuente_financiamiento",
        "anexo": "tipo_anexo",
        "no_observacion": "numero_observacion",
        "no_de_observacion": "numero_observacion",
        "numero_observacion": "numero_observacion",
        "numero_de_observacion": "numero_observacion",
        "concepto_de_irregularidad": "concepto_irregularidad",
        "concepto_irregularidad": "concepto_irregularidad",
        "irregularidad": "concepto_irregularidad",
        "monto_observado": "monto_observado",
        "monto_emitido": "monto_observado",
        "monto_solventado": "monto_solventado",
        "monto_pendiente": "monto_pendiente",
        "estatus": "estatus",
        "estado": "estatus",
    }
    TITULAR_IMPORT_ALLOWED_EXTENSIONS = {".xlsx"}
    TITULAR_IMPORT_PREVIEW_LIMIT = 140
    PDP_DETAIL_IMPORT_ALLOWED_EXTENSIONS = {".xlsx"}
    PDP_DETAIL_REQUIRED_HEADERS = {
        "tipo_fuente": "TIPO DE FUENTE",
        "fuente_nombre": "F.F",
        "periodo": "PERIODO",
        "tipo_auditoria": "SUBTIPO DE AUDITORIA",
        "numeral": "NUMERAL",
        "concepto": "CONCEPTO PDP",
        "monto": "MONTO PDP",
    }
    PDP_DETAIL_HEADER_ALIASES = {
        "tipo_de_fuente": "tipo_fuente",
        "tipo_fuente": "tipo_fuente",
        "origen": "tipo_fuente",
        "origen_fuente": "tipo_fuente",
        "f_f": "fuente_nombre",
        "ff": "fuente_nombre",
        "fuente": "fuente_nombre",
        "fuente_financiamiento": "fuente_nombre",
        "fuente_de_financiamiento": "fuente_nombre",
        "periodo": "periodo",
        "subtipo_de_auditoria": "tipo_auditoria",
        "subtipo_auditoria": "tipo_auditoria",
        "tipo_de_auditoria": "tipo_auditoria",
        "tipo_auditoria": "tipo_auditoria",
        "auditoria": "tipo_auditoria",
        "numeral": "numeral",
        "numero": "numeral",
        "no": "numeral",
        "no_pdp": "numeral",
        "concepto_pdp": "concepto",
        "concepto": "concepto",
        "concepto_de_pdp": "concepto",
        "concepto_de_irregularidad": "concepto",
        "monto_pdp": "monto",
        "monto": "monto",
        "importe_pdp": "monto",
        "importe": "monto",
    }
    TITULAR_MONTHS_ES = {
        "01": "enero",
        "02": "febrero",
        "03": "marzo",
        "04": "abril",
        "05": "mayo",
        "06": "junio",
        "07": "julio",
        "08": "agosto",
        "09": "septiembre",
        "10": "octubre",
        "11": "noviembre",
        "12": "diciembre",
    }
    def _tipo_auditoria_options(tipo_auditoria: str) -> list[str]:
        clean = " ".join((tipo_auditoria or "").split())
        if clean == "Financiera y Obra Pública":
            return ["Financiera", "Obra Pública"]
        if clean in {"Financiera", "Obra Pública"}:
            return [clean]
        return []

    def _normalize_observacion_estado(value: str) -> str:
        raw = " ".join((value or "").split())
        key = raw.lower()
        if key in {"e", "emitido"}:
            return "Emitido"
        if key in {"p", "pendiente"}:
            return "Pendiente"
        if key in {"s", "solventado"}:
            return "Solventado"
        return raw

    def _readonly_obs_message(ejercicio: str) -> str:
        ejercicio_clean = " ".join((ejercicio or "").split())
        return (
            f"El ejercicio {ejercicio_clean} está concluido y quedó bloqueado "
            "para edición en este módulo."
        )

    def _get_user_username(user=None) -> str:
        if isinstance(user, dict):
            return " ".join((user.get("username") or "").split()).lower()
        current_user = get_current_user()
        if isinstance(current_user, dict):
            return " ".join((current_user.get("username") or "").split()).lower()
        return ""

    def _readonly_ejercicios_for_user(user=None) -> set[str]:
        if _get_user_username(user) == "gabo":
            return set(GABO_READONLY_EJERCICIOS)
        return set()

    def _is_readonly_ejercicio(ejercicio: str, *, user=None) -> bool:
        ejercicio_clean = " ".join((ejercicio or "").split())
        return bool(ejercicio_clean and ejercicio_clean in _readonly_ejercicios_for_user(user))

    def _ensure_editable_ejercicio(ejercicio: str, *, user=None) -> None:
        if _is_readonly_ejercicio(ejercicio, user=user):
            raise ValueError(_readonly_obs_message(ejercicio))

    def _readonly_error_status(exc: ValueError) -> int:
        return 403 if "concluido" in str(exc).lower() else 400

    def _editable_ejercicios(ejercicios: list[str], *, user=None) -> list[str]:
        readonly = _readonly_ejercicios_for_user(user)
        filtered = [item for item in ejercicios if item not in readonly]
        return filtered or ejercicios

    def _require_safe_bulk_scope(scope: dict, *, action_label: str) -> None:
        ejercicio = " ".join((scope.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(scope.get("ente_id", ""))
        oficio = " ".join((scope.get("oficio") or "").split())
        if not ejercicio:
            raise ValueError("Debes seleccionar ejercicio para continuar.")
        _ensure_editable_ejercicio(ejercicio)
        if not ente_id or not oficio:
            raise ValueError(
                f"Para {action_label} define ejercicio, ente y oficio."
            )

    def _first_readonly_observacion_ejercicio(db, ids: list[int], *, user=None) -> str:
        readonly = _readonly_ejercicios_for_user(user)
        if not readonly or not ids:
            return ""
        placeholders = ", ".join(["?"] * len(ids))
        rows = db.execute(
            f"""
            SELECT DISTINCT TRIM(COALESCE(ejercicio, '')) AS ejercicio
            FROM observaciones
            WHERE id IN ({placeholders})
            """,
            ids,
        ).fetchall()
        for row in rows:
            ejercicio = " ".join((row["ejercicio"] or "").split())
            if ejercicio in readonly:
                return ejercicio
        return ""

    def _count_scope_ids(db, ids: list[int], where_clauses: list[str], params: list[str]) -> int:
        if not ids:
            return 0
        placeholders = ", ".join(["?"] * len(ids))
        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        row = db.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM observaciones
            WHERE id IN ({placeholders})
              AND {where_sql}
            """,
            [*ids, *params],
        ).fetchone()
        return int((row["total"] if row else 0) or 0)

    def _require_observaciones_admin_edit_scope(scope: dict, *, action_label: str) -> None:
        ente_id = normalize_ente_id(scope.get("ente_id", "")) if isinstance(scope, dict) else ""
        oficio = " ".join((((scope or {}).get("oficio")) or "").split()) if isinstance(scope, dict) else ""
        if not ente_id or not oficio:
            raise ValueError(
                f"Para {action_label} selecciona un ente y un oficio. La edición es oficio por oficio."
            )

    def _validate_observacion_matches_scope(row, raw_scope) -> None:
        if not isinstance(raw_scope, dict) or not raw_scope:
            return
        scope = {
            "ejercicio": " ".join((raw_scope.get("ejercicio") or "").split()),
            "ente_id": normalize_ente_id(raw_scope.get("ente_id", "")),
            "oficio": " ".join((raw_scope.get("oficio") or "").split()),
        }
        _require_observaciones_admin_edit_scope(scope, action_label="editar observaciones")

        ejercicio_actual = " ".join((row["ejercicio"] or "").split())
        ente_actual = normalize_ente_id(row["ente_id"] or "")
        oficio_actual = " ".join((row["oficio"] or "").split())

        if scope["ejercicio"] and scope["ejercicio"] != ejercicio_actual:
            raise ValueError("La observación no pertenece al ejercicio seleccionado. Refresca la consulta.")
        if scope["ente_id"] and scope["ente_id"] != ente_actual:
            raise ValueError("La observación no pertenece al ente seleccionado. Refresca la consulta.")
        if scope["oficio"] and scope["oficio"].lower() != oficio_actual.lower():
            raise ValueError("La observación no pertenece al oficio seleccionado. Refresca la consulta.")

    def _snapshot_safe_label(raw_label: str) -> str:
        clean = re.sub(r"[^a-z0-9]+", "-", (raw_label or "").strip().lower()).strip("-")
        return clean or "snapshot"

    def _create_db_snapshot(raw_label: str) -> str:
        timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S-%f")
        snapshots_dir = Path(BASE_DIR) / "backups" / "db"
        snapshots_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{timestamp}-{_snapshot_safe_label(raw_label)}.sqlite"
        backup_path = snapshots_dir / file_name
        source_conn = sqlite3.connect(DB_PATH)
        try:
            with sqlite3.connect(str(backup_path)) as backup_conn:
                source_conn.backup(backup_conn)
        finally:
            source_conn.close()
        try:
            prune_backup_dir(
                snapshots_dir,
                keep=DB_SNAPSHOT_KEEP_COUNT,
                patterns=("*.sqlite",),
                dry_run=False,
            )
        except OSError as exc:
            print(f"No se pudieron depurar snapshots antiguos: {exc}", file=sys.stderr)
        try:
            return str(backup_path.relative_to(BASE_DIR))
        except ValueError:
            return str(backup_path)

    def _build_titular_form_data(source, *, default_ejercicio: str) -> dict[str, str]:
        ejercicio = " ".join(
            (
                source.get("titular_ejercicio")
                or source.get("ejercicio")
                or default_ejercicio
                or ""
            ).split()
        )
        tipo_auditoria_source = (
            source.get("titular_tipo_auditoria")
            or source.get("tipo_auditoria")
            or ""
        )
        form_data = {
            "titular_ejercicio": ejercicio,
            "titular_ente_id": normalize_ente_id(
                source.get("titular_ente_id") or source.get("ente_id") or ""
            ),
            "titular_tipo_auditoria": normalize_tipo_auditoria(tipo_auditoria_source),
            "titular_periodo_informe": " ".join(
                (source.get("titular_periodo_informe") or "").split()
            ),
            "titular_nombre": " ".join((source.get("titular_nombre") or "").split()),
            "titular_periodo_administrativo": " ".join(
                (source.get("titular_periodo_administrativo") or "").split()
            ),
            "titular_administrativo": " ".join(
                (source.get("titular_administrativo") or "").split()
            ),
            "titular_fecha_inicio": " ".join(
                (source.get("titular_fecha_inicio") or "").split()
            ),
            "titular_fecha_fin": " ".join(
                (source.get("titular_fecha_fin") or "").split()
            ),
            "titular_admin_fecha_inicio": " ".join(
                (source.get("titular_admin_fecha_inicio") or "").split()
            ),
            "titular_admin_fecha_fin": " ".join(
                (source.get("titular_admin_fecha_fin") or "").split()
            ),
            "titular_admin_mismo_periodo": (
                "1" if (source.get("titular_admin_mismo_periodo") or "").strip() else "0"
            ),
            "titular_cedula_resultados": " ".join(
                (source.get("titular_cedula_resultados") or "").split()
            ),
        }
        if form_data["titular_periodo_informe"] and (
            not form_data["titular_fecha_inicio"] or not form_data["titular_fecha_fin"]
        ):
            fecha_inicio, fecha_fin = parse_periodo_cedula(
                ejercicio,
                form_data["titular_periodo_informe"],
            )
            if fecha_inicio and fecha_fin:
                form_data["titular_fecha_inicio"] = fecha_inicio
                form_data["titular_fecha_fin"] = fecha_fin
        if form_data["titular_periodo_administrativo"] and (
            not form_data["titular_admin_fecha_inicio"] or not form_data["titular_admin_fecha_fin"]
        ):
            fecha_inicio, fecha_fin = parse_periodo_cedula(
                ejercicio,
                form_data["titular_periodo_administrativo"],
            )
            if fecha_inicio and fecha_fin:
                form_data["titular_admin_fecha_inicio"] = fecha_inicio
                form_data["titular_admin_fecha_fin"] = fecha_fin
        if (
            form_data["titular_admin_mismo_periodo"] != "1"
            and form_data["titular_fecha_inicio"]
            and form_data["titular_fecha_fin"]
            and form_data["titular_admin_fecha_inicio"] == form_data["titular_fecha_inicio"]
            and form_data["titular_admin_fecha_fin"] == form_data["titular_fecha_fin"]
        ):
            form_data["titular_admin_mismo_periodo"] = "1"
        return form_data

    def _format_periodo_label_from_dates(fecha_inicio_iso: str, fecha_fin_iso: str) -> str:
        fecha_inicio = parse_historial_date(fecha_inicio_iso)
        fecha_fin = parse_historial_date(fecha_fin_iso)
        if not fecha_inicio or not fecha_fin:
            return ""
        mes_inicio = TITULAR_MONTHS_ES.get(fecha_inicio.strftime("%m"), "")
        mes_fin = TITULAR_MONTHS_ES.get(fecha_fin.strftime("%m"), "")
        if not mes_inicio or not mes_fin:
            return ""
        return (
            f"{fecha_inicio.day:02d} de {mes_inicio} al "
            f"{fecha_fin.day:02d} de {mes_fin}"
        )

    def _resolve_period_range(
        ejercicio: str,
        *,
        fecha_inicio_raw: str,
        fecha_fin_raw: str,
        periodo_raw: str,
        label: str,
    ) -> tuple[str, str, str]:
        fecha_inicio_clean = " ".join((fecha_inicio_raw or "").split())
        fecha_fin_clean = " ".join((fecha_fin_raw or "").split())
        periodo_clean = " ".join((periodo_raw or "").split())
        if fecha_inicio_clean or fecha_fin_clean:
            fecha_inicio = parse_historial_date(fecha_inicio_clean)
            fecha_fin = parse_historial_date(fecha_fin_clean)
            if not fecha_inicio or not fecha_fin:
                raise ValueError(f"Titulares: {label} requiere fecha inicial y final válidas.")
            if fecha_inicio > fecha_fin:
                raise ValueError(f"Titulares: {label} no puede tener fecha inicial mayor a la final.")
            fecha_inicio_iso = fecha_inicio.isoformat()
            fecha_fin_iso = fecha_fin.isoformat()
            return (
                fecha_inicio_iso,
                fecha_fin_iso,
                _format_periodo_label_from_dates(fecha_inicio_iso, fecha_fin_iso),
            )
        if not periodo_clean:
            raise ValueError(f"Titulares: {label} requerido.")
        fecha_inicio_iso, fecha_fin_iso = parse_periodo_cedula(
            ejercicio,
            periodo_clean,
        )
        if not fecha_inicio_iso or not fecha_fin_iso:
            raise ValueError(
                f"Titulares: {label} debe usar formato '01 de enero al 31 de diciembre'."
            )
        return fecha_inicio_iso, fecha_fin_iso, periodo_clean

    def _load_titular_entes(db, ejercicio: str) -> list[dict]:
        if not ejercicio:
            return []
        rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY {ente_numero_sort_sql('ente_numero')}, ente_numero, ente_nombre
            """,
            (ejercicio,),
        ).fetchall()
        return [dict(row) for row in rows]

    def _get_ente_row_by_ejercicio_id(db, ejercicio: str, ente_id_norm: str):
        if not ejercicio or not ente_id_norm:
            return None
        return db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(ente_uid, '')) AS ente_uid
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
            LIMIT 1
            """,
            (ejercicio, ente_id_norm),
        ).fetchone()

    def _list_historial_titulares_rows(
        db,
        *,
        ejercicio: str,
        ente_id_norm: str = "",
        tipo_auditoria: str = "",
    ) -> list[dict]:
        if not ejercicio:
            return []

        where_clauses = ["CAST(h.ejercicio AS TEXT) = ?"]
        params: list[str] = [ejercicio]

        if tipo_auditoria:
            where_clauses.append("TRIM(COALESCE(h.tipo_auditoria, '')) = ?")
            params.append(tipo_auditoria)

        if ente_id_norm:
            ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id_norm)
            if not ente_row:
                return []
            ente_uid = (ente_row["ente_uid"] or "").strip()
            ente_nombre = (ente_row["ente_nombre"] or "").strip()
            if ente_uid:
                where_clauses.append(
                    "(TRIM(COALESCE(h.ente_uid, '')) = ? OR TRIM(COALESCE(h.ente, '')) = ?)"
                )
                params.extend([ente_uid, ente_nombre])
            else:
                where_clauses.append("TRIM(COALESCE(h.ente, '')) = ?")
                params.append(ente_nombre)

        where_sql = " AND ".join(where_clauses)
        rows = db.execute(
            f"""
            SELECT
                h.id,
                CAST(h.ejercicio AS TEXT) AS ejercicio,
                TRIM(COALESCE(ed.ente_id, '')) AS ente_id,
                TRIM(COALESCE(ed.ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ed.ente_nombre, h.ente, '')) AS ente_nombre,
                TRIM(COALESCE(h.tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(h.nombre, '')) AS nombre,
                TRIM(COALESCE(h.cargo, '')) AS cargo,
                TRIM(COALESCE(h.fecha_inicio, '')) AS fecha_inicio,
                TRIM(COALESCE(h.fecha_fin, '')) AS fecha_fin,
                TRIM(COALESCE(h.tipo_registro, '')) AS tipo_registro
            FROM historial_titulares AS h
            LEFT JOIN entes_detalle AS ed
              ON ed.id = (
                SELECT ed2.id
                FROM entes_detalle AS ed2
                WHERE TRIM(COALESCE(ed2.ejercicio, '')) = CAST(h.ejercicio AS TEXT)
                  AND (
                    (
                      TRIM(COALESCE(h.ente_uid, '')) != ''
                      AND TRIM(COALESCE(ed2.ente_uid, '')) = TRIM(COALESCE(h.ente_uid, ''))
                    )
                    OR TRIM(COALESCE(ed2.ente_nombre, '')) = TRIM(COALESCE(h.ente, ''))
                  )
                ORDER BY
                  CASE
                    WHEN TRIM(COALESCE(h.ente_uid, '')) != ''
                     AND TRIM(COALESCE(ed2.ente_uid, '')) = TRIM(COALESCE(h.ente_uid, ''))
                    THEN 0
                    ELSE 1
                  END,
                  {ente_numero_sort_sql('ed2.ente_numero')},
                  ed2.id
                LIMIT 1
              )
            WHERE {where_sql}
            ORDER BY
              {ente_numero_sort_sql('ed.ente_numero')} ASC,
              ed.ente_numero ASC,
              ente_nombre ASC,
              h.tipo_auditoria ASC,
              h.tipo_registro ASC,
              h.fecha_inicio DESC,
              h.id DESC
            """,
            params,
        ).fetchall()
        return [dict(row) for row in rows]

    def _list_cargas_titulares_rows(
        db,
        *,
        ejercicio: str,
        ente_id_norm: str = "",
        tipo_auditoria: str = "",
    ) -> list[dict]:
        if not ejercicio:
            return []

        where_clauses = ["TRIM(COALESCE(ct.ejercicio, '')) = ?"]
        params: list[str] = [ejercicio]

        if ente_id_norm:
            where_clauses.append(f"{normalize_ente_id_sql('ct.ente_id')} = ?")
            params.append(ente_id_norm)
        if tipo_auditoria:
            where_clauses.append("TRIM(COALESCE(ct.tipo_auditoria, '')) = ?")
            params.append(tipo_auditoria)

        rows = db.execute(
            f"""
            SELECT
                ct.id,
                TRIM(COALESCE(ct.ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ct.ente_id, '')) AS ente_id,
                TRIM(COALESCE(ed.ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ed.ente_nombre, ct.ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(ct.tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(ct.periodo_informe, '')) AS periodo_informe,
                TRIM(COALESCE(ct.titular, '')) AS titular,
                TRIM(COALESCE(ct.periodo_administrativo, '')) AS periodo_administrativo,
                TRIM(COALESCE(ct.administrativo, '')) AS administrativo,
                TRIM(COALESCE(ct.cedula_resultados, '')) AS cedula_resultados,
                TRIM(COALESCE(ct.created_at, '')) AS created_at
            FROM cargas_titulares AS ct
            LEFT JOIN entes_detalle AS ed
              ON TRIM(COALESCE(ed.ejercicio, '')) = TRIM(COALESCE(ct.ejercicio, ''))
             AND {normalize_ente_id_sql('ed.ente_id')} = {normalize_ente_id_sql('ct.ente_id')}
            WHERE {" AND ".join(where_clauses)}
            ORDER BY ct.id ASC
            """,
            params,
        ).fetchall()

        payload = []
        for row in rows:
            base_row = dict(row)
            cedula_resultados = (base_row.get("cedula_resultados") or "").strip()
            cedula_periodos = [
                item.strip()
                for item in re.split(r"\s*\|\s*", cedula_resultados)
                if item.strip()
            ] or [""]
            for cedula_index, cedula_periodo in enumerate(cedula_periodos):
                split_row = dict(base_row)
                split_row["cedula_resultados"] = cedula_periodo
                split_row["cedula_orden"] = cedula_index
                payload.append(split_row)

        def sort_key(row: dict) -> tuple[str, str, str, str, int, str, int]:
            periodo_informe = row.get("periodo_informe") or ""
            periodo_admin = row.get("periodo_administrativo") or ""
            cedula_resultados = row.get("cedula_resultados") or ""
            informe_inicio, _ = parse_periodo_cedula(ejercicio, periodo_informe)
            admin_inicio, _ = parse_periodo_cedula(ejercicio, periodo_admin)
            cedula_inicio, _ = parse_periodo_cedula(ejercicio, cedula_resultados)
            return (
                informe_inicio or periodo_informe.lower(),
                admin_inicio or periodo_admin.lower(),
                (row.get("titular") or "").lower(),
                (row.get("administrativo") or "").lower(),
                int(row.get("id") or 0),
                cedula_inicio or cedula_resultados.lower(),
                int(row.get("cedula_orden") or 0),
            )

        return sorted(payload, key=sort_key)

    def _upsert_historial_titular(
        db,
        *,
        ejercicio: str,
        ente_uid: str,
        ente_nombre: str,
        tipo_auditoria: str,
        nombre: str,
        cargo: str,
        fecha_inicio: str,
        fecha_fin: str,
        tipo_registro: str,
    ) -> str:
        if ente_uid:
            existing = db.execute(
                """
                SELECT
                    id,
                    TRIM(COALESCE(ente_uid, '')) AS ente_uid,
                    TRIM(COALESCE(ente, '')) AS ente,
                    TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                    TRIM(COALESCE(nombre, '')) AS nombre,
                    TRIM(COALESCE(cargo, '')) AS cargo
                FROM historial_titulares
                WHERE CAST(ejercicio AS TEXT) = ?
                  AND tipo_auditoria = ?
                  AND tipo_registro = ?
                  AND fecha_inicio = ?
                  AND fecha_fin = ?
                  AND (
                    TRIM(COALESCE(ente_uid, '')) = ?
                    OR TRIM(COALESCE(ente, '')) = ?
                  )
                ORDER BY id DESC
                LIMIT 1
                """,
                (
                    ejercicio,
                    tipo_auditoria,
                    tipo_registro,
                    fecha_inicio,
                    fecha_fin,
                    ente_uid,
                    ente_nombre,
                ),
            ).fetchone()
        else:
            existing = db.execute(
                """
                SELECT
                    id,
                    TRIM(COALESCE(ente_uid, '')) AS ente_uid,
                    TRIM(COALESCE(ente, '')) AS ente,
                    TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                    TRIM(COALESCE(nombre, '')) AS nombre,
                    TRIM(COALESCE(cargo, '')) AS cargo
                FROM historial_titulares
                WHERE CAST(ejercicio AS TEXT) = ?
                  AND tipo_auditoria = ?
                  AND tipo_registro = ?
                  AND fecha_inicio = ?
                  AND fecha_fin = ?
                  AND TRIM(COALESCE(ente, '')) = ?
                ORDER BY id DESC
                LIMIT 1
                """,
                (
                    ejercicio,
                    tipo_auditoria,
                    tipo_registro,
                    fecha_inicio,
                    fecha_fin,
                    ente_nombre,
                ),
            ).fetchone()

        if existing:
            changed = (
                (existing["ente_uid"] or "").strip() != (ente_uid or "").strip()
                or (existing["ente"] or "").strip() != ente_nombre
                or (existing["tipo_auditoria"] or "").strip() != tipo_auditoria
                or (existing["nombre"] or "").strip() != nombre
                or (existing["cargo"] or "").strip() != cargo
            )
            db.execute(
                """
                UPDATE historial_titulares
                SET ente_uid = ?,
                    ente = ?,
                    tipo_auditoria = ?,
                    nombre = ?,
                    cargo = ?
                WHERE id = ?
                """,
                (
                    ente_uid or None,
                    ente_nombre,
                    tipo_auditoria,
                    nombre,
                    cargo,
                    existing["id"],
                ),
            )
            return "updated" if changed else "unchanged"

        db.execute(
            """
            INSERT INTO historial_titulares (
                ejercicio, ente_uid, ente, tipo_auditoria, nombre, cargo, fecha_inicio, fecha_fin, tipo_registro
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(ejercicio),
                ente_uid or None,
                ente_nombre,
                tipo_auditoria,
                nombre,
                cargo,
                fecha_inicio,
                fecha_fin,
                tipo_registro,
            ),
        )
        return "inserted"

    def _save_titulares_capture(db, user, form_data: dict[str, str]) -> dict[str, object]:
        titular_ejercicio = " ".join((form_data.get("titular_ejercicio") or "").split())
        titular_ente_id = normalize_ente_id(form_data.get("titular_ente_id", ""))
        titular_tipo_auditoria = normalize_tipo_auditoria(
            form_data.get("titular_tipo_auditoria", "")
        )
        titular_nombre = " ".join((form_data.get("titular_nombre") or "").split())
        titular_administrativo = " ".join(
            (form_data.get("titular_administrativo") or "").split()
        )
        titular_admin_mismo_periodo = (
            "1" if (form_data.get("titular_admin_mismo_periodo") or "").strip() else "0"
        )
        titular_cedula_raw = " ".join(
            (form_data.get("titular_cedula_resultados") or "").split()
        )

        if not titular_ejercicio:
            raise ValueError("Titulares: ejercicio requerido.")
        _ensure_editable_ejercicio(titular_ejercicio, user=user)
        ejercicio_exists = db.execute(
            """
            SELECT 1
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
            LIMIT 1
            """,
            (titular_ejercicio,),
        ).fetchone()
        if not ejercicio_exists:
            raise ValueError("Titulares: el ejercicio seleccionado no está disponible.")
        if not titular_ente_id:
            raise ValueError("Titulares: selecciona un ente.")
        if titular_tipo_auditoria not in {"Financiera", "Obra Pública"}:
            raise ValueError("Titulares: tipo de auditoría inválido.")
        if not titular_nombre:
            raise ValueError("Titulares: nombre del titular requerido.")
        if not titular_administrativo:
            raise ValueError("Titulares: nombre del administrativo requerido.")

        titular_inicio, titular_fin, titular_periodo_informe = _resolve_period_range(
            titular_ejercicio,
            fecha_inicio_raw=form_data.get("titular_fecha_inicio", ""),
            fecha_fin_raw=form_data.get("titular_fecha_fin", ""),
            periodo_raw=form_data.get("titular_periodo_informe", ""),
            label="periodo del titular",
        )
        admin_fecha_inicio_raw = form_data.get("titular_admin_fecha_inicio", "")
        admin_fecha_fin_raw = form_data.get("titular_admin_fecha_fin", "")
        admin_periodo_raw = form_data.get("titular_periodo_administrativo", "")
        if titular_admin_mismo_periodo == "1":
            admin_fecha_inicio_raw = titular_inicio
            admin_fecha_fin_raw = titular_fin
            admin_periodo_raw = titular_periodo_informe
        admin_inicio, admin_fin, titular_periodo_administrativo = _resolve_period_range(
            titular_ejercicio,
            fecha_inicio_raw=admin_fecha_inicio_raw,
            fecha_fin_raw=admin_fecha_fin_raw,
            periodo_raw=admin_periodo_raw,
            label="periodo del administrativo",
        )

        titular_periodo_informe_key = f"{titular_inicio}|{titular_fin}"
        if titular_cedula_raw:
            titular_cedula_periodos, titular_cedula_keys = parse_cedula_periodos(
                titular_ejercicio,
                titular_cedula_raw,
            )
        else:
            titular_cedula_periodos = [titular_periodo_informe]
            titular_cedula_keys = [titular_periodo_informe_key]
        titular_cedula_resultados = " | ".join(titular_cedula_periodos)

        ente_row = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(ente_uid, '')) AS ente_uid
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
            LIMIT 1
            """,
            (titular_ejercicio, titular_ente_id),
        ).fetchone()
        if not ente_row:
            raise ValueError("Titulares: el ente seleccionado no existe para ese ejercicio.")

        ente_nombre = (ente_row["ente_nombre"] or "").strip()
        ente_uid = (ente_row["ente_uid"] or "").strip()

        historial_states = [
            _upsert_historial_titular(
                db,
                ejercicio=titular_ejercicio,
                ente_uid=ente_uid,
                ente_nombre=ente_nombre,
                tipo_auditoria=titular_tipo_auditoria,
                nombre=titular_nombre,
                cargo="Titular",
                fecha_inicio=titular_inicio,
                fecha_fin=titular_fin,
                tipo_registro="titular",
            ),
            _upsert_historial_titular(
                db,
                ejercicio=titular_ejercicio,
                ente_uid=ente_uid,
                ente_nombre=ente_nombre,
                tipo_auditoria=titular_tipo_auditoria,
                nombre=titular_administrativo,
                cargo="Director Administrativo",
                fecha_inicio=admin_inicio,
                fecha_fin=admin_fin,
                tipo_registro="director_administrativo",
            ),
        ]

        duplicate_capture = db.execute(
            """
            SELECT id, created_at
            FROM cargas_titulares
            WHERE ejercicio = ?
              AND ente_id = ?
              AND tipo_auditoria = ?
              AND periodo_informe = ?
              AND titular = ?
              AND periodo_administrativo = ?
              AND administrativo = ?
              AND cedula_resultados = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (
                titular_ejercicio,
                titular_ente_id,
                titular_tipo_auditoria,
                titular_periodo_informe,
                titular_nombre,
                titular_periodo_administrativo,
                titular_administrativo,
                titular_cedula_resultados,
            ),
        ).fetchone()

        if duplicate_capture and all(state == "unchanged" for state in historial_states):
            return {
                "ok": False,
                "level": "info",
                "message": (
                    f"Titulares: ya existe una captura idéntica "
                    f"(ID {duplicate_capture['id']}, fecha {duplicate_capture['created_at']})."
                ),
            }

        if not duplicate_capture:
            db.execute(
                """
                INSERT INTO cargas_titulares (
                    ejercicio,
                    ente_id,
                    ente_nombre,
                    tipo_auditoria,
                    periodo_informe,
                    titular,
                    periodo_administrativo,
                    administrativo,
                    cedula_resultados,
                    created_by,
                    created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    titular_ejercicio,
                    titular_ente_id,
                    ente_nombre,
                    titular_tipo_auditoria,
                    titular_periodo_informe,
                    titular_nombre,
                    titular_periodo_administrativo,
                    titular_administrativo,
                    titular_cedula_resultados,
                    user["username"],
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                ),
            )

        db.commit()

        if duplicate_capture:
            return {
                "ok": True,
                "level": "success",
                "message": "Titulares: historial actualizado; la captura idéntica ya existía en bitácora.",
            }

        return {
            "ok": True,
            "level": "success",
            "message": "Titulares: registro guardado correctamente en historial y bitácora.",
        }

    def _clean_titular_excel_text(value) -> str:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return " ".join(str(value).replace("\xa0", " ").replace("\n", " ").split()).strip()

    def _titular_excel_merged_value(ws, row_index: int, column_index: int):
        cell = ws.cell(row_index, column_index)
        if cell.value is not None:
            return cell.value
        coordinate = cell.coordinate
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return None

    def _normalize_titular_excel_period(raw_value: str) -> str:
        clean = _clean_titular_excel_text(raw_value)
        if not clean:
            return ""
        clean = re.sub(r"\*+", "", clean).strip()
        clean = re.sub(r"\bseptimbre\b", "septiembre", clean, flags=re.IGNORECASE)
        clean = re.sub(r"\bsetiembre\b", "septiembre", clean, flags=re.IGNORECASE)
        clean = re.sub(r"\s+de\s+20\d{2}\b", "", clean, flags=re.IGNORECASE)
        return " ".join(clean.split())

    def _parse_titular_excel_period(ejercicio: str, raw_value: str, *, label: str) -> dict[str, str]:
        periodo = _normalize_titular_excel_period(raw_value)
        if not periodo:
            raise ValueError(f"{label}: periodo vacío.")
        fecha_inicio, fecha_fin = parse_periodo_cedula(ejercicio, periodo)
        if not fecha_inicio or not fecha_fin:
            raise ValueError(
                f"{label}: periodo inválido {periodo!r}. Usa formato '01 de enero al 31 de diciembre'."
            )
        return {
            "periodo": _format_periodo_label_from_dates(fecha_inicio, fecha_fin) or periodo,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        }

    def _is_titular_excel_note_row(values: list[str]) -> bool:
        joined = normalize_text_key(" ".join(values))
        if not joined:
            return True
        note_markers = (
            "fecha de entrega",
            "nota",
            "se autorizo",
            "revision a",
            "revisión a",
        )
        if any(marker in joined for marker in note_markers):
            return True
        first_value = (values[0] or "").strip()
        return first_value.startswith("*")

    def _titular_import_file_entries(upload) -> list[dict[str, object]]:
        file_name = Path(upload.filename or "").name
        suffix = Path(file_name).suffix.lower()
        if suffix not in TITULAR_IMPORT_ALLOWED_EXTENSIONS:
            raise ValueError("Carga titulares: sube un archivo .xlsx.")
        data = upload.read()
        if not data:
            raise ValueError("Carga titulares: el archivo está vacío.")
        return [{"file_name": file_name, "data": data}]

    def _find_titular_excel_header_row(ws) -> int:
        for row_index in range(1, min(ws.max_row, 20) + 1):
            values = [
                normalize_text_key(_clean_titular_excel_text(_titular_excel_merged_value(ws, row_index, col)))
                for col in range(1, 6)
            ]
            if "periodos informe" in values[0] and "titular" in values[1]:
                return row_index
        return 4

    def _titular_import_tipo_options(raw_value: str) -> list[str]:
        clean = normalize_tipo_auditoria(raw_value or "")
        key = normalize_text_key(clean)
        if key in {
            "ambas",
            "financiera y obra publica",
            "obra publica y financiera",
            "financiera obra publica",
        }:
            return ["Financiera", "Obra Pública"]
        if clean in {"Financiera", "Obra Pública"}:
            return [clean]
        return ["Financiera", "Obra Pública"]

    def _clean_pdp_detail_excel_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return " ".join(str(value).replace("\u00a0", " ").split())

    def _pdp_detail_header_key(value) -> str:
        clean = normalize_text_key(_clean_pdp_detail_excel_text(value))
        return re.sub(r"[^a-z0-9]+", "_", clean).strip("_")

    def _pdp_detail_file_entry(upload) -> dict[str, object]:
        if not upload or not (upload.filename or "").strip():
            raise ValueError("Detalle PDP: selecciona un archivo .xlsx.")
        file_name = Path(upload.filename or "").name
        suffix = Path(file_name).suffix.lower()
        if suffix not in PDP_DETAIL_IMPORT_ALLOWED_EXTENSIONS:
            raise ValueError("Detalle PDP: sube un archivo .xlsx.")
        data = upload.read()
        if not data:
            raise ValueError("Detalle PDP: el archivo está vacío.")
        return {"file_name": file_name, "data": data}

    def _parse_pdp_detail_numeral(value, *, row_number: int) -> int:
        raw = _clean_pdp_detail_excel_text(value)
        if not raw:
            raise ValueError(f"Fila {row_number}: NUMERAL requerido.")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            parsed = int(value)
            if abs(float(value) - parsed) > 0.0001:
                raise ValueError(f"Fila {row_number}: NUMERAL debe ser entero.")
        else:
            match = re.search(r"\d+", raw)
            if not match:
                raise ValueError(f"Fila {row_number}: NUMERAL debe ser entero.")
            parsed = int(match.group(0))
        if parsed <= 0:
            raise ValueError(f"Fila {row_number}: NUMERAL debe ser mayor a cero.")
        return parsed

    def _parse_pdp_detail_amount(value, *, row_number: int) -> float:
        if value is None or _clean_pdp_detail_excel_text(value) == "":
            raise ValueError(f"Fila {row_number}: MONTO PDP requerido.")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            parsed = float(value)
        else:
            raw = _clean_pdp_detail_excel_text(value)
            cleaned = (
                raw.replace("$", "")
                .replace(",", "")
                .replace("MXN", "")
                .replace("mxn", "")
                .strip()
            )
            try:
                parsed = float(cleaned)
            except ValueError as exc:
                raise ValueError(f"Fila {row_number}: MONTO PDP inválido.") from exc
        if parsed < 0:
            raise ValueError(f"Fila {row_number}: MONTO PDP no puede ser negativo.")
        return parsed

    def _normalize_pdp_detail_tipo_fuente(value: str) -> dict[str, object]:
        clean = _clean_pdp_detail_excel_text(value)
        key = normalize_text_key(clean)
        compact_key = re.sub(r"[^a-z0-9]+", "", key)
        modalidad = "Convenio" if "convenio" in compact_key else "Fuente"
        es_seguimiento = "seguimiento" in compact_key
        es_remanente = "remanente" in compact_key
        es_del_ejercicio = "delejercicio" in compact_key or compact_key == "ejercicio"
        ejercicio_match = re.search(r"\b(20\d{2})\b", clean)
        ejercicio_fuente = ejercicio_match.group(1) if ejercicio_match else ""
        origen_fuente = "Del Ejercicio"
        if es_remanente or es_seguimiento:
            origen_fuente = "Remanentes"
        elif es_del_ejercicio:
            origen_fuente = "Del Ejercicio"
        tipo_parts = []
        if modalidad == "Convenio":
            tipo_parts.append("convenio")
        if es_seguimiento:
            tipo_parts.append("seguimiento")
        if es_remanente:
            tipo_parts.append("remanente")
        elif es_del_ejercicio:
            tipo_parts.append("del_ejercicio")
        if ejercicio_fuente:
            tipo_parts.append(ejercicio_fuente)
        return {
            "tipo_fuente": clean,
            "tipo_fuente_clave": "_".join(tipo_parts) or compact_key,
            "ejercicio_fuente": ejercicio_fuente,
            "es_seguimiento": es_seguimiento,
            "modalidad": modalidad,
            "origen_fuente": origen_fuente,
        }

    def _split_pdp_detail_convenio_fuente(value, *, modalidad: str) -> dict[str, str]:
        raw = "" if value is None else str(value).replace("\r\n", "\n").replace("\r", "\n")
        clean = _clean_pdp_detail_excel_text(raw)
        fuente_original = normalize_fuente_financiamiento(clean)
        if modalidad != "Convenio":
            return {
                "fuente_nombre": fuente_original,
                "fuente_nombre_original": fuente_original,
                "convenio_nombre": "",
            }

        convenio_nombre = ""
        fuente_text = ""
        lines = [
            _clean_pdp_detail_excel_text(line)
            for line in raw.split("\n")
            if _clean_pdp_detail_excel_text(line)
        ]
        if lines and normalize_text_key(lines[0]).startswith("convenio"):
            convenio_part = re.sub(r"^convenio\s*:\s*", "", lines[0], flags=re.IGNORECASE).strip()
            convenio_nombre = normalize_convenio_text(convenio_part)
            fuente_text = " ".join(lines[1:]).strip()

        if not fuente_text and clean:
            markers = (
                "REMANENTES DE EJERCICIOS ANTERIORES:",
                "DE EJERCICIOS ANTERIORES:",
                "FONDO DE ",
                "RECURSOS RECAUDADOS",
            )
            upper_clean = clean.upper()
            for marker in markers:
                marker_index = upper_clean.find(marker)
                if marker_index <= 0:
                    continue
                convenio_part = clean[:marker_index].strip()
                fuente_text = clean[marker_index:].strip()
                convenio_part = re.sub(r"^convenio\s*:\s*", "", convenio_part, flags=re.IGNORECASE).strip()
                convenio_nombre = normalize_convenio_text(convenio_part)
                break

        return {
            "fuente_nombre": normalize_fuente_financiamiento(fuente_text or clean),
            "fuente_nombre_original": fuente_original,
            "convenio_nombre": convenio_nombre,
        }

    def _find_pdp_detail_header_row(ws) -> tuple[int, dict[str, int]]:
        max_header_row = min(ws.max_row or 1, 20)
        max_col = min(ws.max_column or 1, 40)
        best_mapping: dict[str, int] = {}
        for row_index in range(1, max_header_row + 1):
            mapping: dict[str, int] = {}
            for col_index in range(1, max_col + 1):
                alias = PDP_DETAIL_HEADER_ALIASES.get(
                    _pdp_detail_header_key(ws.cell(row_index, col_index).value)
                )
                if alias and alias not in mapping:
                    mapping[alias] = col_index
            if len(mapping) > len(best_mapping):
                best_mapping = mapping
            if all(field in mapping for field in PDP_DETAIL_REQUIRED_HEADERS):
                return row_index, mapping
        missing = [
            label
            for field, label in PDP_DETAIL_REQUIRED_HEADERS.items()
            if field not in best_mapping
        ]
        raise ValueError(
            "Detalle PDP: no se encontraron columnas requeridas: "
            + ", ".join(missing)
            + "."
        )

    def _parse_pdp_detail_excel_upload(upload) -> dict[str, object]:
        entry = _pdp_detail_file_entry(upload)
        try:
            workbook = load_workbook(BytesIO(entry["data"]), data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValueError("Detalle PDP: el archivo no se pudo abrir como Excel.") from exc

        entries: list[dict[str, object]] = []
        warnings: list[str] = []
        sheet_summaries: list[dict[str, object]] = []

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            try:
                header_row, header_map = _find_pdp_detail_header_row(ws)
            except ValueError as exc:
                warnings.append(f"Hoja {sheet_name}: {exc}")
                continue

            sheet_count = 0
            for row_number in range(header_row + 1, (ws.max_row or header_row) + 1):
                values = {
                    field: ws.cell(row_number, column_index).value
                    for field, column_index in header_map.items()
                }
                if not any(_clean_pdp_detail_excel_text(value) for value in values.values()):
                    continue
                try:
                    tipo_fuente_info = _normalize_pdp_detail_tipo_fuente(
                        values.get("tipo_fuente")
                    )
                    fuente_info = _split_pdp_detail_convenio_fuente(
                        values.get("fuente_nombre"),
                        modalidad=tipo_fuente_info["modalidad"],
                    )
                    fuente_nombre = fuente_info["fuente_nombre"]
                    periodo = _clean_pdp_detail_excel_text(values.get("periodo"))
                    tipo_auditoria = normalize_tipo_auditoria(
                        _clean_pdp_detail_excel_text(values.get("tipo_auditoria"))
                    )
                    numeral = _parse_pdp_detail_numeral(
                        values.get("numeral"),
                        row_number=row_number,
                    )
                    concepto = normalize_irregularidad_concepto(
                        _clean_pdp_detail_excel_text(values.get("concepto")),
                        strict=True,
                    )
                    monto = _parse_pdp_detail_amount(
                        values.get("monto"),
                        row_number=row_number,
                    )
                    if not fuente_nombre:
                        raise ValueError(f"Fila {row_number}: F.F requerida.")
                    if not tipo_fuente_info["tipo_fuente"]:
                        raise ValueError(f"Fila {row_number}: TIPO DE FUENTE requerido.")
                    if not periodo:
                        raise ValueError(f"Fila {row_number}: PERIODO requerido.")
                    if tipo_auditoria not in {"Financiera", "Obra Pública"}:
                        raise ValueError(
                            f"Fila {row_number}: SUBTIPO DE AUDITORIA debe ser Financiero/Financiera u Obra Pública."
                        )
                except ValueError as exc:
                    warnings.append(f"Hoja {sheet_name}: {exc}")
                    continue

                entries.append(
                    {
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "tipo_fuente": tipo_fuente_info["tipo_fuente"],
                        "tipo_fuente_clave": tipo_fuente_info["tipo_fuente_clave"],
                        "ejercicio_fuente": tipo_fuente_info["ejercicio_fuente"],
                        "es_seguimiento": tipo_fuente_info["es_seguimiento"],
                        "modalidad": tipo_fuente_info["modalidad"],
                        "origen_fuente": tipo_fuente_info["origen_fuente"],
                        "fuente_nombre": fuente_nombre,
                        "fuente_nombre_original": fuente_info["fuente_nombre_original"],
                        "convenio_nombre": fuente_info["convenio_nombre"],
                        "periodo": periodo,
                        "tipo_auditoria": tipo_auditoria,
                        "numeral": numeral,
                        "concepto": concepto,
                        "subconcepto": "",
                        "monto": round(monto, 2),
                    }
                )
                sheet_count += 1
            sheet_summaries.append(
                {
                    "sheet_name": sheet_name,
                    "header_row": header_row,
                    "entries": sheet_count,
                }
            )

        if not sheet_summaries:
            raise ValueError("Detalle PDP: ninguna hoja contiene la plantilla esperada.")
        if not entries:
            raise ValueError("Detalle PDP: el Excel no contiene filas válidas.")

        return {
            "ok": True,
            "file_name": entry["file_name"],
            "entries": entries,
            "warnings": warnings,
            "sheets": sheet_summaries,
            "summary": {
                "entries": len(entries),
                "warnings": len(warnings),
                "sheets": len(sheet_summaries),
            },
        }

    def _parse_titular_excel_workbook(
        *,
        file_name: str,
        data: bytes,
        ejercicio: str,
        ente_row: dict,
    ) -> dict[str, object]:
        file_result = {
            "file_name": file_name,
            "sheet_name": "",
            "ejercicio": "",
            "ente_excel": "",
            "ente_id": "",
            "ente_numero": "",
            "ente_nombre": "",
            "resolution": "",
            "status": "ok",
            "warnings": [],
            "errors": [],
            "source_rows": 0,
            "titulares": 0,
            "administrativos": 0,
            "capturas": 0,
        }
        history_rows: list[dict[str, str]] = []
        capture_rows: list[dict[str, object]] = []

        try:
            workbook = load_workbook(BytesIO(data), data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            file_result["status"] = "error"
            file_result["errors"].append("El archivo no se pudo abrir como Excel.")
            return {"file": file_result, "history_rows": history_rows, "capture_rows": capture_rows}

        ws = workbook[workbook.sheetnames[0]]
        file_result["sheet_name"] = ws.title
        ejercicio_text = _clean_titular_excel_text(_titular_excel_merged_value(ws, 1, 1))
        ejercicio_match = re.search(r"(20\d{2})", ejercicio_text)
        ejercicio_excel = ejercicio_match.group(1) if ejercicio_match else ""
        file_result["ejercicio"] = ejercicio
        ente_excel = _clean_titular_excel_text(_titular_excel_merged_value(ws, 2, 2))
        file_result["ente_excel"] = ente_excel
        if ejercicio_excel and ejercicio_excel != ejercicio:
            file_result["warnings"].append(
                f"El Excel indica ejercicio {ejercicio_excel}, pero se cargará en {ejercicio}."
            )

        file_result.update(
            {
                "ente_id": ente_row["ente_id"],
                "ente_numero": ente_row["ente_numero"],
                "ente_nombre": ente_row["ente_nombre"],
                "resolution": "contexto seleccionado",
            }
        )

        header_row = _find_titular_excel_header_row(ws)
        titular_seen: dict[tuple[str, str, str], dict[str, str]] = {}
        admin_seen: dict[tuple[str, str, str], dict[str, str]] = {}
        capture_groups: dict[tuple[str, str, str, str, str, str], dict[str, object]] = {}

        for row_index in range(header_row + 1, ws.max_row + 1):
            raw_values = [
                _clean_titular_excel_text(_titular_excel_merged_value(ws, row_index, column_index))
                for column_index in range(1, 6)
            ]
            if not any(raw_values) or _is_titular_excel_note_row(raw_values):
                continue

            file_result["source_rows"] += 1
            try:
                periodo_informe = _parse_titular_excel_period(
                    ejercicio,
                    raw_values[0],
                    label=f"{file_name}, fila {row_index}, periodo informe",
                )
                cedula = None
                if _normalize_titular_excel_period(raw_values[4]):
                    cedula = _parse_titular_excel_period(
                        ejercicio,
                        raw_values[4],
                        label=f"{file_name}, fila {row_index}, cédula",
                    )
                admin_period_raw = raw_values[2]
                if (
                    not _normalize_titular_excel_period(admin_period_raw)
                    and _clean_titular_excel_text(raw_values[3])
                    and cedula
                ):
                    admin_period_raw = cedula["periodo"]
                periodo_admin = _parse_titular_excel_period(
                    ejercicio,
                    admin_period_raw,
                    label=f"{file_name}, fila {row_index}, periodo administrativo",
                )
            except ValueError as exc:
                file_result["warnings"].append(str(exc))
                continue

            titular_nombre = _clean_titular_excel_text(raw_values[1])
            admin_nombre = _clean_titular_excel_text(raw_values[3])
            if not titular_nombre or not admin_nombre:
                file_result["warnings"].append(
                    f"{file_name}, fila {row_index}: faltan titular o administrativo."
                )
                continue

            base_context = {
                "ejercicio": ejercicio,
                "ente_id": ente_row["ente_id"],
                "ente_numero": ente_row["ente_numero"],
                "ente_nombre": ente_row["ente_nombre"],
                "ente_excel": ente_excel,
                "file_name": file_name,
                "sheet_name": ws.title,
            }
            titular_key = (
                periodo_informe["fecha_inicio"],
                periodo_informe["fecha_fin"],
                titular_nombre.lower(),
            )
            if titular_key not in titular_seen:
                titular_seen[titular_key] = {
                    **base_context,
                    "tipo_registro": "titular",
                    "nombre": titular_nombre,
                    "cargo": "Titular",
                    "periodo": periodo_informe["periodo"],
                    "fecha_inicio": periodo_informe["fecha_inicio"],
                    "fecha_fin": periodo_informe["fecha_fin"],
                }
            admin_key = (
                periodo_admin["fecha_inicio"],
                periodo_admin["fecha_fin"],
                admin_nombre.lower(),
            )
            if admin_key not in admin_seen:
                admin_seen[admin_key] = {
                    **base_context,
                    "tipo_registro": "director_administrativo",
                    "nombre": admin_nombre,
                    "cargo": "Director Administrativo",
                    "periodo": periodo_admin["periodo"],
                    "fecha_inicio": periodo_admin["fecha_inicio"],
                    "fecha_fin": periodo_admin["fecha_fin"],
                }

            capture_key = (
                periodo_informe["fecha_inicio"],
                periodo_informe["fecha_fin"],
                titular_nombre.lower(),
                periodo_admin["fecha_inicio"],
                periodo_admin["fecha_fin"],
                admin_nombre.lower(),
            )
            group = capture_groups.setdefault(
                capture_key,
                {
                    **base_context,
                    "periodo_informe": periodo_informe["periodo"],
                    "periodo_informe_inicio": periodo_informe["fecha_inicio"],
                    "periodo_informe_fin": periodo_informe["fecha_fin"],
                    "titular": titular_nombre,
                    "periodo_administrativo": periodo_admin["periodo"],
                    "periodo_administrativo_inicio": periodo_admin["fecha_inicio"],
                    "periodo_administrativo_fin": periodo_admin["fecha_fin"],
                    "administrativo": admin_nombre,
                    "cedulas": {},
                    "source_rows": [],
                },
            )
            group["source_rows"].append(row_index)
            if cedula:
                group["cedulas"][(cedula["fecha_inicio"], cedula["fecha_fin"])] = cedula["periodo"]

        history_rows = list(titular_seen.values()) + list(admin_seen.values())
        capture_rows = []
        for group in capture_groups.values():
            cedula_items = [
                {"fecha_inicio": start, "fecha_fin": end, "periodo": label}
                for (start, end), label in sorted(group.pop("cedulas").items())
            ]
            group["cedula_periodos"] = cedula_items
            group["cedula_resultados"] = " | ".join(item["periodo"] for item in cedula_items)
            group["source_rows"] = ", ".join(str(item) for item in group["source_rows"])
            capture_rows.append(group)

        file_result["titulares"] = sum(1 for row in history_rows if row["tipo_registro"] == "titular")
        file_result["administrativos"] = sum(
            1 for row in history_rows if row["tipo_registro"] == "director_administrativo"
        )
        file_result["capturas"] = len(capture_rows)
        if file_result["warnings"]:
            file_result["status"] = "warning"
        if not capture_rows and not history_rows and not file_result["warnings"]:
            file_result["status"] = "warning"
            file_result["warnings"].append("No se encontraron registros de titulares en el formato esperado.")

        return {"file": file_result, "history_rows": history_rows, "capture_rows": capture_rows}

    def _build_titulares_excel_preview(
        db,
        upload,
        *,
        ejercicio: str,
        ente_id_norm: str,
        tipo_auditoria_destino: str,
    ) -> dict[str, object]:
        if not ejercicio:
            raise ValueError("Carga titulares: selecciona un ejercicio.")
        _ensure_editable_ejercicio(ejercicio, user=get_current_user())
        if not ente_id_norm:
            raise ValueError("Carga titulares: selecciona un ente.")
        tipo_options = _titular_import_tipo_options(tipo_auditoria_destino)
        if len(tipo_options) != 1:
            raise ValueError("Carga titulares: selecciona un solo tipo de auditoría.")
        ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id_norm)
        if not ente_row:
            raise ValueError("Carga titulares: el ente seleccionado no existe para ese ejercicio.")

        entries = _titular_import_file_entries(upload)
        files: list[dict[str, object]] = []
        history_rows: list[dict[str, str]] = []
        capture_rows: list[dict[str, object]] = []

        for entry in entries:
            parsed = _parse_titular_excel_workbook(
                file_name=str(entry["file_name"]),
                data=entry["data"],
                ejercicio=ejercicio,
                ente_row=dict(ente_row),
            )
            files.append(parsed["file"])
            history_rows.extend(parsed["history_rows"])
            capture_rows.extend(parsed["capture_rows"])

        titular_count = sum(1 for row in history_rows if row["tipo_registro"] == "titular")
        admin_count = sum(1 for row in history_rows if row["tipo_registro"] == "director_administrativo")
        warning_count = sum(len(item.get("warnings") or []) for item in files)
        error_count = sum(len(item.get("errors") or []) for item in files)

        return {
            "ok": True,
            "tipo_auditoria_destino": tipo_options[0],
            "tipo_auditoria_options": tipo_options,
            "files": files,
            "history_rows": history_rows,
            "capture_rows": capture_rows,
            "preview_capture_rows": capture_rows[:TITULAR_IMPORT_PREVIEW_LIMIT],
            "preview_limit": TITULAR_IMPORT_PREVIEW_LIMIT,
            "summary": {
                "files_total": len(files),
                "files_ok": sum(1 for item in files if item.get("status") == "ok"),
                "files_warning": sum(1 for item in files if item.get("status") == "warning"),
                "files_error": sum(1 for item in files if item.get("status") == "error"),
                "titulares": titular_count,
                "administrativos": admin_count,
                "capturas": len(capture_rows),
                "cedulas": sum(len(row.get("cedula_periodos") or []) for row in capture_rows),
                "tipos_destino": len(tipo_options),
                "warnings": warning_count,
                "errors": error_count,
            },
        }

    def _insert_titulares_import_capture(
        db,
        *,
        user,
        ejercicio: str,
        ente_id: str,
        ente_nombre: str,
        tipo_auditoria: str,
        periodo_informe: str,
        titular: str,
        periodo_administrativo: str,
        administrativo: str,
        cedula_resultados: str,
    ) -> str:
        duplicate_capture = db.execute(
            """
            SELECT id
            FROM cargas_titulares
            WHERE ejercicio = ?
              AND ente_id = ?
              AND tipo_auditoria = ?
              AND periodo_informe = ?
              AND titular = ?
              AND periodo_administrativo = ?
              AND administrativo = ?
              AND cedula_resultados = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (
                ejercicio,
                ente_id,
                tipo_auditoria,
                periodo_informe,
                titular,
                periodo_administrativo,
                administrativo,
                cedula_resultados,
            ),
        ).fetchone()
        if duplicate_capture:
            return "unchanged"

        db.execute(
            """
            INSERT INTO cargas_titulares (
                ejercicio,
                ente_id,
                ente_nombre,
                tipo_auditoria,
                periodo_informe,
                titular,
                periodo_administrativo,
                administrativo,
                cedula_resultados,
                created_by,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                ejercicio,
                ente_id,
                ente_nombre,
                tipo_auditoria,
                periodo_informe,
                titular,
                periodo_administrativo,
                administrativo,
                cedula_resultados,
                user["username"],
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ),
        )
        return "inserted"

    def _apply_titulares_excel_import(db, user, payload: dict) -> dict[str, object]:
        history_rows = payload.get("history_rows")
        capture_rows = payload.get("capture_rows")
        if not isinstance(history_rows, list) or not isinstance(capture_rows, list):
            raise ValueError("Carga titulares: analiza el Excel antes de aplicar.")
        tipo_options = _titular_import_tipo_options(payload.get("tipo_auditoria_destino") or "")
        if len(tipo_options) != 1:
            raise ValueError("Carga titulares: selecciona un solo tipo de auditoría.")
        if not history_rows and not capture_rows:
            raise ValueError("Carga titulares: no hay registros válidos para guardar.")

        for raw_row in [*history_rows, *capture_rows]:
            if not isinstance(raw_row, dict):
                continue
            ejercicio = " ".join((raw_row.get("ejercicio") or "").split())
            if ejercicio:
                _ensure_editable_ejercicio(ejercicio, user=user)

        backup_path = _create_db_snapshot("titulares-excel-import")
        history_counts = {"inserted": 0, "updated": 0, "unchanged": 0}
        capture_counts = {"inserted": 0, "unchanged": 0}
        history_seen: set[tuple[str, ...]] = set()
        capture_seen: set[tuple[str, ...]] = set()

        for raw_row in history_rows:
            if not isinstance(raw_row, dict):
                continue
            ejercicio = " ".join((raw_row.get("ejercicio") or "").split())
            ente_id = normalize_ente_id(raw_row.get("ente_id") or "")
            tipo_registro = " ".join((raw_row.get("tipo_registro") or "").split())
            nombre = " ".join((raw_row.get("nombre") or "").split())
            cargo = " ".join((raw_row.get("cargo") or "").split())
            fecha_inicio = " ".join((raw_row.get("fecha_inicio") or "").split())
            fecha_fin = " ".join((raw_row.get("fecha_fin") or "").split())
            if not ejercicio or not ente_id:
                raise ValueError("Carga titulares: cada registro debe tener ejercicio y ente.")
            _ensure_editable_ejercicio(ejercicio, user=user)
            if tipo_registro not in {"titular", "director_administrativo"}:
                raise ValueError("Carga titulares: tipo de registro inválido.")
            if not nombre:
                raise ValueError("Carga titulares: cada registro debe tener nombre.")
            inicio_date = parse_historial_date(fecha_inicio)
            fin_date = parse_historial_date(fecha_fin)
            if not inicio_date or not fin_date or inicio_date > fin_date:
                raise ValueError("Carga titulares: hay fechas inválidas en el Excel.")

            ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
            if not ente_row:
                raise ValueError(f"Carga titulares: ente {ente_id} no existe para {ejercicio}.")
            ente_nombre = (ente_row["ente_nombre"] or "").strip()
            ente_uid = (ente_row["ente_uid"] or "").strip()
            for tipo_auditoria in tipo_options:
                key = (
                    ejercicio,
                    ente_id,
                    tipo_auditoria,
                    tipo_registro,
                    fecha_inicio,
                    fecha_fin,
                    nombre,
                    cargo,
                )
                if key in history_seen:
                    continue
                history_seen.add(key)
                state = _upsert_historial_titular(
                    db,
                    ejercicio=ejercicio,
                    ente_uid=ente_uid,
                    ente_nombre=ente_nombre,
                    tipo_auditoria=tipo_auditoria,
                    nombre=nombre,
                    cargo=cargo or ("Titular" if tipo_registro == "titular" else "Director Administrativo"),
                    fecha_inicio=inicio_date.isoformat(),
                    fecha_fin=fin_date.isoformat(),
                    tipo_registro=tipo_registro,
                )
                history_counts[state] = history_counts.get(state, 0) + 1

        for raw_row in capture_rows:
            if not isinstance(raw_row, dict):
                continue
            ejercicio = " ".join((raw_row.get("ejercicio") or "").split())
            ente_id = normalize_ente_id(raw_row.get("ente_id") or "")
            periodo_informe = " ".join((raw_row.get("periodo_informe") or "").split())
            titular = " ".join((raw_row.get("titular") or "").split())
            periodo_administrativo = " ".join((raw_row.get("periodo_administrativo") or "").split())
            administrativo = " ".join((raw_row.get("administrativo") or "").split())
            cedula_resultados = " ".join((raw_row.get("cedula_resultados") or "").split())
            if not all([ejercicio, ente_id, periodo_informe, titular, periodo_administrativo, administrativo]):
                continue
            _ensure_editable_ejercicio(ejercicio, user=user)
            ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
            if not ente_row:
                raise ValueError(f"Carga titulares: ente {ente_id} no existe para {ejercicio}.")
            ente_nombre = (ente_row["ente_nombre"] or "").strip()
            for tipo_auditoria in tipo_options:
                key = (
                    ejercicio,
                    ente_id,
                    tipo_auditoria,
                    periodo_informe,
                    titular,
                    periodo_administrativo,
                    administrativo,
                    cedula_resultados,
                )
                if key in capture_seen:
                    continue
                capture_seen.add(key)
                state = _insert_titulares_import_capture(
                    db,
                    user=user,
                    ejercicio=ejercicio,
                    ente_id=ente_id,
                    ente_nombre=ente_nombre,
                    tipo_auditoria=tipo_auditoria,
                    periodo_informe=periodo_informe,
                    titular=titular,
                    periodo_administrativo=periodo_administrativo,
                    administrativo=administrativo,
                    cedula_resultados=cedula_resultados,
                )
                capture_counts[state] = capture_counts.get(state, 0) + 1

        db.commit()
        return {
            "ok": True,
            "message": "Carga titulares: Excel guardado en historial y bitácora.",
            "backup_path": backup_path,
            "history": history_counts,
            "captures": capture_counts,
            "tipo_auditoria_options": tipo_options,
        }

    def resolve_fuente_catalogo(
        db,
        fuente_nombre: str,
        *,
        create_missing: bool = False,
    ) -> tuple[int | None, str]:
        return financiamiento_service.resolve_fuente_catalogo(
            db,
            fuente_nombre,
            normalizer=normalize_fuente_financiamiento,
            create_missing=create_missing,
        )

    def get_fuente_clasificacion(db, fuente_nombre: str, fuente_id: int | None = None) -> dict:
        return financiamiento_service.get_fuente_clasificacion(
            db,
            fuente_nombre,
            fuente_id=fuente_id,
            normalizer=normalize_fuente_financiamiento,
            si_no_normalizer=normalize_manual_si_no,
            origen_normalizer=normalize_manual_origen_fuente,
        )

    def list_fuentes_financiamiento_admin(db, ejercicio: str) -> list[dict]:
        return financiamiento_service.list_fuentes_financiamiento_admin(
            db,
            ejercicio,
            normalizer=normalize_fuente_financiamiento,
            si_no_normalizer=normalize_manual_si_no,
            origen_normalizer=normalize_manual_origen_fuente,
        )

    def register_fuente_for_ente(
        db,
        *,
        ejercicio: str,
        ente_id_norm: str,
        fuente_id: int,
        tipo_auditoria: str,
        created_by: str = "",
    ) -> None:
        ejercicio_clean = " ".join((ejercicio or "").split())
        ente_id_clean = normalize_ente_id(ente_id_norm)
        tipo_clean = normalize_tipo_auditoria(tipo_auditoria)
        if not ejercicio_clean or not ente_id_clean or not tipo_clean:
            return
        if tipo_clean not in {"Financiera", "Obra Pública"}:
            return
        try:
            fuente_id_int = int(fuente_id)
        except (TypeError, ValueError):
            return
        if fuente_id_int <= 0:
            return
        existing = db.execute(
            f"""
            SELECT 1
            FROM entes_fuentes
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
              AND fuente_id = ?
              AND TRIM(COALESCE(tipo_auditoria, '')) = ?
            LIMIT 1
            """,
            (
                ejercicio_clean,
                ente_id_clean,
                fuente_id_int,
                tipo_clean,
            ),
        ).fetchone()
        if existing:
            return
        db.execute(
            """
            INSERT INTO entes_fuentes (
                ejercicio,
                ente_id,
                fuente_id,
                tipo_auditoria,
                created_by,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                ejercicio_clean,
                ente_id_clean,
                fuente_id_int,
                tipo_clean,
                (created_by or "").strip() or None,
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ),
        )

    def fuentes_por_ente(
        db,
        ejercicio: str,
        ente_id_norm: str,
        tipo_auditoria: str = "",
    ) -> list[dict]:
        if not ejercicio or not ente_id_norm:
            return []
        tipo_options = _tipo_auditoria_options(tipo_auditoria)
        tipo_filter_sql = ""
        tipo_filter_params: list[str] = []
        if tipo_options:
            placeholders = ", ".join(["?"] * len(tipo_options))
            tipo_filter_sql = f" AND TRIM(COALESCE(o.tipo_auditoria, '')) IN ({placeholders})"
            tipo_filter_params = tipo_options.copy()

        ente_base = db.execute(
            f"""
            SELECT TRIM(COALESCE(ente_uid, '')) AS ente_uid
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
            LIMIT 1
            """,
            (ejercicio, ente_id_norm),
        ).fetchone()
        ente_uid = (ente_base["ente_uid"] or "").strip() if ente_base else ""

        catalogo_rows = db.execute(
            """
            SELECT id, TRIM(COALESCE(nombre, '')) AS nombre
            FROM fuentes_financiamiento
            WHERE TRIM(COALESCE(nombre, '')) != ''
            ORDER BY nombre ASC
            """
        ).fetchall()
        catalogo_por_nombre = {row["nombre"].lower(): dict(row) for row in catalogo_rows}
        entes_fuentes_rows = db.execute(
            f"""
            SELECT DISTINCT ff.id, TRIM(COALESCE(ff.nombre, '')) AS nombre
            FROM entes_fuentes AS ef
            JOIN fuentes_financiamiento AS ff
              ON ef.fuente_id = ff.id
            WHERE TRIM(COALESCE(ef.ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ef.ente_id')} = ?
              {tipo_filter_sql.replace('o.', 'ef.')}
            ORDER BY ff.nombre ASC
            """,
            [ejercicio, ente_id_norm, *tipo_filter_params],
        ).fetchall()

        if ente_uid:
            registros_rows = db.execute(
                f"""
                SELECT DISTINCT ff.id, ff.nombre
                FROM registros AS r
                JOIN fuentes_financiamiento AS ff
                  ON r.fuente_id = ff.id
                JOIN entes_detalle AS ed
                  ON TRIM(COALESCE(r.ejercicio, '')) = TRIM(COALESCE(ed.ejercicio, ''))
                 AND {normalize_ente_id_sql('r.ente_id')} = {normalize_ente_id_sql('ed.ente_id')}
                WHERE TRIM(COALESCE(ed.ente_uid, '')) = ?
                ORDER BY ff.nombre ASC
                """,
                (ente_uid,),
            ).fetchall()
            observaciones_rows = db.execute(
                f"""
                SELECT DISTINCT TRIM(COALESCE(o.fuente_financiamiento, '')) AS nombre
                FROM observaciones AS o
                JOIN entes_detalle AS ed
                  ON TRIM(COALESCE(o.ejercicio, '')) = TRIM(COALESCE(ed.ejercicio, ''))
                 AND {normalize_ente_id_sql('o.ente_id')} = {normalize_ente_id_sql('ed.ente_id')}
                WHERE TRIM(COALESCE(ed.ente_uid, '')) = ?
                  AND TRIM(COALESCE(o.fuente_financiamiento, '')) != ''
                  {tipo_filter_sql}
                ORDER BY nombre ASC
                """,
                [ente_uid, *tipo_filter_params],
            ).fetchall()
        else:
            registros_rows = db.execute(
                f"""
                SELECT DISTINCT ff.id, ff.nombre
                FROM registros AS r
                JOIN fuentes_financiamiento AS ff
                  ON r.fuente_id = ff.id
                WHERE TRIM(COALESCE(r.ejercicio, '')) = ?
                  AND {normalize_ente_id_sql('r.ente_id')} = ?
                ORDER BY ff.nombre ASC
                """,
                (ejercicio, ente_id_norm),
            ).fetchall()
            observaciones_rows = db.execute(
                f"""
                SELECT DISTINCT TRIM(COALESCE(fuente_financiamiento, '')) AS nombre
                FROM observaciones
                WHERE TRIM(COALESCE(ejercicio, '')) = ?
                  AND {normalize_ente_id_sql('ente_id')} = ?
                  AND TRIM(COALESCE(fuente_financiamiento, '')) != ''
                  {tipo_filter_sql.replace("o.", "")}
                ORDER BY nombre ASC
                """,
                [ejercicio, ente_id_norm, *tipo_filter_params],
            ).fetchall()

        resultado = []
        seen_keys = set()
        for row in entes_fuentes_rows:
            nombre = (row["nombre"] or "").strip()
            if not nombre:
                continue
            key = nombre.lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)
            resultado.append({"id": str(row["id"]), "nombre": nombre})

        for row in registros_rows:
            nombre = (row["nombre"] or "").strip()
            if not nombre:
                continue
            key = nombre.lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)
            resultado.append({"id": str(row["id"]), "nombre": nombre})

        for row in observaciones_rows:
            nombre = (row["nombre"] or "").strip()
            if not nombre:
                continue
            key = nombre.lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)
            catalogo_row = catalogo_por_nombre.get(key)
            if catalogo_row:
                resultado.append({"id": str(catalogo_row["id"]), "nombre": catalogo_row["nombre"]})
            else:
                resultado.append({"id": f"__obs__:{nombre}", "nombre": nombre})

        if not resultado:
            for row in catalogo_rows:
                resultado.append({"id": str(row["id"]), "nombre": row["nombre"]})

        if not resultado:
            observaciones_global_rows = db.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(fuente_financiamiento, '')) AS nombre
                FROM observaciones
                WHERE TRIM(COALESCE(fuente_financiamiento, '')) != ''
                  {tipo_filter_sql}
                ORDER BY nombre ASC
                """.format(tipo_filter_sql=tipo_filter_sql.replace("o.", ""))
                if tipo_filter_sql
                else """
                SELECT DISTINCT TRIM(COALESCE(fuente_financiamiento, '')) AS nombre
                FROM observaciones
                WHERE TRIM(COALESCE(fuente_financiamiento, '')) != ''
                ORDER BY nombre ASC
                """,
                tipo_filter_params,
            ).fetchall()
            for row in observaciones_global_rows:
                nombre = (row["nombre"] or "").strip()
                if not nombre:
                    continue
                key = nombre.lower()
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                catalogo_row = catalogo_por_nombre.get(key)
                if catalogo_row:
                    resultado.append({"id": str(catalogo_row["id"]), "nombre": catalogo_row["nombre"]})
                else:
                    resultado.append({"id": f"__obs__:{nombre}", "nombre": nombre})

        return sorted(resultado, key=lambda item: item["nombre"].lower())

    def parse_pdp_amounts(raw_value: str) -> list[float]:
        tokens = re.split(r"(?:\r?\n|;|\|)+", (raw_value or "").strip())
        parsed: list[float] = []
        for token in tokens:
            clean = (token or "").strip()
            if not clean:
                continue
            clean = clean.replace("$", "").replace(",", "").strip()
            if clean in {"-", "—"}:
                parsed.append(0.0)
                continue
            try:
                amount = float(clean)
            except ValueError as exc:
                raise ValueError(f"Monto PDP inválido en detalle: '{token}'.") from exc
            if amount < 0:
                raise ValueError("Los montos PDP no pueden ser negativos.")
            parsed.append(amount)
        return parsed

    def normalize_convenio_text(value: str) -> str:
        clean = " ".join((value or "").split())
        clean = re.sub(r"\bCONVENI\s+O\b", "CONVENIO", clean, flags=re.IGNORECASE)
        return clean

    def normalize_observacion_modalidad(value: str) -> str:
        clean = " ".join((value or "").split()).lower()
        return "Convenio" if clean == "convenio" else "Fuente"

    def normalize_manual_si_no(value: str, *, default: str = "No") -> str:
        return financiamiento_service.normalize_si_no(value, default=default)

    def normalize_manual_origen_fuente(
        fuente_nombre: str,
        raw_origen: str = "",
        raw_remanente=None,
    ) -> str:
        if raw_origen:
            return infer_origen_fuente(fuente_nombre, str(raw_origen))
        if isinstance(raw_remanente, bool):
            return "Remanentes" if raw_remanente else "Del Ejercicio"
        remanente_key = normalize_text_key(str(raw_remanente or ""))
        if remanente_key in {"si", "sí", "s", "1", "true", "verdadero", "remanente", "remanentes"}:
            return "Remanentes"
        if remanente_key in {"no", "n", "0", "false", "falso", "del ejercicio", "ejercicio"}:
            return "Del Ejercicio"
        return infer_origen_fuente(fuente_nombre, "")

    def _convenio_match_key(value: str) -> str:
        clean = unicodedata.normalize("NFD", value or "")
        clean = "".join(ch for ch in clean if unicodedata.category(ch) != "Mn")
        clean = re.sub(r"[^a-z0-9]+", " ", clean.lower())
        return " ".join(clean.split())

    def resolve_convenio_ente_id(db, ejercicio: str, convenio_ente_nombre: str) -> str:
        target = _convenio_match_key(convenio_ente_nombre)
        if not target:
            return ""
        rows = db.execute(
            """
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
            """,
            (ejercicio,),
        ).fetchall()
        for row in rows:
            name_key = _convenio_match_key(row["ente_nombre"] or "")
            if target == name_key or target in name_key or name_key in target:
                return normalize_ente_id(row["ente_id"] or "")
        return ""

    def materialize_observaciones_from_manual(
        db,
        *,
        ejercicio: str,
        ente_id: str,
        ente_numero: str,
        ente_nombre: str,
        tipo_auditoria: str,
        fuente_nombre: str,
        ramo_33: str,
        ramo_28: str,
        origen_fuente: str,
        estado: str,
        periodo_cedula: str,
        periodo_titular: str,
        oficio: str,
        fecha_notificacion: str,
        cantidad_sa: int,
        cantidad_pdp: int,
        cantidad_pras: int,
        cantidad_pefcf: int,
        cantidad_r: int,
        monto_pdp_solventado: float,
        monto_pdp_pendiente: float,
        pdp_amounts: list[float],
        modalidad: str = "Fuente",
        convenio_nombre: str = "",
        convenio_ente_nombre: str = "",
        convenio_ente_id: str = "",
        pdp_details: list[dict] | None = None,
        solventacion_totales_by_anexo: dict[str, dict] | None = None,
        replace_scope: bool = False,
    ) -> None:
        fuente_nombre = normalize_fuente_financiamiento(fuente_nombre)
        ramo_33 = normalize_manual_si_no(ramo_33)
        ramo_28 = normalize_manual_si_no(ramo_28)
        origen_fuente = normalize_manual_origen_fuente(fuente_nombre, origen_fuente)
        modalidad = normalize_observacion_modalidad(modalidad)
        convenio_nombre = normalize_convenio_text(convenio_nombre) if modalidad == "Convenio" else ""
        convenio_ente_nombre = normalize_convenio_text(convenio_ente_nombre) if modalidad == "Convenio" else ""
        convenio_ente_id = normalize_ente_id(convenio_ente_id) if modalidad == "Convenio" else ""
        if replace_scope:
            db.execute(
                """
                DELETE FROM observaciones
                WHERE TRIM(COALESCE(ejercicio, '')) = TRIM(COALESCE(?, ''))
                  AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                  AND TRIM(COALESCE(tipo_auditoria, '')) = TRIM(COALESCE(?, ''))
                  AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
                  AND LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))
                  AND TRIM(COALESCE(modalidad, 'Fuente')) = TRIM(COALESCE(?, 'Fuente'))
                  AND LOWER(TRIM(COALESCE(convenio_nombre, ''))) = LOWER(TRIM(COALESCE(?, '')))
                  AND LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))
                """,
                (
                    ejercicio,
                    ente_id,
                    tipo_auditoria,
                    oficio,
                    fuente_nombre,
                    modalidad,
                    convenio_nombre,
                    periodo_cedula,
                ),
            )

        counts = {
            "SA": cantidad_sa,
            "PDP": cantidad_pdp,
            "PRAS": cantidad_pras,
            "PEFCF": cantidad_pefcf,
            "R": cantidad_r,
        }
        pdp_index = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for tipo_anexo, total in counts.items():
            for numero_observacion in range(1, total + 1):
                monto_emitido = None
                monto_solventado = None
                monto_pendiente = None
                pdp_concepto = None
                pdp_subconcepto = None
                fuente_row_nombre = fuente_nombre
                totales_tipo = (
                    (solventacion_totales_by_anexo or {}).get(tipo_anexo)
                    if isinstance(solventacion_totales_by_anexo, dict)
                    else None
                )
                estado_row = estado
                if isinstance(totales_tipo, dict):
                    solventadas_idx = {
                        int(item)
                        for item in (totales_tipo.get("solventadas_indices") or [])
                        if str(item).strip().isdigit()
                    }
                    pendientes_idx = {
                        int(item)
                        for item in (totales_tipo.get("pendientes_indices") or [])
                        if str(item).strip().isdigit()
                    }
                    emitidas_total = int(totales_tipo.get("emitidas", 0) or 0)
                    solventadas_total = int(totales_tipo.get("solventadas", 0) or 0)
                    pendientes_total = int(totales_tipo.get("pendientes", 0) or 0)
                    if numero_observacion in solventadas_idx:
                        estado_row = "Solventado"
                    elif numero_observacion in pendientes_idx:
                        estado_row = "Pendiente"
                    elif emitidas_total > 0 and solventadas_total == emitidas_total:
                        estado_row = "Solventado"
                    elif emitidas_total > 0 and pendientes_total == emitidas_total:
                        estado_row = "Pendiente"
                if tipo_anexo == "PDP":
                    detalle = pdp_details[pdp_index] if pdp_details and pdp_index < len(pdp_details) else {}
                    monto_emitido = detalle.get("monto")
                    if monto_emitido is None:
                        monto_emitido = pdp_amounts[pdp_index] if pdp_index < len(pdp_amounts) else 0.0
                    pdp_concepto = (detalle.get("concepto") or "").strip() or None
                    pdp_subconcepto = (detalle.get("subconcepto") or "").strip() or None
                    fuente_detalle = normalize_fuente_financiamiento(
                        " ".join(str(detalle.get("fuente") or "").split())
                    )
                    if fuente_detalle:
                        fuente_row_nombre = fuente_detalle
                    if totales_tipo and numero_observacion == 1:
                        monto_solventado = float(totales_tipo.get("solventado", 0.0) or 0.0)
                        monto_pendiente = float(totales_tipo.get("pendiente", 0.0) or 0.0)
                    elif numero_observacion == 1:
                        monto_solventado = monto_pdp_solventado
                        monto_pendiente = monto_pdp_pendiente
                    else:
                        monto_solventado = 0.0
                        monto_pendiente = 0.0
                    pdp_index += 1
                elif totales_tipo:
                    monto_emitido = float(totales_tipo.get("emitido", 0.0) or 0.0) if numero_observacion == 1 else 0.0
                    monto_solventado = float(totales_tipo.get("solventado", 0.0) or 0.0) if numero_observacion == 1 else 0.0
                    monto_pendiente = float(totales_tipo.get("pendiente", 0.0) or 0.0) if numero_observacion == 1 else 0.0
                db.execute(
                    """
                    INSERT INTO observaciones (
                        ejercicio,
                        auditoria,
                        ente_id,
                        ente_numero,
                        ente_numero_sort,
                        ente_nombre,
                        tipo_auditoria,
                        modalidad,
                        fuente_financiamiento,
                        convenio_nombre,
                        convenio_ente_nombre,
                        convenio_ente_id,
                        ramo_33,
                        ramo_28,
                        origen_fuente,
                        periodo,
                        periodo_cedula,
                        periodo_titular,
                        oficio,
                        fecha_notificacion,
                        tipo_anexo,
                        numero_observacion,
                        estatus,
                        estado,
                        monto,
                        monto_pdp_emitido,
                        monto_pdp_solventado,
                        monto_pdp_pendiente,
                        pdp_concepto_irregularidad,
                        pdp_subconcepto_irregularidad,
                        created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ejercicio,
                        tipo_auditoria,
                        ente_id,
                        ente_numero,
                        parse_ente_numero_sort(ente_numero),
                        ente_nombre,
                        tipo_auditoria,
                        modalidad,
                        fuente_row_nombre,
                        convenio_nombre,
                        convenio_ente_nombre,
                        convenio_ente_id,
                        ramo_33,
                        ramo_28,
                        origen_fuente,
                        periodo_cedula,
                        periodo_cedula,
                        periodo_titular or periodo_cedula,
                        oficio,
                        fecha_notificacion,
                        tipo_anexo,
                        numero_observacion,
                        estado_row,
                        estado_row,
                        monto_emitido,
                        monto_emitido,
                        monto_solventado,
                        monto_pendiente,
                        pdp_concepto if tipo_anexo == "PDP" else None,
                        pdp_subconcepto if tipo_anexo == "PDP" else None,
                        now,
                    ),
                )

    def serialize_manual_snapshot(payload) -> str:
        return json.dumps(payload, ensure_ascii=False)

    def parse_pdp_detalle_snapshot(raw_value: str) -> list[dict]:
        if not (raw_value or "").strip():
            return []
        try:
            payload = json.loads(raw_value)
        except json.JSONDecodeError:
            return []
        if not isinstance(payload, list):
            return []
        details = []
        for item in payload:
            data = item if isinstance(item, dict) else {}
            monto_raw = data.get("monto")
            details.append(
                {
                    "concepto": normalize_irregularidad_concepto(
                        data.get("concepto") or "",
                        allow_blank=True,
                    ),
                    "subconcepto": normalize_irregularidad_subconcepto(
                        data.get("subconcepto") or ""
                    ),
                    "fuente": normalize_fuente_financiamiento(
                        " ".join(str(data.get("fuente") or "").split())
                    ),
                    "monto": "" if monto_raw is None else str(monto_raw).strip(),
                }
            )
        return details

    def infer_manual_estado(fuente_nombre: str) -> str:
        return "Emitido"

    def build_fuente_detalle_snapshot(
        *,
        tipo_auditoria: str,
        fuente_nombre: str,
        cantidad_sa: int,
        cantidad_pdp: int,
        cantidad_pras: int,
        cantidad_pefcf: int,
        cantidad_r: int,
        modalidad: str = "Fuente",
        convenio_nombre: str = "",
        convenio_ente_nombre: str = "",
        convenio_ente_id: str = "",
        ramo_33: str = "No",
        ramo_28: str = "No",
        origen_fuente: str = "",
        solventacion_totales_by_anexo: dict[str, object] | None = None,
    ) -> dict[str, object]:
        modalidad = normalize_observacion_modalidad(modalidad)
        fuente_nombre = normalize_fuente_financiamiento(fuente_nombre)
        return {
            "tipo_auditoria": tipo_auditoria,
            "fuente_nombre": fuente_nombre,
            "modalidad": modalidad,
            "convenio_nombre": normalize_convenio_text(convenio_nombre) if modalidad == "Convenio" else "",
            "convenio_ente_nombre": normalize_convenio_text(convenio_ente_nombre) if modalidad == "Convenio" else "",
            "convenio_ente_id": normalize_ente_id(convenio_ente_id) if modalidad == "Convenio" else "",
            "ramo_33": normalize_manual_si_no(ramo_33),
            "ramo_28": normalize_manual_si_no(ramo_28),
            "origen_fuente": normalize_manual_origen_fuente(fuente_nombre, origen_fuente),
            "cantidad_sa": int(cantidad_sa),
            "cantidad_pdp": int(cantidad_pdp),
            "cantidad_pras": int(cantidad_pras),
            "cantidad_pefcf": int(cantidad_pefcf),
            "cantidad_r": int(cantidad_r),
            "solventacion_totales_by_anexo": solventacion_totales_by_anexo or {},
        }

    def count_observaciones_for_manual_scope(
        db,
        *,
        ejercicio: str,
        ente_id: str,
        tipo_auditoria: str,
        oficio: str,
        fuente_nombre: str,
        periodo_cedula: str,
        modalidad: str = "Fuente",
        convenio_nombre: str = "",
    ) -> int:
        modalidad = normalize_observacion_modalidad(modalidad)
        convenio_nombre = normalize_convenio_text(convenio_nombre) if modalidad == "Convenio" else ""
        row = db.execute(
            """
            SELECT COUNT(*) AS total
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = TRIM(COALESCE(?, ''))
              AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
              AND TRIM(COALESCE(tipo_auditoria, '')) = TRIM(COALESCE(?, ''))
              AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND TRIM(COALESCE(modalidad, 'Fuente')) = TRIM(COALESCE(?, 'Fuente'))
              AND LOWER(TRIM(COALESCE(convenio_nombre, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))
            """,
            (
                ejercicio,
                ente_id,
                tipo_auditoria,
                oficio,
                fuente_nombre,
                modalidad,
                convenio_nombre,
                periodo_cedula,
            ),
        ).fetchone()
        return int(row["total"] or 0) if row else 0

    def summarize_existing_manual_rows(rows) -> str:
        return ", ".join(
            f"{row['tipo_auditoria']} (ID {row['id']}, {row['created_at']})"
            for row in (rows or [])
        )

    def repair_existing_manual_rows(db, rows) -> dict[str, int]:
        repair_ids: list[int] = []
        for row in rows or []:
            try:
                item_id = int(row["id"])
            except (TypeError, ValueError, KeyError):
                continue
            if item_id > 0:
                repair_ids.append(item_id)
        if not repair_ids:
            return {"repaired": 0, "observaciones": 0}
        return repair_missing_observaciones_from_cargas(
            db,
            carga_ids=sorted(set(repair_ids)),
        )

    def repair_missing_observaciones_from_cargas(
        db,
        *,
        carga_ids: list[int] | None = None,
    ) -> dict[str, int]:
        where_sql = ""
        params: list[object] = []
        if carga_ids:
            placeholders = ",".join("?" for _ in carga_ids)
            where_sql = f"WHERE cm.id IN ({placeholders})"
            params.extend(int(item) for item in carga_ids)
        rows = db.execute(
            f"""
            SELECT
                cm.*,
                TRIM(COALESCE(cm.fuente_nombre, ff.nombre, '')) AS fuente_nombre_resolved,
                TRIM(COALESCE(ed.ente_numero, '')) AS ente_numero_resolved,
                TRIM(COALESCE(ed.ente_nombre, cm.ente_nombre, '')) AS ente_nombre_resolved
            FROM cargas_manuales AS cm
            LEFT JOIN fuentes_financiamiento AS ff
              ON ff.id = cm.fuente_id
            LEFT JOIN entes_detalle AS ed
              ON TRIM(COALESCE(ed.ejercicio, '')) = TRIM(COALESCE(cm.ejercicio, ''))
             AND TRIM(COALESCE(ed.ente_id, '')) = TRIM(COALESCE(cm.ente_id, ''))
            {where_sql}
            ORDER BY cm.id ASC
            """,
            params,
        ).fetchall()
        repaired = 0
        created_rows = 0
        for row in rows:
            if (row["asunto"] or "").strip() != "Notificación de Cédula de Resultados":
                continue
            expected_total = sum(
                int(row[column] or 0)
                for column in ("cantidad_sa", "cantidad_pdp", "cantidad_pras", "cantidad_pefcf", "cantidad_r")
            )
            if expected_total <= 0:
                continue
            fuente_nombre = (row["fuente_nombre_resolved"] or "").strip()
            if not fuente_nombre:
                continue
            existing_total = count_observaciones_for_manual_scope(
                db,
                ejercicio=(row["ejercicio"] or "").strip(),
                ente_id=normalize_ente_id(row["ente_id"] or ""),
                tipo_auditoria=(row["tipo_auditoria"] or "").strip(),
                oficio=(row["numero_oficio"] or "").strip(),
                fuente_nombre=fuente_nombre,
                periodo_cedula=" ".join((row["periodo"] or "").split()),
                modalidad=(row["modalidad"] or "Fuente").strip(),
                convenio_nombre=(row["convenio_nombre"] or "").strip(),
            )
            if existing_total == expected_total:
                continue
            pdp_details = parse_pdp_detalle_snapshot(row["pdp_detalle_json"] or "")
            pdp_amounts = [
                float(str(item.get("monto") or "0").replace("$", "").replace(",", "").strip() or 0.0)
                for item in pdp_details
            ]
            cantidad_pdp = int(row["cantidad_pdp"] or 0)
            if cantidad_pdp > 0 and not pdp_amounts:
                monto_total = float(row["monto_pdp_emitido"] or 0.0)
                pdp_amounts = [monto_total] + [0.0] * max(0, cantidad_pdp - 1)
            materialize_observaciones_from_manual(
                db,
                ejercicio=(row["ejercicio"] or "").strip(),
                ente_id=normalize_ente_id(row["ente_id"] or ""),
                ente_numero=(row["ente_numero_resolved"] or "").strip(),
                ente_nombre=(row["ente_nombre_resolved"] or "").strip(),
                tipo_auditoria=(row["tipo_auditoria"] or "").strip(),
                fuente_nombre=fuente_nombre,
                ramo_33=(row["ramo_33"] or "").strip() or "No",
                ramo_28=(row["ramo_28"] or "").strip() or "No",
                origen_fuente=(row["origen_fuente"] or "").strip() or infer_origen_fuente(fuente_nombre, ""),
                estado=(row["estado"] or "").strip() or infer_manual_estado(fuente_nombre),
                periodo_cedula=" ".join((row["periodo"] or "").split()),
                periodo_titular=" ".join((row["periodo_titular"] or "").split()),
                oficio=(row["numero_oficio"] or "").strip(),
                fecha_notificacion=(row["fecha_notificacion"] or "").strip(),
                cantidad_sa=int(row["cantidad_sa"] or 0),
                cantidad_pdp=cantidad_pdp,
                cantidad_pras=int(row["cantidad_pras"] or 0),
                cantidad_pefcf=int(row["cantidad_pefcf"] or 0),
                cantidad_r=int(row["cantidad_r"] or 0),
                monto_pdp_solventado=float(row["monto_pdp_solventado"] or 0.0),
                monto_pdp_pendiente=float(row["monto_pdp_pendiente"] or 0.0),
                pdp_amounts=pdp_amounts,
                modalidad=(row["modalidad"] or "Fuente").strip(),
                convenio_nombre=(row["convenio_nombre"] or "").strip(),
                convenio_ente_nombre=(row["convenio_ente_nombre"] or "").strip(),
                convenio_ente_id=(row["convenio_ente_id"] or "").strip(),
                pdp_details=pdp_details,
                solventacion_totales_by_anexo={},
                replace_scope=True,
            )
            repaired += 1
            created_rows += expected_total
        return {"repaired": repaired, "observaciones": created_rows}

    def parse_manual_pdp_details(raw_value: str, cantidad_pdp: int) -> list[dict]:
        if cantidad_pdp <= 0:
            return []
        if not (raw_value or "").strip():
            raise ValueError(
                "Debes capturar concepto y monto en todas las observaciones PDP."
            )
        try:
            payload = json.loads(raw_value)
        except json.JSONDecodeError as exc:
            raise ValueError("El detalle PDP tiene un formato inválido.") from exc
        if not isinstance(payload, list):
            raise ValueError("El detalle PDP tiene un formato inválido.")
        if len(payload) != cantidad_pdp:
            raise ValueError(
                "Debes capturar concepto y monto para cada observación PDP."
            )

        details: list[dict] = []
        for item in payload[:cantidad_pdp]:
            data = item if isinstance(item, dict) else {}
            concepto = normalize_irregularidad_concepto(
                data.get("concepto") or "",
                strict=True,
            )
            subconcepto = normalize_irregularidad_subconcepto(data.get("subconcepto") or "")
            fuente = " ".join(str(data.get("fuente") or "").split())
            monto_field = data.get("monto")
            monto_raw = "" if monto_field is None else str(monto_field).strip()
            monto_val = None
            if monto_raw:
                monto_clean = monto_raw.replace("$", "").replace(",", "").strip()
                try:
                    monto_val = float(monto_clean)
                except ValueError as exc:
                    raise ValueError("Monto PDP inválido en el detalle de observaciones.") from exc
                if monto_val < 0:
                    raise ValueError("Los montos PDP no pueden ser negativos.")
            if not concepto:
                raise ValueError("Debes capturar concepto en todas las observaciones PDP.")
            if monto_val is None:
                raise ValueError("Debes capturar monto en todas las observaciones PDP.")
            details.append(
                {
                    "concepto": concepto,
                    "subconcepto": subconcepto,
                    "monto": monto_val,
                    "fuente": normalize_fuente_financiamiento(fuente),
                }
            )

        return details

    def parse_manual_fuentes_detalle(raw_value: str) -> list[dict]:
        raw = (raw_value or "").strip()
        if not raw:
            return []
        try:
            payload = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise ValueError("El JSON del oficio no es válido.") from exc
        if not isinstance(payload, list):
            raise ValueError("El JSON del oficio debe contener una lista de fuentes.")
        rows: list[dict] = []
        for index, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                raise ValueError(f"La fuente {index} no tiene un formato válido.")
            fuente_nombre = normalize_fuente_financiamiento(
                " ".join(str(item.get("fuente_nombre") or "").split())
            )
            tipo_auditoria = " ".join(str(item.get("tipo_auditoria") or "").split())
            if not fuente_nombre:
                raise ValueError(f"La fuente {index} no tiene nombre.")
            if tipo_auditoria not in {"Financiera", "Obra Pública"}:
                tipo_auditoria = "Financiera"
            modalidad = normalize_observacion_modalidad(str(item.get("modalidad") or "Fuente"))
            convenio_nombre = normalize_convenio_text(str(item.get("convenio_nombre") or ""))
            convenio_ente_nombre = normalize_convenio_text(str(item.get("convenio_ente_nombre") or ""))
            convenio_ente_id = normalize_ente_id(str(item.get("convenio_ente_id") or ""))
            ramo_33 = normalize_manual_si_no(str(item.get("ramo_33") or "No"))
            ramo_28 = normalize_manual_si_no(str(item.get("ramo_28") or "No"))
            raw_remanente = (
                item["es_remanente"]
                if "es_remanente" in item
                else item["remanente"]
                if "remanente" in item
                else None
            )
            origen_fuente = normalize_manual_origen_fuente(
                fuente_nombre,
                str(item.get("origen_fuente") or ""),
                raw_remanente,
            )
            if modalidad != "Convenio":
                convenio_nombre = ""
                convenio_ente_nombre = ""
                convenio_ente_id = ""
            elif not convenio_nombre:
                raise ValueError(f"La fuente {index} marcada como convenio no tiene nombre de convenio.")
            cantidad_sa = parse_non_negative_int(str(item.get("cantidad_sa", "0")), "Cantidad SA")
            cantidad_pdp = parse_non_negative_int(str(item.get("cantidad_pdp", "0")), "Cantidad PDP")
            cantidad_pras = parse_non_negative_int(str(item.get("cantidad_pras", "0")), "Cantidad PRAS")
            cantidad_pefcf = parse_non_negative_int(str(item.get("cantidad_pefcf", "0")), "Cantidad PEFCF")
            cantidad_r = parse_non_negative_int(str(item.get("cantidad_r", "0")), "Cantidad R")
            periodo = " ".join(str(item.get("periodo") or "").split())
            if not periodo:
                raise ValueError(f"La fuente {index} no tiene periodo.")
            if (cantidad_sa + cantidad_pdp + cantidad_pras + cantidad_pefcf + cantidad_r) <= 0:
                raise ValueError(f"La fuente {index} no tiene observaciones capturadas.")
            solventacion_totales_by_anexo = item.get("solventacion_totales_by_anexo")
            if solventacion_totales_by_anexo is None:
                solventacion_totales_by_anexo = {}
            if not isinstance(solventacion_totales_by_anexo, dict):
                raise ValueError(f"La fuente {index} tiene totales de solventación inválidos.")
            rows.append(
                {
                    "fuente_nombre": fuente_nombre,
                    "tipo_auditoria": tipo_auditoria,
                    "modalidad": modalidad,
                    "convenio_nombre": convenio_nombre,
                    "convenio_ente_nombre": convenio_ente_nombre,
                    "convenio_ente_id": convenio_ente_id,
                    "ramo_33": ramo_33,
                    "ramo_28": ramo_28,
                    "origen_fuente": origen_fuente,
                    "periodo": periodo,
                    "cantidad_sa": cantidad_sa,
                    "cantidad_pdp": cantidad_pdp,
                    "cantidad_pras": cantidad_pras,
                    "cantidad_pefcf": cantidad_pefcf,
                    "cantidad_r": cantidad_r,
                    "solventacion_totales_by_anexo": solventacion_totales_by_anexo,
                }
            )
        return rows

    SOLVENTACION_ASUNTO = "Resultados de Solventación"
    SOLVENTACION_SIGLA_PAREN_RE = re.compile(r"\(([A-Z0-9.\-]+)\)")
    SOLVENTACION_FILENAME_RE = re.compile(
        r"^(?P<ente_id>[\d.]+)\.-\s+(?P<sigla>.+?)_OFS_(?P<numero>\d{4})_(?P<anio>\d{4})(?:_(?P<suffix>.+?))?\.(?P<ext>pdf|docx)$",
        re.IGNORECASE,
    )

    def _normalize_solventacion_sigla(value: str) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        normalized = unicodedata.normalize("NFKD", raw)
        normalized = "".join(char for char in normalized if not unicodedata.combining(char))
        normalized = re.sub(r"[^A-Za-z0-9]+", "", normalized)
        return normalized.upper()

    def _extract_solventacion_sigla(value: str) -> str:
        raw = str(value or "").strip()
        if not raw:
            return ""
        match = SOLVENTACION_SIGLA_PAREN_RE.search(raw)
        if match:
            return _normalize_solventacion_sigla(match.group(1))
        return ""

    def _normalize_solventacion_suffix(raw_suffix: str) -> tuple[str, str]:
        clean = " ".join(str(raw_suffix or "").replace("_", " ").split())
        if not clean:
            return "", ""
        suffix_key = clean.lower()
        if suffix_key.startswith("convenio "):
            return "", clean
        if " convenio " in suffix_key:
            period_part, convenio_part = re.split(r"\s+convenio\s+", clean, maxsplit=1, flags=re.IGNORECASE)
            return period_part.strip(), f"Convenio {convenio_part.strip()}"
        return clean, ""

    def _parse_solventacion_filename(filename: str) -> dict:
        basename = Path(filename or "").name
        match = SOLVENTACION_FILENAME_RE.match(basename)
        if not match:
            return {
                "basename": basename,
                "ente_id": "",
                "sigla": "",
                "oficio": "",
                "periodo_hint": "",
                "convenio_hint": "",
                "extension": Path(basename).suffix.lower().lstrip("."),
                "filename_valid": False,
            }
        period_hint, convenio_hint = _normalize_solventacion_suffix(match.group("suffix") or "")
        return {
            "basename": basename,
            "ente_id": normalize_ente_id(match.group("ente_id") or ""),
            "sigla": " ".join((match.group("sigla") or "").replace("_", " ").split()),
            "oficio": f"OFS/{match.group('numero')}/{match.group('anio')}",
            "periodo_hint": period_hint,
            "convenio_hint": convenio_hint,
            "extension": (match.group("ext") or "").lower(),
            "filename_valid": True,
        }

    def _build_solventacion_import_payload(parsed: dict, *, entry_meta: dict | None = None) -> dict:
        rows: list[dict] = []
        totals = {"emitidas": 0, "solventadas": 0, "pendientes": 0}
        tipos_importados: list[str] = []
        for auditoria in parsed.get("auditorias") or []:
            tipo = " ".join(str(auditoria.get("tipo") or "").split()) or "Financiera"
            if tipo not in tipos_importados:
                tipos_importados.append(tipo)
            for fuente in auditoria.get("fuentes") or []:
                modalidad = normalize_observacion_modalidad(fuente.get("modalidad") or "Fuente")
                for registro in fuente.get("registros") or []:
                    solventacion = registro.get("solventacion") if isinstance(registro.get("solventacion"), dict) else {}
                    row = {
                        "tipo_auditoria": tipo,
                        "fuente_nombre": normalize_fuente_financiamiento(
                            " ".join(str(fuente.get("nombre") or "").split())
                        ),
                        "modalidad": modalidad,
                        "convenio_nombre": normalize_convenio_text(fuente.get("convenio_nombre") or ""),
                        "convenio_ente_nombre": normalize_convenio_text(fuente.get("convenio_ente_nombre") or ""),
                        "convenio_ente_id": normalize_ente_id(fuente.get("convenio_ente_id") or ""),
                        "periodo": " ".join(str(registro.get("periodo") or "").split()),
                        "cantidad_sa": len(registro.get("SA") or []),
                        "cantidad_pdp": len(registro.get("PDP") or []),
                        "cantidad_pras": len(registro.get("PRAS") or []),
                        "cantidad_pefcf": len(registro.get("PEFCF") or []),
                        "cantidad_r": len(registro.get("R") or []),
                        "solventacion_totales_by_anexo": {},
                    }
                    for anexo in ("SA", "PDP", "PRAS", "PEFCF", "R"):
                        item = (
                            solventacion.get(anexo)
                            if isinstance(solventacion, dict) and isinstance(solventacion.get(anexo), dict)
                            else {}
                        )
                        row["solventacion_totales_by_anexo"][anexo] = {
                            "emitidas": int(item.get("emitidas", len(registro.get(anexo) or [])) or 0),
                            "solventadas": int(item.get("solventadas", 0) or 0),
                            "pendientes": int(item.get("pendientes", 0) or 0),
                            "emitido": float(item.get("emitidas", len(registro.get(anexo) or [])) or 0),
                            "solventado": float(item.get("solventadas", 0) or 0),
                            "pendiente": float(item.get("pendientes", 0) or 0),
                            "solventadas_indices": item.get("solventadas_indices") or [],
                            "pendientes_indices": item.get("pendientes_indices") or [],
                        }
                        totals["emitidas"] += int(item.get("emitidas", len(registro.get(anexo) or [])) or 0)
                        totals["solventadas"] += int(item.get("solventadas", 0) or 0)
                        totals["pendientes"] += int(item.get("pendientes", 0) or 0)
                    rows.append(row)
        tipo_auditoria = tipos_importados[0] if len(tipos_importados) == 1 else ""
        return {
            "ok": True,
            "mode": "solventacion",
            "asunto": SOLVENTACION_ASUNTO,
            "oficio": parsed.get("oficio") or (entry_meta or {}).get("oficio") or "",
            "fecha": parsed.get("fecha") or "",
            "ejercicio": parsed.get("ejercicio") or "",
            "periodo": parsed.get("periodo") or "",
            "oficio_base": parsed.get("oficio_base") or "",
            "destinatario": parsed.get("destinatario") or "",
            "tipo_auditoria": tipo_auditoria,
            "rows": rows,
            "totals": totals,
            "catalog_entry": entry_meta or {},
        }

    def count_existing_observaciones_scope(
        db,
        *,
        ejercicio: str,
        ente_id: str,
        tipo_auditoria: str,
        fuente_nombre: str,
        periodo: str,
        oficio: str,
    ) -> tuple[int, dict[str, int]]:
        rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                COUNT(*) AS total
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
              AND TRIM(COALESCE(tipo_auditoria, '')) = ?
              AND LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
            GROUP BY TRIM(COALESCE(tipo_anexo, ''))
            """,
            (ejercicio, ente_id, tipo_auditoria, fuente_nombre, periodo, oficio),
        ).fetchall()
        counts: dict[str, int] = {}
        total = 0
        for row in rows:
            tipo = (row["tipo_anexo"] or "").strip().upper()
            qty = int(row["total"] or 0)
            if not tipo or qty <= 0:
                continue
            counts[tipo] = qty
            total += qty
        return total, counts

    def count_existing_observaciones_by_clave(
        db,
        *,
        ejercicio: str,
        ente_id: str,
        periodo: str,
        oficio: str,
    ) -> tuple[int, dict[str, int]]:
        rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                COUNT(*) AS total
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
              AND LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))
              AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
            GROUP BY TRIM(COALESCE(tipo_anexo, ''))
            """,
            (ejercicio, ente_id, periodo, oficio),
        ).fetchall()
        counts: dict[str, int] = {}
        total = 0
        for row in rows:
            tipo = (row["tipo_anexo"] or "").strip().upper()
            qty = int(row["total"] or 0)
            if not tipo or qty <= 0:
                continue
            counts[tipo] = qty
            total += qty
        return total, counts

    def build_observaciones_admin_scope(raw_scope) -> tuple[list[str], list[str], dict, int]:
        ejercicio = " ".join((raw_scope.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(raw_scope.get("ente_id", ""))
        tipo_auditoria = " ".join((raw_scope.get("tipo_auditoria") or "").split())
        fuente = " ".join((raw_scope.get("fuente") or "").split())
        periodo = " ".join((raw_scope.get("periodo") or "").split())
        oficio = " ".join((raw_scope.get("oficio") or "").split())
        estado_raw = " ".join((raw_scope.get("estado") or "").split())
        tipo_anexo = " ".join((raw_scope.get("tipo_anexo") or "").split()).upper()

        if tipo_anexo == "PEFCT":
            tipo_anexo = "PEFCF"

        where_clauses: list[str] = []
        params: list[str] = []
        extra_filters = 0

        if ejercicio:
            where_clauses.append("TRIM(COALESCE(ejercicio, '')) = ?")
            params.append(ejercicio)

        if ente_id:
            where_clauses.append(f"{normalize_ente_id_sql('ente_id')} = ?")
            params.append(ente_id)
            extra_filters += 1

        if tipo_auditoria:
            tipo_options = _tipo_auditoria_options(tipo_auditoria)
            if not tipo_options:
                tipo_options = [tipo_auditoria]
            placeholders = ", ".join(["?"] * len(tipo_options))
            where_clauses.append(
                f"TRIM(COALESCE(tipo_auditoria, '')) IN ({placeholders})"
            )
            params.extend(tipo_options)
            extra_filters += 1

        if fuente:
            where_clauses.append(
                "LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))"
            )
            params.append(fuente)
            extra_filters += 1

        if periodo:
            where_clauses.append(
                "LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))"
            )
            params.append(periodo)
            extra_filters += 1

        if oficio:
            where_clauses.append(
                "LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))"
            )
            params.append(oficio)
            extra_filters += 1

        if estado_raw:
            estado = _normalize_observacion_estado(estado_raw)
            if estado not in OBSERVACION_ESTADOS_VALIDOS:
                raise ValueError("Estado inválido para filtrar observaciones.")
            if estado == "Emitido":
                where_clauses.append(
                    "LOWER(TRIM(COALESCE(estado, ''))) IN ('emitido', 'e')"
                )
            elif estado == "Pendiente":
                where_clauses.append(
                    "LOWER(TRIM(COALESCE(estado, ''))) IN ('pendiente', 'p')"
                )
            else:
                where_clauses.append(
                    "LOWER(TRIM(COALESCE(estado, ''))) IN ('solventado', 's')"
                )
            extra_filters += 1

        if tipo_anexo:
            if tipo_anexo not in {"SA", "PDP", "PRAS", "PEFCF", "R"}:
                raise ValueError("Tipo de anexo inválido para filtrar observaciones.")
            where_clauses.append("UPPER(TRIM(COALESCE(tipo_anexo, ''))) = ?")
            params.append(tipo_anexo)
            extra_filters += 1

        return (
            where_clauses,
            params,
            {
                "ejercicio": ejercicio,
                "ente_id": ente_id,
                "tipo_auditoria": tipo_auditoria,
                "fuente": fuente,
                "periodo": periodo,
                "oficio": oficio,
                "estado": estado_raw,
                "tipo_anexo": tipo_anexo,
            },
            extra_filters,
        )

    def _normalize_mass_upload_header(value: str) -> str:
        key = _normalize_mass_upload_text_key(value)
        return re.sub(r"[^a-z0-9]+", "_", key).strip("_")

    def _normalize_mass_upload_text_key(value: str) -> str:
        clean = " ".join(str(value or "").strip().lower().split())
        if not clean:
            return ""
        clean = unicodedata.normalize("NFKD", clean)
        clean = "".join(char for char in clean if not unicodedata.combining(char))
        return re.sub(r"\s+", " ", clean).strip()

    def _resolve_mass_upload_header_map(fieldnames) -> dict[str, str]:
        resolved: dict[str, str] = {}
        duplicates: list[str] = []
        for original in fieldnames or []:
            target = MASS_UPLOAD_HEADER_ALIASES.get(_normalize_mass_upload_header(original or ""))
            if not target:
                continue
            if target in resolved:
                duplicates.append(MASS_UPLOAD_REQUIRED_HEADERS.get(target, target))
                continue
            resolved[target] = original
        if duplicates:
            repeated = ", ".join(sorted(set(duplicates)))
            raise ValueError(f"El archivo repite columnas requeridas: {repeated}.")
        missing = [
            label
            for key, label in MASS_UPLOAD_REQUIRED_HEADERS.items()
            if key not in resolved
        ]
        if missing:
            raise ValueError(
                "Faltan columnas requeridas en el archivo: " + ", ".join(missing) + "."
            )
        return resolved

    def _detect_mass_upload_delimiter(raw_text: str) -> str:
        sample = (raw_text or "").strip()
        if not sample:
            return ","
        first_line = sample.splitlines()[0] if sample.splitlines() else sample
        tab_count = first_line.count("\t")
        comma_count = first_line.count(",")
        if tab_count > comma_count:
            return "\t"
        try:
            dialect = csv.Sniffer().sniff(sample[:4096], delimiters=",;\t|")
            return dialect.delimiter or ","
        except csv.Error:
            return ","

    def _parse_optional_mass_upload_money(raw_value: str, *, line_number: int, label: str) -> float | None:
        clean = str(raw_value or "").strip()
        if not clean or clean in {"-", "—"}:
            return None
        clean = clean.replace("$", "").replace(",", "").strip()
        try:
            amount = float(clean)
        except ValueError as exc:
            raise ValueError(f"Línea {line_number}: {label} inválido.") from exc
        if amount < 0:
            raise ValueError(f"Línea {line_number}: {label} no puede ser negativo.")
        return amount

    def _parse_mass_upload_numero_observacion(raw_value: str, *, line_number: int) -> int:
        clean = str(raw_value or "").strip()
        if not clean:
            raise ValueError(f"Línea {line_number}: NO. OBSERVACIÓN requerido.")
        try:
            number = int(clean)
        except ValueError as exc:
            raise ValueError(f"Línea {line_number}: NO. OBSERVACIÓN inválido.") from exc
        if number <= 0:
            raise ValueError(f"Línea {line_number}: NO. OBSERVACIÓN debe ser mayor a 0.")
        return number

    def _normalize_mass_upload_tipo_anexo(value: str, *, line_number: int | None = None) -> str:
        clean = " ".join((value or "").split()).upper()
        if clean == "PEFCT":
            clean = "PEFCF"
        if clean not in {"SA", "PDP", "PRAS", "PEFCF", "R"}:
            if line_number is None:
                raise ValueError("Tipo de anexo inválido.")
            raise ValueError(f"Línea {line_number}: ANEXO inválido.")
        return clean

    def _normalize_mass_upload_estado(value: str, *, line_number: int | None = None) -> str:
        key = _normalize_mass_upload_text_key(value)
        if key in {"e", "emitido", "emitida"}:
            return "Emitido"
        if key in {"p", "pendiente"}:
            return "Pendiente"
        if key in {"s", "solventado", "solventada"}:
            return "Solventado"
        if key.startswith("solventad"):
            return "Solventado"
        if key.startswith("pendient"):
            return "Pendiente"
        if key.startswith("emitid"):
            return "Emitido"
        if line_number is None:
            raise ValueError("Estado inválido.")
        raise ValueError(f"Línea {line_number}: ESTATUS inválido.")

    def _normalize_mass_upload_ramo_33(value: str) -> str:
        key = _normalize_mass_upload_text_key(value)
        if not key or key in {"no", "n", "0", "false"}:
            return "No"
        return "Si"

    def _normalize_mass_upload_periodo_text(value: str) -> str:
        clean = " ".join((value or "").replace("—", "-").replace("–", "-").split())
        clean = re.sub(r"\s*-\s*", " al ", clean)
        clean = re.sub(r"\bde\b", "de", clean, flags=re.IGNORECASE)
        clean = re.sub(r"\bal\b", "al", clean, flags=re.IGNORECASE)
        clean = re.sub(r"\s+", " ", clean).strip()
        return clean

    def _mass_upload_periodo_key(ejercicio: str, value: str) -> str:
        clean = _normalize_mass_upload_periodo_text(value)
        if not clean:
            return ""
        fecha_inicio, fecha_fin = parse_periodo_cedula(ejercicio, clean)
        if fecha_inicio and fecha_fin:
            return f"{fecha_inicio}|{fecha_fin}"
        return _normalize_mass_upload_text_key(clean)

    def _normalize_mass_upload_concept(value: str, *, line_number: int | None = None) -> str:
        clean = " ".join((value or "").split())
        clean = re.sub(r"^\s*\d+\s*[-.)]?\s*", "", clean).strip()
        try:
            return normalize_irregularidad_concepto(clean, strict=True)
        except ValueError as exc:
            if line_number is None:
                raise
            raise ValueError(f"Línea {line_number}: {exc}") from exc

    def _read_mass_upload_file_rows(upload_file) -> tuple[list[dict], str]:
        if upload_file is None:
            raise ValueError("Selecciona un archivo CSV para analizar.")
        file_name = " ".join((getattr(upload_file, "filename", "") or "").split())
        if not file_name:
            raise ValueError("Selecciona un archivo CSV para analizar.")
        raw_bytes = upload_file.read()
        if not raw_bytes:
            raise ValueError("El archivo está vacío.")
        try:
            raw_text = raw_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raw_text = raw_bytes.decode("latin-1")
        delimiter = _detect_mass_upload_delimiter(raw_text)
        reader = csv.DictReader(StringIO(raw_text), delimiter=delimiter)
        header_map = _resolve_mass_upload_header_map(reader.fieldnames or [])

        rows: list[dict] = []
        errors: list[str] = []
        for line_number, raw_row in enumerate(reader, start=2):
            row_data = raw_row if isinstance(raw_row, dict) else {}
            if not any(str(value or "").strip() for value in row_data.values()):
                continue
            try:
                tipo_auditoria = normalize_tipo_auditoria(
                    row_data.get(header_map["tipo_auditoria"], "")
                )
                if tipo_auditoria not in {"Financiera", "Obra Pública"}:
                    raise ValueError(f"Línea {line_number}: TIPO DE AUDITORÍA inválido.")

                periodo = _normalize_mass_upload_periodo_text(
                    row_data.get(header_map["periodo"], "")
                )
                if not periodo:
                    raise ValueError(f"Línea {line_number}: PERIODO requerido.")

                fuente_financiamiento = normalize_fuente_financiamiento(
                    " ".join(
                        str(row_data.get(header_map["fuente_financiamiento"], "") or "").split()
                    )
                )
                if not fuente_financiamiento:
                    raise ValueError(
                        f"Línea {line_number}: FUENTE DE FINANCIAMIENTO requerida."
                    )

                tipo_anexo = _normalize_mass_upload_tipo_anexo(
                    row_data.get(header_map["tipo_anexo"], ""),
                    line_number=line_number,
                )
                numero_observacion = _parse_mass_upload_numero_observacion(
                    row_data.get(header_map["numero_observacion"], ""),
                    line_number=line_number,
                )
                estado = _normalize_mass_upload_estado(
                    row_data.get(header_map["estatus"], ""),
                    line_number=line_number,
                )
                concepto_irregularidad = _normalize_mass_upload_concept(
                    row_data.get(header_map["concepto_irregularidad"], ""),
                    line_number=line_number,
                )
                rows.append(
                    {
                        "line_number": line_number,
                        "ente_nombre_csv": " ".join(
                            str(row_data.get(header_map["ente_nombre_csv"], "") or "").split()
                        ),
                        "ente_nombre_key": _normalize_mass_upload_text_key(
                            row_data.get(header_map["ente_nombre_csv"], "")
                        ),
                        "ramo_33": _normalize_mass_upload_ramo_33(
                            row_data.get(header_map["asf"], "")
                        ),
                        "periodo": periodo,
                        "periodo_key": _mass_upload_periodo_key(
                            MASS_UPLOAD_EJERCICIO_FIJO,
                            periodo,
                        ),
                        "tipo_auditoria": tipo_auditoria,
                        "fuente_financiamiento": fuente_financiamiento,
                        "fuente_key": _normalize_mass_upload_text_key(fuente_financiamiento),
                        "tipo_anexo": tipo_anexo,
                        "numero_observacion": numero_observacion,
                        "concepto_irregularidad": concepto_irregularidad,
                        "concepto_key": _normalize_mass_upload_text_key(concepto_irregularidad),
                        "monto_observado": _parse_optional_mass_upload_money(
                            row_data.get(header_map["monto_observado"], ""),
                            line_number=line_number,
                            label="MONTO OBSERVADO",
                        ),
                        "monto_solventado": _parse_optional_mass_upload_money(
                            row_data.get(header_map["monto_solventado"], ""),
                            line_number=line_number,
                            label="MONTO SOLVENTADO",
                        ),
                        "monto_pendiente": _parse_optional_mass_upload_money(
                            row_data.get(header_map["monto_pendiente"], ""),
                            line_number=line_number,
                            label="MONTO PENDIENTE",
                        ),
                        "estado": estado,
                    }
                )
            except ValueError as exc:
                errors.append(str(exc))
                if len(errors) >= 25:
                    break

        if errors:
            raise ValueError(" ".join(errors))
        if not rows:
            raise ValueError("El archivo no contiene registros válidos.")
        return rows, file_name

    def _mass_upload_observacion_key(
        *,
        ejercicio: str,
        tipo_auditoria: str,
        fuente_financiamiento: str,
        periodo: str,
        tipo_anexo: str,
        numero_observacion: int,
    ) -> tuple[str, str, str, str, int]:
        return (
            normalize_tipo_auditoria(tipo_auditoria) or "Financiera",
            _normalize_mass_upload_text_key(fuente_financiamiento),
            _mass_upload_periodo_key(ejercicio, periodo),
            _normalize_mass_upload_tipo_anexo(tipo_anexo),
            int(numero_observacion or 0),
        )

    def _current_ramo_33_label(value: str) -> str:
        key = _normalize_mass_upload_text_key(value)
        if not key or key in {"no", "n", "0", "false"}:
            return "No"
        return "Si"

    def _build_mass_upload_preview(db, *, ejercicio: str, ente_id: str, csv_rows: list[dict]) -> dict[str, object]:
        ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
        if not ente_row:
            raise ValueError("El ente seleccionado no existe para el ejercicio 2025.")

        alias_keys = {
            _normalize_mass_upload_text_key(alias)
            for alias in get_ente_aliases_by_uid(
                db,
                ejercicio,
                ente_id,
                fallback_names=[ente_row["ente_nombre"] or ""],
            )
            if _normalize_mass_upload_text_key(alias)
        }

        existing_rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(fuente_financiamiento, '')) AS fuente_financiamiento,
                TRIM(COALESCE(ramo_33, '')) AS ramo_33,
                TRIM(COALESCE(periodo_cedula, '')) AS periodo_cedula,
                TRIM(COALESCE(oficio, '')) AS oficio,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                COALESCE(numero_observacion, 0) AS numero_observacion,
                TRIM(COALESCE(estado, '')) AS estado,
                COALESCE(monto_pdp_emitido, 0) AS monto_pdp_emitido,
                COALESCE(monto_pdp_solventado, 0) AS monto_pdp_solventado,
                COALESCE(monto_pdp_pendiente, 0) AS monto_pdp_pendiente,
                TRIM(COALESCE(pdp_concepto_irregularidad, '')) AS pdp_concepto_irregularidad
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
            ORDER BY id ASC
            """,
            (ejercicio, ente_id),
        ).fetchall()

        existing_by_key: dict[tuple[str, str, str, str, int], dict] = {}
        duplicate_db: list[str] = []
        for row in existing_rows:
            item = dict(row)
            key = _mass_upload_observacion_key(
                ejercicio=ejercicio,
                tipo_auditoria=item["tipo_auditoria"],
                fuente_financiamiento=item["fuente_financiamiento"],
                periodo=item["periodo_cedula"],
                tipo_anexo=item["tipo_anexo"],
                numero_observacion=int(item["numero_observacion"] or 0),
            )
            if key in existing_by_key:
                duplicate_db.append(
                    f"{item['tipo_auditoria']} / {item['fuente_financiamiento']} / "
                    f"{item['periodo_cedula']} / {item['tipo_anexo']} #{item['numero_observacion']}"
                )
                continue
            item["estado"] = _normalize_observacion_estado(item.get("estado", ""))
            existing_by_key[key] = item
        if duplicate_db:
            raise ValueError(
                "Ya existen observaciones duplicadas en el sistema para este ente. "
                "Revisa primero: " + "; ".join(duplicate_db[:10]) + "."
            )

        seen_csv_keys: dict[tuple[str, str, str, str, int], int] = {}
        duplicate_csv: list[str] = []
        changes: list[dict] = []
        unchanged: list[dict] = []
        projection_rows: list[dict] = []
        unmatched: list[dict] = []
        skipped_other_ente: list[dict] = []
        warning_rows: list[str] = []

        summary = {
            "total_rows": len(csv_rows),
            "matched_rows": 0,
            "changed_rows": 0,
            "unchanged_rows": 0,
            "unmatched_rows": 0,
            "skipped_other_ente_rows": 0,
            "estado_changes": 0,
            "ramo_33_changes": 0,
            "pdp_amount_changes": 0,
            "warning_rows": 0,
        }

        for csv_row in csv_rows:
            csv_ente_key = csv_row.get("ente_nombre_key", "")
            if csv_ente_key and alias_keys and csv_ente_key not in alias_keys:
                skipped_other_ente.append(
                    {
                        "line_number": csv_row["line_number"],
                        "ente_nombre_csv": csv_row["ente_nombre_csv"] or "-",
                        "tipo_auditoria": csv_row["tipo_auditoria"],
                        "fuente_financiamiento": csv_row["fuente_financiamiento"],
                        "periodo": csv_row["periodo"],
                        "tipo_anexo": csv_row["tipo_anexo"],
                        "numero_observacion": csv_row["numero_observacion"],
                        "reason": "El nombre del ente no coincide con el ente seleccionado.",
                    }
                )
                continue

            key = _mass_upload_observacion_key(
                ejercicio=ejercicio,
                tipo_auditoria=csv_row["tipo_auditoria"],
                fuente_financiamiento=csv_row["fuente_financiamiento"],
                periodo=csv_row["periodo"],
                tipo_anexo=csv_row["tipo_anexo"],
                numero_observacion=csv_row["numero_observacion"],
            )
            if key in seen_csv_keys:
                duplicate_csv.append(
                    f"Líneas {seen_csv_keys[key]} y {csv_row['line_number']}: "
                    f"{csv_row['tipo_auditoria']} / {csv_row['fuente_financiamiento']} / "
                    f"{csv_row['periodo']} / {csv_row['tipo_anexo']} #{csv_row['numero_observacion']}"
                )
                continue
            seen_csv_keys[key] = csv_row["line_number"]

            current = existing_by_key.get(key)
            if not current:
                unmatched.append(
                    {
                        "line_number": csv_row["line_number"],
                        "tipo_auditoria": csv_row["tipo_auditoria"],
                        "fuente_financiamiento": csv_row["fuente_financiamiento"],
                        "periodo": csv_row["periodo"],
                        "tipo_anexo": csv_row["tipo_anexo"],
                        "numero_observacion": csv_row["numero_observacion"],
                        "estado": csv_row["estado"],
                        "reason": "No se encontró coincidencia en observaciones 2025 para el ente seleccionado.",
                    }
                )
                continue

            summary["matched_rows"] += 1
            is_pdp = (current["tipo_anexo"] or "").strip().upper() == "PDP"
            current_estado = _normalize_observacion_estado(current.get("estado", ""))
            current_ramo_33 = _current_ramo_33_label(current.get("ramo_33", ""))
            target_estado = csv_row["estado"]
            target_ramo_33 = csv_row["ramo_33"]
            current_emitido = float(current.get("monto_pdp_emitido") or 0.0)
            current_solventado = float(current.get("monto_pdp_solventado") or 0.0)
            current_pendiente = float(current.get("monto_pdp_pendiente") or 0.0)
            target_emitido = current_emitido
            target_solventado = current_solventado
            target_pendiente = current_pendiente
            warnings: list[str] = []

            if is_pdp:
                if csv_row["monto_observado"] is not None:
                    target_emitido = float(csv_row["monto_observado"])
                if target_estado == "Solventado":
                    if (
                        csv_row["monto_solventado"] is not None
                        and abs(float(csv_row["monto_solventado"]) - target_emitido) > 0.009
                    ):
                        raise ValueError(
                            f"Línea {csv_row['line_number']}: una observación PDP en Solventado "
                            "debe tener MONTO SOLVENTADO igual al MONTO OBSERVADO."
                        )
                    target_solventado = target_emitido
                elif csv_row["monto_solventado"] is not None:
                    target_solventado = float(csv_row["monto_solventado"])
                if target_solventado > target_emitido + 0.009:
                    raise ValueError(
                        f"Línea {csv_row['line_number']}: MONTO SOLVENTADO no puede ser mayor al MONTO OBSERVADO."
                    )
                target_pendiente = max(0.0, target_emitido - target_solventado)
                if (
                    csv_row["monto_pendiente"] is not None
                    and abs(float(csv_row["monto_pendiente"]) - target_pendiente) > 0.009
                ):
                    warnings.append("MONTO PENDIENTE del archivo no coincide con el cálculo emitido-solventado.")
                current_concept_key = _normalize_mass_upload_text_key(
                    current.get("pdp_concepto_irregularidad", "")
                )
                csv_concept_key = csv_row.get("concepto_key", "")
                if csv_concept_key and current_concept_key and csv_concept_key != current_concept_key:
                    warnings.append("El concepto PDP del archivo no coincide con el concepto actual.")
            elif any(
                value is not None
                for value in (
                    csv_row.get("monto_observado"),
                    csv_row.get("monto_solventado"),
                    csv_row.get("monto_pendiente"),
                )
            ):
                warnings.append("Los montos del archivo se ignoraron porque el anexo no es PDP.")

            changed_fields: list[str] = []
            if target_estado != current_estado:
                changed_fields.append("estado")
                summary["estado_changes"] += 1
            if target_ramo_33 != current_ramo_33:
                changed_fields.append("ramo_33")
                summary["ramo_33_changes"] += 1
            if is_pdp and (
                abs(target_emitido - current_emitido) > 0.009
                or abs(target_solventado - current_solventado) > 0.009
                or abs(target_pendiente - current_pendiente) > 0.009
            ):
                changed_fields.append("pdp_montos")
                summary["pdp_amount_changes"] += 1

            item = {
                "line_number": csv_row["line_number"],
                "id": int(current["id"]),
                "ente_numero": (current["ente_numero"] or "").strip(),
                "ente_nombre": (current["ente_nombre"] or "").strip(),
                "tipo_auditoria": current["tipo_auditoria"],
                "fuente_financiamiento": current["fuente_financiamiento"],
                "periodo_cedula": current["periodo_cedula"],
                "oficio": (current["oficio"] or "").strip(),
                "tipo_anexo": (current["tipo_anexo"] or "").strip().upper(),
                "numero_observacion": int(current["numero_observacion"] or 0),
                "estado_before": current_estado,
                "estado_after": target_estado,
                "ramo_33_before": current_ramo_33,
                "ramo_33_after": target_ramo_33,
                "monto_emitido_before": current_emitido,
                "monto_emitido_after": target_emitido,
                "monto_solventado_before": current_solventado,
                "monto_solventado_after": target_solventado,
                "monto_pendiente_before": current_pendiente,
                "monto_pendiente_after": target_pendiente,
                "warnings": warnings,
                "changed_fields": changed_fields,
            }
            projection_rows.append(item)
            if warnings:
                summary["warning_rows"] += 1
                if len(warning_rows) < 25:
                    warning_rows.append(
                        f"Línea {csv_row['line_number']} / ID {current['id']}: " + " ".join(warnings)
                    )
            if changed_fields:
                changes.append(item)
            else:
                unchanged.append(item)

        if duplicate_csv:
            raise ValueError(
                "El archivo contiene observaciones duplicadas. "
                + " ".join(duplicate_csv[:20])
            )

        summary["changed_rows"] = len(changes)
        summary["unchanged_rows"] = len(unchanged)
        summary["unmatched_rows"] = len(unmatched)
        summary["skipped_other_ente_rows"] = len(skipped_other_ente)

        return {
            "ente": {
                "ente_id": (ente_row["ente_id"] or "").strip(),
                "ente_numero": (ente_row["ente_numero"] or "").strip(),
                "ente_nombre": (ente_row["ente_nombre"] or "").strip(),
            },
            "summary": summary,
            "changes": changes,
            "projection_rows": projection_rows,
            "changes_preview": changes[:MASS_UPLOAD_PREVIEW_LIMIT],
            "changes_truncated": max(0, len(changes) - MASS_UPLOAD_PREVIEW_LIMIT),
            "unchanged_preview": unchanged[:MASS_UPLOAD_PREVIEW_LIMIT],
            "unchanged_truncated": max(0, len(unchanged) - MASS_UPLOAD_PREVIEW_LIMIT),
            "unmatched_preview": unmatched[:MASS_UPLOAD_PREVIEW_LIMIT],
            "unmatched_truncated": max(0, len(unmatched) - MASS_UPLOAD_PREVIEW_LIMIT),
            "skipped_preview": skipped_other_ente[:MASS_UPLOAD_PREVIEW_LIMIT],
            "skipped_truncated": max(0, len(skipped_other_ente) - MASS_UPLOAD_PREVIEW_LIMIT),
            "warnings_preview": warning_rows,
        }

    @app.get("/carga/entes")
    @gabo_required
    def carga_entes_por_ejercicio():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        if not ejercicio:
            return jsonify([])

        db = get_db()
        rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY {ente_numero_sort_sql('ente_numero')}, ente_numero, ente_nombre
            """,
            (ejercicio,),
        ).fetchall()
        return jsonify([dict(row) for row in rows])

    @app.get("/carga/fuentes-ente")
    @gabo_required
    def carga_fuentes_por_ente():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = (request.args.get("tipo_auditoria") or "").strip()
        if not ejercicio or not ente_id:
            return jsonify([])

        db = get_db()
        rows = fuentes_por_ente(db, ejercicio, ente_id, tipo_auditoria=tipo_auditoria)
        return jsonify([dict(row) for row in rows])

    @app.post("/carga/fuentes-ente/nueva")
    @gabo_required
    def carga_fuente_nueva_por_ente():
        user = get_current_user()
        payload = request.get_json(silent=True) if request.is_json else None
        source = payload if isinstance(payload, dict) else request.form
        ejercicio = " ".join((source.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(source.get("ente_id") or "")
        tipo_auditoria = normalize_tipo_auditoria(source.get("tipo_auditoria") or "")
        fuente_nombre = normalize_fuente_financiamiento(
            " ".join((source.get("nombre") or "").split())
        )

        if not ejercicio:
            return jsonify({"ok": False, "message": "Debes seleccionar el ejercicio."}), 400
        try:
            _ensure_editable_ejercicio(ejercicio, user=user)
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 403
        if not ente_id:
            return jsonify({"ok": False, "message": "Debes seleccionar el ente."}), 400
        if tipo_auditoria not in {"Financiera", "Obra Pública"}:
            return jsonify({"ok": False, "message": "Debes seleccionar el tipo de auditoría."}), 400
        if not fuente_nombre:
            return jsonify({"ok": False, "message": "Debes escribir la nueva fuente."}), 400

        db = get_db()
        ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
        if not ente_row:
            return jsonify({"ok": False, "message": "El ente seleccionado no existe para ese ejercicio."}), 404

        existing_fuente = db.execute(
            """
            SELECT id, TRIM(COALESCE(nombre, '')) AS nombre
            FROM fuentes_financiamiento
            WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?))
            LIMIT 1
            """,
            (fuente_nombre,),
        ).fetchone()
        if existing_fuente:
            fuente_id_value = str(int(existing_fuente["id"]))
            fuente_nombre_final = (existing_fuente["nombre"] or "").strip()
        else:
            fuente_id_value = f"__obs__:{fuente_nombre}"
            fuente_nombre_final = fuente_nombre

        return jsonify(
            {
                "ok": True,
                "id": fuente_id_value,
                "nombre": fuente_nombre_final,
                "ente_id": ente_id,
                "tipo_auditoria": tipo_auditoria,
                "deferred_save": True,
            }
        )

    def find_manual_cargas_repair_candidates(
        db,
        *,
        ejercicio: str,
        ente_id: str = "",
        tipo_auditoria: str = "",
    ) -> list[int]:
        repair_where = [
            "TRIM(COALESCE(cm.ejercicio, '')) = ?",
            "TRIM(COALESCE(cm.asunto, '')) = 'Notificación de Cédula de Resultados'",
            """(
                COALESCE(cm.cantidad_sa, 0)
                + COALESCE(cm.cantidad_pdp, 0)
                + COALESCE(cm.cantidad_pras, 0)
                + COALESCE(cm.cantidad_pefcf, 0)
                + COALESCE(cm.cantidad_r, 0)
            ) > 0""",
        ]
        repair_params: list[object] = [ejercicio]
        if ente_id:
            repair_where.append(f"{normalize_ente_id_sql('cm.ente_id')} = ?")
            repair_params.append(ente_id)
        if tipo_auditoria:
            tipo_options = _tipo_auditoria_options(tipo_auditoria)
            if not tipo_options:
                tipo_options = [tipo_auditoria]
            tipo_placeholders = ", ".join(["?"] * len(tipo_options))
            repair_where.append(
                f"TRIM(COALESCE(cm.tipo_auditoria, '')) IN ({tipo_placeholders})"
            )
            repair_params.extend(tipo_options)
        repair_rows = db.execute(
            f"""
            SELECT
                cm.id,
                TRIM(COALESCE(cm.ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(cm.ente_id, '')) AS ente_id,
                TRIM(COALESCE(cm.tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(cm.fuente_nombre, ff.nombre, '')) AS fuente_nombre,
                TRIM(COALESCE(cm.periodo, '')) AS periodo,
                TRIM(COALESCE(cm.numero_oficio, '')) AS numero_oficio,
                TRIM(COALESCE(cm.modalidad, 'Fuente')) AS modalidad,
                TRIM(COALESCE(cm.convenio_nombre, '')) AS convenio_nombre
            FROM cargas_manuales AS cm
            LEFT JOIN fuentes_financiamiento AS ff ON ff.id = cm.fuente_id
            WHERE {" AND ".join(repair_where)}
            ORDER BY cm.id ASC
            """,
            repair_params,
        ).fetchall()
        missing_ids: list[int] = []
        for row in repair_rows:
            existing_total = count_observaciones_for_manual_scope(
                db,
                ejercicio=row["ejercicio"] or ejercicio,
                ente_id=normalize_ente_id(row["ente_id"] or ""),
                tipo_auditoria=row["tipo_auditoria"] or "",
                fuente_nombre=normalize_fuente_financiamiento(row["fuente_nombre"] or ""),
                periodo_cedula=row["periodo"] or "",
                oficio=row["numero_oficio"] or "",
                modalidad=row["modalidad"] or "Fuente",
                convenio_nombre=row["convenio_nombre"] or "",
            )
            if existing_total <= 0:
                missing_ids.append(int(row["id"]))
        return missing_ids

    def _empty_oficios_totals() -> dict:
        return {
            "sa": 0,
            "pdp": 0,
            "pras": 0,
            "r": 0,
            "pefcf": 0,
            "total": 0,
            "monto_emitido": 0.0,
            "monto_solventado": 0.0,
            "monto_pendiente": 0.0,
        }

    def _query_carga_oficios_resumen(
        *,
        ejercicio: str,
        ente_id: str = "",
        tipo_auditoria: str = "",
        limit: int | None = 500,
    ) -> tuple[list[dict], dict]:
        params: list[str] = [ejercicio]
        where_extra: list[str] = []
        if ente_id:
            where_extra.append(f"{normalize_ente_id_sql('ente_id')} = ?")
            params.append(ente_id)
        if tipo_auditoria:
            tipo_options = _tipo_auditoria_options(tipo_auditoria)
            if not tipo_options:
                tipo_options = [tipo_auditoria]
            tipo_placeholders = ", ".join(["?"] * len(tipo_options))
            where_extra.append(
                f"TRIM(COALESCE(tipo_auditoria, '')) IN ({tipo_placeholders})"
            )
            params.extend(tipo_options)

        where_sql = ""
        if where_extra:
            where_sql = " AND " + " AND ".join(where_extra)

        db = get_db()
        limit_sql = ""
        if limit is not None and int(limit) > 0:
            limit_sql = f"LIMIT {int(limit)}"
        rows = db.execute(
            f"""
            SELECT
                {normalize_ente_id_sql('ente_id')} AS ente_id,
                MIN(COALESCE(ente_numero_sort, 0)) AS ente_numero_sort,
                MIN(TRIM(COALESCE(ente_numero, ''))) AS ente_numero,
                MIN(TRIM(COALESCE(ente_nombre, ''))) AS ente_nombre,
                MIN(TRIM(COALESCE(oficio, ''))) AS oficio,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipos_auditoria,
                MIN(NULLIF(TRIM(COALESCE(fecha_notificacion, '')), '')) AS fecha_notificacion,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'SA' THEN 1 ELSE 0 END) AS sa,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP' THEN 1 ELSE 0 END) AS pdp,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PRAS' THEN 1 ELSE 0 END) AS pras,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'R' THEN 1 ELSE 0 END) AS r,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PEFCF' THEN 1 ELSE 0 END) AS pefcf,
                COUNT(*) AS total,
                SUM(CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN COALESCE(monto_pdp_emitido, 0)
                    ELSE 0
                END) AS monto_emitido,
                SUM(CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN CASE
                        WHEN LOWER(TRIM(COALESCE(estado, ''))) IN ('solventado', 'solventada', 's')
                        THEN COALESCE(monto_pdp_emitido, 0)
                        WHEN COALESCE(monto_pdp_solventado, 0) > COALESCE(monto_pdp_emitido, 0)
                        THEN COALESCE(monto_pdp_emitido, 0)
                        ELSE COALESCE(monto_pdp_solventado, 0)
                    END
                    ELSE 0
                END) AS monto_solventado,
                SUM(CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN MAX(
                        0,
                        COALESCE(monto_pdp_emitido, 0)
                        - CASE
                            WHEN LOWER(TRIM(COALESCE(estado, ''))) IN ('solventado', 'solventada', 's')
                            THEN COALESCE(monto_pdp_emitido, 0)
                            WHEN COALESCE(monto_pdp_solventado, 0) > COALESCE(monto_pdp_emitido, 0)
                            THEN COALESCE(monto_pdp_emitido, 0)
                            ELSE COALESCE(monto_pdp_solventado, 0)
                        END
                    )
                    ELSE 0
                END) AS monto_pendiente
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(oficio, '')) != ''
              {where_sql}
            GROUP BY
              {normalize_ente_id_sql('ente_id')},
              LOWER(TRIM(COALESCE(oficio, ''))),
              TRIM(COALESCE(tipo_auditoria, ''))
            ORDER BY
              MIN(COALESCE(ente_numero_sort, 0)) ASC,
              MIN(TRIM(COALESCE(ente_numero, ''))) ASC,
              MIN(TRIM(COALESCE(ente_nombre, ''))) ASC,
              LOWER(TRIM(COALESCE(oficio, ''))) ASC,
              CASE TRIM(COALESCE(tipo_auditoria, ''))
                WHEN 'Financiera' THEN 0
                WHEN 'Obra Pública' THEN 1
                ELSE 2
              END ASC,
              TRIM(COALESCE(tipo_auditoria, '')) ASC
            {limit_sql}
            """,
            params,
        ).fetchall()

        payload_rows = []
        totals = _empty_oficios_totals()
        for row in rows:
            item = dict(row)
            tipos = [
                " ".join(tipo.split())
                for tipo in str(item.get("tipos_auditoria") or "").split(",")
                if " ".join(tipo.split())
            ]
            item["tipos_auditoria"] = ", ".join(tipos)
            for key in ("sa", "pdp", "pras", "r", "pefcf", "total"):
                item[key] = int(item.get(key) or 0)
                totals[key] += item[key]
            for key in ("monto_emitido", "monto_solventado", "monto_pendiente"):
                item[key] = float(item.get(key) or 0.0)
                totals[key] += item[key]
            payload_rows.append(item)
        return payload_rows, totals

    @app.get("/carga/oficios-resumen")
    @gabo_required
    def carga_oficios_resumen():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = (request.args.get("tipo_auditoria") or "").strip()
        if not ejercicio:
            return jsonify({"rows": [], "totals": {}, "repair_candidates": 0})

        db = get_db()
        repair_candidate_ids = find_manual_cargas_repair_candidates(
            db,
            ejercicio=ejercicio,
            ente_id=ente_id,
            tipo_auditoria=tipo_auditoria,
        )
        payload_rows, totals = _query_carga_oficios_resumen(
            ejercicio=ejercicio,
            ente_id=ente_id,
            tipo_auditoria=tipo_auditoria,
            limit=500,
        )
        return jsonify(
            {
                "rows": payload_rows,
                "totals": totals,
                "repair_candidates": len(repair_candidate_ids),
            }
        )

    @app.get("/carga/oficios-resumen/exportar")
    @gabo_required
    def carga_oficios_resumen_exportar():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = (request.args.get("tipo_auditoria") or "").strip()
        if not ejercicio:
            return jsonify({"ok": False, "error": "Selecciona un ejercicio para exportar."}), 400

        rows, totals = _query_carga_oficios_resumen(
            ejercicio=ejercicio,
            ente_id=ente_id,
            tipo_auditoria=tipo_auditoria,
            limit=None,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Oficios"
        headers = [
            "Oficio",
            "Ente",
            "Alcance",
            "Fecha",
            "SA",
            "PDP",
            "PRAS",
            "R",
            "PEFCF",
            "Total",
            "Monto emitido",
            "Monto solventado",
            "Monto pendiente",
        ]
        sheet.append(headers)

        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        header_fill = PatternFill("solid", fgColor="174C3A")
        total_fill = PatternFill("solid", fgColor="DCEBE3")
        subtotal_fill = PatternFill("solid", fgColor="F2F8F5")
        zebra_fill = PatternFill("solid", fgColor="F8FBF9")
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True, color="1E5139")
        normal_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        money_columns = {11, 12, 13}
        count_columns = {5, 6, 7, 8, 9, 10}

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = normal_alignment

        def ente_label(row: dict) -> str:
            return " - ".join(
                part for part in [
                    str(row.get("ente_numero") or "").strip(),
                    str(row.get("ente_nombre") or "").strip(),
                ] if part
            ) or "-"

        def append_data_row(row: dict) -> None:
            sheet.append([
                row.get("oficio") or "-",
                ente_label(row),
                row.get("tipos_auditoria") or "-",
                row.get("fecha_notificacion") or "-",
                int(row.get("sa") or 0),
                int(row.get("pdp") or 0),
                int(row.get("pras") or 0),
                int(row.get("r") or 0),
                int(row.get("pefcf") or 0),
                int(row.get("total") or 0),
                float(row.get("monto_emitido") or 0.0),
                float(row.get("monto_solventado") or 0.0),
                float(row.get("monto_pendiente") or 0.0),
            ])

        def add_totals(target: dict, row: dict) -> None:
            for key in ("sa", "pdp", "pras", "r", "pefcf", "total"):
                target[key] += int(row.get(key) or 0)
            for key in ("monto_emitido", "monto_solventado", "monto_pendiente"):
                target[key] += float(row.get(key) or 0.0)

        def append_total_row(label: str, target: dict, *, fill: PatternFill) -> None:
            sheet.append([
                label,
                "",
                "",
                "",
                int(target.get("sa") or 0),
                int(target.get("pdp") or 0),
                int(target.get("pras") or 0),
                int(target.get("r") or 0),
                int(target.get("pefcf") or 0),
                int(target.get("total") or 0),
                float(target.get("monto_emitido") or 0.0),
                float(target.get("monto_solventado") or 0.0),
                float(target.get("monto_pendiente") or 0.0),
            ])
            row_idx = sheet.max_row
            for cell in sheet[row_idx]:
                cell.fill = fill
                cell.font = bold_font

        groups: list[dict] = []
        groups_by_ente: dict[str, dict] = {}
        for item in rows:
            group_key = str(item.get("ente_id") or "").strip() or ente_label(item).lower()
            if group_key not in groups_by_ente:
                group = {"ente": ente_label(item), "rows": [], "totals": _empty_oficios_totals()}
                groups_by_ente[group_key] = group
                groups.append(group)
            group = groups_by_ente[group_key]
            group["rows"].append(item)
            add_totals(group["totals"], item)

        for group in groups:
            for item in group["rows"]:
                append_data_row(item)
            if len(groups) > 1:
                append_total_row(f"Subtotal {group['ente']}", group["totals"], fill=subtotal_fill)
        append_total_row("Total general", totals, fill=total_fill)

        for row_idx in range(2, sheet.max_row + 1):
            is_total_row = str(sheet.cell(row=row_idx, column=1).value or "").startswith(("Subtotal", "Total general"))
            for col_idx in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx in {1, 2, 3} else normal_alignment
                if col_idx in money_columns:
                    cell.number_format = '$#,##0.00'
                if col_idx in count_columns:
                    cell.number_format = '#,##0'
                if row_idx % 2 == 0 and not is_total_row:
                    cell.fill = zebra_fill

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:M{max(sheet.max_row, 1)}"
        widths = [18, 46, 22, 16, 9, 9, 9, 9, 9, 10, 18, 18, 18]
        for col_idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        scope_bits = [ejercicio]
        if ente_id:
            scope_bits.append(ente_id.replace(".", "_"))
        if tipo_auditoria:
            scope_bits.append(tipo_auditoria.lower().replace(" ", "_").replace("ú", "u"))
        filename = f"oficios_capturados_{'_'.join(scope_bits)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/carga/oficios-resumen/reparar")
    @gabo_required
    def carga_oficios_resumen_reparar():
        user = get_current_user()
        payload = request.get_json(silent=True) if request.is_json else None
        source = payload if isinstance(payload, dict) else request.form
        ejercicio = (source.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(source.get("ente_id", ""))
        tipo_auditoria = (source.get("tipo_auditoria") or "").strip()
        if not ejercicio:
            return jsonify({"ok": False, "message": "Debes seleccionar el ejercicio."}), 400
        try:
            _ensure_editable_ejercicio(ejercicio, user=user)
        except ValueError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 403

        db = get_db()
        carga_ids = find_manual_cargas_repair_candidates(
            db,
            ejercicio=ejercicio,
            ente_id=ente_id,
            tipo_auditoria=tipo_auditoria,
        )
        if not carga_ids:
            return jsonify({"ok": True, "repaired": 0, "observaciones": 0})

        repair_result = repair_missing_observaciones_from_cargas(db, carga_ids=carga_ids)
        if int(repair_result.get("repaired") or 0) > 0:
            db.commit()
        return jsonify(
            {
                "ok": True,
                "repaired": int(repair_result.get("repaired") or 0),
                "observaciones": int(repair_result.get("observaciones") or 0),
            }
        )

    @app.get("/carga/pdp-catalogo")
    @gabo_required
    def carga_pdp_catalogo():
        db = get_db()
        rows = db.execute(
            """
            SELECT DISTINCT
                TRIM(COALESCE(pdp_concepto_irregularidad, '')) AS concepto,
                TRIM(COALESCE(pdp_subconcepto_irregularidad, '')) AS subconcepto
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) IN ('2023', '2024')
              AND TRIM(COALESCE(tipo_anexo, '')) = 'PDP'
              AND TRIM(COALESCE(pdp_concepto_irregularidad, '')) != ''
            ORDER BY concepto ASC, subconcepto ASC
            """
        ).fetchall()

        conceptos: list[str] = []
        subconceptos_by_concept: dict[str, list[str]] = {}
        seen_conceptos = set()
        for row in rows:
            concepto = (row["concepto"] or "").strip()
            subconcepto = (row["subconcepto"] or "").strip()
            if not concepto:
                continue
            if concepto not in seen_conceptos:
                seen_conceptos.add(concepto)
                conceptos.append(concepto)
            if subconcepto:
                subconceptos_by_concept.setdefault(concepto, [])
                if subconcepto not in subconceptos_by_concept[concepto]:
                    subconceptos_by_concept[concepto].append(subconcepto)

        return jsonify(
            {
                "conceptos": conceptos,
                "subconceptos_by_concepto": subconceptos_by_concept,
            }
        )

    @app.post("/carga/pdp-detalle-excel/preview")
    @gabo_required
    def carga_pdp_detalle_excel_preview():
        upload = request.files.get("pdp_file") or request.files.get("file")
        try:
            preview = _parse_pdp_detail_excel_upload(upload)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        return jsonify(preview)

    @app.get("/carga/convenios-pdp-referencias")
    @gabo_required
    def carga_convenios_pdp_referencias():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        oficio = " ".join((request.args.get("oficio") or "").split())
        if not ejercicio:
            return jsonify({"ok": True, "rows": []})

        where = [
            "TRIM(COALESCE(ejercicio, '')) = ?",
            "TRIM(COALESCE(tipo_anexo, '')) = 'PDP'",
            "TRIM(COALESCE(modalidad, 'Fuente')) = 'Convenio'",
            "COALESCE(monto_pdp_emitido, 0) > 0",
            "TRIM(COALESCE(pdp_concepto_irregularidad, '')) != ''",
        ]
        params: list[object] = [ejercicio]
        if oficio:
            where.append("LOWER(TRIM(COALESCE(oficio, ''))) != LOWER(TRIM(COALESCE(?, '')))")
            params.append(oficio)

        db = get_db()
        rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(fuente_financiamiento, '')) AS fuente_financiamiento,
                TRIM(COALESCE(modalidad, 'Fuente')) AS modalidad,
                TRIM(COALESCE(convenio_nombre, '')) AS convenio_nombre,
                TRIM(COALESCE(convenio_ente_nombre, '')) AS convenio_ente_nombre,
                TRIM(COALESCE(convenio_ente_id, '')) AS convenio_ente_id,
                TRIM(COALESCE(periodo_cedula, '')) AS periodo_cedula,
                TRIM(COALESCE(oficio, '')) AS oficio,
                COALESCE(numero_observacion, 0) AS numero_observacion,
                COALESCE(monto_pdp_emitido, 0) AS monto_pdp_emitido,
                TRIM(COALESCE(pdp_concepto_irregularidad, '')) AS pdp_concepto_irregularidad,
                TRIM(COALESCE(pdp_subconcepto_irregularidad, '')) AS pdp_subconcepto_irregularidad,
                TRIM(COALESCE(created_at, '')) AS created_at
            FROM observaciones
            WHERE {" AND ".join(where)}
            ORDER BY id DESC
            LIMIT 5000
            """,
            params,
        ).fetchall()

        payload: list[dict[str, object]] = []
        for row in rows:
            monto = float(row["monto_pdp_emitido"] or 0)
            if monto <= 0:
                continue
            convenio_nombre = normalize_convenio_text(row["convenio_nombre"] or "")
            fuente_financiamiento = " ".join((row["fuente_financiamiento"] or "").split())
            periodo = " ".join((row["periodo_cedula"] or "").split())
            payload.append(
                {
                    "id": int(row["id"]),
                    "ejercicio": row["ejercicio"] or "",
                    "ente_id": normalize_ente_id(row["ente_id"] or ""),
                    "ente_nombre": row["ente_nombre"] or "",
                    "tipo_auditoria": row["tipo_auditoria"] or "Obra Pública",
                    "modalidad": "Convenio",
                    "fuente_nombre": fuente_financiamiento,
                    "fuente_financiamiento": fuente_financiamiento,
                    "convenio_nombre": convenio_nombre,
                    "convenio_ente_nombre": normalize_convenio_text(row["convenio_ente_nombre"] or ""),
                    "convenio_ente_id": normalize_ente_id(row["convenio_ente_id"] or ""),
                    "periodo": periodo,
                    "periodo_cedula": periodo,
                    "periodo_key": normalize_solventacion_periodo_key(ejercicio, periodo),
                    "oficio": row["oficio"] or "",
                    "numero_observacion": int(row["numero_observacion"] or 0),
                    "monto": monto,
                    "concepto": row["pdp_concepto_irregularidad"] or "",
                    "subconcepto": row["pdp_subconcepto_irregularidad"] or "",
                    "created_at": row["created_at"] or "",
                    "convenio_key": _convenio_match_key(convenio_nombre),
                    "fuente_key": _convenio_match_key(fuente_financiamiento),
                }
            )

        return jsonify({"ok": True, "rows": payload})

    @app.get("/carga/observaciones-cargadas")
    @gabo_required
    def carga_observaciones_cargadas():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = (request.args.get("tipo_auditoria") or "").strip()
        fuente = " ".join((request.args.get("fuente") or "").split())
        periodo = " ".join((request.args.get("periodo") or "").split())
        oficio = " ".join((request.args.get("oficio") or "").split())
        if not ejercicio or not ente_id:
            return jsonify([])

        db = get_db()
        params: list[str] = [ejercicio, ente_id]
        tipo_clause = ""
        if tipo_auditoria:
            tipo_options = _tipo_auditoria_options(tipo_auditoria)
            if not tipo_options:
                tipo_options = [tipo_auditoria]
            tipo_placeholders = ", ".join(["?"] * len(tipo_options))
            tipo_clause = f" AND TRIM(COALESCE(tipo_auditoria, '')) IN ({tipo_placeholders})"
            params.extend(tipo_options)
        where_extra: list[str] = []
        if fuente:
            where_extra.append("LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))")
            params.append(fuente)
        if periodo:
            where_extra.append("LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))")
            params.append(periodo)
        if oficio:
            where_extra.append("LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))")
            params.append(oficio)

        where_sql = ""
        if where_extra:
            where_sql = " AND " + " AND ".join(where_extra)

        rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                numero_observacion,
                TRIM(COALESCE(estado, '')) AS estado,
                COALESCE(reclasificada, 0) AS reclasificada,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(fuente_financiamiento, '')) AS fuente_financiamiento,
                TRIM(COALESCE(modalidad, 'Fuente')) AS modalidad,
                TRIM(COALESCE(convenio_nombre, '')) AS convenio_nombre,
                TRIM(COALESCE(convenio_ente_nombre, '')) AS convenio_ente_nombre,
                TRIM(COALESCE(convenio_ente_id, '')) AS convenio_ente_id,
                monto_pdp_emitido,
                monto_pdp_solventado,
                monto_pdp_pendiente,
                TRIM(COALESCE(pdp_concepto_irregularidad, '')) AS pdp_concepto_irregularidad,
                TRIM(COALESCE(pdp_subconcepto_irregularidad, '')) AS pdp_subconcepto_irregularidad
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
              {tipo_clause}
              {where_sql}
            ORDER BY
              CASE TRIM(COALESCE(tipo_anexo, ''))
                WHEN 'SA' THEN 1
                WHEN 'PDP' THEN 2
                WHEN 'PRAS' THEN 3
                WHEN 'PEFCF' THEN 4
                WHEN 'R' THEN 5
                ELSE 9
              END,
              COALESCE(numero_observacion, 0)
            LIMIT 1200
            """,
            params,
        ).fetchall()
        payload = []
        for row in rows:
            item = dict(row)
            item["estado"] = _normalize_observacion_estado(item.get("estado", ""))
            item["reclasificada"] = 1 if int(item.get("reclasificada") or 0) else 0
            payload.append(item)
        return jsonify(payload)

    @app.get("/carga/observaciones-claves")
    @gabo_required
    def carga_observaciones_claves():
        ejercicio = (request.args.get("ejercicio") or "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = (request.args.get("tipo_auditoria") or "").strip()
        fuente = " ".join((request.args.get("fuente") or "").split())
        periodo = " ".join((request.args.get("periodo") or "").split())
        if not ejercicio:
            return jsonify([])

        db = get_db()
        params: list[str] = [ejercicio]
        tipo_clause = ""
        if tipo_auditoria:
            tipo_options = _tipo_auditoria_options(tipo_auditoria)
            if not tipo_options:
                tipo_options = [tipo_auditoria]
            tipo_placeholders = ", ".join(["?"] * len(tipo_options))
            tipo_clause = f" AND TRIM(COALESCE(tipo_auditoria, '')) IN ({tipo_placeholders})"
            params.extend(tipo_options)

        where_extra: list[str] = []
        if ente_id:
            where_extra.append(f"{normalize_ente_id_sql('ente_id')} = ?")
            params.append(ente_id)
        if fuente:
            where_extra.append(
                "LOWER(TRIM(COALESCE(fuente_financiamiento, ''))) = LOWER(TRIM(COALESCE(?, '')))"
            )
            params.append(fuente)
        if periodo:
            where_extra.append(
                "LOWER(TRIM(COALESCE(periodo_cedula, ''))) = LOWER(TRIM(COALESCE(?, '')))"
            )
            params.append(periodo)

        where_sql = ""
        if where_extra:
            where_sql = " AND " + " AND ".join(where_extra)

        rows = db.execute(
            f"""
            SELECT DISTINCT
                TRIM(COALESCE(periodo_cedula, '')) AS periodo,
                TRIM(COALESCE(oficio, '')) AS oficio
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              {tipo_clause}
              {where_sql}
              AND TRIM(COALESCE(periodo_cedula, '')) != ''
              AND TRIM(COALESCE(oficio, '')) != ''
            ORDER BY
              LOWER(TRIM(COALESCE(periodo_cedula, ''))) ASC,
              LOWER(TRIM(COALESCE(oficio, ''))) ASC
            LIMIT 1200
            """,
            params,
        ).fetchall()

        payload = []
        for row in rows:
            periodo = (row["periodo"] or "").strip()
            oficio = (row["oficio"] or "").strip()
            if not periodo or not oficio:
                continue
            payload.append({"periodo": periodo, "oficio": oficio})
        payload.sort(
            key=lambda item: (
                parse_periodo_cedula(ejercicio, item["periodo"])[0] or "9999-12-31",
                parse_periodo_cedula(ejercicio, item["periodo"])[1] or "9999-12-31",
                item["periodo"].lower(),
                item["oficio"].lower(),
            )
        )
        return jsonify(payload)

    @app.post("/carga/observaciones-cargadas/<int:observacion_id>/actualizar")
    @gabo_required
    def carga_observacion_actualizar(observacion_id: int):
        payload = request.get_json(silent=True) or {}
        accion = " ".join(str(payload.get("accion", "") or "").lower().split())
        estado = _normalize_observacion_estado(payload.get("estado", ""))
        monto_emitido_raw = payload.get("monto_pdp_emitido", "")
        monto_solventado_raw = payload.get("monto_pdp_solventado", "")
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}

        db = get_db()
        current = db.execute(
            """
            SELECT
                id,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(oficio, '')) AS oficio
            FROM observaciones
            WHERE id = ?
            LIMIT 1
            """,
            (observacion_id,),
        ).fetchone()
        if not current:
            return jsonify({"ok": False, "error": "Observación no encontrada."}), 404

        tipo_anexo = (current["tipo_anexo"] or "").strip().upper()
        ejercicio = " ".join((current["ejercicio"] or "").split())
        try:
            _ensure_editable_ejercicio(ejercicio)
            _validate_observacion_matches_scope(current, raw_scope)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        if accion in {"reclasificar_pdp_pras", "reclasificar_pras", "pdp_a_pras"}:
            if tipo_anexo != "PDP":
                return jsonify(
                    {"ok": False, "error": "Solo se puede reclasificar observaciones PDP."},
                    400,
                )
            backup_path = _create_db_snapshot(f"observacion-{observacion_id}-actualizar")
            db.execute(
                """
                UPDATE observaciones
                SET tipo_anexo = 'PRAS',
                    estado = 'Pendiente',
                    reclasificada = 1,
                    pdp_concepto_irregularidad = '',
                    pdp_subconcepto_irregularidad = '',
                    monto_pdp_emitido = 0,
                    monto_pdp_solventado = 0,
                    monto_pdp_pendiente = 0,
                    monto = 0
                WHERE id = ?
                """,
                (observacion_id,),
            )
            db.commit()
            return jsonify(
                {
                    "ok": True,
                    "id": observacion_id,
                    "estado": "Pendiente",
                    "tipo_anexo": "PRAS",
                    "reclasificada": 1,
                    "accion": "reclasificar_pdp_pras",
                    "backup_path": backup_path,
                }
            )

        if estado not in OBSERVACION_ESTADOS_VALIDOS:
            return jsonify({"ok": False, "error": "Estado inválido."}), 400

        if tipo_anexo == "PDP":
            try:
                monto_emitido = parse_non_negative_float(str(monto_emitido_raw), "Monto PDP emitido")
                monto_solventado = parse_non_negative_float(str(monto_solventado_raw), "Monto PDP solventado")
            except ValueError:
                return jsonify({"ok": False, "error": "Montos PDP inválidos."}), 400
            if monto_solventado > monto_emitido:
                return jsonify(
                    {"ok": False, "error": "Monto solventado no puede ser mayor a emitido."},
                    400,
                )
            monto_pendiente = monto_emitido - monto_solventado
            backup_path = _create_db_snapshot(f"observacion-{observacion_id}-actualizar")
            db.execute(
                """
                UPDATE observaciones
                SET estado = ?,
                    monto_pdp_emitido = ?,
                    monto_pdp_solventado = ?,
                    monto_pdp_pendiente = ?,
                    monto = ?
                WHERE id = ?
                """,
                (
                    estado,
                    monto_emitido,
                    monto_solventado,
                    monto_pendiente,
                    monto_emitido,
                    observacion_id,
                ),
            )
        else:
            backup_path = _create_db_snapshot(f"observacion-{observacion_id}-actualizar")
            db.execute(
                """
                UPDATE observaciones
                SET estado = ?
                WHERE id = ?
                """,
                (estado, observacion_id),
            )
        db.commit()
        return jsonify(
            {
                "ok": True,
                "id": observacion_id,
                "estado": estado,
                "tipo_anexo": tipo_anexo,
                "backup_path": backup_path,
            }
        )

    @app.get("/carga/observaciones-admin")
    @gabo_required
    def carga_observaciones_admin():
        user = get_current_user()
        db = get_db()
        return_vista = (request.args.get("return_vista") or "").strip().lower()
        if return_vista not in {"manual", "titulares"}:
            return_vista = "manual"
        ejercicios_rows = db.execute(
            """
            SELECT DISTINCT TRIM(COALESCE(ejercicio, '')) AS ejercicio
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) != ''
            ORDER BY CAST(TRIM(COALESCE(ejercicio, '')) AS INTEGER) DESC, ejercicio DESC
            """
        ).fetchall()
        ejercicios = [row["ejercicio"] for row in ejercicios_rows if (row["ejercicio"] or "").strip()]
        if not ejercicios:
            ejercicios = [TITULAR_EJERCICIO_FIJO]
        initial_scope = {
            "ejercicio": " ".join((request.args.get("ejercicio") or "").split()),
            "ente_id": normalize_ente_id(request.args.get("ente_id", "")),
            "tipo_auditoria": normalize_tipo_auditoria(request.args.get("tipo_auditoria", "")),
            "fuente": " ".join((request.args.get("fuente") or "").split()),
            "periodo": " ".join((request.args.get("periodo") or "").split()),
            "oficio": " ".join((request.args.get("oficio") or "").split()),
            "estado": _normalize_observacion_estado(request.args.get("estado", "")),
            "tipo_anexo": " ".join((request.args.get("tipo_anexo") or "").split()).upper(),
        }
        if initial_scope["tipo_anexo"] == "PEFCT":
            initial_scope["tipo_anexo"] = "PEFCF"
        if initial_scope["tipo_anexo"] not in {"", "SA", "PDP", "PRAS", "PEFCF", "R"}:
            initial_scope["tipo_anexo"] = ""
        if initial_scope["estado"] not in {"", *OBSERVACION_ESTADOS_VALIDOS}:
            initial_scope["estado"] = ""
        locked_scope = None
        if all(
            [
                return_vista == "manual",
                initial_scope["ejercicio"],
                initial_scope["ente_id"],
                initial_scope["tipo_auditoria"],
                initial_scope["oficio"],
            ]
        ):
            ente_row = db.execute(
                f"""
                SELECT
                    TRIM(COALESCE(ente_id, '')) AS ente_id,
                    TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                    TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
                FROM entes_detalle
                WHERE TRIM(COALESCE(ejercicio, '')) = ?
                  AND {normalize_ente_id_sql('ente_id')} = ?
                LIMIT 1
                """,
                (initial_scope["ejercicio"], initial_scope["ente_id"]),
            ).fetchone()
            oficio_rows = db.execute(
                """
                SELECT DISTINCT TRIM(COALESCE(periodo_cedula, '')) AS periodo
                FROM observaciones
                WHERE TRIM(COALESCE(ejercicio, '')) = ?
                  AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
                """,
                (initial_scope["ejercicio"], initial_scope["oficio"]),
            ).fetchall()
            periodos = [
                (row["periodo"] or "").strip()
                for row in oficio_rows
                if (row["periodo"] or "").strip()
            ]
            periodos_unicos = []
            for periodo in periodos:
                if periodo not in periodos_unicos:
                    periodos_unicos.append(periodo)
            oficio_label = initial_scope["oficio"]
            if len(periodos_unicos) > 1:
                oficio_label = f"{oficio_label} · varios periodos"
            elif periodos_unicos:
                oficio_label = f"{oficio_label} · {periodos_unicos[0]}"
            locked_scope = {
                "active": True,
                "ejercicio": initial_scope["ejercicio"],
                "ente_id": initial_scope["ente_id"],
                "ente_label": (
                    f"{(ente_row['ente_numero'] or '').strip()} - {(ente_row['ente_nombre'] or '').strip()}".strip(" -")
                    if ente_row
                    else initial_scope["ente_id"]
                ),
                "tipo_auditoria": initial_scope["tipo_auditoria"],
                "oficio": initial_scope["oficio"],
                "oficio_label": oficio_label,
            }
        return render_template(
            "carga_observaciones_admin.html",
            user=user,
            ejercicios=ejercicios,
            ejercicio_default=initial_scope["ejercicio"] if initial_scope["ejercicio"] in ejercicios else ejercicios[0],
            return_vista=return_vista,
            initial_scope=initial_scope,
            locked_scope=locked_scope,
            read_only_ejercicios=sorted(_readonly_ejercicios_for_user(user)),
        )

    def _parse_solventacion_progressive_numbers(raw_value) -> list[int]:
        if isinstance(raw_value, list):
            tokens = [str(item or "").strip() for item in raw_value]
        else:
            tokens = re.split(r"[\s,;|]+", str(raw_value or "").strip())
        values: list[int] = []
        seen: set[int] = set()
        for token in tokens:
            token_clean = token.strip()
            if not token_clean:
                continue
            try:
                parsed = int(token_clean)
            except (TypeError, ValueError):
                continue
            if parsed <= 0 or parsed in seen:
                continue
            seen.add(parsed)
            values.append(parsed)
        values.sort()
        return values

    @app.get("/carga/herramientas")
    @gabo_required
    def carga_herramientas():
        user = get_current_user()
        db = get_db()
        ejercicios_rows = db.execute(
            """
            SELECT DISTINCT ejercicio
            FROM (
                SELECT TRIM(COALESCE(ejercicio, '')) AS ejercicio
                FROM entes_detalle
                UNION
                SELECT TRIM(COALESCE(ejercicio, '')) AS ejercicio
                FROM observaciones
                UNION
                SELECT TRIM(COALESCE(ejercicio, '')) AS ejercicio
                FROM cargas_manuales
            )
            WHERE ejercicio != ''
            ORDER BY CAST(ejercicio AS INTEGER) DESC, ejercicio DESC
            """
        ).fetchall()
        ejercicios = [
            row["ejercicio"]
            for row in ejercicios_rows
            if (row["ejercicio"] or "").strip()
            and row["ejercicio"] not in GABO_READONLY_EJERCICIOS
        ]
        if not ejercicios:
            ejercicios = [TITULAR_EJERCICIO_FIJO]
        requested_ejercicio = " ".join((request.args.get("ejercicio") or "").split())
        ejercicio_default = requested_ejercicio if requested_ejercicio in ejercicios else ejercicios[0]
        fuentes_admin = list_fuentes_financiamiento_admin(db, ejercicio_default)
        return render_template(
            "carga_herramientas.html",
            user=user,
            ejercicios=ejercicios,
            ejercicio_default=ejercicio_default,
            fuentes_admin=fuentes_admin,
            read_only_ejercicios=sorted(_readonly_ejercicios_for_user(user)),
        )

    @app.post("/carga/fuentes-financiamiento/clasificacion")
    @gabo_required
    def carga_fuente_financiamiento_clasificacion():
        user = get_current_user()
        db = get_db()
        payload = request.get_json(silent=True) or request.form
        try:
            ejercicio = " ".join(str(payload.get("ejercicio") or "").split())
            if not ejercicio:
                raise ValueError("Selecciona un ejercicio editable para actualizar la fuente.")
            _ensure_editable_ejercicio(ejercicio, user=user)

            fuente_id_raw = " ".join(str(payload.get("fuente_id") or "").split())
            fuente_id = None
            if fuente_id_raw:
                try:
                    fuente_id = int(fuente_id_raw)
                except ValueError as exc:
                    raise ValueError("La fuente seleccionada no es válida.") from exc
            result = financiamiento_service.update_fuente_clasificacion(
                db,
                fuente_id=fuente_id,
                fuente_nombre=" ".join(str(payload.get("fuente_nombre") or "").split()),
                normalizer=normalize_fuente_financiamiento,
                ejercicio=ejercicio,
                ramo_33=str(payload.get("ramo_33") or "No"),
                ramo_28=str(payload.get("ramo_28") or "No"),
                origen_fuente=str(payload.get("origen_fuente") or ""),
            )
            db.commit()
            return jsonify({"ok": True, **result})
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)

    @app.get("/carga/observaciones-admin/carga-masiva")
    @gabo_required
    def carga_observaciones_admin_carga_masiva():
        user = get_current_user()
        db = get_db()
        return_vista = (request.args.get("return_vista") or "").strip().lower()
        if return_vista not in {"manual", "titulares"}:
            return_vista = "manual"
        entes = _load_titular_entes(db, MASS_UPLOAD_EJERCICIO_FIJO)
        ente_default = normalize_ente_id(request.args.get("ente_id", ""))
        if ente_default and not any(
            normalize_ente_id(item.get("ente_id", "")) == ente_default for item in entes
        ):
            ente_default = ""
        return render_template(
            "carga_observaciones_masiva.html",
            user=user,
            entes=entes,
            ejercicio_fijo=MASS_UPLOAD_EJERCICIO_FIJO,
            ente_default=ente_default,
            return_vista=return_vista,
        )

    @app.post("/carga/observaciones-admin/carga-masiva/preview")
    @gabo_required
    def carga_observaciones_admin_carga_masiva_preview():
        ejercicio = MASS_UPLOAD_EJERCICIO_FIJO
        ente_id = normalize_ente_id(request.form.get("ente_id", ""))
        upload_file = request.files.get("csv_file")

        if not ente_id:
            return jsonify({"ok": False, "error": "Selecciona un ente antes de analizar el archivo."}), 400
        try:
            _ensure_editable_ejercicio(ejercicio)
            csv_rows, file_name = _read_mass_upload_file_rows(upload_file)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)

        db = get_db()
        try:
            preview = _build_mass_upload_preview(
                db,
                ejercicio=ejercicio,
                ente_id=ente_id,
                csv_rows=csv_rows,
            )
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

        return jsonify(
            {
                "ok": True,
                "file_name": file_name,
                "ejercicio": ejercicio,
                "can_apply": bool(preview["changes"]),
                **preview,
            }
        )

    @app.post("/carga/observaciones-admin/carga-masiva/aplicar")
    @gabo_required
    def carga_observaciones_admin_carga_masiva_aplicar():
        payload = request.get_json(silent=True) or {}
        ejercicio = MASS_UPLOAD_EJERCICIO_FIJO
        ente_id = normalize_ente_id(payload.get("ente_id", ""))
        raw_updates = payload.get("updates")

        if not ente_id:
            return jsonify({"ok": False, "error": "Selecciona un ente antes de aplicar cambios."}), 400
        if not isinstance(raw_updates, list) or not raw_updates:
            return jsonify({"ok": False, "error": "No se recibieron cambios para aplicar."}), 400
        try:
            _ensure_editable_ejercicio(ejercicio)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 403

        sanitized_updates: list[dict] = []
        seen_ids: set[int] = set()
        for raw in raw_updates:
            if not isinstance(raw, dict):
                continue
            try:
                obs_id = int(raw.get("id"))
            except (TypeError, ValueError):
                continue
            if obs_id <= 0 or obs_id in seen_ids:
                continue
            seen_ids.add(obs_id)
            sanitized_updates.append(
                {
                    "id": obs_id,
                    "estado": _normalize_mass_upload_estado(
                        raw.get("estado_after") or raw.get("estado") or ""
                    ),
                    "ramo_33": _normalize_mass_upload_ramo_33(
                        raw.get("ramo_33_after") or raw.get("ramo_33") or ""
                    ),
                    "monto_emitido": float(raw.get("monto_emitido_after", raw.get("monto_emitido", 0)) or 0),
                    "monto_solventado": float(
                        raw.get("monto_solventado_after", raw.get("monto_solventado", 0)) or 0
                    ),
                    "monto_pendiente": float(
                        raw.get("monto_pendiente_after", raw.get("monto_pendiente", 0)) or 0
                    ),
                }
            )
        if not sanitized_updates:
            return jsonify({"ok": False, "error": "La lista de cambios es inválida."}), 400

        db = get_db()
        ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
        if not ente_row:
            return jsonify({"ok": False, "error": "El ente seleccionado no existe para el ejercicio 2025."}), 400

        ids = [item["id"] for item in sanitized_updates]
        placeholders = ", ".join(["?"] * len(ids))
        found_rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo
            FROM observaciones
            WHERE id IN ({placeholders})
              AND TRIM(COALESCE(ejercicio, '')) = ?
              AND {normalize_ente_id_sql('ente_id')} = ?
            """,
            [*ids, ejercicio, ente_id],
        ).fetchall()
        if len(found_rows) != len(ids):
            return jsonify(
                {
                    "ok": False,
                    "error": "Refresca el análisis: una o más observaciones ya no coinciden con el ente seleccionado.",
                }
            ), 400

        rows_by_id = {int(row["id"]): dict(row) for row in found_rows}
        for item in sanitized_updates:
            if item["monto_emitido"] < 0 or item["monto_solventado"] < 0 or item["monto_pendiente"] < 0:
                return jsonify({"ok": False, "error": "Los montos no pueden ser negativos."}), 400
            current = rows_by_id[item["id"]]
            tipo_anexo = (current["tipo_anexo"] or "").strip().upper()
            if tipo_anexo == "PDP":
                emitido = float(item["monto_emitido"])
                solventado = float(item["monto_solventado"])
                if item["estado"] == "Solventado" and abs(solventado - emitido) > 0.009:
                    return jsonify(
                        {
                            "ok": False,
                            "error": f"ID {item['id']}: una observación PDP en Solventado debe quedar totalmente solventada.",
                        }
                    ), 400
                if solventado > emitido + 0.009:
                    return jsonify(
                        {
                            "ok": False,
                            "error": f"ID {item['id']}: monto solventado no puede ser mayor al observado.",
                        }
                    ), 400
                item["monto_pendiente"] = max(0.0, emitido - solventado)

        backup_path = _create_db_snapshot(f"observaciones-carga-masiva-{ejercicio}-ente-{ente_id}")
        updated = 0
        pdp_updated = 0
        for item in sanitized_updates:
            current = rows_by_id[item["id"]]
            tipo_anexo = (current["tipo_anexo"] or "").strip().upper()
            if tipo_anexo == "PDP":
                db.execute(
                    """
                    UPDATE observaciones
                    SET estado = ?,
                        ramo_33 = ?,
                        monto_pdp_emitido = ?,
                        monto_pdp_solventado = ?,
                        monto_pdp_pendiente = ?,
                        monto = ?
                    WHERE id = ?
                    """,
                    (
                        item["estado"],
                        item["ramo_33"],
                        item["monto_emitido"],
                        item["monto_solventado"],
                        item["monto_pendiente"],
                        item["monto_emitido"],
                        item["id"],
                    ),
                )
                pdp_updated += 1
            else:
                db.execute(
                    """
                    UPDATE observaciones
                    SET estado = ?,
                        ramo_33 = ?
                    WHERE id = ?
                    """,
                    (
                        item["estado"],
                        item["ramo_33"],
                        item["id"],
                    ),
                )
            updated += 1
        db.commit()

        return jsonify(
            {
                "ok": True,
                "updated": updated,
                "pdp_updated": pdp_updated,
                "backup_path": backup_path,
                "ente": {
                    "ente_id": (ente_row["ente_id"] or "").strip(),
                    "ente_numero": (ente_row["ente_numero"] or "").strip(),
                    "ente_nombre": (ente_row["ente_nombre"] or "").strip(),
                },
            }
        )

    @app.route("/carga/titulares", methods=["GET", "POST"])
    @gabo_required
    def carga_titulares():
        user = get_current_user()
        db = get_db()
        ejercicios_rows = db.execute(
            """
            SELECT DISTINCT TRIM(COALESCE(ejercicio, '')) AS ejercicio
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) != ''
            ORDER BY CAST(TRIM(COALESCE(ejercicio, '')) AS INTEGER) DESC, ejercicio DESC
            """
        ).fetchall()
        ejercicios = [row["ejercicio"] for row in ejercicios_rows if (row["ejercicio"] or "").strip()]
        if not ejercicios:
            ejercicios = [TITULAR_EJERCICIO_FIJO]
        ejercicios = _editable_ejercicios(ejercicios, user=user)

        form_source = request.form if request.method == "POST" else request.args
        requested_default = (
            (request.form.get("titular_ejercicio") if request.method == "POST" else request.args.get("ejercicio"))
            or TITULAR_EJERCICIO_FIJO
        )
        default_ejercicio = " ".join((requested_default or "").split())
        if request.method != "POST" and default_ejercicio not in ejercicios:
            default_ejercicio = ejercicios[0]
        elif request.method == "POST" and not default_ejercicio:
            default_ejercicio = ejercicios[0]

        form_data = _build_titular_form_data(
            form_source,
            default_ejercicio=default_ejercicio,
        )
        if request.method != "POST" and form_data["titular_ejercicio"] not in ejercicios:
            form_data["titular_ejercicio"] = default_ejercicio

        titular_result = None
        if request.method == "POST":
            try:
                _ensure_editable_ejercicio(form_data["titular_ejercicio"], user=user)
                if form_data["titular_ejercicio"] not in ejercicios:
                    raise ValueError("Titulares: el ejercicio seleccionado no está disponible.")
                backup_path = _create_db_snapshot("titulares-save")
                titular_result = _save_titulares_capture(db, user, form_data)
                if titular_result.get("ok"):
                    form_data["titular_nombre"] = ""
                    form_data["titular_administrativo"] = ""
                    titular_result["message"] = (
                        f"{titular_result.get('message', 'Titulares guardados correctamente.')} "
                        f"Respaldo: {backup_path}."
                    ).strip()
            except ValueError as exc:
                titular_result = {
                    "ok": False,
                    "level": "error",
                    "message": str(exc),
                }

        entes = _load_titular_entes(db, form_data["titular_ejercicio"])

        return render_template(
            "carga_titulares.html",
            user=user,
            ejercicios=ejercicios,
            entes=entes,
            form_data=form_data,
            titular_result=titular_result,
            read_only_ejercicios=sorted(_readonly_ejercicios_for_user(user)),
        )

    @app.post("/carga/titulares/excel/preview")
    @gabo_required
    def carga_titulares_excel_preview():
        upload = request.files.get("titulares_file")
        if not upload or not (upload.filename or "").strip():
            return jsonify({"ok": False, "error": "Selecciona un archivo .xlsx."}), 400
        ejercicio = " ".join((request.form.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(request.form.get("ente_id") or "")
        tipo_auditoria_destino = normalize_tipo_auditoria(
            request.form.get("tipo_auditoria_destino") or request.form.get("tipo_auditoria") or ""
        )
        db = get_db()
        try:
            preview = _build_titulares_excel_preview(
                db,
                upload,
                ejercicio=ejercicio,
                ente_id_norm=ente_id,
                tipo_auditoria_destino=tipo_auditoria_destino,
            )
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        return jsonify(preview)

    @app.post("/carga/titulares/excel/aplicar")
    @gabo_required
    def carga_titulares_excel_aplicar():
        user = get_current_user()
        payload = request.get_json(silent=True) or {}
        db = get_db()
        try:
            result = _apply_titulares_excel_import(db, user, payload)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        return jsonify(result)

    @app.get("/carga/titulares/historial")
    @gabo_required
    def carga_titulares_historial():
        ejercicio = " ".join((request.args.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = normalize_tipo_auditoria(request.args.get("tipo_auditoria", ""))
        if not ejercicio:
            return jsonify({"ok": False, "error": "Selecciona un ejercicio para consultar."}), 400

        db = get_db()
        rows = _list_historial_titulares_rows(
            db,
            ejercicio=ejercicio,
            ente_id_norm=ente_id,
            tipo_auditoria=tipo_auditoria,
        )
        capture_rows = _list_cargas_titulares_rows(
            db,
            ejercicio=ejercicio,
            ente_id_norm=ente_id,
            tipo_auditoria=tipo_auditoria,
        )
        return jsonify({"ok": True, "rows": rows, "capture_rows": capture_rows})

    @app.get("/carga/titulares/exportar")
    @gabo_required
    def carga_titulares_exportar():
        ejercicio = " ".join((request.args.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = normalize_tipo_auditoria(request.args.get("tipo_auditoria", ""))
        if not ejercicio:
            return jsonify({"ok": False, "error": "Selecciona un ejercicio para exportar."}), 400

        db = get_db()
        rows = _list_cargas_titulares_rows(
            db,
            ejercicio=ejercicio,
            ente_id_norm=ente_id,
            tipo_auditoria=tipo_auditoria,
        )

        def ente_label(row: dict) -> str:
            numero = " ".join((row.get("ente_numero") or "").split())
            nombre = " ".join((row.get("ente_nombre") or "").split())
            if numero and nombre:
                return f"{numero} - {nombre}"
            return nombre or numero or "Sin ente"

        def sort_key(row: dict) -> tuple:
            row_ejercicio = row.get("ejercicio") or ejercicio
            informe_inicio, _ = parse_periodo_cedula(row_ejercicio, row.get("periodo_informe") or "")
            admin_inicio, _ = parse_periodo_cedula(row_ejercicio, row.get("periodo_administrativo") or "")
            cedula_inicio, _ = parse_periodo_cedula(row_ejercicio, row.get("cedula_resultados") or "")
            return (
                normalize_text_key(ente_label(row)),
                normalize_text_key(row.get("tipo_auditoria") or ""),
                informe_inicio or normalize_text_key(row.get("periodo_informe") or ""),
                normalize_text_key(row.get("titular") or ""),
                admin_inicio or normalize_text_key(row.get("periodo_administrativo") or ""),
                normalize_text_key(row.get("administrativo") or ""),
                cedula_inicio or normalize_text_key(row.get("cedula_resultados") or ""),
                int(row.get("cedula_orden") or 0),
            )

        sorted_rows = sorted(rows, key=sort_key)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Titulares"
        headers = [
            "Ente",
            "Alcance",
            "Periodos Informe",
            "Titular",
            "Administrativo",
            "Director Administrativo a cargo",
            "Cédulas de resultados",
        ]
        worksheet.append(headers)

        header_fill = PatternFill("solid", fgColor="174C3A")
        header_font = Font(color="FFFFFF", bold=True)
        thin_side = Side(style="thin", color="D9E5DD")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        merged_fill = PatternFill("solid", fgColor="F7FBF8")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        data_start_row = 2
        prepared = []
        for row in sorted_rows:
            ente = ente_label(row)
            alcance = " ".join((row.get("tipo_auditoria") or "").split()) or "-"
            periodo = " ".join((row.get("periodo_informe") or "").split()) or "-"
            titular = " ".join((row.get("titular") or "").split()) or "-"
            admin_periodo = " ".join((row.get("periodo_administrativo") or "").split()) or "-"
            administrativo = " ".join((row.get("administrativo") or "").split()) or "-"
            cedula = " ".join((row.get("cedula_resultados") or "").split()) or "-"
            worksheet.append([ente, alcance, periodo, titular, admin_periodo, administrativo, cedula])
            ente_key = normalize_text_key(ente)
            alcance_key = (ente_key, normalize_text_key(alcance))
            period_key = (*alcance_key, normalize_text_key(periodo))
            titular_key = (*period_key, normalize_text_key(titular))
            admin_period_key = (*titular_key, normalize_text_key(admin_periodo))
            admin_name_key = (*admin_period_key, normalize_text_key(administrativo))
            cedula_key = (*admin_name_key, normalize_text_key(cedula))
            prepared.append(
                {
                    "keys": [
                        ente_key,
                        alcance_key,
                        period_key,
                        titular_key,
                        admin_period_key,
                        admin_name_key,
                        cedula_key,
                    ]
                }
            )

        def merge_column_by_key(column_index: int, key_index: int) -> None:
            if not prepared:
                return
            group_start = data_start_row
            current_key = prepared[0]["keys"][key_index]
            for offset, item in enumerate(prepared[1:], start=1):
                row_number = data_start_row + offset
                if item["keys"][key_index] == current_key:
                    continue
                if row_number - group_start > 1:
                    worksheet.merge_cells(
                        start_row=group_start,
                        start_column=column_index,
                        end_row=row_number - 1,
                        end_column=column_index,
                    )
                group_start = row_number
                current_key = item["keys"][key_index]
            last_row = data_start_row + len(prepared) - 1
            if last_row - group_start >= 1:
                worksheet.merge_cells(
                    start_row=group_start,
                    start_column=column_index,
                    end_row=last_row,
                    end_column=column_index,
                )

        widths = [34, 16, 24, 28, 24, 30, 24]
        for column_index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        if prepared:
            for row in worksheet.iter_rows(
                min_row=data_start_row,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=len(headers),
            ):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if cell.column < len(headers):
                        cell.fill = merged_fill

        for index in range(7):
            merge_column_by_key(index + 1, index)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        scope_bits = [ejercicio]
        if ente_id:
            scope_bits.append(ente_id.replace(".", "_"))
        if tipo_auditoria:
            scope_bits.append(normalize_text_key(tipo_auditoria).replace(" ", "_") or "alcance")
        filename = f"titulares_{'_'.join(scope_bits)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/carga/titulares/historial/<int:historial_id>/actualizar")
    @gabo_required
    def carga_titulares_historial_actualizar(historial_id: int):
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = request.form

        ejercicio = " ".join((payload.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(payload.get("ente_id", ""))
        tipo_auditoria = normalize_tipo_auditoria(payload.get("tipo_auditoria", ""))
        tipo_registro = " ".join((payload.get("tipo_registro") or "").split()).lower()
        nombre = " ".join((payload.get("nombre") or "").split())
        cargo = " ".join((payload.get("cargo") or "").split())
        fecha_inicio_raw = " ".join((payload.get("fecha_inicio") or "").split())
        fecha_fin_raw = " ".join((payload.get("fecha_fin") or "").split())

        if not ejercicio:
            return jsonify({"ok": False, "error": "Selecciona un ejercicio."}), 400
        try:
            _ensure_editable_ejercicio(ejercicio)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 403
        if not ente_id:
            return jsonify({"ok": False, "error": "Selecciona un ente."}), 400
        if tipo_auditoria not in {"Financiera", "Obra Pública"}:
            return jsonify({"ok": False, "error": "Tipo de auditoría inválido."}), 400
        if tipo_registro not in {"titular", "director_administrativo"}:
            return jsonify({"ok": False, "error": "Tipo de registro inválido."}), 400
        if not nombre:
            return jsonify({"ok": False, "error": "Captura el nombre del responsable."}), 400

        if not cargo:
            cargo = "Director Administrativo" if tipo_registro == "director_administrativo" else "Titular"

        fecha_inicio = parse_historial_date(fecha_inicio_raw)
        fecha_fin = parse_historial_date(fecha_fin_raw)
        if not fecha_inicio or not fecha_fin:
            return jsonify({"ok": False, "error": "Captura fechas válidas en formato ISO."}), 400
        if fecha_inicio > fecha_fin:
            return jsonify({"ok": False, "error": "La fecha inicial no puede ser mayor que la final."}), 400

        db = get_db()
        existing_row = db.execute(
            """
            SELECT id, CAST(ejercicio AS TEXT) AS ejercicio
            FROM historial_titulares
            WHERE id = ?
            LIMIT 1
            """,
            (historial_id,),
        ).fetchone()
        if not existing_row:
            return jsonify({"ok": False, "error": "El registro solicitado no existe."}), 404
        try:
            _ensure_editable_ejercicio(existing_row["ejercicio"])
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)

        ente_row = _get_ente_row_by_ejercicio_id(db, ejercicio, ente_id)
        if not ente_row:
            return jsonify({"ok": False, "error": "El ente seleccionado no existe en ese ejercicio."}), 400

        fecha_inicio_iso = fecha_inicio.isoformat()
        fecha_fin_iso = fecha_fin.isoformat()
        ente_nombre = (ente_row["ente_nombre"] or "").strip()
        ente_uid = (ente_row["ente_uid"] or "").strip()

        if ente_uid:
            duplicate_row = db.execute(
                """
                SELECT id
                FROM historial_titulares
                WHERE id != ?
                  AND CAST(ejercicio AS TEXT) = ?
                  AND tipo_auditoria = ?
                  AND TRIM(COALESCE(nombre, '')) = ?
                  AND TRIM(COALESCE(cargo, '')) = ?
                  AND fecha_inicio = ?
                  AND fecha_fin = ?
                  AND tipo_registro = ?
                  AND (
                    TRIM(COALESCE(ente_uid, '')) = ?
                    OR TRIM(COALESCE(ente, '')) = ?
                  )
                LIMIT 1
                """,
                (
                    historial_id,
                    ejercicio,
                    tipo_auditoria,
                    nombre,
                    cargo,
                    fecha_inicio_iso,
                    fecha_fin_iso,
                    tipo_registro,
                    ente_uid,
                    ente_nombre,
                ),
            ).fetchone()
        else:
            duplicate_row = db.execute(
                """
                SELECT id
                FROM historial_titulares
                WHERE id != ?
                  AND CAST(ejercicio AS TEXT) = ?
                  AND TRIM(COALESCE(ente, '')) = ?
                  AND tipo_auditoria = ?
                  AND TRIM(COALESCE(nombre, '')) = ?
                  AND TRIM(COALESCE(cargo, '')) = ?
                  AND fecha_inicio = ?
                  AND fecha_fin = ?
                  AND tipo_registro = ?
                LIMIT 1
                """,
                (
                    historial_id,
                    ejercicio,
                    ente_nombre,
                    tipo_auditoria,
                    nombre,
                    cargo,
                    fecha_inicio_iso,
                    fecha_fin_iso,
                    tipo_registro,
                ),
            ).fetchone()
        if duplicate_row:
            return jsonify({"ok": False, "error": "Ya existe otro registro idéntico en historial."}), 400

        backup_path = _create_db_snapshot(f"historial-titulares-{historial_id}-actualizar")
        db.execute(
            """
            UPDATE historial_titulares
            SET ejercicio = ?,
                ente_uid = ?,
                ente = ?,
                tipo_auditoria = ?,
                nombre = ?,
                cargo = ?,
                fecha_inicio = ?,
                fecha_fin = ?,
                tipo_registro = ?
            WHERE id = ?
            """,
            (
                int(ejercicio),
                ente_uid or None,
                ente_nombre,
                tipo_auditoria,
                nombre,
                cargo,
                fecha_inicio_iso,
                fecha_fin_iso,
                tipo_registro,
                historial_id,
            ),
        )
        db.commit()

        updated_rows = _list_historial_titulares_rows(
            db,
            ejercicio=ejercicio,
            ente_id_norm=ente_id,
            tipo_auditoria=tipo_auditoria,
        )
        updated_row = next(
            (item for item in updated_rows if int(item["id"]) == historial_id),
            None,
        )
        if updated_row is None:
            updated_row = {
                "id": historial_id,
                "ejercicio": ejercicio,
                "ente_id": (ente_row["ente_id"] or "").strip(),
                "ente_numero": (ente_row["ente_numero"] or "").strip(),
                "ente_nombre": ente_nombre,
                "tipo_auditoria": tipo_auditoria,
                "nombre": nombre,
                "cargo": cargo,
                "fecha_inicio": fecha_inicio_iso,
                "fecha_fin": fecha_fin_iso,
                "tipo_registro": tipo_registro,
            }

        return jsonify(
            {
                "ok": True,
                "message": "Historial de titulares actualizado correctamente.",
                "row": updated_row,
                "backup_path": backup_path,
            }
        )

    @app.get("/carga/observaciones-admin/datos")
    @gabo_required
    def carga_observaciones_admin_datos():
        try:
            where_clauses, params, scope, _ = build_observaciones_admin_scope(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if not scope["ejercicio"]:
            return jsonify({"ok": False, "error": "Selecciona un ejercicio para consultar."}), 400

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        db = get_db()
        rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(fuente_financiamiento, '')) AS fuente_financiamiento,
                TRIM(COALESCE(modalidad, 'Fuente')) AS modalidad,
                TRIM(COALESCE(convenio_nombre, '')) AS convenio_nombre,
                TRIM(COALESCE(convenio_ente_nombre, '')) AS convenio_ente_nombre,
                TRIM(COALESCE(convenio_ente_id, '')) AS convenio_ente_id,
                TRIM(COALESCE(periodo_cedula, '')) AS periodo_cedula,
                TRIM(COALESCE(oficio, '')) AS oficio,
                TRIM(COALESCE(fecha_notificacion, '')) AS fecha_notificacion,
                TRIM(COALESCE(created_at, '')) AS created_at,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                COALESCE(numero_observacion, 0) AS numero_observacion,
                TRIM(COALESCE(estado, '')) AS estado,
                COALESCE(reclasificada, 0) AS reclasificada,
                COALESCE(monto_pdp_emitido, 0) AS monto_pdp_emitido,
                COALESCE(monto_pdp_solventado, 0) AS monto_pdp_solventado,
                COALESCE(monto_pdp_pendiente, 0) AS monto_pdp_pendiente,
                TRIM(COALESCE(pdp_concepto_irregularidad, '')) AS pdp_concepto_irregularidad,
                TRIM(COALESCE(pdp_subconcepto_irregularidad, '')) AS pdp_subconcepto_irregularidad
            FROM observaciones
            WHERE {where_sql}
            ORDER BY
                id DESC
            LIMIT 2500
            """,
            params,
        ).fetchall()
        payload = []
        for row in rows:
            item = dict(row)
            item["estado"] = _normalize_observacion_estado(item.get("estado", ""))
            item["reclasificada"] = 1 if int(item.get("reclasificada") or 0) else 0
            payload.append(item)
        return jsonify({"ok": True, "rows": payload})

    @app.post("/carga/observaciones-admin/solventacion-importar")
    @gabo_required
    def carga_observaciones_admin_solventacion_importar():
        payload = request.get_json(silent=True) or {}
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}
        raw_rows = payload.get("rows")
        if not isinstance(raw_rows, list) or not raw_rows:
            return jsonify({"ok": False, "error": "No se recibieron bloques de solventación para aplicar."}), 400
        try:
            where_clauses, params, scope, _ = build_observaciones_admin_scope(raw_scope)
            _require_safe_bulk_scope(scope, action_label="importar solventación")
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)

        db = get_db()
        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        rows = db.execute(
            f"""
            SELECT
                id,
                TRIM(COALESCE(tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(fuente_financiamiento, '')) AS fuente_financiamiento,
                TRIM(COALESCE(modalidad, 'Fuente')) AS modalidad,
                TRIM(COALESCE(convenio_nombre, '')) AS convenio_nombre,
                TRIM(COALESCE(periodo_cedula, '')) AS periodo_cedula,
                TRIM(COALESCE(tipo_anexo, '')) AS tipo_anexo,
                COALESCE(numero_observacion, 0) AS numero_observacion,
                TRIM(COALESCE(estado, '')) AS estado,
                COALESCE(monto_pdp_emitido, 0) AS monto_pdp_emitido,
                COALESCE(monto_pdp_solventado, 0) AS monto_pdp_solventado
            FROM observaciones
            WHERE {where_sql}
            ORDER BY
                LOWER(TRIM(COALESCE(tipo_auditoria, ''))),
                LOWER(TRIM(COALESCE(fuente_financiamiento, ''))),
                LOWER(TRIM(COALESCE(periodo_cedula, ''))),
                LOWER(TRIM(COALESCE(tipo_anexo, ''))),
                COALESCE(numero_observacion, 0),
                id
            """,
            params,
        ).fetchall()
        if not rows:
            return jsonify({"ok": False, "error": "No hay observaciones del oficio actual para aplicar la solventación."}), 404

        grouped_rows: dict[tuple[str, str, str, str], list[sqlite3.Row]] = {}
        scope_ejercicio = str(scope.get("ejercicio") or "").strip()
        for row in rows:
            modalidad = normalize_observacion_modalidad(row["modalidad"] or "Fuente")
            convenio_nombre = normalize_convenio_text(row["convenio_nombre"] or "") if modalidad == "Convenio" else ""
            key = (
                modalidad.lower(),
                " ".join(str(row["tipo_auditoria"] or "").split()).lower(),
                " ".join(str(row["fuente_financiamiento"] or "").split()).lower(),
                convenio_nombre.lower(),
                normalize_solventacion_periodo_key(scope_ejercicio, str(row["periodo_cedula"] or "")),
                " ".join(str(row["tipo_anexo"] or "").split()).upper(),
            )
            grouped_rows.setdefault(key, []).append(row)

        normalized_import_rows: list[dict[str, object]] = []
        for index, item in enumerate(raw_rows, start=1):
            if not isinstance(item, dict):
                return jsonify({"ok": False, "error": f"El bloque {index} es inválido."}), 400
            tipo_auditoria = " ".join(str(item.get("tipo_auditoria") or item.get("tipoAuditoria") or "").split())
            fuente = " ".join(str(item.get("fuente_financiamiento") or item.get("fuente_nombre") or "").split())
            modalidad = normalize_observacion_modalidad(item.get("modalidad") or "Fuente")
            convenio_nombre = normalize_convenio_text(item.get("convenio_nombre") or "") if modalidad == "Convenio" else ""
            periodo = " ".join(str(item.get("periodo") or "").split())
            tipo_anexo = " ".join(str(item.get("tipo_anexo") or "").split()).upper()
            emitidas = int(item.get("emitidas") or 0)
            solventadas = _parse_solventacion_progressive_numbers(item.get("solventadas_indices") or [])
            pendientes = _parse_solventacion_progressive_numbers(item.get("pendientes_indices") or [])
            if not tipo_auditoria or not fuente or not periodo or not tipo_anexo:
                return jsonify({"ok": False, "error": f"El bloque {index} debe incluir tipo, fuente, periodo y anexo."}), 400
            if modalidad == "Convenio" and not convenio_nombre:
                return jsonify({"ok": False, "error": f"El bloque {index} de convenio debe incluir el nombre del convenio."}), 400
            if emitidas < 0:
                return jsonify({"ok": False, "error": f"El bloque {index} incluye una cantidad emitida inválida."}), 400
            if emitidas == 0 and (solventadas or pendientes):
                return jsonify({"ok": False, "error": f"El bloque {index} no puede incluir progresivos si las emitidas son 0."}), 400
            overlap = sorted(set(solventadas) & set(pendientes))
            if overlap:
                return jsonify({"ok": False, "error": f"El bloque {index} repite números en solventadas y pendientes: {', '.join(map(str, overlap))}."}), 400
            if emitidas > 0:
                expected_numbers = set(range(1, emitidas + 1))
                received_numbers = set(solventadas) | set(pendientes)
                if received_numbers != expected_numbers:
                    missing = sorted(expected_numbers - received_numbers)
                    extras = sorted(received_numbers - expected_numbers)
                    parts = []
                    if missing:
                        parts.append(f"faltan {', '.join(map(str, missing))}")
                    if extras:
                        parts.append(f"sobran {', '.join(map(str, extras))}")
                    return jsonify({"ok": False, "error": f"El bloque {index} no cubre exactamente las emitidas: {'; '.join(parts)}."}), 400
            normalized_import_rows.append(
                {
                    "tipo_auditoria": tipo_auditoria,
                    "fuente": fuente,
                    "modalidad": modalidad,
                    "convenio_nombre": convenio_nombre,
                    "periodo": periodo,
                    "tipo_anexo": tipo_anexo,
                    "emitidas": emitidas,
                    "solventadas": solventadas,
                    "pendientes": pendientes,
                }
            )

        updates: list[tuple[str, int, int]] = []
        touched_ids: list[int] = []
        for item in normalized_import_rows:
            modalidad = str(item["modalidad"])
            convenio_nombre = str(item["convenio_nombre"])
            key = (
                modalidad.lower(),
                str(item["tipo_auditoria"]).lower(),
                str(item["fuente"]).lower(),
                convenio_nombre.lower(),
                normalize_solventacion_periodo_key(scope_ejercicio, str(item["periodo"])),
                str(item["tipo_anexo"]).upper(),
            )
            target_rows = list(grouped_rows.get(key) or [])
            if not target_rows and modalidad == "Convenio":
                convenio_fallback_key = (
                    modalidad.lower(),
                    "obra pública",
                    str(item["fuente"]).lower(),
                    convenio_nombre.lower(),
                    normalize_solventacion_periodo_key(scope_ejercicio, str(item["periodo"])),
                    str(item["tipo_anexo"]).upper(),
                )
                target_rows = list(grouped_rows.get(convenio_fallback_key) or [])
            if not target_rows:
                if int(item["emitidas"]) == 0:
                    continue
                return jsonify(
                    {
                        "ok": False,
                        "error": (
                            "No se encontró el bloque "
                            f"{item['tipo_auditoria']} / {item['fuente']} / {item['periodo']} / {item['tipo_anexo']}"
                            + (f" / {convenio_nombre}" if modalidad == "Convenio" else "")
                        ),
                    }
                ), 400
            if int(item["emitidas"]) == 0:
                return jsonify(
                    {
                        "ok": False,
                        "error": (
                            "El bloque "
                            f"{item['tipo_auditoria']} / {item['fuente']} / {item['periodo']} / {item['tipo_anexo']} "
                            f"marca 0 emitidas, pero el oficio actual tiene {len(target_rows)} observaciones."
                        ),
                    }
                ), 400
            if len(target_rows) != int(item["emitidas"]):
                return jsonify(
                    {
                        "ok": False,
                        "error": (
                            "El bloque "
                            f"{item['tipo_auditoria']} / {item['fuente']} / {item['periodo']} / {item['tipo_anexo']} "
                            f"espera {item['emitidas']} observaciones y el oficio actual tiene {len(target_rows)}."
                        ),
                    }
                ), 400
            solventadas_set = set(item["solventadas"])
            for number, row in enumerate(target_rows, start=1):
                estado = "Solventado" if number in solventadas_set else "Pendiente"
                updates.append((estado, number, int(row["id"])))
                touched_ids.append(int(row["id"]))

        if not updates:
            return jsonify({"ok": False, "error": "No se generaron cambios para aplicar."}), 400

        backup_path = _create_db_snapshot("observaciones-solventacion-importar")
        current_rows_by_id = {int(row["id"]): row for row in rows}
        for estado, numero_observacion, observacion_id in updates:
            current = current_rows_by_id.get(observacion_id)
            if current is None:
                continue
            if str(current["tipo_anexo"] or "").strip().upper() == "PDP":
                monto_emitido = float(current["monto_pdp_emitido"] or 0.0)
                monto_solventado_actual = float(current["monto_pdp_solventado"] or 0.0)
                monto_solventado = (
                    monto_emitido
                    if estado == "Solventado"
                    else min(max(monto_solventado_actual, 0.0), monto_emitido)
                )
                monto_pendiente = max(0.0, monto_emitido - monto_solventado)
                db.execute(
                    """
                    UPDATE observaciones
                    SET estado = ?,
                        numero_observacion = ?,
                        monto_pdp_solventado = ?,
                        monto_pdp_pendiente = ?,
                        monto = ?
                    WHERE id = ?
                    """,
                    (
                        estado,
                        numero_observacion,
                        monto_solventado,
                        monto_pendiente,
                        monto_emitido,
                        observacion_id,
                    ),
                )
            else:
                db.execute(
                    """
                    UPDATE observaciones
                    SET estado = ?,
                        numero_observacion = ?
                    WHERE id = ?
                    """,
                    (estado, numero_observacion, observacion_id),
                )
        db.commit()
        return jsonify(
            {
                "ok": True,
                "updated": len(touched_ids),
                "blocks": len(normalized_import_rows),
                "backup_path": backup_path,
            }
        )

    @app.post("/carga/observaciones-admin/<int:observacion_id>/borrar")
    @gabo_required
    def carga_observaciones_admin_borrar(observacion_id: int):
        payload = request.get_json(silent=True) or {}
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}
        db = get_db()
        found = db.execute(
            """
            SELECT
                id,
                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(oficio, '')) AS oficio
            FROM observaciones
            WHERE id = ?
            LIMIT 1
            """,
            (observacion_id,),
        ).fetchone()
        if not found:
            return jsonify({"ok": False, "error": "Observación no encontrada."}), 404
        try:
            _ensure_editable_ejercicio(found["ejercicio"])
            _validate_observacion_matches_scope(found, raw_scope)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        backup_path = _create_db_snapshot(f"observacion-{observacion_id}-borrar")
        db.execute("DELETE FROM observaciones WHERE id = ?", (observacion_id,))
        db.commit()
        return jsonify({"ok": True, "deleted": 1, "id": observacion_id, "backup_path": backup_path})

    @app.post("/carga/observaciones-admin/borrar")
    @gabo_required
    def carga_observaciones_admin_borrar_multiples():
        payload = request.get_json(silent=True) or {}
        raw_ids = payload.get("ids")
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}
        if not isinstance(raw_ids, list) or not raw_ids:
            return jsonify({"ok": False, "error": "No se recibieron observaciones para borrar."}), 400

        ids: list[int] = []
        for raw_id in raw_ids:
            try:
                item_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if item_id > 0:
                ids.append(item_id)
        ids = sorted(set(ids))
        if not ids:
            return jsonify({"ok": False, "error": "Lista de observaciones inválida."}), 400

        db = get_db()
        if raw_scope:
            try:
                where_clauses, params, scope, _ = build_observaciones_admin_scope(raw_scope)
                _require_observaciones_admin_edit_scope(scope, action_label="borrar observaciones")
            except ValueError as exc:
                return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
            if _count_scope_ids(db, ids, where_clauses, params) != len(ids):
                return jsonify(
                    {
                        "ok": False,
                        "error": "Refresca la consulta: las observaciones seleccionadas ya no coinciden con el oficio activo.",
                    }
                ), 400
        blocked_ejercicio = _first_readonly_observacion_ejercicio(db, ids)
        if blocked_ejercicio:
            return jsonify({"ok": False, "error": _readonly_obs_message(blocked_ejercicio)}), 403
        placeholders = ", ".join(["?"] * len(ids))
        count_row = db.execute(
            f"SELECT COUNT(*) AS total FROM observaciones WHERE id IN ({placeholders})",
            ids,
        ).fetchone()
        total = int((count_row["total"] if count_row else 0) or 0)
        if total <= 0:
            return jsonify({"ok": False, "error": "No se encontraron observaciones para borrar."}), 404

        backup_path = _create_db_snapshot("observaciones-borrar-seleccion")
        db.execute(f"DELETE FROM observaciones WHERE id IN ({placeholders})", ids)
        db.commit()
        return jsonify({"ok": True, "deleted": total, "backup_path": backup_path})

    @app.post("/carga/observaciones-admin/solventar")
    @gabo_required
    def carga_observaciones_admin_solventar_multiples():
        payload = request.get_json(silent=True) or {}
        raw_ids = payload.get("ids")
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}
        if not isinstance(raw_ids, list) or not raw_ids:
            return jsonify({"ok": False, "error": "No se recibieron observaciones para solventar."}), 400

        ids: list[int] = []
        for raw_id in raw_ids:
            try:
                item_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if item_id > 0:
                ids.append(item_id)
        ids = sorted(set(ids))
        if not ids:
            return jsonify({"ok": False, "error": "Lista de observaciones inválida."}), 400

        db = get_db()
        try:
            where_clauses, params, scope, _ = build_observaciones_admin_scope(raw_scope)
            _require_safe_bulk_scope(scope, action_label="solventar en bloque")
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        if _count_scope_ids(db, ids, where_clauses, params) != len(ids):
            return jsonify(
                {
                    "ok": False,
                    "error": "Refresca la consulta: el conjunto visible ya no coincide con los filtros activos.",
                }
            ), 400
        blocked_ejercicio = _first_readonly_observacion_ejercicio(db, ids)
        if blocked_ejercicio:
            return jsonify({"ok": False, "error": _readonly_obs_message(blocked_ejercicio)}), 403
        placeholders = ", ".join(["?"] * len(ids))
        count_row = db.execute(
            f"SELECT COUNT(*) AS total FROM observaciones WHERE id IN ({placeholders})",
            ids,
        ).fetchone()
        total = int((count_row["total"] if count_row else 0) or 0)
        if total <= 0:
            return jsonify({"ok": False, "error": "No se encontraron observaciones para solventar."}), 404

        pdp_row = db.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM observaciones
            WHERE id IN ({placeholders})
              AND UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
            """,
            ids,
        ).fetchone()
        total_pdp = int((pdp_row["total"] if pdp_row else 0) or 0)

        backup_path = _create_db_snapshot("observaciones-solventar-todo")
        db.execute(
            f"""
            UPDATE observaciones
            SET estado = 'Solventado',
                monto_pdp_solventado = CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN COALESCE(monto_pdp_emitido, 0)
                    ELSE monto_pdp_solventado
                END,
                monto_pdp_pendiente = CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN 0
                    ELSE monto_pdp_pendiente
                END
            WHERE id IN ({placeholders})
            """,
            ids,
        )
        db.commit()
        return jsonify({"ok": True, "updated": total, "pdp": total_pdp, "backup_path": backup_path})

    @app.post("/carga/observaciones-admin/pendiente")
    @gabo_required
    def carga_observaciones_admin_pendiente_multiples():
        payload = request.get_json(silent=True) or {}
        raw_ids = payload.get("ids")
        raw_scope = payload.get("scope") if isinstance(payload.get("scope"), dict) else {}
        if not isinstance(raw_ids, list) or not raw_ids:
            return jsonify({"ok": False, "error": "No se recibieron observaciones para dejar en pendiente."}), 400

        ids: list[int] = []
        for raw_id in raw_ids:
            try:
                item_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if item_id > 0:
                ids.append(item_id)
        ids = sorted(set(ids))
        if not ids:
            return jsonify({"ok": False, "error": "Lista de observaciones inválida."}), 400

        db = get_db()
        try:
            where_clauses, params, scope, _ = build_observaciones_admin_scope(raw_scope)
            _require_safe_bulk_scope(scope, action_label="dejar pendientes en bloque")
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)
        if _count_scope_ids(db, ids, where_clauses, params) != len(ids):
            return jsonify(
                {
                    "ok": False,
                    "error": "Refresca la consulta: el conjunto visible ya no coincide con los filtros activos.",
                }
            ), 400
        blocked_ejercicio = _first_readonly_observacion_ejercicio(db, ids)
        if blocked_ejercicio:
            return jsonify({"ok": False, "error": _readonly_obs_message(blocked_ejercicio)}), 403
        placeholders = ", ".join(["?"] * len(ids))
        count_row = db.execute(
            f"SELECT COUNT(*) AS total FROM observaciones WHERE id IN ({placeholders})",
            ids,
        ).fetchone()
        total = int((count_row["total"] if count_row else 0) or 0)
        if total <= 0:
            return jsonify({"ok": False, "error": "No se encontraron observaciones para dejar en pendiente."}), 404

        pdp_row = db.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM observaciones
            WHERE id IN ({placeholders})
              AND UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
            """,
            ids,
        ).fetchone()
        total_pdp = int((pdp_row["total"] if pdp_row else 0) or 0)

        backup_path = _create_db_snapshot("observaciones-pendiente-todo")
        db.execute(
            f"""
            UPDATE observaciones
            SET estado = 'Pendiente',
                monto_pdp_solventado = CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN 0
                    ELSE monto_pdp_solventado
                END,
                monto_pdp_pendiente = CASE
                    WHEN UPPER(TRIM(COALESCE(tipo_anexo, ''))) = 'PDP'
                    THEN COALESCE(monto_pdp_emitido, 0)
                    ELSE monto_pdp_pendiente
                END
            WHERE id IN ({placeholders})
            """,
            ids,
        )
        db.commit()
        return jsonify({"ok": True, "updated": total, "pdp": total_pdp, "backup_path": backup_path})

    @app.post("/carga/observaciones-admin/borrar-todo")
    @gabo_required
    def carga_observaciones_admin_borrar_todo():
        payload = request.get_json(silent=True) or {}
        try:
            where_clauses, params, scope, extra_filters = build_observaciones_admin_scope(payload)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

        if not scope["ejercicio"]:
            return jsonify({"ok": False, "error": "Debes seleccionar ejercicio para borrar."}), 400
        try:
            _require_safe_bulk_scope(scope, action_label="borrar por filtros")
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), _readonly_error_status(exc)

        where_sql = " AND ".join(where_clauses)
        db = get_db()
        count_row = db.execute(
            f"SELECT COUNT(*) AS total FROM observaciones WHERE {where_sql}",
            params,
        ).fetchone()
        total = int((count_row["total"] if count_row else 0) or 0)
        if total <= 0:
            return jsonify({"ok": True, "deleted": 0})

        backup_path = _create_db_snapshot("observaciones-borrar-filtro")
        db.execute(
            f"DELETE FROM observaciones WHERE {where_sql}",
            params,
        )
        db.commit()
        return jsonify({"ok": True, "deleted": total, "backup_path": backup_path})

    @app.post("/api/cedulas/procesar")
    @gabo_required
    def api_procesar_cedula():
        if "file" not in request.files:
            return jsonify({"error": "No se recibió ningún archivo."}), 400
        f = request.files["file"]
        if not f.filename:
            return jsonify({"error": "Nombre de archivo vacío."}), 400
        if not f.filename.lower().endswith(".pdf"):
            return jsonify({"error": "Solo se aceptan archivos PDF."}), 400
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_path = tmp.name
            f.save(tmp_path)
        try:
            result = parse_cedula(tmp_path)
        except Exception as e:
            return jsonify({"error": f"Error al procesar el PDF: {e}"}), 500
        finally:
            os.unlink(tmp_path)
        return jsonify(result)

    @app.post("/api/solventacion/procesar")
    @gabo_required
    def api_procesar_solventacion():
        if "file" not in request.files:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo."}), 400
        upload = request.files["file"]
        if not upload.filename:
            return jsonify({"ok": False, "error": "Nombre de archivo vacío."}), 400
        if not upload.filename.lower().endswith(".pdf"):
            return jsonify({"ok": False, "error": "Solo se aceptan archivos PDF."}), 400

        ejercicio = " ".join((request.form.get("ejercicio") or "").split())
        ente_id = normalize_ente_id(request.form.get("ente_id", ""))
        if not ejercicio or not ente_id:
            return jsonify({"ok": False, "error": "Selecciona ejercicio y ente antes de procesar el oficio."}), 400

        file_meta = _parse_solventacion_filename(upload.filename or "")
        if not file_meta.get("filename_valid"):
            return jsonify(
                {
                    "ok": False,
                    "error": "El nombre del PDF no cumple el formato esperado 'ente.- SIGLA_OFS_0000_2026[_periodo].pdf'.",
                }
            ), 400
        if normalize_ente_id(file_meta.get("ente_id") or "") != ente_id:
            return jsonify(
                {
                    "ok": False,
                    "error": (
                        f"El PDF corresponde al ente {file_meta.get('ente_id') or '-'} "
                        f"y el contexto activo está en {ente_id}."
                    ),
                }
            ), 400

        db = get_db()
        ente_row = db.execute(
            """
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(ente_id, '')) = ?
            LIMIT 1
            """,
            (ejercicio, ente_id),
        ).fetchone()
        if not ente_row:
            return jsonify({"ok": False, "error": "El ente activo no existe para el ejercicio seleccionado."}), 400

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_path = tmp.name
            upload.save(tmp_path)
        try:
            parsed = parse_solventacion(tmp_path)
        except Exception as exc:
            return jsonify({"ok": False, "error": f"No se pudo procesar el oficio de solventación: {exc}"}), 500
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

        payload = _build_solventacion_import_payload(
            parsed,
            entry_meta={
                **file_meta,
                "ente_numero": (ente_row["ente_numero"] or "").strip(),
                "ente_nombre": (ente_row["ente_nombre"] or "").strip(),
            },
        )
        ente_sigla = _extract_solventacion_sigla(ente_row["ente_nombre"] or "")
        destinatario_sigla = _extract_solventacion_sigla(payload.get("destinatario") or "")
        filename_sigla = _normalize_solventacion_sigla(file_meta.get("sigla") or "")
        if ente_sigla and filename_sigla and ente_sigla != filename_sigla:
            return jsonify(
                {
                    "ok": False,
                    "error": (
                        f"El nombre del archivo identifica la sigla {file_meta.get('sigla') or '-'} "
                        f"y el ente activo usa {ente_sigla}."
                    ),
                }
            ), 400
        if ente_sigla and destinatario_sigla and ente_sigla != destinatario_sigla:
            return jsonify(
                {
                    "ok": False,
                    "error": (
                        f"El PDF está dirigido a la sigla {destinatario_sigla} "
                        f"y el ente activo usa {ente_sigla}."
                    ),
                }
            ), 400
        if not destinatario_sigla:
            payload.setdefault("warnings", []).append(
                "No se pudo corroborar la sigla del ente dentro del PDF; se validó con el nombre del archivo."
            )
        if payload.get("ejercicio") and payload["ejercicio"] != ejercicio:
            return jsonify(
                {
                    "ok": False,
                    "error": (
                        f"El PDF corresponde al ejercicio {payload['ejercicio']} "
                        f"y el contexto activo está en {ejercicio}."
                    ),
                }
            ), 400
        return jsonify(payload)

    @app.route("/carga", methods=["GET", "POST"])
    @gabo_required
    def carga():
        user = get_current_user()
        db = get_db()
        requested_loader_view = (request.args.get("vista") or "").strip().lower()
        if requested_loader_view == "titulares":
            initial_loader_mode = "titulares"
        elif requested_loader_view == "solventacion":
            initial_loader_mode = "solventacion"
        else:
            initial_loader_mode = "manual"
        manual_ejercicios_rows = db.execute(
            """
            SELECT DISTINCT TRIM(COALESCE(ejercicio, '')) AS ejercicio
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) != ''
            ORDER BY CAST(TRIM(COALESCE(ejercicio, '')) AS INTEGER) DESC, ejercicio DESC
            """
        ).fetchall()
        manual_ejercicios = [row["ejercicio"] for row in manual_ejercicios_rows]
        if not manual_ejercicios:
            manual_ejercicios = [TITULAR_EJERCICIO_FIJO]
        manual_ejercicios = _editable_ejercicios(manual_ejercicios, user=user)
        manual_ejercicio_default = manual_ejercicios[0] if manual_ejercicios else TITULAR_EJERCICIO_FIJO
        fuentes_rows = db.execute(
            """
            SELECT id, nombre
            FROM fuentes_financiamiento
            ORDER BY nombre ASC
            """
        ).fetchall()
        fuentes = [dict(row) for row in fuentes_rows]
    
        titular_ejercicios = manual_ejercicios.copy()

        script_result = None
        manual_result = None
        titular_result = None
        form_data = {
            "template_ejercicio": "",
            "template_out": "bases/historial_titulares_template.csv",
            "template_tipo_auditoria": "Financiera",
            "csv_path": "",
            "csv_ejercicio": "",
            "csv_replace": "0",
            "json_path": "",
            "json_ejercicio": "",
            "json_tipo_auditoria": "Financiera",
            "manual_id": "",
            "manual_ente_id": "",
            "manual_tipo_auditoria": "",
            "manual_tipo_responsable": "Titular",
            "manual_titular_nombre": "",
            "manual_administrativo_nombre": "",
            "manual_numero_oficio": "",
            "manual_oficio_base": "",
            "manual_asunto": "Notificación de Cédula de Resultados",
            "manual_ejercicio": manual_ejercicio_default,
            "manual_fuente_id": "",
            "manual_fuente_nueva": "",
            "manual_fuentes_detalle_json": "",
            "manual_periodo": "",
            "manual_periodo_titular": "",
            "manual_ramo_33": "No",
            "manual_ramo_28": "No",
            "manual_estado": "Emitido",
            "manual_fecha_notificacion": "",
            "manual_cantidad_sa": "0",
            "manual_cantidad_pdp": "0",
            "manual_cantidad_pras": "0",
            "manual_cantidad_pefcf": "0",
            "manual_cantidad_r": "0",
            "manual_monto_pdp_emitido": "0",
            "manual_monto_pdp_solventado": "0",
            "manual_monto_pdp_pendiente": "0",
            "manual_montos_pdp": "",
            "manual_pdp_detalle_json": "",
            "titular_ejercicio": manual_ejercicio_default,
            "titular_ente_id": "",
            "titular_tipo_auditoria": "Financiera",
            "titular_periodo_informe": "",
            "titular_nombre": "",
            "titular_periodo_administrativo": "",
            "titular_administrativo": "",
            "titular_cedula_resultados": "",
        }
    
        action = ""
        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            manual_estado_raw = (request.form.get("manual_estado") or "").strip()
            form_data.update(
                {
                    "template_ejercicio": (request.form.get("template_ejercicio") or "").strip(),
                    "template_out": (request.form.get("template_out") or "").strip() or "bases/historial_titulares_template.csv",
                    "template_tipo_auditoria": (request.form.get("template_tipo_auditoria") or "").strip() or "Financiera",
                    "csv_path": (request.form.get("csv_path") or "").strip(),
                    "csv_ejercicio": (request.form.get("csv_ejercicio") or "").strip(),
                    "csv_replace": "1" if request.form.get("csv_replace") else "0",
                    "json_path": (request.form.get("json_path") or "").strip(),
                    "json_ejercicio": (request.form.get("json_ejercicio") or "").strip(),
                    "json_tipo_auditoria": (request.form.get("json_tipo_auditoria") or "").strip() or "Financiera",
                    "manual_id": (request.form.get("manual_id") or "").strip(),
                    "manual_ente_id": (request.form.get("manual_ente_id") or "").strip(),
                    "manual_tipo_auditoria": (request.form.get("manual_tipo_auditoria") or "").strip(),
                    "manual_tipo_responsable": (request.form.get("manual_tipo_responsable") or "").strip() or "Titular",
                    "manual_titular_nombre": (request.form.get("manual_titular_nombre") or "").strip(),
                    "manual_administrativo_nombre": (request.form.get("manual_administrativo_nombre") or "").strip(),
                    "manual_numero_oficio": (request.form.get("manual_numero_oficio") or "").strip(),
                    "manual_oficio_base": " ".join((request.form.get("manual_oficio_base") or "").split()),
                    "manual_asunto": (request.form.get("manual_asunto") or "").strip() or "Notificación de Cédula de Resultados",
                    "manual_ejercicio": (request.form.get("manual_ejercicio") or "").strip() or manual_ejercicio_default,
                    "manual_fuente_id": (request.form.get("manual_fuente_id") or "").strip(),
                    "manual_fuente_nueva": (request.form.get("manual_fuente_nueva") or "").strip(),
                    "manual_fuentes_detalle_json": (request.form.get("manual_fuentes_detalle_json") or "").strip(),
                    "manual_periodo": (request.form.get("manual_periodo") or "").strip(),
                    "manual_periodo_titular": (request.form.get("manual_periodo_titular") or "").strip(),
                    "manual_ramo_33": (request.form.get("manual_ramo_33") or "").strip() or "No",
                    "manual_ramo_28": (request.form.get("manual_ramo_28") or "").strip() or "No",
                    "manual_estado": _normalize_observacion_estado(manual_estado_raw) or "Emitido",
                    "manual_fecha_notificacion": (request.form.get("manual_fecha_notificacion") or "").strip(),
                    "manual_cantidad_sa": (request.form.get("manual_cantidad_sa") or "").strip() or "0",
                    "manual_cantidad_pdp": (request.form.get("manual_cantidad_pdp") or "").strip() or "0",
                    "manual_cantidad_pras": (request.form.get("manual_cantidad_pras") or "").strip() or "0",
                    "manual_cantidad_pefcf": (request.form.get("manual_cantidad_pefcf") or "").strip() or "0",
                    "manual_cantidad_r": (request.form.get("manual_cantidad_r") or "").strip() or "0",
                    "manual_monto_pdp_emitido": (request.form.get("manual_monto_pdp_emitido") or "").strip() or "0",
                    "manual_monto_pdp_solventado": (request.form.get("manual_monto_pdp_solventado") or "").strip() or "0",
                    "manual_monto_pdp_pendiente": (request.form.get("manual_monto_pdp_pendiente") or "").strip() or "0",
                    "manual_montos_pdp": (request.form.get("manual_montos_pdp") or "").strip(),
                    "manual_pdp_detalle_json": (request.form.get("manual_pdp_detalle_json") or "").strip(),
                }
            )
            form_data.update(
                _build_titular_form_data(
                    request.form,
                    default_ejercicio=form_data["manual_ejercicio"],
                )
            )
    
            try:
                if action == "titular_save":
                    _ensure_editable_ejercicio(form_data["titular_ejercicio"], user=user)
                    backup_path = _create_db_snapshot("titulares-save")
                    titular_result = _save_titulares_capture(db, user, form_data)
                    if titular_result.get("ok"):
                        titular_result["message"] = (
                            f"{titular_result.get('message', 'Titulares guardados correctamente.')} "
                            f"Respaldo: {backup_path}."
                        ).strip()
                elif action in {"manual_check", "manual_save"}:
                    manual_id_raw = form_data["manual_id"]
                    manual_ente_id = normalize_ente_id(form_data["manual_ente_id"])
                    form_data["manual_ente_id"] = manual_ente_id
                    tipo_auditoria = form_data["manual_tipo_auditoria"]
                    tipo_responsable = "Titular"
                    titular_nombre = ""
                    administrativo_nombre = ""
                    numero_oficio = form_data["manual_numero_oficio"]
                    oficio_base = " ".join(form_data["manual_oficio_base"].split())
                    asunto = form_data["manual_asunto"]
                    ejercicio = form_data["manual_ejercicio"]
                    fuente_id_raw = form_data["manual_fuente_id"]
                    fuente_nueva = " ".join(form_data["manual_fuente_nueva"].split()) if fuente_id_raw == "__new__" else ""
                    raw_periodo = form_data["manual_periodo"]
                    periodo = raw_periodo if user and user.get("username") == "gabo" else " ".join(raw_periodo.split())
                    periodo_titular = form_data["manual_periodo_titular"]
                    ramo_33 = normalize_manual_si_no(form_data["manual_ramo_33"])
                    ramo_28 = normalize_manual_si_no(form_data["manual_ramo_28"])
                    origen_fuente = "Del Ejercicio"
                    estado = "Emitido"
                    fecha_notificacion = form_data["manual_fecha_notificacion"]
                    raw_montos_pdp = form_data["manual_montos_pdp"]
                    raw_pdp_detalle_json = form_data["manual_pdp_detalle_json"]
                    fuentes_detalle_rows = parse_manual_fuentes_detalle(form_data["manual_fuentes_detalle_json"])
                    modalidad = "Fuente"
                    convenio_nombre = ""
                    convenio_ente_nombre = ""
                    convenio_ente_id = ""
                    usa_fuentes_detalle = len(fuentes_detalle_rows) > 0
                    if usa_fuentes_detalle and not periodo:
                        periodo = " ".join((fuentes_detalle_rows[0].get("periodo") or "").split())
                        form_data["manual_periodo"] = periodo
                    manual_edit_id = None
                    if manual_id_raw:
                        try:
                            manual_edit_id = int(manual_id_raw)
                        except ValueError as exc:
                            raise ValueError("ID de edición manual inválido.") from exc

                    if not manual_ente_id:
                        raise ValueError("Debes seleccionar un ente.")
                    if not numero_oficio:
                        raise ValueError("Debes capturar el número de oficio.")
                    if asunto not in ASUNTOS_MANUALES:
                        raise ValueError("Debes seleccionar un asunto válido.")
                    if not ejercicio:
                        raise ValueError("Debes capturar el ejercicio.")
                    _ensure_editable_ejercicio(ejercicio, user=user)
                    if ejercicio not in manual_ejercicios:
                        raise ValueError("El ejercicio seleccionado no está disponible.")
                    if not fecha_notificacion:
                        raise ValueError("Debes capturar la fecha de notificación.")
                    try:
                        int(ejercicio)
                    except ValueError as exc:
                        raise ValueError("Ejercicio inválido.") from exc
                    if not periodo:
                        raise ValueError("Debes capturar el periodo.")

                    if not tipo_auditoria:
                        raise ValueError("Debes seleccionar el tipo de auditoría.")
                    if tipo_auditoria not in {"Financiera", "Obra Pública"}:
                        raise ValueError("Tipo de auditoría inválido.")
                    backup_path = ""
                    if action == "manual_save":
                        backup_path = _create_db_snapshot("carga-manual-save")
                    fuente_id = None
                    fuente_nombre = ""
                    if manual_edit_id:
                        edit_scope_row = db.execute(
                            """
                            SELECT
                                id,
                                TRIM(COALESCE(ente_id, '')) AS ente_id,
                                TRIM(COALESCE(numero_oficio, '')) AS numero_oficio,
                                TRIM(COALESCE(asunto, '')) AS asunto,
                                TRIM(COALESCE(ejercicio, '')) AS ejercicio,
                                TRIM(COALESCE(periodo, '')) AS periodo,
                                TRIM(COALESCE(created_by, '')) AS created_by
                            FROM cargas_manuales
                            WHERE id = ?
                            LIMIT 1
                            """,
                            (manual_edit_id,),
                        ).fetchone()
                        same_scope = bool(
                            edit_scope_row
                            and (edit_scope_row["created_by"] or "").strip() == user["username"]
                            and normalize_ente_id(edit_scope_row["ente_id"] or "") == normalize_ente_id(manual_ente_id)
                            and " ".join((edit_scope_row["numero_oficio"] or "").split()).lower() == numero_oficio.lower()
                            and (edit_scope_row["asunto"] or "").strip() == asunto
                            and (edit_scope_row["ejercicio"] or "").strip() == ejercicio
                            and " ".join((edit_scope_row["periodo"] or "").split()).lower() == periodo.lower()
                        )
                        if not same_scope:
                            manual_edit_id = None
                            form_data["manual_id"] = ""
                    if manual_edit_id and len(fuentes_detalle_rows) > 1:
                        raise ValueError("La edición con múltiples fuentes no está soportada.")
                    if (
                        action == "manual_save"
                        and usa_fuentes_detalle
                        and not manual_edit_id
                        and asunto in {"Notificación de Cédula de Resultados", SOLVENTACION_ASUNTO}
                    ):
                        db.execute(
                            """
                            DELETE FROM observaciones
                            WHERE TRIM(COALESCE(ejercicio, '')) = TRIM(COALESCE(?, ''))
                              AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                              AND LOWER(TRIM(COALESCE(oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
                            """,
                            (ejercicio, manual_ente_id, numero_oficio),
                        )
                        db.execute(
                            """
                            DELETE FROM cargas_manuales
                            WHERE LOWER(TRIM(COALESCE(numero_oficio, ''))) = LOWER(TRIM(COALESCE(?, '')))
                              AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                              AND asunto = ?
                              AND ejercicio = ?
                            """,
                            (numero_oficio, manual_ente_id, asunto, ejercicio),
                        )
                    if not usa_fuentes_detalle and not fuente_id_raw:
                        raise ValueError("Debes seleccionar una fuente.")
                    if not usa_fuentes_detalle and fuente_id_raw == "__new__" and not fuente_nueva:
                        raise ValueError("Debes escribir la nueva fuente.")
                    if usa_fuentes_detalle:
                        fuente_nombre = fuentes_detalle_rows[0]["fuente_nombre"]
                        fuente_id_resolved, fuente_nombre = resolve_fuente_catalogo(
                            db,
                            fuente_nombre,
                            create_missing=action == "manual_save",
                        )
                        fuente_id = (
                            int(fuente_id_resolved)
                            if fuente_id_resolved is not None
                            else -1
                        )
                    elif fuente_id_raw == "__new__":
                        fuente_nombre = fuente_nueva
                        fuente_id_resolved, fuente_nombre = resolve_fuente_catalogo(
                            db,
                            fuente_nueva,
                            create_missing=action == "manual_save",
                        )
                        fuente_id = (
                            int(fuente_id_resolved)
                            if fuente_id_resolved is not None
                            else -1
                        )
                    else:
                        if fuente_id_raw.startswith("__obs__:"):
                            fuente_obs = " ".join(fuente_id_raw.replace("__obs__:", "", 1).split())
                            if not fuente_obs:
                                raise ValueError("Debes seleccionar una fuente válida.")
                            fuente_nombre = fuente_obs
                            fuente_id_resolved, fuente_nombre = resolve_fuente_catalogo(
                                db,
                                fuente_obs,
                                create_missing=action == "manual_save",
                            )
                            fuente_id = (
                                int(fuente_id_resolved)
                                if fuente_id_resolved is not None
                                else -1
                            )
                        else:
                            try:
                                fuente_id = int(fuente_id_raw)
                            except ValueError as exc:
                                raise ValueError("Debes seleccionar una fuente válida.") from exc
                            fuente_exists = db.execute(
                                "SELECT 1 FROM fuentes_financiamiento WHERE id = ? LIMIT 1",
                                (fuente_id,),
                            ).fetchone()
                            if not fuente_exists:
                                raise ValueError("La fuente seleccionada no existe.")
                    ente_row = db.execute(
                        """
                        SELECT
                            TRIM(COALESCE(ente_id, '')) AS ente_id,
                            TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                            TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
                        FROM entes_detalle
                        WHERE TRIM(COALESCE(ejercicio, '')) = ?
                          AND TRIM(COALESCE(ente_id, '')) = ?
                        LIMIT 1
                        """,
                        (ejercicio, manual_ente_id),
                    ).fetchone()
                    if not ente_row:
                        raise ValueError(f"El ente seleccionado no existe para ejercicio {ejercicio}.")
                    if fuente_id is not None and fuente_id >= 0:
                        fuente_nombre_row = db.execute(
                            """
                            SELECT TRIM(COALESCE(nombre, '')) AS nombre
                            FROM fuentes_financiamiento
                            WHERE id = ?
                            LIMIT 1
                            """,
                            (fuente_id,),
                        ).fetchone()
                        if not fuente_nombre_row or not (fuente_nombre_row["nombre"] or "").strip():
                            raise ValueError("No se pudo resolver el nombre de la fuente seleccionada.")
                        fuente_nombre = (fuente_nombre_row["nombre"] or "").strip()
                    elif not fuente_nombre:
                        raise ValueError("No se pudo resolver el nombre de la fuente seleccionada.")

                    fuente_clasificacion = get_fuente_clasificacion(db, fuente_nombre, fuente_id)
                    ramo_33 = fuente_clasificacion["ramo_33"]
                    ramo_28 = fuente_clasificacion["ramo_28"]
                    origen_fuente = fuente_clasificacion["origen_fuente"]
                    # Regla automática para carga Gabo: el estado inicial queda como Emitido.
                    estado = "Emitido"
                    if asunto == "Notificación de Cédula de Resultados":
                        pdp_details_by_fuente: list[list[dict]] = []
                        if usa_fuentes_detalle:
                            first_row = fuentes_detalle_rows[0]
                            tipo_auditoria = str(first_row["tipo_auditoria"])
                            modalidad = first_row.get("modalidad") or "Fuente"
                            convenio_nombre = first_row.get("convenio_nombre") or ""
                            convenio_ente_nombre = first_row.get("convenio_ente_nombre") or ""
                            convenio_ente_id = first_row.get("convenio_ente_id") or ""
                            if modalidad == "Convenio" and not convenio_ente_id:
                                convenio_ente_id = resolve_convenio_ente_id(
                                    db,
                                    ejercicio,
                                    convenio_ente_nombre,
                                )
                            form_data["manual_tipo_auditoria"] = tipo_auditoria
                            periodo_fuente = " ".join((first_row.get("periodo") or "").split())
                            if not periodo_fuente:
                                raise ValueError("Cada fuente debe incluir periodo.")
                            periodo = periodo_fuente
                            form_data["manual_periodo"] = periodo_fuente
                            periodo_titular = periodo_fuente
                            cantidad_sa = int(first_row["cantidad_sa"])
                            cantidad_pdp = int(first_row["cantidad_pdp"])
                            cantidad_pras = int(first_row["cantidad_pras"])
                            cantidad_pefcf = int(first_row["cantidad_pefcf"])
                            cantidad_r = int(first_row["cantidad_r"])
                        else:
                            cantidad_sa = parse_non_negative_int(form_data["manual_cantidad_sa"], "Cantidad SA")
                            cantidad_pdp = parse_non_negative_int(form_data["manual_cantidad_pdp"], "Cantidad PDP")
                            cantidad_pras = parse_non_negative_int(form_data["manual_cantidad_pras"], "Cantidad PRAS")
                            cantidad_pefcf = parse_non_negative_int(form_data["manual_cantidad_pefcf"], "Cantidad PEFCF")
                            cantidad_r = parse_non_negative_int(form_data["manual_cantidad_r"], "Cantidad R")
                        monto_pdp_emitido = parse_non_negative_float(form_data["manual_monto_pdp_emitido"], "Monto PDP emitido")
                        monto_pdp_solventado = parse_non_negative_float(form_data["manual_monto_pdp_solventado"], "Monto PDP solventado")
                        monto_pdp_pendiente = parse_non_negative_float(form_data["manual_monto_pdp_pendiente"], "Monto PDP pendiente")
                        if usa_fuentes_detalle:
                            total_pdp_fuentes = sum(
                                int(row["cantidad_pdp"]) for row in fuentes_detalle_rows
                            )
                            pdp_details = parse_manual_pdp_details(
                                raw_pdp_detalle_json, total_pdp_fuentes
                            )
                            pdp_amounts = [
                                float(item.get("monto") or 0.0) for item in pdp_details
                            ]
                            pdp_offset = 0
                            for fuente_row in fuentes_detalle_rows:
                                cantidad_row = int(fuente_row["cantidad_pdp"])
                                next_offset = pdp_offset + cantidad_row
                                pdp_details_by_fuente.append(pdp_details[pdp_offset:next_offset])
                                pdp_offset = next_offset
                        else:
                            pdp_amounts = parse_pdp_amounts(raw_montos_pdp)
                            pdp_details = parse_manual_pdp_details(
                                raw_pdp_detalle_json, cantidad_pdp
                            )
                        solventacion_totales_by_anexo = {}
                        if not usa_fuentes_detalle and cantidad_pdp == 0 and pdp_amounts:
                            raise ValueError("Capturaste montos PDP pero la cantidad PDP es 0.")
                        if not usa_fuentes_detalle and pdp_amounts and len(pdp_amounts) != cantidad_pdp:
                            raise ValueError(
                                "La cantidad de montos PDP no coincide con 'Cantidad PDP'. "
                                "Captura un monto por línea."
                            )
                        if not usa_fuentes_detalle and not pdp_amounts and cantidad_pdp > 0:
                            # Sin detalle por observación: concentrar el total en la primera PDP.
                            pdp_amounts = [monto_pdp_emitido] + [0.0] * (cantidad_pdp - 1)
                        if not usa_fuentes_detalle and cantidad_pdp > 0 and pdp_details:
                            has_detail_amount = any((item.get("monto") is not None) for item in pdp_details)
                            if has_detail_amount:
                                pdp_amounts = [(item.get("monto") or 0.0) for item in pdp_details]
                    elif asunto == SOLVENTACION_ASUNTO:
                        if not usa_fuentes_detalle:
                            raise ValueError("Debes procesar un PDF de resultados de solventación antes de guardar.")
                        first_row = fuentes_detalle_rows[0]
                        tipo_auditoria = str(first_row["tipo_auditoria"])
                        modalidad = first_row.get("modalidad") or "Fuente"
                        convenio_nombre = first_row.get("convenio_nombre") or ""
                        convenio_ente_nombre = first_row.get("convenio_ente_nombre") or ""
                        convenio_ente_id = first_row.get("convenio_ente_id") or ""
                        if modalidad == "Convenio" and not convenio_ente_id:
                            convenio_ente_id = resolve_convenio_ente_id(
                                db,
                                ejercicio,
                                convenio_ente_nombre,
                            )
                        form_data["manual_tipo_auditoria"] = tipo_auditoria
                        periodo_fuente = " ".join((first_row.get("periodo") or "").split())
                        if not periodo_fuente:
                            raise ValueError("Cada fuente del oficio de solventación debe incluir periodo.")
                        periodo = periodo_fuente
                        form_data["manual_periodo"] = periodo_fuente
                        periodo_titular = periodo_fuente
                        cantidad_sa = int(first_row["cantidad_sa"])
                        cantidad_pdp = int(first_row["cantidad_pdp"])
                        cantidad_pras = int(first_row["cantidad_pras"])
                        cantidad_pefcf = int(first_row["cantidad_pefcf"])
                        cantidad_r = int(first_row["cantidad_r"])
                        monto_pdp_emitido = 0.0
                        monto_pdp_solventado = 0.0
                        monto_pdp_pendiente = 0.0
                        pdp_details = []
                        pdp_amounts = [0.0] * max(0, cantidad_pdp)
                        pdp_details_by_fuente = []
                        solventacion_totales_by_anexo = first_row.get("solventacion_totales_by_anexo") or {}
                        estado = "Pendiente"
                    fuente_detalle_snapshot = build_fuente_detalle_snapshot(
                        tipo_auditoria=tipo_auditoria,
                        fuente_nombre=fuente_nombre,
                        cantidad_sa=cantidad_sa,
                        cantidad_pdp=cantidad_pdp,
                        cantidad_pras=cantidad_pras,
                        cantidad_pefcf=cantidad_pefcf,
                        cantidad_r=cantidad_r,
                        modalidad=modalidad,
                        convenio_nombre=convenio_nombre,
                        convenio_ente_nombre=convenio_ente_nombre,
                        convenio_ente_id=convenio_ente_id,
                        ramo_33=ramo_33,
                        ramo_28=ramo_28,
                        origen_fuente=origen_fuente,
                        solventacion_totales_by_anexo=solventacion_totales_by_anexo,
                    )
                    fuente_detalle_json = serialize_manual_snapshot(fuente_detalle_snapshot)
                    pdp_detalle_json = serialize_manual_snapshot(pdp_details if cantidad_pdp > 0 else [])
                    tipos_auditoria = [tipo_auditoria]
                    if manual_edit_id and len(tipos_auditoria) > 1:
                        raise ValueError(
                            "Para editar una captura existente, selecciona solo 'Financiero'."
                        )
                    if manual_edit_id:
                        owned_row = db.execute(
                            """
                            SELECT id
                            FROM cargas_manuales
                            WHERE id = ? AND created_by = ?
                            LIMIT 1
                            """,
                            (manual_edit_id, user["username"]),
                        ).fetchone()
                        if not owned_row:
                            raise ValueError("La captura manual a editar no existe o no te pertenece.")
                    existing_rows = []
                    for tipo_item in tipos_auditoria:
                        if manual_edit_id:
                            found = db.execute(
                                """
                                SELECT id, created_at, created_by, tipo_auditoria
                                FROM cargas_manuales
                                WHERE id != ?
                                  AND LOWER(TRIM(numero_oficio)) = LOWER(TRIM(?))
                                  AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                                  AND asunto = ?
                                  AND ejercicio = ?
                                  AND fuente_id = ?
                                  AND TRIM(COALESCE(modalidad, 'Fuente')) = TRIM(COALESCE(?, 'Fuente'))
                                  AND LOWER(TRIM(COALESCE(convenio_nombre, ''))) = LOWER(TRIM(COALESCE(?, '')))
                                  AND LOWER(TRIM(periodo)) = LOWER(TRIM(?))
                                  AND tipo_auditoria = ?
                                ORDER BY id DESC
                                LIMIT 1
                                """,
                                (
                                    manual_edit_id,
                                    numero_oficio,
                                    manual_ente_id,
                                    asunto,
                                    ejercicio,
                                    fuente_id,
                                    modalidad,
                                    convenio_nombre,
                                    periodo,
                                    tipo_item,
                                ),
                            ).fetchone()
                        else:
                            found = db.execute(
                                """
                                SELECT id, created_at, created_by, tipo_auditoria
                                FROM cargas_manuales
                                WHERE LOWER(TRIM(numero_oficio)) = LOWER(TRIM(?))
                                  AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                                  AND asunto = ?
                                  AND ejercicio = ?
                                  AND fuente_id = ?
                                  AND TRIM(COALESCE(modalidad, 'Fuente')) = TRIM(COALESCE(?, 'Fuente'))
                                  AND LOWER(TRIM(COALESCE(convenio_nombre, ''))) = LOWER(TRIM(COALESCE(?, '')))
                                  AND LOWER(TRIM(periodo)) = LOWER(TRIM(?))
                                  AND tipo_auditoria = ?
                                ORDER BY id DESC
                                LIMIT 1
                                """,
                                (
                                    numero_oficio,
                                    manual_ente_id,
                                    asunto,
                                    ejercicio,
                                    fuente_id,
                                    modalidad,
                                    convenio_nombre,
                                    periodo,
                                    tipo_item,
                                ),
                            ).fetchone()
                        if found:
                            existing_rows.append(found)

                    if action == "manual_check":
                        if existing_rows:
                            resumen = summarize_existing_manual_rows(existing_rows)
                            manual_result = {
                                "ok": False,
                                "level": "info",
                                "message": f"Ya existe registro para: {resumen}.",
                            }
                        else:
                            manual_result = {
                                "ok": True,
                                "level": "success",
                                "message": "No se encontró registro previo con esta clave. Puedes guardar.",
                            }
                    else:
                        if manual_edit_id:
                            if existing_rows:
                                resumen = summarize_existing_manual_rows(existing_rows)
                                repair_result = repair_existing_manual_rows(db, existing_rows)
                                if repair_result["repaired"] > 0:
                                    db.commit()
                                manual_result = {
                                    "ok": False,
                                    "level": "info",
                                    "message": (
                                        f"No se puede actualizar por duplicado: {resumen}. "
                                        f"Se reparó la captura existente y se regeneraron "
                                        f"{repair_result['observaciones']} observaciones para visibilidad."
                                        if repair_result["repaired"] > 0
                                        else f"No se puede actualizar por duplicado: {resumen}."
                                    ),
                                }
                            else:
                                db.execute(
                                    """
                                    UPDATE cargas_manuales
                                    SET ente_id = ?,
                                        ente_nombre = ?,
                                        tipo_auditoria = ?,
                                        tipo_responsable = ?,
                                        titular_nombre = ?,
                                        administrativo_nombre = ?,
                                        numero_oficio = ?,
                                        asunto = ?,
                                        ejercicio = ?,
                                        fuente_id = ?,
                                        fuente_nombre = ?,
                                        modalidad = ?,
                                        convenio_nombre = ?,
                                        convenio_ente_nombre = ?,
                                        convenio_ente_id = ?,
                                        periodo = ?,
                                        periodo_titular = ?,
                                        fecha_notificacion = ?,
                                        ramo_33 = ?,
                                        ramo_28 = ?,
                                        origen_fuente = ?,
                                        estado = ?,
                                        cantidad_sa = ?,
                                        cantidad_pdp = ?,
                                        cantidad_pras = ?,
                                        cantidad_pefcf = ?,
                                        cantidad_r = ?,
                                        monto_pdp_emitido = ?,
                                        monto_pdp_solventado = ?,
                                        monto_pdp_pendiente = ?,
                                        fuente_detalle_json = ?,
                                        pdp_detalle_json = ?,
                                        created_at = ?
                                    WHERE id = ? AND created_by = ?
                                    """,
                                    (
                                        manual_ente_id,
                                        (ente_row["ente_nombre"] or "").strip(),
                                        tipos_auditoria[0],
                                        tipo_responsable,
                                        titular_nombre or None,
                                        administrativo_nombre or None,
                                        numero_oficio,
                                        asunto,
                                        ejercicio,
                                        fuente_id,
                                        fuente_nombre,
                                        modalidad,
                                        convenio_nombre,
                                        convenio_ente_nombre,
                                        convenio_ente_id,
                                        periodo,
                                        periodo_titular,
                                        fecha_notificacion,
                                        ramo_33,
                                        ramo_28,
                                        origen_fuente,
                                        estado,
                                        cantidad_sa,
                                        cantidad_pdp,
                                        cantidad_pras,
                                        cantidad_pefcf,
                                        cantidad_r,
                                        monto_pdp_emitido,
                                        monto_pdp_solventado,
                                        monto_pdp_pendiente,
                                        fuente_detalle_json,
                                        pdp_detalle_json,
                                        datetime.now().strftime("%Y-%m-%d %H:%M"),
                                        manual_edit_id,
                                        user["username"],
                                    ),
                                )
                                register_fuente_for_ente(
                                    db,
                                    ejercicio=ejercicio,
                                    ente_id_norm=manual_ente_id,
                                    fuente_id=fuente_id,
                                    tipo_auditoria=tipos_auditoria[0],
                                    created_by=user["username"],
                                )
                                materialize_observaciones_from_manual(
                                    db,
                                    ejercicio=ejercicio,
                                    ente_id=manual_ente_id,
                                    ente_numero=(ente_row["ente_numero"] or "").strip(),
                                    ente_nombre=(ente_row["ente_nombre"] or "").strip(),
                                    tipo_auditoria=tipos_auditoria[0],
                                    fuente_nombre=fuente_nombre,
                                    ramo_33=ramo_33,
                                    ramo_28=ramo_28,
                                    origen_fuente=origen_fuente,
                                    estado=estado,
                                    periodo_cedula=periodo,
                                    periodo_titular=periodo_titular,
                                    oficio=numero_oficio,
                                    fecha_notificacion=fecha_notificacion,
                                    cantidad_sa=cantidad_sa,
                                    cantidad_pdp=cantidad_pdp,
                                    cantidad_pras=cantidad_pras,
                                    cantidad_pefcf=cantidad_pefcf,
                                    cantidad_r=cantidad_r,
                                    monto_pdp_solventado=monto_pdp_solventado,
                                    monto_pdp_pendiente=monto_pdp_pendiente,
                                    pdp_amounts=pdp_amounts,
                                    modalidad=modalidad,
                                    convenio_nombre=convenio_nombre,
                                    convenio_ente_nombre=convenio_ente_nombre,
                                    convenio_ente_id=convenio_ente_id,
                                    pdp_details=pdp_details,
                                    solventacion_totales_by_anexo=solventacion_totales_by_anexo,
                                    replace_scope=True,
                                )
                                db.commit()
                                repair_result = repair_missing_observaciones_from_cargas(
                                    db, carga_ids=[manual_edit_id]
                                )
                                if repair_result["repaired"] > 0:
                                    db.commit()
                                form_data["manual_id"] = str(manual_edit_id)
                                manual_result = {
                                    "ok": True,
                                    "level": "success",
                                    "message": "Captura manual actualizada. Puedes seguir editando.",
                                }
                        else:
                            tipos_existentes = {row["tipo_auditoria"] for row in existing_rows}
                            tipos_por_insertar = [tipo for tipo in tipos_auditoria if tipo not in tipos_existentes]
                            if not tipos_por_insertar:
                                resumen = summarize_existing_manual_rows(existing_rows)
                                repair_result = repair_existing_manual_rows(db, existing_rows)
                                if repair_result["repaired"] > 0:
                                    db.commit()
                                    manual_result = {
                                        "ok": True,
                                        "level": "success",
                                        "message": (
                                            f"El registro ya existía para: {resumen}. "
                                            f"Se reparó la captura previa y se regeneraron "
                                            f"{repair_result['observaciones']} observaciones."
                                        ),
                                    }
                                else:
                                    manual_result = {
                                        "ok": False,
                                        "level": "info",
                                        "message": f"Registro duplicado detectado para: {resumen}.",
                                    }
                            else:
                                inserted_ids = []
                                for tipo_item in tipos_por_insertar:
                                    pdp_details_tipo_item = pdp_details
                                    pdp_amounts_tipo_item = pdp_amounts
                                    if usa_fuentes_detalle:
                                        pdp_details_tipo_item = (
                                            pdp_details_by_fuente[0] if pdp_details_by_fuente else []
                                        )
                                        pdp_amounts_tipo_item = [
                                            float(item.get("monto") or 0.0)
                                            for item in pdp_details_tipo_item
                                        ]
                                    register_fuente_for_ente(
                                        db,
                                        ejercicio=ejercicio,
                                        ente_id_norm=manual_ente_id,
                                        fuente_id=fuente_id,
                                        tipo_auditoria=tipo_item,
                                        created_by=user["username"],
                                    )
                                    cursor = db.execute(
                                        """
                                        INSERT INTO cargas_manuales (
                                            ente_id,
                                            ente_nombre,
                                            tipo_auditoria,
                                            tipo_responsable,
                                            titular_nombre,
                                            administrativo_nombre,
                                            numero_oficio,
                                            asunto,
                                            ejercicio,
                                            fuente_id,
                                            fuente_nombre,
                                            modalidad,
                                            convenio_nombre,
                                            convenio_ente_nombre,
                                            convenio_ente_id,
                                            periodo,
                                            periodo_titular,
                                            fecha_notificacion,
                                            ramo_33,
                                            ramo_28,
                                            origen_fuente,
                                            estado,
                                            cantidad_sa,
                                            cantidad_pdp,
                                            cantidad_pras,
                                            cantidad_pefcf,
                                            cantidad_r,
                                            monto_pdp_emitido,
                                            monto_pdp_solventado,
                                            monto_pdp_pendiente,
                                            fuente_detalle_json,
                                            pdp_detalle_json,
                                            created_by,
                                            created_at
                                        )
                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                        """,
                                        (
                                            manual_ente_id,
                                            (ente_row["ente_nombre"] or "").strip(),
                                            tipo_item,
                                            tipo_responsable,
                                            titular_nombre or None,
                                            administrativo_nombre or None,
                                            numero_oficio,
                                            asunto,
                                            ejercicio,
                                            fuente_id,
                                            fuente_nombre,
                                            modalidad,
                                            convenio_nombre,
                                            convenio_ente_nombre,
                                            convenio_ente_id,
                                            periodo,
                                            periodo_titular,
                                            fecha_notificacion,
                                            ramo_33,
                                            ramo_28,
                                            origen_fuente,
                                            estado,
                                            cantidad_sa,
                                            cantidad_pdp,
                                            cantidad_pras,
                                            cantidad_pefcf,
                                            cantidad_r,
                                            monto_pdp_emitido,
                                            monto_pdp_solventado,
                                            monto_pdp_pendiente,
                                            fuente_detalle_json,
                                            pdp_detalle_json,
                                            user["username"],
                                            datetime.now().strftime("%Y-%m-%d %H:%M"),
                                        ),
                                    )
                                    inserted_ids.append(int(cursor.lastrowid))
                                    materialize_observaciones_from_manual(
                                        db,
                                        ejercicio=ejercicio,
                                        ente_id=manual_ente_id,
                                        ente_numero=(ente_row["ente_numero"] or "").strip(),
                                        ente_nombre=(ente_row["ente_nombre"] or "").strip(),
                                        tipo_auditoria=tipo_item,
                                        fuente_nombre=fuente_nombre,
                                        ramo_33=ramo_33,
                                        ramo_28=ramo_28,
                                        origen_fuente=origen_fuente,
                                        estado=estado,
                                        periodo_cedula=periodo,
                                        periodo_titular=periodo_titular,
                                        oficio=numero_oficio,
                                        fecha_notificacion=fecha_notificacion,
                                        cantidad_sa=cantidad_sa,
                                        cantidad_pdp=cantidad_pdp,
                                        cantidad_pras=cantidad_pras,
                                        cantidad_pefcf=cantidad_pefcf,
                                        cantidad_r=cantidad_r,
                                        monto_pdp_solventado=monto_pdp_solventado,
                                        monto_pdp_pendiente=monto_pdp_pendiente,
                                        pdp_amounts=pdp_amounts_tipo_item,
                                        modalidad=modalidad,
                                        convenio_nombre=convenio_nombre,
                                        convenio_ente_nombre=convenio_ente_nombre,
                                        convenio_ente_id=convenio_ente_id,
                                        pdp_details=pdp_details_tipo_item,
                                        solventacion_totales_by_anexo=solventacion_totales_by_anexo,
                                        replace_scope=False,
                                    )
                                db.commit()
                                repair_result = repair_missing_observaciones_from_cargas(
                                    db, carga_ids=inserted_ids
                                )
                                if repair_result["repaired"] > 0:
                                    db.commit()
                                _save_summary = {
                                    "ente": (ente_row["ente_nombre"] or "").strip(),
                                    "oficio": numero_oficio,
                                    "ejercicio": ejercicio,
                                    "fuente": fuente_nombre,
                                    "periodo": periodo,
                                    "sa": cantidad_sa,
                                    "pdp": cantidad_pdp,
                                    "pras": cantidad_pras,
                                    "pefcf": cantidad_pefcf,
                                    "r": cantidad_r,
                                    "monto_emitido": monto_pdp_emitido,
                                    "monto_solventado": monto_pdp_solventado,
                                    "monto_pendiente": monto_pdp_pendiente,
                                }
                                if existing_rows:
                                    existentes = ", ".join(sorted(tipos_existentes))
                                    insertados = ", ".join(tipos_por_insertar)
                                    manual_result = {
                                        "ok": True,
                                        "level": "success",
                                        "message": (
                                            f"Registro manual guardado para: {insertados}. "
                                            f"Ya existían: {existentes}."
                                        ),
                                        "summary": _save_summary,
                                    }
                                    form_data["manual_id"] = ""
                                else:
                                    form_data["manual_id"] = str(inserted_ids[0]) if len(inserted_ids) == 1 else ""
                                    manual_result = {
                                        "ok": True,
                                        "level": "success",
                                        "message": (
                                            "Registro manual guardado correctamente."
                                            if len(inserted_ids) == 1
                                            else "Registros manuales guardados correctamente."
                                        ),
                                        "summary": _save_summary,
                                    }
                                if (
                                    action == "manual_save"
                                    and asunto in {"Notificación de Cédula de Resultados", SOLVENTACION_ASUNTO}
                                    and usa_fuentes_detalle
                                    and len(fuentes_detalle_rows) > 1
                                    and manual_result
                                    and manual_result.get("ok")
                                ):
                                    extra_inserted = 0
                                    extra_inserted_ids = []
                                    extra_skipped = 0
                                    for extra_idx, extra_row in enumerate(
                                        fuentes_detalle_rows[1:], start=1
                                    ):
                                        extra_fuente_nombre = normalize_fuente_financiamiento(
                                            " ".join(str(extra_row["fuente_nombre"]).split())
                                        )
                                        extra_periodo = " ".join(str(extra_row.get("periodo") or "").split())
                                        if not extra_fuente_nombre:
                                            continue
                                        if not extra_periodo:
                                            continue
                                        extra_fuente_db = db.execute(
                                            """
                                            SELECT id
                                            FROM fuentes_financiamiento
                                            WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(?))
                                            LIMIT 1
                                            """,
                                            (extra_fuente_nombre,),
                                        ).fetchone()
                                        if extra_fuente_db:
                                            extra_fuente_id = int(extra_fuente_db["id"])
                                        else:
                                            extra_fuente_id, extra_fuente_nombre = resolve_fuente_catalogo(
                                                db,
                                                extra_fuente_nombre,
                                                create_missing=True,
                                            )
                                            if extra_fuente_id is None:
                                                continue
                                        extra_tipo = extra_row["tipo_auditoria"]
                                        extra_modalidad = extra_row.get("modalidad") or "Fuente"
                                        extra_convenio_nombre = extra_row.get("convenio_nombre") or ""
                                        extra_convenio_ente_nombre = extra_row.get("convenio_ente_nombre") or ""
                                        extra_convenio_ente_id = extra_row.get("convenio_ente_id") or ""
                                        if extra_modalidad == "Convenio" and not extra_convenio_ente_id:
                                            extra_convenio_ente_id = resolve_convenio_ente_id(
                                                db,
                                                ejercicio,
                                                extra_convenio_ente_nombre,
                                            )
                                        extra_clasificacion = get_fuente_clasificacion(
                                            db,
                                            extra_fuente_nombre,
                                            extra_fuente_id,
                                        )
                                        extra_ramo_33 = extra_clasificacion["ramo_33"]
                                        extra_ramo_28 = extra_clasificacion["ramo_28"]
                                        extra_origen_fuente = extra_clasificacion["origen_fuente"]
                                        extra_tipos = [extra_tipo]
                                        for extra_tipo_item in extra_tipos:
                                            register_fuente_for_ente(
                                                db,
                                                ejercicio=ejercicio,
                                                ente_id_norm=manual_ente_id,
                                                fuente_id=extra_fuente_id,
                                                tipo_auditoria=extra_tipo_item,
                                                created_by=user["username"],
                                            )
                                            duplicate_extra = db.execute(
                                                """
                                                SELECT id
                                                FROM cargas_manuales
                                                WHERE LOWER(TRIM(numero_oficio)) = LOWER(TRIM(?))
                                                  AND TRIM(COALESCE(ente_id, '')) = TRIM(COALESCE(?, ''))
                                                  AND asunto = ?
                                                  AND ejercicio = ?
                                                  AND fuente_id = ?
                                                  AND TRIM(COALESCE(modalidad, 'Fuente')) = TRIM(COALESCE(?, 'Fuente'))
                                                  AND LOWER(TRIM(COALESCE(convenio_nombre, ''))) = LOWER(TRIM(COALESCE(?, '')))
                                                  AND LOWER(TRIM(periodo)) = LOWER(TRIM(?))
                                                  AND tipo_auditoria = ?
                                                LIMIT 1
                                                """,
                                                (
                                                    numero_oficio,
                                                    manual_ente_id,
                                                    asunto,
                                                    ejercicio,
                                                    extra_fuente_id,
                                                    extra_modalidad,
                                                    extra_convenio_nombre,
                                                    extra_periodo,
                                                    extra_tipo_item,
                                                ),
                                            ).fetchone()
                                            if duplicate_extra:
                                                extra_skipped += 1
                                                continue
                                            cantidad_sa_extra = int(extra_row["cantidad_sa"])
                                            cantidad_pdp_extra = int(extra_row["cantidad_pdp"])
                                            cantidad_pras_extra = int(extra_row["cantidad_pras"])
                                            cantidad_pefcf_extra = int(extra_row["cantidad_pefcf"])
                                            cantidad_r_extra = int(extra_row["cantidad_r"])
                                            if asunto == SOLVENTACION_ASUNTO:
                                                pdp_details_extra = []
                                                pdp_amounts_extra = [0.0] * max(0, cantidad_pdp_extra)
                                                extra_solventacion_totales = extra_row.get("solventacion_totales_by_anexo") or {}
                                                extra_estado = "Pendiente"
                                            else:
                                                pdp_details_extra = (
                                                    pdp_details_by_fuente[extra_idx]
                                                    if extra_idx < len(pdp_details_by_fuente)
                                                    else []
                                                )
                                                pdp_amounts_extra = [
                                                    float(item.get("monto") or 0.0)
                                                    for item in pdp_details_extra
                                                ]
                                                extra_solventacion_totales = {}
                                                extra_estado = estado
                                            extra_fuente_detalle_json = serialize_manual_snapshot(
                                                build_fuente_detalle_snapshot(
                                                    tipo_auditoria=extra_tipo_item,
                                                    fuente_nombre=extra_fuente_nombre,
                                                    cantidad_sa=cantidad_sa_extra,
                                                    cantidad_pdp=cantidad_pdp_extra,
                                                    cantidad_pras=cantidad_pras_extra,
                                                    cantidad_pefcf=cantidad_pefcf_extra,
                                                    cantidad_r=cantidad_r_extra,
                                                    modalidad=extra_modalidad,
                                                    convenio_nombre=extra_convenio_nombre,
                                                    convenio_ente_nombre=extra_convenio_ente_nombre,
                                                    convenio_ente_id=extra_convenio_ente_id,
                                                    ramo_33=extra_ramo_33,
                                                    ramo_28=extra_ramo_28,
                                                    origen_fuente=extra_origen_fuente,
                                                    solventacion_totales_by_anexo=extra_solventacion_totales,
                                                )
                                            )
                                            extra_pdp_detalle_json = serialize_manual_snapshot(
                                                pdp_details_extra if cantidad_pdp_extra > 0 else []
                                            )
                                            cursor_extra = db.execute(
                                                """
                                                INSERT INTO cargas_manuales (
                                                    ente_id,
                                                    ente_nombre,
                                                    tipo_auditoria,
                                                    tipo_responsable,
                                                    titular_nombre,
                                                    administrativo_nombre,
                                                    numero_oficio,
                                                    asunto,
                                                    ejercicio,
                                                    fuente_id,
                                                    fuente_nombre,
                                                    modalidad,
                                                    convenio_nombre,
                                                    convenio_ente_nombre,
                                                    convenio_ente_id,
                                                    periodo,
                                                    periodo_titular,
                                                    fecha_notificacion,
                                                    ramo_33,
                                                    ramo_28,
                                                    origen_fuente,
                                                    estado,
                                                    cantidad_sa,
                                                    cantidad_pdp,
                                                    cantidad_pras,
                                                    cantidad_pefcf,
                                                    cantidad_r,
                                                    monto_pdp_emitido,
                                                    monto_pdp_solventado,
                                                    monto_pdp_pendiente,
                                                    fuente_detalle_json,
                                                    pdp_detalle_json,
                                                    created_by,
                                                    created_at
                                                )
                                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                                """,
                                                (
                                                    manual_ente_id,
                                                    (ente_row["ente_nombre"] or "").strip(),
                                                    extra_tipo_item,
                                                    tipo_responsable,
                                                    titular_nombre or None,
                                                    administrativo_nombre or None,
                                                    numero_oficio,
                                                    asunto,
                                                    ejercicio,
                                                    extra_fuente_id,
                                                    extra_fuente_nombre,
                                                    extra_modalidad,
                                                    extra_convenio_nombre,
                                                    extra_convenio_ente_nombre,
                                                    extra_convenio_ente_id,
                                                    extra_periodo,
                                                    extra_periodo,
                                                    fecha_notificacion,
                                                    extra_ramo_33,
                                                    extra_ramo_28,
                                                    extra_origen_fuente,
                                                    extra_estado,
                                                    cantidad_sa_extra,
                                                    cantidad_pdp_extra,
                                                    cantidad_pras_extra,
                                                    cantidad_pefcf_extra,
                                                    cantidad_r_extra,
                                                    0.0,
                                                    0.0,
                                                    0.0,
                                                    extra_fuente_detalle_json,
                                                    extra_pdp_detalle_json,
                                                    user["username"],
                                                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                                                ),
                                            )
                                            extra_inserted_ids.append(int(cursor_extra.lastrowid))
                                            materialize_observaciones_from_manual(
                                                db,
                                                ejercicio=ejercicio,
                                                ente_id=manual_ente_id,
                                                ente_numero=(ente_row["ente_numero"] or "").strip(),
                                                ente_nombre=(ente_row["ente_nombre"] or "").strip(),
                                                tipo_auditoria=extra_tipo_item,
                                                fuente_nombre=extra_fuente_nombre,
                                                ramo_33=extra_ramo_33,
                                                ramo_28=extra_ramo_28,
                                                origen_fuente=extra_origen_fuente,
                                                estado=extra_estado,
                                                periodo_cedula=extra_periodo,
                                                periodo_titular=extra_periodo,
                                                oficio=numero_oficio,
                                                fecha_notificacion=fecha_notificacion,
                                                cantidad_sa=cantidad_sa_extra,
                                                cantidad_pdp=cantidad_pdp_extra,
                                                cantidad_pras=cantidad_pras_extra,
                                                cantidad_pefcf=cantidad_pefcf_extra,
                                                cantidad_r=cantidad_r_extra,
                                                monto_pdp_solventado=0.0,
                                                monto_pdp_pendiente=0.0,
                                                pdp_amounts=pdp_amounts_extra,
                                                modalidad=extra_modalidad,
                                                convenio_nombre=extra_convenio_nombre,
                                                convenio_ente_nombre=extra_convenio_ente_nombre,
                                                convenio_ente_id=extra_convenio_ente_id,
                                                pdp_details=pdp_details_extra,
                                                solventacion_totales_by_anexo=extra_solventacion_totales,
                                                replace_scope=False,
                                            )
                                            extra_inserted += 1
                                    if extra_inserted > 0:
                                        db.commit()
                                    extra_repair_ids = inserted_ids + extra_inserted_ids
                                    if extra_repair_ids:
                                        repair_result = repair_missing_observaciones_from_cargas(
                                            db, carga_ids=extra_repair_ids
                                        )
                                        if repair_result["repaired"] > 0:
                                            db.commit()
                                    if manual_result and manual_result.get("ok"):
                                        manual_result["message"] = (
                                            f"{manual_result['message']} "
                                            f"Fuentes adicionales: {extra_inserted} agregadas"
                                            + (f", {extra_skipped} omitidas por duplicado." if extra_skipped else ".")
                                        )
                    if action == "manual_save" and manual_result and manual_result.get("ok") and backup_path:
                        manual_result["message"] = (
                            f"{manual_result.get('message', '').strip()} Respaldo: {backup_path}."
                        ).strip()
                    if action == "manual_save" and manual_result and manual_result.get("ok"):
                        kept_ejercicio = form_data["manual_ejercicio"]
                        for key in list(form_data):
                            if key.startswith("manual_"):
                                form_data[key] = ""
                        form_data["manual_ejercicio"] = kept_ejercicio
                        form_data["manual_asunto"] = "Notificación de Cédula de Resultados"
                        form_data["manual_tipo_responsable"] = "Titular"
                        form_data["manual_ramo_33"] = "No"
                        form_data["manual_ramo_28"] = "No"
                        form_data["manual_estado"] = "Emitido"
                        form_data["manual_cantidad_sa"] = "0"
                        form_data["manual_cantidad_pdp"] = "0"
                        form_data["manual_cantidad_pras"] = "0"
                        form_data["manual_cantidad_pefcf"] = "0"
                        form_data["manual_cantidad_r"] = "0"
                        form_data["manual_monto_pdp_emitido"] = "0"
                        form_data["manual_monto_pdp_solventado"] = "0"
                        form_data["manual_monto_pdp_pendiente"] = "0"
                else:
                    command = [sys.executable]
                    if action == "template_generate":
                        if not form_data["template_ejercicio"]:
                            raise ValueError("Debes indicar el ejercicio para generar plantilla.")
                        out_path = resolve_project_path(form_data["template_out"], must_exist=False)
                        out_path.parent.mkdir(parents=True, exist_ok=True)
                        command.extend(
                            [
                                "scripts/make_historial_titulares_template.py",
                                "--db",
                                DB_PATH,
                                "--ejercicio",
                                str(int(form_data["template_ejercicio"])),
                                "--tipo-auditoria",
                                form_data["template_tipo_auditoria"],
                                "--out",
                                str(out_path),
                            ]
                        )
                    elif action in {"csv_validate", "csv_import"}:
                        if action == "csv_import":
                            if not form_data["csv_ejercicio"]:
                                raise ValueError(
                                    "Debes indicar un ejercicio editable para importar titulares por CSV."
                                )
                            _ensure_editable_ejercicio(form_data["csv_ejercicio"], user=user)
                        csv_path = resolve_project_path(form_data["csv_path"], must_exist=True)
                        command.extend(
                            [
                                "scripts/import_historial_titulares.py",
                                "--db",
                                DB_PATH,
                                "--csv",
                                str(csv_path),
                            ]
                        )
                        if form_data["csv_ejercicio"]:
                            command.extend(["--ejercicio", str(int(form_data["csv_ejercicio"]))])
                        if form_data["csv_replace"] == "1":
                            command.append("--replace")
                        if action == "csv_validate":
                            command.append("--dry-run")
                    elif action in {"json_validate", "json_import"}:
                        if action == "json_import":
                            if not form_data["json_ejercicio"]:
                                raise ValueError(
                                    "Debes indicar un ejercicio editable para importar titulares por JSON."
                                )
                            _ensure_editable_ejercicio(form_data["json_ejercicio"], user=user)
                        json_path = resolve_project_path(form_data["json_path"], must_exist=True)
                        command.extend(
                            [
                                "scripts/import_historial_titulares_json.py",
                                "--db",
                                DB_PATH,
                                "--json",
                                str(json_path),
                                "--tipo-auditoria",
                                form_data["json_tipo_auditoria"],
                            ]
                        )
                        if form_data["json_ejercicio"]:
                            command.extend(["--ejercicio", str(int(form_data["json_ejercicio"]))])
                        if action == "json_validate":
                            command.append("--dry-run")
                    else:
                        raise ValueError("Acción de carga no soportada.")
                    script_result = run_loader_command(command)
            except ValueError as exc:
                try:
                    db.rollback()
                except Exception:
                    pass
                if action == "manual_save":
                    manual_result = {
                        "ok": False,
                        "level": "error",
                        "message": str(exc),
                    }
                elif action == "manual_check":
                    manual_result = {
                        "ok": False,
                        "level": "error",
                        "message": str(exc),
                    }
                elif action == "titular_save":
                    titular_result = {
                        "ok": False,
                        "level": "error",
                        "message": str(exc),
                    }
                else:
                    script_result = {
                        "ok": False,
                        "returncode": 1,
                        "command": "",
                        "stdout": "",
                        "stderr": str(exc),
                    }
            except Exception as exc:
                try:
                    db.rollback()
                except Exception:
                    pass
                error_message = f"Ocurrió un error inesperado: {exc}"
                if action == "manual_save":
                    manual_result = {
                        "ok": False,
                        "level": "error",
                        "message": error_message,
                    }
                elif action == "manual_check":
                    manual_result = {
                        "ok": False,
                        "level": "error",
                        "message": error_message,
                    }
                elif action == "titular_save":
                    titular_result = {
                        "ok": False,
                        "level": "error",
                        "message": error_message,
                    }
                else:
                    script_result = {
                        "ok": False,
                        "returncode": 1,
                        "command": "",
                        "stdout": "",
                        "stderr": error_message,
                    }

        if action == "titular_save" or titular_result is not None:
            initial_loader_mode = "titulares"
        elif action in {"manual_save", "manual_check"} or manual_result is not None:
            initial_loader_mode = "manual"
    
        titular_entes_rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY {ente_numero_sort_sql('ente_numero')}, ente_numero, ente_nombre
            """,
            (form_data["titular_ejercicio"],),
        ).fetchall()

        fuentes_rows = db.execute(
            """
            SELECT id, nombre
            FROM fuentes_financiamiento
            ORDER BY nombre ASC
            """
        ).fetchall()
        fuentes = [dict(row) for row in fuentes_rows]
        manual_ente_id_norm = normalize_ente_id(form_data["manual_ente_id"])
        manual_entes_rows = db.execute(
            f"""
            SELECT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM entes_detalle
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY {ente_numero_sort_sql('ente_numero')}, ente_numero, ente_nombre
            """,
            (form_data["manual_ejercicio"],),
        ).fetchall()
        if manual_ente_id_norm:
            manual_fuentes = fuentes_por_ente(
                db,
                form_data["manual_ejercicio"],
                manual_ente_id_norm,
                tipo_auditoria=form_data["manual_tipo_auditoria"],
            )
        else:
            manual_fuentes = []
    
        return render_template(
            "carga.html",
            user=user,
            result=script_result,
            manual_result=manual_result,
            titular_result=titular_result,
            initial_loader_mode=initial_loader_mode,
            form_data=form_data,
            fuentes=fuentes,
            manual_fuentes=manual_fuentes,
            manual_ejercicios=manual_ejercicios,
            titular_ejercicios=titular_ejercicios,
            read_only_ejercicios=sorted(_readonly_ejercicios_for_user(user)),
            titular_entes=[dict(row) for row in titular_entes_rows],
            manual_entes=[dict(row) for row in manual_entes_rows],
            asuntos=[
                "Notificación de Cédula de Resultados",
                SOLVENTACION_ASUNTO,
            ],
            tipos_responsable=["Titular", "Administrativo", "Ambos"],
        )
    
