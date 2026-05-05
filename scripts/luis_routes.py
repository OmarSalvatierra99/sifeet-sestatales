from datetime import datetime, timedelta
from io import BytesIO
import time

from flask import jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scripts.financiamiento import normalize_origen_fuente, origen_fuente_sql


def register_luis_routes(app, deps):
    globals().update(deps)
    dashboard_cache: dict[str, tuple[float, dict]] = {}
    dashboard_cache_ttl_seconds = 45
    comparison_cache: dict[str, tuple[float, dict]] = {}
    comparison_cache_ttl_seconds = 45
    filter_order = (
        "ente_id",
        "tipo_auditoria",
        "tipo_anexo",
        "estado",
        "fuente_financiamiento",
        "modalidad",
        "convenio_ente_nombre",
        "origen_fuente",
        "ramo_33",
        "ramo_28",
        "concepto_irregularidad",
        "periodo_cedula",
    )
    filter_labels = {
        "ente_id": "Ente",
        "tipo_auditoria": "Tipo de auditoria",
        "tipo_anexo": "Tipo de anexo",
        "estado": "Estado",
        "fuente_financiamiento": "Fuente de Financiamiento",
        "modalidad": "Modalidad",
        "convenio_ente_nombre": "Ente hijo / convenio",
        "origen_fuente": "Del Ejercicio / Remanentes",
        "ramo_33": "Ramo 33",
        "ramo_28": "Ramo 28",
        "concepto_irregularidad": "Concepto de irregularidad",
        "periodo_cedula": "Cedula de resultados",
    }
    comparison_filter_order = (
        "ente_uid",
        "tipo_auditoria",
        "tipo_anexo",
        "estado",
        "fuente_financiamiento",
        "origen_fuente",
        "ramo_33",
        "ramo_28",
    )
    comparison_filter_labels = {
        "ente_uid": "Ente",
        "tipo_auditoria": "Tipo de auditoria",
        "tipo_anexo": "Tipo de anexo",
        "estado": "Estado",
        "fuente_financiamiento": "Fuente de Financiamiento",
        "origen_fuente": "Del Ejercicio / Remanentes",
        "ramo_33": "Ramo 33",
        "ramo_28": "Ramo 28",
    }

    def parse_multi_values(param_name: str, normalizer=None):
        values: list[str] = []
        seen = set()
        for raw in request.args.getlist(param_name):
            clean = (raw or "").strip()
            if not clean:
                continue
            if normalizer:
                clean = normalizer(clean)
            clean = (clean or "").strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            values.append(clean)
        return values

    def parse_selected_filters():
        return {
            "ente_id": parse_multi_values("ente_id", normalize_ente_id),
            "tipo_auditoria": parse_multi_values("tipo_auditoria", normalize_tipo_auditoria),
            "tipo_anexo": parse_multi_values("tipo_anexo"),
            "estado": parse_multi_values("estado"),
            "fuente_financiamiento": parse_multi_values("fuente_financiamiento"),
            "modalidad": parse_multi_values("modalidad"),
            "convenio_ente_nombre": parse_multi_values("convenio_ente_nombre"),
            "origen_fuente": parse_multi_values("origen_fuente", normalize_origen_fuente),
            "ramo_33": parse_multi_values("ramo_33"),
            "ramo_28": parse_multi_values("ramo_28"),
            "concepto_irregularidad": parse_multi_values("concepto_irregularidad"),
            "periodo_cedula": parse_multi_values("periodo_cedula"),
        }

    def column_sql(column: str, alias: str = "") -> str:
        return f"{alias}.{column}" if alias else column

    def apply_filter_clause(
        clauses: list[str],
        params: list,
        key: str,
        values: list[str],
        alias: str = "",
    ):
        if not values:
            return
        placeholders = ", ".join(["?"] * len(values))
        if key == "ente_id":
            clauses.append(f"{normalize_ente_id_sql(column_sql('ente_id', alias))} IN ({placeholders})")
            params.extend(values)
            return
        if key == "concepto_irregularidad":
            concepto_col = column_sql("pdp_concepto_irregularidad", alias)
            subconcepto_col = column_sql("pdp_subconcepto_irregularidad", alias)
            clauses.append(
                f"({concepto_col} IN ({placeholders}) OR {subconcepto_col} IN ({placeholders}))"
            )
            params.extend(values)
            params.extend(values)
            return
        if key == "origen_fuente":
            clauses.append(f"{origen_fuente_sql(alias)} IN ({placeholders})")
            params.extend(values)
            return
        key_to_column = {
            "tipo_auditoria": "tipo_auditoria",
            "tipo_anexo": "tipo_anexo",
            "estado": "estado",
            "fuente_financiamiento": "fuente_financiamiento",
            "modalidad": "modalidad",
            "convenio_ente_nombre": "convenio_ente_nombre",
            "ramo_33": "ramo_33",
            "ramo_28": "ramo_28",
            "periodo_cedula": "periodo_cedula",
        }
        column = key_to_column.get(key)
        if not column:
            return
        clauses.append(f"{column_sql(column, alias)} IN ({placeholders})")
        params.extend(values)

    def build_observaciones_scope(
        ejercicio: str,
        selected_filters: dict[str, list[str]],
        *,
        exclude_key: str = "",
        include_ente: bool = True,
        alias: str = "",
    ):
        clauses = [f"{column_sql('ejercicio', alias)} = ?"]
        params: list = [ejercicio]
        for key in filter_order:
            if key == exclude_key:
                continue
            if key == "ente_id" and not include_ente:
                continue
            apply_filter_clause(clauses, params, key, selected_filters.get(key, []), alias)
        return " AND ".join(clauses), params

    def selected_values_for_key(selected_filters: dict[str, list[str]], key: str):
        return selected_filters.get(key, []) or []

    def format_periodo_display(start_date, end_date):
        if not start_date or not end_date:
            return "—"
        months = (
            "",
            "enero",
            "febrero",
            "marzo",
            "abril",
            "mayo",
            "junio",
            "julio",
            "agosto",
            "septiembre",
            "octubre",
            "noviembre",
            "diciembre",
        )
        return (
            f"{start_date.day:02d} de {months[start_date.month]} al "
            f"{end_date.day:02d} de {months[end_date.month]}"
        )

    def merge_responsable_periods(historial_rows):
        normalized_rows = []
        for item in historial_rows:
            nombre = (item["nombre"] or "").strip()
            tipo_registro = (item["tipo_registro"] or "").strip()
            inicio = parse_historial_date(item["fecha_inicio"])
            fin = parse_historial_date(item["fecha_fin"])
            if not nombre or not tipo_registro or not inicio or not fin:
                continue
            normalized_rows.append(
                {
                    "tipo_registro": tipo_registro,
                    "nombre": nombre,
                    "inicio": inicio,
                    "fin": fin,
                }
            )

        normalized_rows.sort(
            key=lambda item: (
                item["tipo_registro"],
                item["nombre"],
                item["inicio"],
                item["fin"],
            )
        )

        merged = []
        for item in normalized_rows:
            if (
                merged
                and merged[-1]["tipo_registro"] == item["tipo_registro"]
                and merged[-1]["nombre"] == item["nombre"]
                and item["inicio"] <= (merged[-1]["fin"] + timedelta(days=1))
            ):
                if item["fin"] > merged[-1]["fin"]:
                    merged[-1]["fin"] = item["fin"]
                continue
            merged.append(item.copy())

        merged.sort(
            key=lambda item: (
                item["tipo_registro"],
                item["inicio"],
                item["fin"],
                item["nombre"],
            )
        )
        return merged

    anexos_orden = ("R", "SA", "PDP", "PRAS", "PEFCF")
    comparison_chart_anexos_orden = ("SA", "PDP", "PRAS", "PEFCF", "R")
    comparison_chart_anexos_rank = {
        anexo: idx for idx, anexo in enumerate(comparison_chart_anexos_orden)
    }
    anexos_alias = {
        "PEFCT": "PEFCF",
        "PEFCE": "PEFCF",
    }

    def normalize_anexo_bucket(value: str) -> str:
        clean = (value or "").strip().upper()
        if clean in anexos_alias:
            return anexos_alias[clean]
        return clean

    def build_status_metrics():
        return {
            **{anexo: 0 for anexo in anexos_orden},
            "total": 0,
            "monto_dano": 0.0,
        }

    def append_status_metric(metrics: dict, anexo: str, monto: float = 0.0):
        if anexo not in anexos_orden:
            return
        metrics[anexo] += 1
        metrics["total"] += 1
        metrics["monto_dano"] += float(monto or 0)

    def get_available_comparison_years(db) -> list[str]:
        rows = db.execute(
            """
            SELECT DISTINCT TRIM(COALESCE(ejercicio, '')) AS ejercicio
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) != ''
            ORDER BY ejercicio
            """
        ).fetchall()
        return [row["ejercicio"] for row in rows if (row["ejercicio"] or "").strip()]

    def parse_comparison_years(available_years: list[str]) -> list[str]:
        requested = []
        seen = set()
        for param_name in ("anios", "ejercicio"):
            for raw in request.args.getlist(param_name):
                clean = (raw or "").strip()
                if not clean or clean in seen:
                    continue
                seen.add(clean)
                requested.append(clean)
        valid_years = [year for year in requested if year in available_years]
        if valid_years:
            valid_set = set(valid_years)
            return [year for year in available_years if year in valid_set]
        if not available_years:
            return []
        default_size = 2 if len(available_years) >= 2 else 1
        return available_years[:default_size]

    def parse_comparison_filters():
        return {
            "ente_uid": parse_multi_values("ente_uid"),
            "tipo_auditoria": parse_multi_values("tipo_auditoria", normalize_tipo_auditoria),
            "tipo_anexo": parse_multi_values("tipo_anexo"),
            "estado": parse_multi_values("estado"),
            "fuente_financiamiento": parse_multi_values("fuente_financiamiento"),
            "origen_fuente": parse_multi_values("origen_fuente", normalize_origen_fuente),
            "ramo_33": parse_multi_values("ramo_33"),
            "ramo_28": parse_multi_values("ramo_28"),
            "universo": (
                "complete"
                if (request.args.get("universo", "").strip().lower() in {"common", "complete", "presentes"})
                else "all"
            ),
        }

    def build_comparison_base_sql(selected_years: list[str]):
        if not selected_years:
            return (
                """
                SELECT
                    '' AS ejercicio,
                    '' AS ente_uid,
                    '' AS ente_numero,
                    '' AS ente_nombre,
                    '' AS tipo_auditoria,
                    '' AS tipo_anexo,
                    '' AS estado,
                    '' AS fuente_financiamiento,
                    '' AS origen_fuente,
                    '' AS ramo_33,
                    '' AS ramo_28,
                    0 AS monto_pdp_emitido,
                    0 AS monto_pdp_solventado,
                    0 AS monto_pdp_pendiente
                WHERE 1 = 0
                """,
                [],
            )

        placeholders = ", ".join(["?"] * len(selected_years))
        return (
            f"""
            SELECT
                TRIM(COALESCE(o.ejercicio, '')) AS ejercicio,
                COALESCE(
                    NULLIF(TRIM(COALESCE(ed.ente_uid, '')), ''),
                    '__NOUID__:' || TRIM(COALESCE(o.ente_id, ''))
                ) AS ente_uid,
                TRIM(
                    COALESCE(
                        NULLIF(TRIM(COALESCE(ed.ente_numero, '')), ''),
                        NULLIF(TRIM(COALESCE(o.ente_numero, '')), ''),
                        ''
                    )
                ) AS ente_numero,
                TRIM(
                    COALESCE(
                        NULLIF(TRIM(COALESCE(ed.ente_nombre, '')), ''),
                        NULLIF(TRIM(COALESCE(o.ente_nombre, '')), ''),
                        NULLIF(TRIM(COALESCE(o.ente_id, '')), ''),
                        'Sin ente'
                    )
                ) AS ente_nombre,
                TRIM(COALESCE(o.tipo_auditoria, '')) AS tipo_auditoria,
                TRIM(COALESCE(o.tipo_anexo, '')) AS tipo_anexo,
                TRIM(COALESCE(o.estado, '')) AS estado,
                TRIM(COALESCE(o.fuente_financiamiento, '')) AS fuente_financiamiento,
                {origen_fuente_sql("o")} AS origen_fuente,
                TRIM(COALESCE(o.ramo_33, '')) AS ramo_33,
                TRIM(COALESCE(o.ramo_28, '')) AS ramo_28,
                COALESCE(o.monto_pdp_emitido, 0) AS monto_pdp_emitido,
                COALESCE(o.monto_pdp_solventado, 0) AS monto_pdp_solventado,
                COALESCE(o.monto_pdp_pendiente, 0) AS monto_pdp_pendiente
            FROM observaciones AS o
            LEFT JOIN entes_detalle AS ed
              ON TRIM(COALESCE(ed.ejercicio, '')) = TRIM(COALESCE(o.ejercicio, ''))
             AND ed.ente_id = o.ente_id
            WHERE TRIM(COALESCE(o.ejercicio, '')) IN ({placeholders})
            """,
            list(selected_years),
        )

    def build_comparison_scope_cache_key(
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
    ):
        return (
            exclude_key,
            selected_filters.get("universo", "all"),
            tuple(selected_filters.get("ente_uid", []) or []),
            tuple(selected_filters.get("tipo_auditoria", []) or []),
            tuple(selected_filters.get("tipo_anexo", []) or []),
            tuple(selected_filters.get("estado", []) or []),
            tuple(selected_filters.get("fuente_financiamiento", []) or []),
            tuple(selected_filters.get("origen_fuente", []) or []),
            tuple(selected_filters.get("ramo_33", []) or []),
            tuple(selected_filters.get("ramo_28", []) or []),
        )

    def fetch_comparison_base_rows(db, selected_years: list[str]) -> list[dict]:
        if not selected_years:
            return []
        base_sql, base_params = build_comparison_base_sql(selected_years)
        rows = db.execute(base_sql, base_params).fetchall()
        return [dict(row) for row in rows]

    def row_matches_comparison_filters(
        row: dict,
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
    ) -> bool:
        for key in comparison_filter_order:
            if key == exclude_key:
                continue
            values = selected_filters.get(key, [])
            if not isinstance(values, list) or not values:
                continue
            if (row.get(key) or "") not in values:
                return False
        return True

    def compute_common_comparison_entity_uids(
        rows: list[dict],
        selected_years: list[str],
    ) -> set[str]:
        if not selected_years:
            return set()
        required_years = set(selected_years)
        years_by_uid: dict[str, set[str]] = {}
        for row in rows:
            ente_uid = (row.get("ente_uid") or "").strip()
            ejercicio = (row.get("ejercicio") or "").strip()
            if not ente_uid or not ejercicio:
                continue
            if ejercicio not in required_years:
                continue
            years_by_uid.setdefault(ente_uid, set()).add(ejercicio)
        return {
            ente_uid
            for ente_uid, years_present in years_by_uid.items()
            if years_present == required_years
        }

    def build_comparison_scope_from_rows(
        base_rows: list[dict],
        selected_years: list[str],
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
        scope_cache: dict | None = None,
    ) -> dict:
        cache_key = build_comparison_scope_cache_key(
            selected_filters,
            exclude_key=exclude_key,
        )
        if scope_cache is not None and cache_key in scope_cache:
            return scope_cache[cache_key]

        filtered_rows = [
            row
            for row in base_rows
            if row_matches_comparison_filters(
                row,
                selected_filters,
                exclude_key=exclude_key,
            )
        ]

        common_uids: set[str] = set()
        if selected_filters.get("universo") == "complete" and selected_years:
            common_uids = compute_common_comparison_entity_uids(filtered_rows, selected_years)
            if common_uids:
                filtered_rows = [
                    row
                    for row in filtered_rows
                    if (row.get("ente_uid") or "").strip() in common_uids
                ]
            else:
                filtered_rows = []

        payload = {
            "rows": filtered_rows,
            "common_uids": common_uids,
        }
        if scope_cache is not None:
            scope_cache[cache_key] = payload
        return payload

    def collect_comparison_distinct_values(rows: list[dict], column: str) -> list[str]:
        values = sorted(
            {
                (row.get(column) or "").strip()
                for row in rows
                if (row.get(column) or "").strip()
            }
        )
        return values

    def aggregate_comparison_rows_by_entity(scope_rows: list[dict]) -> list[dict]:
        grouped: dict[tuple[str, str], dict] = {}
        for row in scope_rows:
            ente_uid = (row.get("ente_uid") or "").strip()
            ejercicio = (row.get("ejercicio") or "").strip()
            if not ente_uid or not ejercicio:
                continue
            key = (ente_uid, ejercicio)
            item = grouped.setdefault(
                key,
                {
                    "ente_uid": ente_uid,
                    "ejercicio": ejercicio,
                    "ente_numero": "",
                    "ente_nombre": "",
                    "total_observaciones": 0,
                },
            )
            ente_numero = (row.get("ente_numero") or "").strip()
            ente_nombre = (row.get("ente_nombre") or "").strip()
            if ente_numero and not item["ente_numero"]:
                item["ente_numero"] = ente_numero
            if ente_nombre:
                item["ente_nombre"] = ente_nombre
            item["total_observaciones"] += 1
        return list(grouped.values())

    def summarize_comparison_scope(scope_rows: list[dict], selected_years: list[str]) -> dict:
        year_rank = {year: idx for idx, year in enumerate(selected_years)}
        metrics_by_year = {
            year: {
                "emitidas": 0,
                "solventadas": 0,
                "pendientes": 0,
                "monto_pdp_emitido": 0.0,
                "monto_pdp_solventado": 0.0,
                "monto_pdp_pendiente": 0.0,
            }
            for year in selected_years
        }
        status_totals: dict[tuple[str, str], int] = {}
        anexo_totals: dict[tuple[str, str], int] = {}
        stacked_by_anexo: dict[tuple[str, str], dict] = {}

        for row in scope_rows:
            ejercicio = (row.get("ejercicio") or "").strip()
            if ejercicio not in metrics_by_year:
                continue

            metrics = metrics_by_year[ejercicio]
            metrics["emitidas"] += 1
            estado = (row.get("estado") or "").strip()
            estado_key = estado.lower()
            if estado_key == "solventado":
                metrics["solventadas"] += 1
            elif estado_key == "pendiente":
                metrics["pendientes"] += 1

            metrics["monto_pdp_emitido"] += float(row.get("monto_pdp_emitido") or 0)
            metrics["monto_pdp_solventado"] += float(row.get("monto_pdp_solventado") or 0)
            metrics["monto_pdp_pendiente"] += float(row.get("monto_pdp_pendiente") or 0)

            if estado:
                status_key = (ejercicio, estado)
                status_totals[status_key] = status_totals.get(status_key, 0) + 1

            tipo_anexo = normalize_anexo_bucket(row.get("tipo_anexo") or "")
            if tipo_anexo not in comparison_chart_anexos_rank:
                continue

            anexo_key = (ejercicio, tipo_anexo)
            anexo_totals[anexo_key] = anexo_totals.get(anexo_key, 0) + 1

            stacked_item = stacked_by_anexo.setdefault(
                anexo_key,
                {
                    "ejercicio": ejercicio,
                    "tipo_anexo": tipo_anexo,
                    "solventadas": 0,
                    "pendientes": 0,
                },
            )
            if estado_key == "solventado":
                stacked_item["solventadas"] += 1
            elif estado_key == "pendiente":
                stacked_item["pendientes"] += 1

        kpis_by_year = []
        totals_by_year = []
        pdp_amounts_by_year = []
        for year in selected_years:
            row = metrics_by_year.get(year, {})
            emitidas = int(row.get("emitidas", 0))
            solventadas = int(row.get("solventadas", 0))
            pendientes = int(row.get("pendientes", 0))
            monto_pdp_emitido = float(row.get("monto_pdp_emitido", 0))
            monto_pdp_solventado = float(row.get("monto_pdp_solventado", 0))
            monto_pdp_pendiente = float(row.get("monto_pdp_pendiente", 0))
            porcentaje_solventacion = (solventadas / emitidas * 100) if emitidas else 0.0
            kpis_by_year.append(
                {
                    "ejercicio": year,
                    "emitidas": emitidas,
                    "solventadas": solventadas,
                    "pendientes": pendientes,
                    "porcentaje_solventacion": porcentaje_solventacion,
                    "monto_pdp_emitido": monto_pdp_emitido,
                    "monto_pdp_solventado": monto_pdp_solventado,
                    "monto_pdp_pendiente": monto_pdp_pendiente,
                }
            )
            totals_by_year.append(
                {
                    "ejercicio": year,
                    "total_observaciones": emitidas,
                }
            )
            pdp_amounts_by_year.append(
                {
                    "ejercicio": year,
                    "emitido": monto_pdp_emitido,
                    "solventado": monto_pdp_solventado,
                    "pendiente": monto_pdp_pendiente,
                }
            )

        status_by_year = [
            {
                "ejercicio": ejercicio,
                "estado": estado,
                "total": total,
            }
            for (ejercicio, estado), total in sorted(status_totals.items())
        ]
        anexo_totals_by_year = [
            {
                "ejercicio": ejercicio,
                "tipo_anexo": tipo_anexo,
                "total": total,
            }
            for (ejercicio, tipo_anexo), total in sorted(
                anexo_totals.items(),
                key=lambda item: (
                    year_rank.get(item[0][0], len(year_rank)),
                    comparison_chart_anexos_rank.get(
                        item[0][1],
                        len(comparison_chart_anexos_rank),
                    ),
                    item[0][1],
                ),
            )
        ]
        stacked_by_anexo_rows = [
            {
                "ejercicio": item["ejercicio"],
                "tipo_anexo": item["tipo_anexo"],
                "solventadas": int(item["solventadas"]),
                "pendientes": int(item["pendientes"]),
            }
            for item in sorted(
                stacked_by_anexo.values(),
                key=lambda row: (
                    year_rank.get(row["ejercicio"], len(year_rank)),
                    comparison_chart_anexos_rank.get(
                        row["tipo_anexo"],
                        len(comparison_chart_anexos_rank),
                    ),
                    row["tipo_anexo"],
                ),
            )
        ]
        return {
            "kpis_by_year": kpis_by_year,
            "totals_by_year": totals_by_year,
            "status_by_year": status_by_year,
            "anexo_totals_by_year": anexo_totals_by_year,
            "stacked_by_anexo": stacked_by_anexo_rows,
            "pdp_amounts_by_year": pdp_amounts_by_year,
        }

    def apply_comparison_filter_clause(
        clauses: list[str],
        params: list,
        key: str,
        values: list[str],
    ):
        if not values:
            return
        placeholders = ", ".join(["?"] * len(values))
        column_map = {
            "ente_uid": "ente_uid",
            "tipo_auditoria": "tipo_auditoria",
            "tipo_anexo": "tipo_anexo",
            "estado": "estado",
            "fuente_financiamiento": "fuente_financiamiento",
            "origen_fuente": "origen_fuente",
            "ramo_33": "ramo_33",
            "ramo_28": "ramo_28",
        }
        column = column_map.get(key)
        if not column:
            return
        clauses.append(f"{column} IN ({placeholders})")
        params.extend(values)

    def build_comparison_where(
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
    ):
        clauses: list[str] = []
        params: list = []
        for key in comparison_filter_order:
            if key == exclude_key:
                continue
            values = selected_filters.get(key, [])
            if isinstance(values, list):
                apply_comparison_filter_clause(clauses, params, key, values)
        return " AND ".join(clauses), params

    def fetch_common_comparison_entity_uids(
        db,
        selected_years: list[str],
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
    ) -> list[str]:
        base_sql, base_params = build_comparison_base_sql(selected_years)
        where_sql, where_params = build_comparison_where(
            selected_filters,
            exclude_key=exclude_key,
        )
        where_clause = f"WHERE {where_sql}" if where_sql else ""
        rows = db.execute(
            f"""
            SELECT ente_uid
            FROM ({base_sql}) AS comparison_base
            {where_clause}
            GROUP BY ente_uid
            HAVING COUNT(DISTINCT ejercicio) = ?
            ORDER BY ente_uid
            """,
            [*base_params, *where_params, len(selected_years)],
        ).fetchall()
        return [row["ente_uid"] for row in rows if (row["ente_uid"] or "").strip()]

    def build_comparison_scope_query(
        db,
        selected_years: list[str],
        selected_filters: dict[str, list[str] | str],
        *,
        exclude_key: str = "",
    ):
        base_sql, base_params = build_comparison_base_sql(selected_years)
        where_sql, where_params = build_comparison_where(
            selected_filters,
            exclude_key=exclude_key,
        )
        clauses: list[str] = []
        params = [*base_params]
        if where_sql:
            clauses.append(where_sql)
            params.extend(where_params)

        if selected_filters.get("universo") == "complete" and selected_years:
            common_uids = fetch_common_comparison_entity_uids(
                db,
                selected_years,
                selected_filters,
                exclude_key=exclude_key,
            )
            if common_uids:
                placeholders = ", ".join(["?"] * len(common_uids))
                clauses.append(f"ente_uid IN ({placeholders})")
                params.extend(common_uids)
            else:
                clauses.append("1 = 0")

        scope_sql = f"SELECT * FROM ({base_sql}) AS comparison_base_scope"
        if clauses:
            scope_sql += f" WHERE {' AND '.join(clauses)}"
        return scope_sql, params

    def build_comparison_ente_options(rows, selected_years: list[str]):
        year_priority = {year: idx for idx, year in enumerate(selected_years)}
        grouped: dict[str, dict] = {}
        for row in rows:
            ente_uid = (row["ente_uid"] or "").strip()
            if not ente_uid:
                continue
            item = grouped.setdefault(
                ente_uid,
                {
                    "ente_uid": ente_uid,
                    "ente_numero": "",
                    "ente_nombre": "",
                    "aliases": [],
                    "years_present": set(),
                    "label_rank": -1,
                },
            )
            ejercicio = (row["ejercicio"] or "").strip()
            ente_nombre = (row["ente_nombre"] or "").strip()
            ente_numero = (row["ente_numero"] or "").strip()
            if ejercicio:
                item["years_present"].add(ejercicio)
            if ente_nombre and ente_nombre not in item["aliases"]:
                item["aliases"].append(ente_nombre)
            if ente_numero and not item["ente_numero"]:
                item["ente_numero"] = ente_numero
            rank = year_priority.get(ejercicio, -1)
            if rank >= item["label_rank"]:
                if ente_nombre:
                    item["ente_nombre"] = ente_nombre
                if ente_numero:
                    item["ente_numero"] = ente_numero
                item["label_rank"] = rank

        options = []
        for item in grouped.values():
            ente_numero = item["ente_numero"]
            ente_nombre = item["ente_nombre"] or (item["aliases"][0] if item["aliases"] else item["ente_uid"])
            label = f"{ente_numero} - {ente_nombre}" if ente_numero else ente_nombre
            options.append(
                {
                    "ente_uid": item["ente_uid"],
                    "ente_numero": ente_numero,
                    "ente_nombre": ente_nombre,
                    "label": label,
                    "aliases": item["aliases"],
                    "has_historical_names": len(item["aliases"]) > 1,
                    "years_present": sorted(item["years_present"]),
                }
            )
        options.sort(
            key=lambda item: (
                parse_ente_numero_sort(item.get("ente_numero") or ""),
                item.get("ente_numero") or "",
                item.get("ente_nombre") or item.get("label") or "",
            )
        )
        return options

    def build_comparison_table_rows(rows, selected_years: list[str]):
        year_priority = {year: idx for idx, year in enumerate(selected_years)}
        first_year = selected_years[0] if selected_years else ""
        last_year = selected_years[-1] if selected_years else ""
        grouped: dict[str, dict] = {}

        for row in rows:
            ente_uid = (row["ente_uid"] or "").strip()
            if not ente_uid:
                continue
            item = grouped.setdefault(
                ente_uid,
                {
                    "ente_uid": ente_uid,
                    "ente_numero": "",
                    "ente_nombre": "",
                    "aliases": [],
                    "years_present": set(),
                    "counts_by_year": {year: 0 for year in selected_years},
                    "label_rank": -1,
                },
            )
            ejercicio = (row["ejercicio"] or "").strip()
            ente_numero = (row["ente_numero"] or "").strip()
            ente_nombre = (row["ente_nombre"] or "").strip()
            total_observaciones = int(row["total_observaciones"] or 0)
            if ejercicio in item["counts_by_year"]:
                item["counts_by_year"][ejercicio] = total_observaciones
                item["years_present"].add(ejercicio)
            if ente_nombre and ente_nombre not in item["aliases"]:
                item["aliases"].append(ente_nombre)
            if ente_numero and not item["ente_numero"]:
                item["ente_numero"] = ente_numero
            rank = year_priority.get(ejercicio, -1)
            if rank >= item["label_rank"]:
                if ente_nombre:
                    item["ente_nombre"] = ente_nombre
                if ente_numero:
                    item["ente_numero"] = ente_numero
                item["label_rank"] = rank

        table_rows = []
        for item in grouped.values():
            counts_by_year = item["counts_by_year"]
            base_value = counts_by_year.get(first_year, 0)
            compare_value = counts_by_year.get(last_year, 0)
            delta_abs = compare_value - base_value
            if base_value > 0:
                delta_pct = (delta_abs / base_value) * 100
            elif compare_value == 0:
                delta_pct = 0.0
            else:
                delta_pct = None
            if base_value == 0 and compare_value > 0:
                change_label = "Nuevo"
            elif delta_abs > 0:
                change_label = "Subio"
            elif delta_abs < 0:
                change_label = "Bajo"
            else:
                change_label = "Sin cambio"
            ente_nombre = item["ente_nombre"] or (item["aliases"][0] if item["aliases"] else item["ente_uid"])
            label = f"{item['ente_numero']} - {ente_nombre}" if item["ente_numero"] else ente_nombre
            table_rows.append(
                {
                    "ente_uid": item["ente_uid"],
                    "ente_numero": item["ente_numero"],
                    "ente_nombre": ente_nombre,
                    "label": label,
                    "aliases": item["aliases"],
                    "has_historical_names": len(item["aliases"]) > 1,
                    "years_present": sorted(item["years_present"]),
                    "counts_by_year": counts_by_year,
                    "delta_abs": delta_abs,
                    "delta_pct": delta_pct,
                    "change_label": change_label,
                }
            )

        table_rows.sort(
            key=lambda item: (
                -abs(int(item["delta_abs"] or 0)),
                -(item["counts_by_year"].get(last_year, 0) if last_year else 0),
                item.get("ente_nombre") or item.get("label") or "",
            )
        )
        return table_rows

    def build_multi_filter_comparison(
        db,
        ejercicio: str,
        selected_filters: dict[str, list[str]],
        entes_catalog: list[dict],
    ):
        comparison_priority = (
            "ente_id",
            "periodo_cedula",
            "fuente_financiamiento",
            "origen_fuente",
            "tipo_anexo",
            "tipo_auditoria",
            "estado",
            "concepto_irregularidad",
        )
        chosen_key = ""
        for key in comparison_priority:
            selected_values = selected_values_for_key(selected_filters, key)
            if len(selected_values) > 1:
                chosen_key = key
                break
        if not chosen_key:
            return {
                "active": False,
                "filter_key": "",
                "filter_label": "",
                "items": [],
            }

        selected_values = selected_values_for_key(selected_filters, chosen_key)
        base_where, base_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            exclude_key=chosen_key,
            include_ente=True,
        )

        ente_labels = {}
        for ente in entes_catalog:
            ente_id = (ente.get("ente_id") or "").strip()
            if not ente_id:
                continue
            numero = (ente.get("ente_numero") or "").strip()
            nombre = (ente.get("ente_nombre") or "").strip()
            if numero:
                label = f"{numero} - {nombre or ente_id}"
            else:
                label = nombre or ente_id
            ente_labels[ente_id] = label

        items = []
        for value in selected_values:
            value_clauses = []
            value_params: list = []
            apply_filter_clause(value_clauses, value_params, chosen_key, [value])
            if not value_clauses:
                continue
            metrics_row = db.execute(
                f"""
                SELECT
                    COUNT(*) AS emitidas,
                    SUM(CASE WHEN LOWER(TRIM(COALESCE(estado, ''))) = 'solventado' THEN 1 ELSE 0 END) AS solventadas,
                    SUM(CASE WHEN LOWER(TRIM(COALESCE(estado, ''))) = 'pendiente' THEN 1 ELSE 0 END) AS pendientes
                FROM observaciones
                WHERE {base_where}
                  AND {" AND ".join(value_clauses)}
                """,
                [*base_params, *value_params],
            ).fetchone()
            emitidas = int((metrics_row["emitidas"] or 0) if metrics_row else 0)
            solventadas = int((metrics_row["solventadas"] or 0) if metrics_row else 0)
            pendientes = int((metrics_row["pendientes"] or 0) if metrics_row else 0)
            if chosen_key == "ente_id":
                label = ente_labels.get(value, value)
            else:
                label = value
            items.append(
                {
                    "value": value,
                    "label": label,
                    "emitidas": emitidas,
                    "solventadas": solventadas,
                    "pendientes": pendientes,
                }
            )

        return {
            "active": len(items) > 1,
            "filter_key": chosen_key,
            "filter_label": filter_labels.get(chosen_key, chosen_key),
            "items": items,
        }

    def build_pendientes_por_periodo_summary(
        db,
        ejercicio: str,
        scope_sql: str,
        scope_params: list,
        selected_entes: list[str],
    ):
        pendientes_por_periodo = {
            "requires_ente": True,
            "groups": [],
        }
        if not selected_entes:
            return pendientes_por_periodo

        def normalize_numero_observacion(value) -> str:
            if value is None:
                return ""
            if isinstance(value, int):
                return str(value)
            raw = str(value).strip()
            if not raw:
                return ""
            try:
                parsed = float(raw.replace(",", ""))
                if parsed.is_integer():
                    return str(int(parsed))
            except ValueError:
                pass
            return raw

        def numero_sort_key(value: str):
            return (parse_ente_numero_sort(value), value or "")

        pendientes_rows = db.execute(
            f"""
            SELECT
                id,
                ente_id,
                ente_numero,
                ente_nombre,
                tipo_auditoria,
                periodo_cedula,
                periodo_titular,
                tipo_anexo,
                numero_observacion,
                estado,
                monto_pdp_pendiente,
                ente_numero_sort
            FROM observaciones
            WHERE {scope_sql}
            ORDER BY
                COALESCE(ente_numero_sort, 0) ASC,
                ente_numero ASC,
                ente_id ASC,
                tipo_auditoria ASC,
                periodo_cedula ASC,
                tipo_anexo ASC,
                numero_observacion ASC
            """,
            scope_params,
        ).fetchall()

        groups_index = {}
        for row in pendientes_rows:
            tipo_anexo_row = normalize_anexo_bucket(row["tipo_anexo"])
            if tipo_anexo_row not in anexos_orden:
                continue

            group_key = (
                (row["ente_id"] or "").strip(),
                (row["ente_numero"] or "").strip(),
                (row["ente_nombre"] or "").strip(),
                normalize_tipo_auditoria(row["tipo_auditoria"] or ""),
            )
            if group_key not in groups_index:
                groups_index[group_key] = {
                    "ente_id": group_key[0],
                    "ente_numero": group_key[1],
                    "ente_nombre": group_key[2] or "—",
                    "tipo_auditoria": group_key[3] or "—",
                    "ente_numero_sort": float(row["ente_numero_sort"] or 0),
                    "period_map": {},
                }
            group_payload = groups_index[group_key]

            periodo_key = (
                (row["periodo_cedula"] or "").strip(),
                (row["periodo_titular"] or "").strip(),
            )
            if periodo_key not in group_payload["period_map"]:
                group_payload["period_map"][periodo_key] = {
                    "periodo_cedula": periodo_key[0] or "—",
                    "periodo_titular": periodo_key[1] or "—",
                    "all_by_tipo": {anexo: set() for anexo in anexos_orden},
                    "has_any_by_tipo": {anexo: False for anexo in anexos_orden},
                    "pending_by_tipo": {anexo: set() for anexo in anexos_orden},
                    "pending_without_numero": {anexo: 0 for anexo in anexos_orden},
                    "pdp_pending_montos": {},
                    "pdp_pending_sin_numero": 0.0,
                }
            period_payload = group_payload["period_map"][periodo_key]
            period_payload["has_any_by_tipo"][tipo_anexo_row] = True

            numero_obs = normalize_numero_observacion(row["numero_observacion"])
            if numero_obs:
                period_payload["all_by_tipo"][tipo_anexo_row].add(numero_obs)

            estado_norm = (row["estado"] or "").strip().lower()
            if estado_norm != "pendiente":
                continue

            monto_pendiente = float(row["monto_pdp_pendiente"] or 0)
            if numero_obs:
                period_payload["pending_by_tipo"][tipo_anexo_row].add(numero_obs)
            else:
                period_payload["pending_without_numero"][tipo_anexo_row] += 1

            if tipo_anexo_row == "PDP":
                if numero_obs:
                    current = float(period_payload["pdp_pending_montos"].get(numero_obs, 0))
                    period_payload["pdp_pending_montos"][numero_obs] = max(current, monto_pendiente)
                else:
                    period_payload["pdp_pending_sin_numero"] += monto_pendiente

        groups_payload = []
        for group in groups_index.values():
            period_rows = list(group["period_map"].values())
            for period_row in period_rows:
                inicio_sort, _ = parse_periodo_cedula(ejercicio, period_row["periodo_cedula"])
                period_row["periodo_sort"] = inicio_sort or "9999-12-31"

            period_rows.sort(
                key=lambda item: (
                    item["periodo_sort"],
                    item["periodo_cedula"],
                    item["periodo_titular"],
                )
            )

            totales_group = {anexo: 0 for anexo in anexos_orden}
            monto_group = 0.0
            period_payload_rows = []

            for period_row in period_rows:
                pendientes = {}
                numerales = {}
                total_row = 0
                for anexo in anexos_orden:
                    numeros_pendientes = sorted(period_row["pending_by_tipo"][anexo], key=numero_sort_key)
                    pendientes_sin_numero = int(period_row["pending_without_numero"][anexo] or 0)
                    cantidad = len(numeros_pendientes) + pendientes_sin_numero
                    pendientes[anexo] = cantidad
                    total_row += cantidad

                    if numeros_pendientes:
                        joined = ",".join(numeros_pendientes)
                        if pendientes_sin_numero:
                            joined = f"{joined},s/n"
                        numerales[anexo] = joined
                    elif pendientes_sin_numero:
                        numerales[anexo] = "s/n"
                    elif period_row["has_any_by_tipo"][anexo]:
                        numerales[anexo] = "0"
                    else:
                        numerales[anexo] = "-"

                    totales_group[anexo] += cantidad

                monto_row = float(sum(period_row["pdp_pending_montos"].values()) + period_row["pdp_pending_sin_numero"])
                monto_group += monto_row
                period_payload_rows.append(
                    {
                        "periodo_cedula": period_row["periodo_cedula"],
                        "periodo_titular": period_row["periodo_titular"],
                        "pendientes": {
                            **pendientes,
                            "total": total_row,
                            "monto_dano": monto_row,
                        },
                        "numerales_no_solventadas": numerales,
                    }
                )

            groups_payload.append(
                {
                    "ente_id": group["ente_id"],
                    "ente_numero": group["ente_numero"],
                    "ente_nombre": group["ente_nombre"],
                    "tipo_auditoria": group["tipo_auditoria"],
                    "periodos": period_payload_rows,
                    "totales": {
                        **totales_group,
                        "total": sum(totales_group.values()),
                        "monto_dano": monto_group,
                    },
                    "_ente_numero_sort": group["ente_numero_sort"],
                }
            )

        groups_payload.sort(
            key=lambda item: (
                float(item.get("_ente_numero_sort", 0)),
                item.get("ente_numero", ""),
                item.get("ente_nombre", ""),
                item.get("tipo_auditoria", ""),
            )
        )
        for item in groups_payload:
            item.pop("_ente_numero_sort", None)
        pendientes_por_periodo["groups"] = groups_payload
        return pendientes_por_periodo

    @app.route("/entes", methods=["GET", "POST"])
    @luis_required
    def entes():
        if request.method == "POST":
            if get_current_user()["role"] != "editor":
                return redirect(url_for("index", notice="no_permission"))
            ejercicio = request.form.get("ente_ejercicio", "").strip()
            ente_id = request.form.get("ente_id", "").strip()
            ente_numero = request.form.get("ente_numero", "").strip()
            ente_nombre = request.form.get("ente_nombre", "").strip()
            responsable = request.form.get("ente_responsable", "").strip()
            clasificacion = request.form.get("ente_clasificacion", "").strip()
            ramo33 = request.form.get("ente_ramo33", "").strip()
            ramo28 = request.form.get("ente_ramo28", "").strip() or "No"
    
            if not all([ejercicio, ente_id, ente_numero, ente_nombre]):
                return redirect(url_for("index", notice="ente_error"))
    
            db = get_db()
            ente_uid = resolve_ente_uid(db, ente_nombre)
            db.execute(
                """
                INSERT INTO entes_detalle (
                    ente_uid, ente_id, ejercicio, ente_numero, ente_nombre,
                    responsable, clasificacion, ramo33, ramo28, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(ente_id, ejercicio) DO UPDATE SET
                    ente_uid = COALESCE(entes_detalle.ente_uid, excluded.ente_uid),
                    ente_numero = excluded.ente_numero,
                    ente_nombre = excluded.ente_nombre,
                    responsable = excluded.responsable,
                    clasificacion = excluded.clasificacion,
                    ramo33 = excluded.ramo33,
                    ramo28 = excluded.ramo28
                """,
                (
                    ente_uid,
                    ente_id,
                    ejercicio,
                    ente_numero,
                    ente_nombre,
                    responsable or "",
                    clasificacion or "",
                    ramo33 or "",
                    ramo28,
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                ),
            )
            db.commit()
            return redirect(url_for("index", notice="ente_saved"))
    
        ejercicio = request.args.get("ejercicio", "").strip()
        if not ejercicio:
            return jsonify([])
    
        db = get_db()
        rows = db.execute(
            f"""
            SELECT ente_id, ente_numero, ente_nombre, responsable, clasificacion, ramo33, ramo28
            FROM entes_detalle
            WHERE ejercicio = ?
            ORDER BY {ente_numero_sort_sql('ente_numero')} ASC, ente_numero ASC
            """,
            (ejercicio,),
        ).fetchall()
        return jsonify([dict(row) for row in rows])
    
    
    @app.route("/historial", methods=["GET", "POST"])
    @luis_required
    def historial():
        if request.method == "POST":
            if get_current_user()["role"] != "editor":
                return redirect(url_for("index", notice="no_permission"))
            ejercicio = request.form.get("historial_ejercicio", "").strip()
            ente_id = normalize_ente_id(request.form.get("historial_ente_id", ""))
            nombre = request.form.get("historial_nombre", "").strip()
            cargo = request.form.get("historial_cargo", "").strip()
            fecha_inicio = request.form.get("historial_fecha_inicio", "").strip()
            fecha_fin = request.form.get("historial_fecha_fin", "").strip()
            tipo_registro = request.form.get("historial_tipo_registro", "").strip()
            tipo_auditoria = request.form.get("historial_tipo_auditoria", "").strip() or "Financiera"
    
            if not all([ejercicio, ente_id, nombre, cargo, fecha_inicio, fecha_fin, tipo_registro]):
                return redirect(url_for("index", notice="historial_error"))
    
            db = get_db()
            ente_row = db.execute(
                f"""
                SELECT ente_nombre, ente_uid
                FROM entes_detalle
                WHERE ejercicio = ? AND {normalize_ente_id_sql('ente_id')} = ?
                """,
                (ejercicio, ente_id),
            ).fetchone()
            if ente_row is None:
                return redirect(url_for("index", notice="historial_error"))
            ente_nombre = ente_row["ente_nombre"]
            ente_uid = (ente_row["ente_uid"] or "").strip()
            db.execute(
                """
                INSERT INTO historial_titulares (
                    ejercicio, ente_uid, ente, tipo_auditoria, nombre, cargo, fecha_inicio, fecha_fin, tipo_registro
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(ejercicio),
                    ente_uid or None,
                    ente_nombre,
                    tipo_auditoria,
                    nombre,
                    cargo,
                    fecha_inicio,
                    fecha_fin,
                    tipo_registro,
                ),
            )
            db.commit()
            return redirect(url_for("index", notice="historial_saved"))
    
        ejercicio = request.args.get("ejercicio", "").strip()
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        if not ejercicio or not ente_id:
            return jsonify([])
    
        db = get_db()
        ente_info = db.execute(
            f"""
            SELECT ente_uid, ente_nombre
            FROM entes_detalle
            WHERE ejercicio = ? AND {normalize_ente_id_sql('ente_id')} = ?
            LIMIT 1
            """,
            (ejercicio, ente_id),
        ).fetchone()
        if not ente_info:
            return jsonify([])
    
        ente_uid = (ente_info["ente_uid"] or "").strip()
        ente_aliases = get_ente_aliases_by_uid(
            db,
            ejercicio,
            ente_id,
            fallback_names=[ente_info["ente_nombre"]],
        )
    
        filter_clause = ""
        filter_params = []
        if ente_uid and ente_aliases:
            placeholders = ", ".join(["?"] * len(ente_aliases))
            filter_clause = (
                f"AND (TRIM(COALESCE(ente_uid, '')) = ? OR TRIM(COALESCE(ente, '')) IN ({placeholders}))"
            )
            filter_params.extend([ente_uid, *ente_aliases])
        elif ente_uid:
            filter_clause = "AND TRIM(COALESCE(ente_uid, '')) = ?"
            filter_params.append(ente_uid)
        elif ente_aliases:
            placeholders = ", ".join(["?"] * len(ente_aliases))
            filter_clause = f"AND TRIM(COALESCE(ente, '')) IN ({placeholders})"
            filter_params.extend(ente_aliases)
    
        rows = db.execute(
            f"""
            SELECT id, nombre, cargo, fecha_inicio, fecha_fin, tipo_registro, tipo_auditoria
            FROM historial_titulares
            WHERE ejercicio = ?
            {filter_clause}
            ORDER BY id DESC
            """,
            [ejercicio, *filter_params],
        ).fetchall()
        return jsonify([dict(row) for row in rows])
    
    
    @app.post("/oficios")
    @luis_required
    def oficios():
        user = get_current_user()
        if not user or user["username"] != "omar":
            return redirect(url_for("index", notice="no_permission"))
    
        ejercicio = request.form.get("oficio_ejercicio", "").strip()
        ente_id = request.form.get("oficio_ente_id", "").strip()
        oficio_numero = request.form.get("oficio_numero", "").strip()
        tipo_auditoria = request.form.get("tipo_auditoria", "").strip()
        fecha_notificacion = request.form.get("fecha_notificacion", "").strip()
        fuente_id = request.form.get("oficio_fuente_id", "").strip()
    
        if not all(
            [
                ejercicio,
                ente_id,
                oficio_numero,
                tipo_auditoria,
                fecha_notificacion,
                fuente_id,
            ]
        ):
            return redirect(url_for("index", notice="oficio_error"))
    
        db = get_db()
        ente_row = db.execute(
            """
            SELECT ente_id
            FROM entes_detalle
            WHERE ente_id = ? AND ejercicio = ?
            """,
            (ente_id, ejercicio),
        ).fetchone()
    
        if ente_row is None:
            return redirect(url_for("index", notice="oficio_error"))
    
        fuente_row = db.execute(
            """
            SELECT 1
            FROM registros
            WHERE ejercicio = ? AND ente_id = ? AND fuente_id = ?
            LIMIT 1
            """,
            (ejercicio, ente_id, fuente_id),
        ).fetchone()
        if fuente_row is None:
            return redirect(url_for("index", notice="oficio_error"))
    
        db.execute(
            """
            INSERT INTO oficios (
                ente_id, ejercicio, oficio, tipo_auditoria, fecha_notificacion, observaciones,
                fuente_id, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                ente_id,
                ejercicio,
                oficio_numero,
                tipo_auditoria,
                fecha_notificacion,
                "",
                int(fuente_id),
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ),
        )
        db.commit()
        return redirect(url_for("index", notice="oficio_saved"))
    
    
    @app.get("/fuentes-ente")
    @luis_required
    def fuentes_ente():
        ejercicio = request.args.get("ejercicio", "").strip()
        ente_id = request.args.get("ente_id", "").strip()
        if not ejercicio or not ente_id:
            return jsonify([])
    
        db = get_db()
        rows = db.execute(
            """
            SELECT DISTINCT fuentes_financiamiento.id, fuentes_financiamiento.nombre
            FROM registros
            JOIN fuentes_financiamiento
                ON registros.fuente_id = fuentes_financiamiento.id
            WHERE registros.ejercicio = ? AND registros.ente_id = ?
            ORDER BY fuentes_financiamiento.nombre ASC
            """,
            (ejercicio, ente_id),
        ).fetchall()
        return jsonify([dict(row) for row in rows])
    
    
    @app.get("/ejercicios-disponibles")
    @luis_required
    def ejercicios_disponibles():
        db = get_db()
        observaciones_rows = db.execute(
            """
            SELECT ejercicio, COUNT(*) AS total
            FROM observaciones
            GROUP BY ejercicio
            """
        ).fetchall()
        entes_rows = db.execute(
            """
            SELECT ejercicio, COUNT(*) AS total
            FROM entes_detalle
            GROUP BY ejercicio
            """
        ).fetchall()
    
        resumen = {}
        for row in observaciones_rows:
            ejercicio = row["ejercicio"]
            resumen.setdefault(
                ejercicio,
                {"ejercicio": ejercicio, "total_observaciones": 0, "total_entes": 0},
            )
            resumen[ejercicio]["total_observaciones"] = row["total"]
        for row in entes_rows:
            ejercicio = row["ejercicio"]
            resumen.setdefault(
                ejercicio,
                {"ejercicio": ejercicio, "total_observaciones": 0, "total_entes": 0},
            )
            resumen[ejercicio]["total_entes"] = row["total"]
    
        ordered = sorted(resumen.values(), key=lambda item: item["ejercicio"], reverse=True)
        return jsonify(ordered)


    @app.get("/observaciones-dashboard")
    @luis_required
    def observaciones_dashboard():
        ejercicio = request.args.get("ejercicio", "").strip()
        if not ejercicio:
            return jsonify({
                "rows": [],
                "total_rows": 0,
                "page": 1,
                "page_size": 40,
                "total_pages": 1,
                "filtros": {},
                "summary": {
                    "tipos": [],
                    "totals": {"emitidas": 0, "solventadas": 0, "pendientes": 0},
                    "pdp_montos": {"emitido": 0, "solventado": 0, "pendiente": 0},
                    "top_pendientes": [],
                    "pendientes_por_periodo": {
                        "requires_ente": True,
                        "groups": [],
                    },
                    "multi_compare": {
                        "active": False,
                        "filter_key": "",
                        "filter_label": "",
                        "items": [],
                    },
                },
            })

        selected_filters = parse_selected_filters()
        selected_entes = selected_values_for_key(selected_filters, "ente_id")

        try:
            page = max(1, int(request.args.get("page", "1")))
        except ValueError:
            page = 1
        try:
            page_size = int(request.args.get("page_size", "40"))
        except ValueError:
            page_size = 40
        page_size = max(10, min(200, page_size))

        cache_key = request.query_string.decode("utf-8")
        now = time.time()
        cached = dashboard_cache.get(cache_key)
        if cached and now - cached[0] < dashboard_cache_ttl_seconds:
            return jsonify(cached[1])

        db = get_db()
        scope_sql, scope_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            include_ente=True,
        )
        total_rows = db.execute(
            f"SELECT COUNT(*) FROM observaciones WHERE {scope_sql}",
            scope_params,
        ).fetchone()[0]
        total_pages = max(1, (total_rows + page_size - 1) // page_size)
        page = min(page, total_pages)
        offset = (page - 1) * page_size

        pendientes_por_periodo = build_pendientes_por_periodo_summary(
            db,
            ejercicio,
            scope_sql,
            scope_params,
            selected_entes,
        )

        rows = db.execute(
            f"""
            SELECT
                id,
                ejercicio,
                ente_id,
                ente_numero,
                ente_nombre,
                tipo_auditoria,
                fuente_financiamiento,
                ramo_33,
                ramo_28,
                periodo_cedula,
                periodo_titular,
                oficio,
                fecha_notificacion,
                tipo_anexo,
                numero_observacion,
                estado,
                monto_pdp_emitido,
                monto_pdp_solventado,
                monto_pdp_pendiente,
                pdp_no_irregularidad,
                pdp_concepto_irregularidad,
                pdp_subconcepto_irregularidad
            FROM observaciones
            WHERE {scope_sql}
            ORDER BY
                COALESCE(ente_numero_sort, 0) ASC,
                ente_numero ASC,
                ente_id ASC,
                tipo_anexo ASC,
                numero_observacion ASC
            LIMIT ? OFFSET ?
            """,
            [*scope_params, page_size, offset],
        ).fetchall()

        summary_rows = db.execute(
            f"""
            SELECT
                tipo_anexo,
                COUNT(*) AS emitidas,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(estado, ''))) = 'solventado' THEN 1 ELSE 0 END) AS solventadas,
                SUM(CASE WHEN LOWER(TRIM(COALESCE(estado, ''))) = 'pendiente' THEN 1 ELSE 0 END) AS pendientes
            FROM observaciones
            WHERE {scope_sql}
              AND TRIM(COALESCE(tipo_anexo, '')) != ''
            GROUP BY tipo_anexo
            ORDER BY
                CASE tipo_anexo
                    WHEN 'SA' THEN 0
                    WHEN 'PDP' THEN 1
                    WHEN 'PRAS' THEN 2
                    WHEN 'PEFCF' THEN 3
                    WHEN 'R' THEN 4
                    ELSE 5
                END,
                tipo_anexo
            """,
            scope_params,
        ).fetchall()
        pdp_montos_row = db.execute(
            f"""
            SELECT
                SUM(COALESCE(monto_pdp_emitido, 0)) AS emitido,
                SUM(COALESCE(monto_pdp_solventado, 0)) AS solventado,
                SUM(COALESCE(monto_pdp_pendiente, 0)) AS pendiente
            FROM observaciones
            WHERE {scope_sql}
              AND tipo_anexo = 'PDP'
            """,
            scope_params,
        ).fetchone()
        # Este ranking es intencionalmente estático por ejercicio.
        top_pendientes_rows = db.execute(
            """
            SELECT
                TRIM(COALESCE(ente_nombre, 'Sin ente')) AS ente_nombre,
                COUNT(*) AS pendientes
            FROM observaciones
            WHERE ejercicio = ?
              AND LOWER(TRIM(COALESCE(estado, ''))) = 'pendiente'
            GROUP BY ente_id, ente_nombre
            ORDER BY pendientes DESC, ente_nombre ASC
            LIMIT 5
            """,
            [ejercicio],
        ).fetchall()

        def query_distinct(column: str, exclude_key: str):
            where_sql, where_params = build_observaciones_scope(
                ejercicio,
                selected_filters,
                exclude_key=exclude_key,
                include_ente=True,
            )
            value_sql = origen_fuente_sql() if column == "origen_fuente" else column
            return db.execute(
                f"""
                SELECT DISTINCT {value_sql} AS value
                FROM observaciones
                WHERE {where_sql}
                  AND TRIM(COALESCE({value_sql}, '')) != ''
                ORDER BY value
                """,
                where_params,
            ).fetchall()

        entes_where, entes_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            include_ente=False,
        )
        entes = db.execute(
            f"""
            SELECT DISTINCT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM observaciones
            WHERE {entes_where}
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY COALESCE(ente_numero_sort, 0), ente_numero, ente_nombre
            """,
            entes_params,
        ).fetchall()

        concepto_where, concepto_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            exclude_key="concepto_irregularidad",
            include_ente=True,
        )
        conceptos = db.execute(
            f"""
            SELECT DISTINCT concepto
            FROM (
                SELECT pdp_concepto_irregularidad AS concepto
                FROM observaciones
                WHERE {concepto_where}
                  AND pdp_concepto_irregularidad IS NOT NULL AND TRIM(pdp_concepto_irregularidad) != ''
                UNION
                SELECT pdp_subconcepto_irregularidad AS concepto
                FROM observaciones
                WHERE {concepto_where}
                  AND pdp_subconcepto_irregularidad IS NOT NULL AND TRIM(pdp_subconcepto_irregularidad) != ''
            )
            ORDER BY concepto
            """,
            concepto_params + concepto_params,
        ).fetchall()

        filtros = {
            "tipo_anexo": [row[0] for row in query_distinct("tipo_anexo", "tipo_anexo")],
            "tipo_auditoria": [row[0] for row in query_distinct("tipo_auditoria", "tipo_auditoria")],
            "estado": [row[0] for row in query_distinct("estado", "estado")],
            "fuente_financiamiento": [row[0] for row in query_distinct("fuente_financiamiento", "fuente_financiamiento")],
            "modalidad": [row[0] for row in query_distinct("modalidad", "modalidad")],
            "convenios": [row[0] for row in query_distinct("convenio_ente_nombre", "convenio_ente_nombre")],
            "origen_fuente": [row[0] for row in query_distinct("origen_fuente", "origen_fuente")],
            "ramo_33": [row[0] for row in query_distinct("ramo_33", "ramo_33")],
            "ramo_28": [row[0] for row in query_distinct("ramo_28", "ramo_28")],
            "cedulas": [row[0] for row in query_distinct("periodo_cedula", "periodo_cedula")],
            "conceptos_irregularidad": [row[0] for row in conceptos],
            "entes": [dict(row) for row in entes],
        }
        multi_compare = build_multi_filter_comparison(
            db,
            ejercicio,
            selected_filters,
            filtros["entes"],
        )

        tipos_summary = [dict(row) for row in summary_rows]
        totals = {"emitidas": 0, "solventadas": 0, "pendientes": 0}
        for row in tipos_summary:
            totals["emitidas"] += int(row.get("emitidas") or 0)
            totals["solventadas"] += int(row.get("solventadas") or 0)
            totals["pendientes"] += int(row.get("pendientes") or 0)

        payload = {
            "rows": [dict(row) for row in rows],
            "total_rows": total_rows,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "filtros": filtros,
            "summary": {
                "tipos": tipos_summary,
                "totals": totals,
                "pdp_montos": {
                    "emitido": float((pdp_montos_row["emitido"] or 0) if pdp_montos_row else 0),
                    "solventado": float((pdp_montos_row["solventado"] or 0) if pdp_montos_row else 0),
                    "pendiente": float((pdp_montos_row["pendiente"] or 0) if pdp_montos_row else 0),
                },
                "top_pendientes": [
                    {
                        "ente_nombre": row["ente_nombre"],
                        "pendientes": int(row["pendientes"] or 0),
                    }
                    for row in top_pendientes_rows
                ],
                "pendientes_por_periodo": pendientes_por_periodo,
                "multi_compare": multi_compare,
            },
        }
        dashboard_cache[cache_key] = (now, payload)
        if len(dashboard_cache) > 250:
            # Prevent unbounded growth in long-running processes.
            oldest_key = min(dashboard_cache.items(), key=lambda item: item[1][0])[0]
            dashboard_cache.pop(oldest_key, None)
        return jsonify(payload)
    
    
    @app.get("/observaciones")
    @luis_required
    def observaciones_api():
        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        periodo_informe = request.args.get("periodo_informe", "").strip()
        titular = request.args.get("titular", "").strip()
        periodo_admin = request.args.get("periodo_admin", "").strip()
        administrativo = request.args.get("administrativo", "").strip()
        search = request.args.get("search", "").strip()
        if not ejercicio:
            return jsonify([])
    
        db = get_db()
        params = [ejercicio]
        filter_clauses = []
        for key in filter_order:
            apply_filter_clause(
                filter_clauses,
                params,
                key,
                selected_filters.get(key, []),
                alias="observaciones",
            )
        if search:
            filter_clauses.append(
                """
                (
                    observaciones.ente_nombre LIKE ?
                    OR observaciones.oficio LIKE ?
                    OR observaciones.fuente_financiamiento LIKE ?
                    OR observaciones.convenio_nombre LIKE ?
                    OR observaciones.convenio_ente_nombre LIKE ?
                    OR observaciones.pdp_concepto_irregularidad LIKE ?
                    OR observaciones.pdp_subconcepto_irregularidad LIKE ?
                    OR CAST(observaciones.numero_observacion AS TEXT) LIKE ?
                )
                """
            )
            search_term = f"%{search}%"
            params.extend([search_term] * 8)
    
        filter_sql = ""
        if filter_clauses:
            filter_sql = " AND " + " AND ".join(filter_clauses)
        needs_historial_join = bool(periodo_informe or titular or periodo_admin or administrativo)
    
        if needs_historial_join:
            if periodo_informe:
                filter_clauses.append(f"{periodo_sql('resp')} = ?")
                params.append(periodo_informe)
            if titular:
                filter_clauses.append("resp.nombre = ?")
                params.append(titular)
            if periodo_admin:
                filter_clauses.append(f"{periodo_sql('admin')} = ?")
                params.append(periodo_admin)
            if administrativo:
                filter_clauses.append("admin.nombre = ?")
                params.append(administrativo)
    
            join_filter_sql = " AND " + " AND ".join(filter_clauses) if filter_clauses else ""
            rows = db.execute(
                f"""
                SELECT
                    observaciones.id,
                    observaciones.ejercicio,
                    observaciones.ente_id,
                    observaciones.ente_numero,
                    observaciones.ente_nombre,
                    observaciones.tipo_auditoria,
                    observaciones.fuente_financiamiento,
                    observaciones.modalidad,
                    observaciones.convenio_nombre,
                    observaciones.convenio_ente_nombre,
                    observaciones.convenio_ente_id,
                    observaciones.ramo_33,
                    observaciones.ramo_28,
                    observaciones.periodo_cedula,
                    observaciones.periodo_titular,
                    observaciones.oficio,
                    observaciones.fecha_notificacion,
                    observaciones.tipo_anexo,
                    observaciones.numero_observacion,
                    observaciones.estado,
                    observaciones.monto_pdp_emitido,
                    observaciones.monto_pdp_solventado,
                    observaciones.monto_pdp_pendiente,
                    observaciones.pdp_no_irregularidad,
                    observaciones.pdp_concepto_irregularidad,
                    observaciones.pdp_subconcepto_irregularidad
                FROM observaciones
                LEFT JOIN entes_detalle
                    ON {normalize_ente_id_sql("observaciones.ente_id")} = {normalize_ente_id_sql("entes_detalle.ente_id")}
                    AND observaciones.ejercicio = entes_detalle.ejercicio
                LEFT JOIN historial_titulares AS resp
                    ON resp.id = (
                        SELECT id
                        FROM historial_titulares
                        WHERE ejercicio = observaciones.ejercicio
                          AND tipo_auditoria = observaciones.tipo_auditoria
                          AND (
                              (
                                  TRIM(COALESCE(entes_detalle.ente_uid, '')) != ''
                                  AND TRIM(COALESCE(historial_titulares.ente_uid, '')) = TRIM(entes_detalle.ente_uid)
                              )
                              OR ente = COALESCE(observaciones.ente_nombre, entes_detalle.ente_nombre)
                          )
                          AND tipo_registro = 'titular'
                        ORDER BY id DESC
                        LIMIT 1
                    )
                LEFT JOIN historial_titulares AS admin
                    ON admin.id = (
                        SELECT id
                        FROM historial_titulares
                        WHERE ejercicio = observaciones.ejercicio
                          AND tipo_auditoria = observaciones.tipo_auditoria
                          AND (
                              (
                                  TRIM(COALESCE(entes_detalle.ente_uid, '')) != ''
                                  AND TRIM(COALESCE(historial_titulares.ente_uid, '')) = TRIM(entes_detalle.ente_uid)
                              )
                              OR ente = COALESCE(observaciones.ente_nombre, entes_detalle.ente_nombre)
                          )
                          AND tipo_registro = 'director_administrativo'
                        ORDER BY id DESC
                        LIMIT 1
                    )
                WHERE observaciones.ejercicio = ?
                {join_filter_sql}
                ORDER BY
                    COALESCE(observaciones.ente_numero_sort, 0) ASC,
                    observaciones.ente_numero ASC,
                    observaciones.ente_id ASC,
                    observaciones.tipo_anexo ASC,
                    observaciones.numero_observacion ASC
                """,
                params,
            ).fetchall()
        else:
            rows = db.execute(
                f"""
                SELECT
                    observaciones.id,
                    observaciones.ejercicio,
                    observaciones.ente_id,
                    observaciones.ente_numero,
                    observaciones.ente_nombre,
                    observaciones.tipo_auditoria,
                    observaciones.fuente_financiamiento,
                    observaciones.modalidad,
                    observaciones.convenio_nombre,
                    observaciones.convenio_ente_nombre,
                    observaciones.convenio_ente_id,
                    observaciones.ramo_33,
                    observaciones.ramo_28,
                    observaciones.periodo_cedula,
                    observaciones.periodo_titular,
                    observaciones.oficio,
                    observaciones.fecha_notificacion,
                    observaciones.tipo_anexo,
                    observaciones.numero_observacion,
                    observaciones.estado,
                    observaciones.monto_pdp_emitido,
                    observaciones.monto_pdp_solventado,
                    observaciones.monto_pdp_pendiente,
                    observaciones.pdp_no_irregularidad,
                    observaciones.pdp_concepto_irregularidad,
                    observaciones.pdp_subconcepto_irregularidad
                FROM observaciones
                WHERE observaciones.ejercicio = ?
                {filter_sql}
                ORDER BY
                    COALESCE(observaciones.ente_numero_sort, 0) ASC,
                    observaciones.ente_numero ASC,
                    observaciones.ente_id ASC,
                    observaciones.tipo_anexo ASC,
                    observaciones.numero_observacion ASC
                """,
                params,
            ).fetchall()
    
        return jsonify([dict(row) for row in rows])
    
    
    @app.get("/observaciones-filtros")
    @luis_required
    def observaciones_filtros():
        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        ente_values = selected_values_for_key(selected_filters, "ente_id")
        tipo_auditoria_values = selected_values_for_key(selected_filters, "tipo_auditoria")
        ente_id = ente_values[0] if len(ente_values) == 1 else ""
        tipo_auditoria = tipo_auditoria_values[0] if len(tipo_auditoria_values) == 1 else ""
        titular_seleccionado = request.args.get("titular", "").strip()
        administrativo_seleccionado = request.args.get("administrativo", "").strip()
        include_historial = request.args.get("include_historial", "").strip() == "1"
        if not ejercicio:
            return jsonify({})
    
        db = get_db()
        filtros = {}
    
        def query_distinct(column: str, exclude_key: str):
            where_sql, where_params = build_observaciones_scope(
                ejercicio,
                selected_filters,
                exclude_key=exclude_key,
                include_ente=True,
            )
            value_sql = origen_fuente_sql() if column == "origen_fuente" else column
            return db.execute(
                f"""
                SELECT DISTINCT {value_sql} AS value
                FROM observaciones
                WHERE {where_sql}
                  AND TRIM(COALESCE({value_sql}, '')) != ''
                ORDER BY value
                """,
                where_params,
            ).fetchall()
    
        auditorias = query_distinct("tipo_auditoria", "tipo_auditoria")
        tipos = query_distinct("tipo_anexo", "tipo_anexo")
        estados = query_distinct("estado", "estado")
        fuentes = query_distinct("fuente_financiamiento", "fuente_financiamiento")
        modalidades = query_distinct("modalidad", "modalidad")
        convenios = query_distinct("convenio_ente_nombre", "convenio_ente_nombre")
        origenes = query_distinct("origen_fuente", "origen_fuente")
        ramos = query_distinct("ramo_33", "ramo_33")
        ramos_28 = query_distinct("ramo_28", "ramo_28")
        cedulas = query_distinct("periodo_cedula", "periodo_cedula")
        concepto_where, concepto_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            exclude_key="concepto_irregularidad",
            include_ente=True,
        )
        conceptos = db.execute(
            f"""
            SELECT DISTINCT concepto
            FROM (
                SELECT pdp_concepto_irregularidad AS concepto
                FROM observaciones
                WHERE {concepto_where}
                  AND pdp_concepto_irregularidad IS NOT NULL AND TRIM(pdp_concepto_irregularidad) != ''
                UNION
                SELECT pdp_subconcepto_irregularidad AS concepto
                FROM observaciones
                WHERE {concepto_where}
                  AND pdp_subconcepto_irregularidad IS NOT NULL AND TRIM(pdp_subconcepto_irregularidad) != ''
            )
            ORDER BY concepto
            """,
            concepto_params + concepto_params,
        ).fetchall()
        entes_where, entes_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            include_ente=False,
        )
        entes = db.execute(
            f"""
            SELECT DISTINCT
                TRIM(COALESCE(ente_id, '')) AS ente_id,
                TRIM(COALESCE(ente_numero, '')) AS ente_numero,
                TRIM(COALESCE(ente_nombre, '')) AS ente_nombre
            FROM observaciones
            WHERE {entes_where}
              AND TRIM(COALESCE(ente_id, '')) != ''
            ORDER BY COALESCE(ente_numero_sort, 0), ente_numero, ente_nombre
            """,
            entes_params,
        ).fetchall()
        periodos_informe = []
        titulares = []
        periodos_admin = []
        administrativos = []
        if include_historial:
            ente_aliases = get_ente_aliases_by_uid(db, ejercicio, ente_id)
            ente_uid = get_ente_uid_by_ejercicio_id(db, ejercicio, ente_id)
    
            titular_params = [ejercicio]
            titular_clause = ""
            if ente_uid and ente_aliases:
                placeholders = ", ".join(["?"] * len(ente_aliases))
                titular_clause = (
                    f"AND (TRIM(COALESCE(ente_uid, '')) = ? OR TRIM(COALESCE(ente, '')) IN ({placeholders}))"
                )
                titular_params.extend([ente_uid, *ente_aliases])
            elif ente_uid:
                titular_clause = "AND TRIM(COALESCE(ente_uid, '')) = ?"
                titular_params.append(ente_uid)
            elif ente_aliases:
                placeholders = ", ".join(["?"] * len(ente_aliases))
                titular_clause = f"AND TRIM(COALESCE(ente, '')) IN ({placeholders})"
                titular_params.extend(ente_aliases)
    
            periodos_params = titular_params.copy()
            periodo_titular_clause = ""
            if titular_seleccionado:
                periodo_titular_clause = "AND nombre = ?"
                periodos_params.append(titular_seleccionado)
    
            titular_tipo_clause = ""
            titular_tipo_params = []
            if tipo_auditoria:
                titular_tipo_clause = "AND tipo_auditoria = ?"
                titular_tipo_params.append(tipo_auditoria)
    
            periodos_informe = db.execute(
                f"""
                SELECT DISTINCT {periodo_sql("historial_titulares")} AS periodo
                FROM historial_titulares
                WHERE ejercicio = ? {titular_clause} {periodo_titular_clause}
                  {titular_tipo_clause}
                  AND tipo_registro = 'titular'
                  AND fecha_inicio IS NOT NULL AND fecha_fin IS NOT NULL
                ORDER BY fecha_inicio
                """,
                periodos_params + titular_tipo_params,
            ).fetchall()
            titulares = db.execute(
                f"""
                SELECT DISTINCT nombre
                FROM historial_titulares
                WHERE ejercicio = ? {titular_clause}
                  {titular_tipo_clause}
                  AND tipo_registro = 'titular'
                  AND nombre IS NOT NULL AND nombre != ''
                ORDER BY nombre
                """,
                titular_params + titular_tipo_params,
            ).fetchall()
    
            admin_params = [ejercicio]
            admin_clause = ""
            if ente_uid and ente_aliases:
                placeholders = ", ".join(["?"] * len(ente_aliases))
                admin_clause = (
                    f"AND (TRIM(COALESCE(ente_uid, '')) = ? OR TRIM(COALESCE(ente, '')) IN ({placeholders}))"
                )
                admin_params.extend([ente_uid, *ente_aliases])
            elif ente_uid:
                admin_clause = "AND TRIM(COALESCE(ente_uid, '')) = ?"
                admin_params.append(ente_uid)
            elif ente_aliases:
                placeholders = ", ".join(["?"] * len(ente_aliases))
                admin_clause = f"AND TRIM(COALESCE(ente, '')) IN ({placeholders})"
                admin_params.extend(ente_aliases)
    
            admin_periodos_params = admin_params.copy()
            periodo_admin_clause = ""
            if administrativo_seleccionado:
                periodo_admin_clause = "AND nombre = ?"
                admin_periodos_params.append(administrativo_seleccionado)
    
            admin_tipo_clause = ""
            admin_tipo_params = []
            if tipo_auditoria:
                admin_tipo_clause = "AND tipo_auditoria = ?"
                admin_tipo_params.append(tipo_auditoria)
    
            periodos_admin = db.execute(
                f"""
                SELECT DISTINCT {periodo_sql("historial_titulares")} AS periodo
                FROM historial_titulares
                WHERE ejercicio = ? {admin_clause} {periodo_admin_clause}
                  {admin_tipo_clause}
                  AND tipo_registro = 'director_administrativo'
                  AND fecha_inicio IS NOT NULL AND fecha_fin IS NOT NULL
                ORDER BY fecha_inicio
                """,
                admin_periodos_params + admin_tipo_params,
            ).fetchall()
            administrativos = db.execute(
                f"""
                SELECT DISTINCT nombre
                FROM historial_titulares
                WHERE ejercicio = ? {admin_clause}
                  {admin_tipo_clause}
                  AND tipo_registro = 'director_administrativo'
                  AND nombre IS NOT NULL AND nombre != ''
                ORDER BY nombre
                """,
                admin_params + admin_tipo_params,
            ).fetchall()
        filtros["tipo_anexo"] = [row[0] for row in tipos]
        filtros["tipo_auditoria"] = [row[0] for row in auditorias]
        filtros["entes"] = [dict(row) for row in entes]
        filtros["estado"] = [row[0] for row in estados]
        filtros["fuente_financiamiento"] = [row[0] for row in fuentes]
        filtros["modalidad"] = [row[0] for row in modalidades]
        filtros["convenios"] = [row[0] for row in convenios]
        filtros["origen_fuente"] = [row[0] for row in origenes]
        filtros["ramo_33"] = [row[0] for row in ramos]
        filtros["ramo_28"] = [row[0] for row in ramos_28]
        filtros["conceptos_irregularidad"] = [row[0] for row in conceptos]
        filtros["periodo_informe"] = [row[0] for row in periodos_informe]
        filtros["titulares"] = [row[0] for row in titulares]
        filtros["periodo_admin"] = [row[0] for row in periodos_admin]
        filtros["administrativos"] = [row[0] for row in administrativos]
        filtros["cedulas"] = [row[0] for row in cedulas]
    
        return jsonify(filtros)
    
    
    def build_observaciones_responsables_groups(
        db,
        ejercicio: str,
        selected_filters: dict[str, list[str]],
    ) -> list[dict]:
        filter_clauses = ["o.ejercicio = ?"]
        params = [ejercicio]
        for key in filter_order:
            apply_filter_clause(filter_clauses, params, key, selected_filters.get(key, []), alias="o")

        where_sql = " AND ".join(filter_clauses)
        observaciones_rows = db.execute(
            f"""
            SELECT
                o.ejercicio,
                o.ente_id,
                o.ente_numero,
                o.ente_numero_sort,
                o.ente_nombre,
                ed.ente_nombre AS ente_detalle_nombre,
                ed.ente_uid AS ente_uid,
                o.tipo_auditoria,
                o.periodo_cedula,
                o.tipo_anexo,
                o.estado,
                o.monto_pdp_emitido,
                o.monto_pdp_solventado,
                o.monto_pdp_pendiente
            FROM observaciones AS o
            LEFT JOIN entes_detalle AS ed
                ON {normalize_ente_id_sql("o.ente_id")} = {normalize_ente_id_sql("ed.ente_id")}
                AND o.ejercicio = ed.ejercicio
            WHERE {where_sql}
            ORDER BY
                COALESCE(o.ente_numero_sort, 0) ASC,
                o.ente_numero ASC,
                o.ente_id ASC,
                o.tipo_auditoria ASC,
                o.periodo_cedula ASC,
                o.tipo_anexo ASC,
                o.numero_observacion ASC
            """,
            params,
        ).fetchall()

        groups_index = {}
        for row in observaciones_rows:
            periodo_cedula = (row["periodo_cedula"] or "").strip()
            if not periodo_cedula:
                continue
            tipo_anexo_row = normalize_anexo_bucket(row["tipo_anexo"])
            if tipo_anexo_row not in anexos_orden:
                continue
            ente_id_norm = normalize_ente_id(row["ente_id"])
            tipo_auditoria = (
                normalize_tipo_auditoria(row["tipo_auditoria"] or "")
                or (row["tipo_auditoria"] or "").strip()
                or "—"
            )
            ente_nombre = (row["ente_nombre"] or row["ente_detalle_nombre"] or "—").strip() or "—"
            group_key = (
                ente_id_norm,
                (row["ente_numero"] or "").strip(),
                ente_nombre,
                tipo_auditoria,
            )
            if group_key not in groups_index:
                groups_index[group_key] = {
                    "ejercicio": row["ejercicio"],
                    "ente_id": row["ente_id"],
                    "ente_id_norm": ente_id_norm,
                    "ente_numero": (row["ente_numero"] or "").strip(),
                    "ente_numero_sort": float(row["ente_numero_sort"] or 0),
                    "ente_nombre": ente_nombre,
                    "ente_detalle_nombre": (row["ente_detalle_nombre"] or "").strip(),
                    "ente_uid": (row["ente_uid"] or "").strip(),
                    "tipo_auditoria": tipo_auditoria,
                    "rows_by_periodo": {},
                }

            group_payload = groups_index[group_key]
            if periodo_cedula not in group_payload["rows_by_periodo"]:
                group_payload["rows_by_periodo"][periodo_cedula] = {
                    "periodo_cedula": periodo_cedula,
                    "periodo_cedula_inicio": "",
                    "periodo_cedula_fin": "",
                    "titulares": [],
                    "administrativos": [],
                    "emitidas": build_status_metrics(),
                    "solventadas": build_status_metrics(),
                    "pendientes": build_status_metrics(),
                }

            period_payload = group_payload["rows_by_periodo"][periodo_cedula]
            append_status_metric(
                period_payload["emitidas"],
                tipo_anexo_row,
                float(row["monto_pdp_emitido"] or 0),
            )

            estado_norm = (row["estado"] or "").strip().lower()
            if estado_norm.startswith("solvent"):
                append_status_metric(
                    period_payload["solventadas"],
                    tipo_anexo_row,
                    float(row["monto_pdp_solventado"] or 0),
                )
            elif estado_norm.startswith("pendient"):
                append_status_metric(
                    period_payload["pendientes"],
                    tipo_anexo_row,
                    float(row["monto_pdp_pendiente"] or 0),
                )

        groups_payload = []
        for group in groups_index.values():
            nombres_ente = get_ente_aliases_by_uid(
                db,
                group["ejercicio"],
                group["ente_id_norm"],
                fallback_names=[group["ente_nombre"], group["ente_detalle_nombre"]],
            )
            ente_uid = (
                get_ente_uid_by_ejercicio_id(db, group["ejercicio"], group["ente_id_norm"])
                or group["ente_uid"]
            )
            historial_rows = []
            scope_clause = ""
            scope_params = []
            if ente_uid and nombres_ente:
                placeholders = ", ".join(["?"] * len(nombres_ente))
                scope_clause = (
                    f"AND (TRIM(COALESCE(h.ente_uid, '')) = ? OR TRIM(COALESCE(h.ente, '')) IN ({placeholders}))"
                )
                scope_params.extend([ente_uid, *nombres_ente])
            elif ente_uid:
                scope_clause = "AND TRIM(COALESCE(h.ente_uid, '')) = ?"
                scope_params.append(ente_uid)
            elif nombres_ente:
                placeholders = ", ".join(["?"] * len(nombres_ente))
                scope_clause = f"AND TRIM(COALESCE(h.ente, '')) IN ({placeholders})"
                scope_params.extend(nombres_ente)

            if scope_clause:
                historial_rows = db.execute(
                    f"""
                    SELECT
                        h.tipo_registro,
                        h.tipo_auditoria,
                        h.nombre,
                        h.fecha_inicio,
                        h.fecha_fin
                    FROM historial_titulares AS h
                    WHERE h.ejercicio = ?
                      {scope_clause}
                      AND h.tipo_registro IN ('titular', 'director_administrativo')
                      AND h.nombre IS NOT NULL AND h.nombre != ''
                    ORDER BY h.tipo_registro ASC, h.nombre ASC, h.fecha_inicio ASC, h.fecha_fin ASC
                    """,
                    [group["ejercicio"], *scope_params],
                ).fetchall()

            historial_periodos = merge_responsable_periods(
                [
                    item
                    for item in historial_rows
                    if normalize_tipo_auditoria(item["tipo_auditoria"] or "")
                    == normalize_tipo_auditoria(group["tipo_auditoria"] or "")
                ]
            )

            period_rows = list(group["rows_by_periodo"].values())
            for period_row in period_rows:
                cedula_inicio, cedula_fin = parse_periodo_cedula(
                    group["ejercicio"],
                    period_row["periodo_cedula"],
                )
                period_row["periodo_cedula_inicio"] = cedula_inicio or ""
                period_row["periodo_cedula_fin"] = cedula_fin or ""
                cedula_inicio_date = parse_historial_date(cedula_inicio) if cedula_inicio else None
                cedula_fin_date = parse_historial_date(cedula_fin) if cedula_fin else None
                if not cedula_inicio_date or not cedula_fin_date:
                    continue

                titulares = []
                administrativos = []
                titulares_seen = set()
                administrativos_seen = set()
                for item in historial_periodos:
                    inicio = item["inicio"]
                    fin = item["fin"]
                    if inicio > cedula_fin_date or fin < cedula_inicio_date:
                        continue
                    payload = {
                        "nombre": item["nombre"],
                        "periodo": format_periodo_display(inicio, fin),
                        "fecha_inicio": inicio.isoformat(),
                        "fecha_fin": fin.isoformat(),
                    }
                    responsable_key = (payload["nombre"], payload["periodo"])
                    if item["tipo_registro"] == "titular":
                        if responsable_key in titulares_seen:
                            continue
                        titulares_seen.add(responsable_key)
                        titulares.append(payload)
                    elif item["tipo_registro"] == "director_administrativo":
                        if responsable_key in administrativos_seen:
                            continue
                        administrativos_seen.add(responsable_key)
                        administrativos.append(payload)

                period_row["titulares"] = titulares
                period_row["administrativos"] = administrativos

            period_rows.sort(
                key=lambda item: (
                    item.get("periodo_cedula_inicio") or "9999-12-31",
                    item.get("periodo_cedula_fin") or "9999-12-31",
                    (item.get("periodo_cedula") or "").strip(),
                )
            )

            totals = {
                "emitidas": build_status_metrics(),
                "solventadas": build_status_metrics(),
                "pendientes": build_status_metrics(),
            }
            for period_row in period_rows:
                for status_key in ("emitidas", "solventadas", "pendientes"):
                    row_status = period_row.get(status_key, {}) or {}
                    for anexo in anexos_orden:
                        totals[status_key][anexo] += int(row_status.get(anexo) or 0)
                    totals[status_key]["total"] += int(row_status.get("total") or 0)
                    totals[status_key]["monto_dano"] += float(row_status.get("monto_dano") or 0)

            groups_payload.append(
                {
                    "ente_id": group["ente_id"],
                    "ente_numero": group["ente_numero"],
                    "ente_nombre": group["ente_nombre"],
                    "tipo_auditoria": group["tipo_auditoria"],
                    "rows": period_rows,
                    "totals": totals,
                    "_ente_numero_sort": group["ente_numero_sort"],
                }
            )

        groups_payload.sort(
            key=lambda item: (
                float(item.get("_ente_numero_sort", 0)),
                item.get("ente_numero", ""),
                item.get("ente_nombre", ""),
                normalize_tipo_auditoria(item.get("tipo_auditoria", "")),
            )
        )
        for item in groups_payload:
            item.pop("_ente_numero_sort", None)

        return groups_payload

    def build_responsable_row_values(items, total_rows: int) -> list[dict | None]:
        safe_total_rows = max(1, int(total_rows or 0))
        normalized_items = []
        for item in items or []:
            nombre = str((item or {}).get("nombre") or "").strip()
            periodo = str((item or {}).get("periodo") or "").strip()
            if not nombre and not periodo:
                continue
            normalized_items.append({"nombre": nombre, "periodo": periodo})
        if not normalized_items:
            return [None] * safe_total_rows

        values: list[dict | None] = [None] * safe_total_rows
        boundaries = [
            round(index * safe_total_rows / len(normalized_items))
            for index in range(len(normalized_items) + 1)
        ]
        boundaries[0] = 0
        boundaries[-1] = safe_total_rows
        for index, item in enumerate(normalized_items):
            start = boundaries[index]
            end = boundaries[index + 1]
            if end <= start:
                end = min(safe_total_rows, start + 1)
            for row_index in range(start, end):
                values[row_index] = item
        return values

    def build_collapsed_cells(values: list[str]) -> list[dict | None]:
        normalized_values = [str(value or "").strip() or "—" for value in (values or ["—"])]
        cells: list[dict | None] = [None] * len(normalized_values)
        index = 0
        while index < len(normalized_values):
            current_value = normalized_values[index]
            end = index + 1
            while end < len(normalized_values) and normalized_values[end] == current_value:
                end += 1
            cells[index] = {
                "value": current_value,
                "rowspan": end - index,
            }
            index = end
        return cells

    def build_observaciones_responsables_visual_rows(group: dict) -> list[dict]:
        visual_rows: list[dict] = []
        for row in (group.get("rows", []) or []):
            titulares = row.get("titulares", []) or []
            administrativos = row.get("administrativos", []) or []
            visual_row_count = max(len(titulares), len(administrativos), 1)
            titular_values = build_responsable_row_values(titulares, visual_row_count)
            administrativo_values = build_responsable_row_values(administrativos, visual_row_count)
            for visual_index in range(visual_row_count):
                visual_rows.append(
                    {
                        "titular_periodo": ((titular_values[visual_index] or {}).get("periodo") or "—"),
                        "titular_nombre": ((titular_values[visual_index] or {}).get("nombre") or "—"),
                        "administrativo_periodo": ((administrativo_values[visual_index] or {}).get("periodo") or "—"),
                        "administrativo_nombre": ((administrativo_values[visual_index] or {}).get("nombre") or "—"),
                        "periodo_cedula": str(row.get("periodo_cedula") or "").strip() or "—",
                        "emitidas": row.get("emitidas", {}) or {},
                        "solventadas": row.get("solventadas", {}) or {},
                        "pendientes": row.get("pendientes", {}) or {},
                        "show_metrics": visual_index == 0,
                        "metric_rowspan": visual_row_count,
                    }
                )
        return visual_rows

    def add_status_totals(target: dict, source: dict) -> None:
        for anexo in anexos_orden:
            target[anexo] += int((source or {}).get(anexo) or 0)
        target["total"] += int((source or {}).get("total") or 0)
        target["monto_dano"] += float((source or {}).get("monto_dano") or 0)

    def build_observaciones_breakdown(
        db,
        ejercicio: str,
        selected_filters: dict[str, list[str]],
        group_by: str,
    ) -> dict:
        group_config = {
            "general": {
                "column": "ente_nombre",
                "label": "Nombre del Ente",
                "empty": "Sin ente",
            },
            "tipo_auditoria": {
                "column": "tipo_auditoria",
                "label": "Tipo de Auditoría",
                "empty": "Sin tipo de auditoría",
            },
            "fuente_financiamiento": {
                "column": "fuente_financiamiento",
                "label": "Fuente de Financiamiento",
                "empty": "Sin fuente de financiamiento",
            },
        }
        config = group_config.get(group_by)
        if not config:
            group_by = "tipo_auditoria"
            config = group_config[group_by]

        scope_sql, scope_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            include_ente=True,
            alias="o",
        )
        rows = db.execute(
            f"""
            SELECT
                o.{config["column"]} AS group_value,
                o.ente_id,
                o.ente_numero,
                o.ente_numero_sort,
                o.ente_nombre,
                o.tipo_auditoria,
                o.fuente_financiamiento,
                o.periodo_cedula,
                o.periodo_titular,
                o.tipo_anexo,
                o.estado,
                o.monto_pdp_emitido,
                o.monto_pdp_solventado,
                o.monto_pdp_pendiente
            FROM observaciones AS o
            WHERE {scope_sql}
            ORDER BY
                COALESCE(o.ente_numero_sort, 0) ASC,
                o.ente_numero ASC,
                o.ente_id ASC,
                group_value ASC,
                o.periodo_cedula ASC,
                o.periodo_titular ASC,
                o.tipo_anexo ASC,
                o.numero_observacion ASC
            """,
            scope_params,
        ).fetchall()

        groups_index: dict[str, dict] = {}
        general_totals = {
            "emitidas": build_status_metrics(),
            "solventadas": build_status_metrics(),
            "pendientes": build_status_metrics(),
        }
        if group_by == "fuente_financiamiento":
            ente_groups: dict[str, dict] = {}
            for row in rows:
                tipo_anexo = normalize_anexo_bucket(row["tipo_anexo"] or "")
                if tipo_anexo not in anexos_orden:
                    continue
                ente_key = (
                    normalize_ente_id(row["ente_id"] or "")
                    or f"{row['ente_numero'] or ''}|{row['ente_nombre'] or ''}".casefold()
                )
                ente_nombre = " ".join((row["ente_nombre"] or "").split()) or "Sin ente"
                ente_no = (row["ente_numero"] or row["ente_id"] or "").strip() or "—"
                if ente_key not in ente_groups:
                    ente_groups[ente_key] = {
                        "no": ente_no,
                        "label": ente_nombre,
                        "_sort": float(row["ente_numero_sort"] or 0),
                        "items": {},
                        "totals": {
                            "emitidas": build_status_metrics(),
                            "solventadas": build_status_metrics(),
                            "pendientes": build_status_metrics(),
                        },
                    }

                tipo_auditoria = (
                    normalize_tipo_auditoria(row["tipo_auditoria"] or "")
                    or "Sin tipo de auditoría"
                )
                fuente = " ".join((row["fuente_financiamiento"] or "").split()) or "Sin fuente de financiamiento"
                periodo_cedula = " ".join((row["periodo_cedula"] or "").split()) or "—"
                periodo_titular = " ".join((row["periodo_titular"] or "").split()) or "—"
                item_key = (
                    tipo_auditoria.casefold(),
                    fuente.casefold(),
                    periodo_cedula.casefold(),
                    periodo_titular.casefold(),
                )
                if item_key not in ente_groups[ente_key]["items"]:
                    periodo_inicio, periodo_fin = parse_periodo_cedula(ejercicio, periodo_cedula)
                    ente_groups[ente_key]["items"][item_key] = {
                        "tipo_auditoria": tipo_auditoria,
                        "fuente_financiamiento": fuente,
                        "periodo_cedula": periodo_cedula,
                        "periodo_titular": periodo_titular,
                        "_periodo_inicio": periodo_inicio or "9999-12-31",
                        "_periodo_fin": periodo_fin or "9999-12-31",
                        "totals": {
                            "emitidas": build_status_metrics(),
                            "solventadas": build_status_metrics(),
                            "pendientes": build_status_metrics(),
                        },
                    }

                item_totals = ente_groups[ente_key]["items"][item_key]["totals"]
                ente_totals = ente_groups[ente_key]["totals"]
                append_status_metric(item_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))
                append_status_metric(ente_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))
                append_status_metric(general_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))

                estado_norm = (row["estado"] or "").strip().lower()
                if estado_norm.startswith("solvent"):
                    append_status_metric(item_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                    append_status_metric(ente_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                    append_status_metric(general_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                elif estado_norm.startswith("pendient"):
                    append_status_metric(item_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))
                    append_status_metric(ente_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))
                    append_status_metric(general_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))

            payload_rows = []
            for ente_group in sorted(
                ente_groups.values(),
                key=lambda item: (
                    float(item.get("_sort") or 0),
                    str(item.get("no") or ""),
                    str(item.get("label") or "").casefold(),
                ),
            ):
                source_rows = sorted(
                    ente_group["items"].values(),
                    key=lambda item: (
                        str(item.get("tipo_auditoria") or "").casefold(),
                        str(item.get("fuente_financiamiento") or "").casefold(),
                        str(item.get("_periodo_inicio") or "9999-12-31"),
                        str(item.get("_periodo_fin") or "9999-12-31"),
                        str(item.get("periodo_cedula") or "").casefold(),
                        str(item.get("periodo_titular") or "").casefold(),
                    ),
                )
                tipo_rowspans = [0] * len(source_rows)
                fuente_rowspans = [0] * len(source_rows)
                cursor = 0
                while cursor < len(source_rows):
                    current_tipo = source_rows[cursor].get("tipo_auditoria")
                    end = cursor + 1
                    while end < len(source_rows) and source_rows[end].get("tipo_auditoria") == current_tipo:
                        end += 1
                    tipo_rowspans[cursor] = end - cursor
                    cursor = end
                cursor = 0
                while cursor < len(source_rows):
                    current_key = (
                        source_rows[cursor].get("tipo_auditoria"),
                        source_rows[cursor].get("fuente_financiamiento"),
                    )
                    end = cursor + 1
                    while end < len(source_rows) and (
                        source_rows[end].get("tipo_auditoria"),
                        source_rows[end].get("fuente_financiamiento"),
                    ) == current_key:
                        end += 1
                    fuente_rowspans[cursor] = end - cursor
                    cursor = end
                for index, item in enumerate(source_rows):
                    item.pop("_periodo_inicio", None)
                    item.pop("_periodo_fin", None)
                    payload_rows.append(
                        {
                            "no": ente_group["no"],
                            "label": ente_group["label"],
                            "show_entity": index == 0,
                            "entity_rowspan": len(source_rows),
                            "tipo_auditoria": item["tipo_auditoria"],
                            "show_tipo_auditoria": tipo_rowspans[index] > 0,
                            "tipo_auditoria_rowspan": tipo_rowspans[index],
                            "fuente_financiamiento": item["fuente_financiamiento"],
                            "show_fuente_financiamiento": fuente_rowspans[index] > 0,
                            "fuente_financiamiento_rowspan": fuente_rowspans[index],
                            "periodo_cedula": item["periodo_cedula"],
                            "periodo_titular": item["periodo_titular"],
                            "totals": item["totals"],
                        }
                    )
                if source_rows:
                    payload_rows.append(
                        {
                            "row_type": "subtotal",
                            "no": "",
                            "label": "SUBTOTAL",
                            "tipo_auditoria": "",
                            "fuente_financiamiento": "",
                            "periodo_cedula": "",
                            "periodo_titular": "",
                            "totals": ente_group["totals"],
                        }
                    )
            return {
                "group_by": group_by,
                "label": "Nombre del Ente",
                "detail_label": "Fuente de Financiamiento",
                "rows": payload_rows,
                "totals": general_totals,
            }

        if group_by == "tipo_auditoria":
            ente_groups: dict[str, dict] = {}
            for row in rows:
                tipo_anexo = normalize_anexo_bucket(row["tipo_anexo"] or "")
                if tipo_anexo not in anexos_orden:
                    continue
                ente_key = (
                    normalize_ente_id(row["ente_id"] or "")
                    or f"{row['ente_numero'] or ''}|{row['ente_nombre'] or ''}".casefold()
                )
                ente_nombre = " ".join((row["ente_nombre"] or "").split()) or "Sin ente"
                ente_no = (row["ente_numero"] or row["ente_id"] or "").strip() or "—"
                if ente_key not in ente_groups:
                    ente_groups[ente_key] = {
                        "no": ente_no,
                        "label": ente_nombre,
                        "_sort": float(row["ente_numero_sort"] or 0),
                        "items": {},
                        "totals": {
                            "emitidas": build_status_metrics(),
                            "solventadas": build_status_metrics(),
                            "pendientes": build_status_metrics(),
                        },
                    }
                tipo_auditoria = (
                    normalize_tipo_auditoria(row["group_value"] or "")
                    or "Sin tipo de auditoría"
                )
                if tipo_auditoria not in ente_groups[ente_key]["items"]:
                    ente_groups[ente_key]["items"][tipo_auditoria] = {
                        "detail": tipo_auditoria,
                        "totals": {
                            "emitidas": build_status_metrics(),
                            "solventadas": build_status_metrics(),
                            "pendientes": build_status_metrics(),
                        },
                    }

                item_totals = ente_groups[ente_key]["items"][tipo_auditoria]["totals"]
                ente_totals = ente_groups[ente_key]["totals"]
                append_status_metric(item_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))
                append_status_metric(ente_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))
                append_status_metric(general_totals["emitidas"], tipo_anexo, float(row["monto_pdp_emitido"] or 0))

                estado_norm = (row["estado"] or "").strip().lower()
                if estado_norm.startswith("solvent"):
                    append_status_metric(item_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                    append_status_metric(ente_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                    append_status_metric(general_totals["solventadas"], tipo_anexo, float(row["monto_pdp_solventado"] or 0))
                elif estado_norm.startswith("pendient"):
                    append_status_metric(item_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))
                    append_status_metric(ente_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))
                    append_status_metric(general_totals["pendientes"], tipo_anexo, float(row["monto_pdp_pendiente"] or 0))

            payload_rows = []
            for ente_group in sorted(
                ente_groups.values(),
                key=lambda item: (
                    float(item.get("_sort") or 0),
                    str(item.get("no") or ""),
                    str(item.get("label") or "").casefold(),
                ),
            ):
                type_rows = sorted(
                    ente_group["items"].values(),
                    key=lambda item: str(item.get("detail") or "").casefold(),
                )
                for index, item in enumerate(type_rows):
                    payload_rows.append(
                        {
                            "no": ente_group["no"],
                            "label": ente_group["label"],
                            "show_entity": index == 0,
                            "entity_rowspan": len(type_rows),
                            "detail": item["detail"],
                            "totals": item["totals"],
                        }
                    )
                if type_rows:
                    payload_rows.append(
                        {
                            "row_type": "subtotal",
                            "no": "",
                            "label": "SUBTOTAL",
                            "detail": "",
                            "totals": ente_group["totals"],
                        }
                    )
            return {
                "group_by": group_by,
                "label": "Nombre del Ente",
                "detail_label": "Tipo de Auditoría",
                "rows": payload_rows,
                "totals": general_totals,
            }

        for row in rows:
            group_value = " ".join((row["group_value"] or "").split()) or config["empty"]
            if group_by == "general":
                group_key = (
                    normalize_ente_id(row["ente_id"] or "")
                    or f"{row['ente_numero'] or ''}|{group_value}".casefold()
                )
                group_no = (row["ente_numero"] or row["ente_id"] or "").strip() or "—"
                group_sort = float(row["ente_numero_sort"] or 0)
            elif group_by == "tipo_auditoria":
                group_value = normalize_tipo_auditoria(group_value) or group_value
                group_key = group_value.casefold()
                group_no = ""
                group_sort = 0.0
            else:
                group_key = group_value.casefold()
                group_no = ""
                group_sort = 0.0
            if group_key not in groups_index:
                groups_index[group_key] = {
                    "label": group_value,
                    "no": group_no,
                    "_sort": group_sort,
                    "totals": {
                        "emitidas": build_status_metrics(),
                        "solventadas": build_status_metrics(),
                        "pendientes": build_status_metrics(),
                    },
                }

            tipo_anexo = normalize_anexo_bucket(row["tipo_anexo"] or "")
            if tipo_anexo not in anexos_orden:
                continue
            group_totals = groups_index[group_key]["totals"]
            append_status_metric(
                group_totals["emitidas"],
                tipo_anexo,
                float(row["monto_pdp_emitido"] or 0),
            )
            append_status_metric(
                general_totals["emitidas"],
                tipo_anexo,
                float(row["monto_pdp_emitido"] or 0),
            )

            estado_norm = (row["estado"] or "").strip().lower()
            if estado_norm.startswith("solvent"):
                append_status_metric(
                    group_totals["solventadas"],
                    tipo_anexo,
                    float(row["monto_pdp_solventado"] or 0),
                )
                append_status_metric(
                    general_totals["solventadas"],
                    tipo_anexo,
                    float(row["monto_pdp_solventado"] or 0),
                )
            elif estado_norm.startswith("pendient"):
                append_status_metric(
                    group_totals["pendientes"],
                    tipo_anexo,
                    float(row["monto_pdp_pendiente"] or 0),
                )
                append_status_metric(
                    general_totals["pendientes"],
                    tipo_anexo,
                    float(row["monto_pdp_pendiente"] or 0),
                )

        if group_by == "general":
            payload_rows = sorted(
                groups_index.values(),
                key=lambda item: (
                    float(item.get("_sort") or 0),
                    str(item.get("no") or ""),
                    str(item.get("label") or "").casefold(),
                ),
            )
        else:
            payload_rows = sorted(
                groups_index.values(),
                key=lambda item: (str(item.get("label") or "").casefold()),
            )
        for item in payload_rows:
            item.pop("_sort", None)
        return {
            "group_by": group_by,
            "label": config["label"],
            "rows": payload_rows,
            "totals": general_totals,
        }

    @app.get("/observaciones-responsables")
    @luis_required
    def observaciones_responsables():
        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        selected_entes = selected_values_for_key(selected_filters, "ente_id")
        if not ejercicio or not selected_entes:
            return jsonify([])

        groups_payload = build_observaciones_responsables_groups(
            get_db(),
            ejercicio,
            selected_filters,
        )
        return jsonify(groups_payload)

    @app.get("/observaciones-desglose")
    @luis_required
    def observaciones_desglose():
        ejercicio = request.args.get("ejercicio", "").strip()
        group_by = request.args.get("group_by", "tipo_auditoria").strip()
        selected_filters = parse_selected_filters()
        if not ejercicio:
            return jsonify({
                "group_by": group_by,
                "label": "",
                "rows": [],
                "totals": {
                    "emitidas": build_status_metrics(),
                    "solventadas": build_status_metrics(),
                    "pendientes": build_status_metrics(),
                },
            })
        return jsonify(
            build_observaciones_breakdown(
                get_db(),
                ejercicio,
                selected_filters,
                group_by,
            )
        )

    @app.get("/observaciones-responsables-exportar")
    @luis_required
    def observaciones_responsables_exportar():
        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        selected_entes = selected_values_for_key(selected_filters, "ente_id")
        if not ejercicio:
            return jsonify({"error": "ejercicio requerido"}), 400
        if not selected_entes:
            return jsonify({"error": "selecciona al menos un ente"}), 400

        db = get_db()
        groups = build_observaciones_responsables_groups(db, ejercicio, selected_filters)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Responsables por periodo"
        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        header_primary_fill = PatternFill("solid", fgColor="1F3B2C")
        header_secondary_fill = PatternFill("solid", fgColor="2A503D")
        total_fill = PatternFill("solid", fgColor="EDF4EF")
        zebra_fill = PatternFill("solid", fgColor="F8FAF7")
        white_bold_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_alignment = Alignment(horizontal="right", vertical="center")

        sheet.append(
            [
                "No",
                "Ente Fiscalizable",
                f"Periodos Informe {ejercicio}",
                f"Titular {ejercicio}",
                "Periodos Administrativo / Responsable de Obra",
                "Administrativo / Encargado de Obra",
                "Periodo Cédula",
                "EMITIDAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "SOLVENTADAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "PENDIENTES",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.append(
            [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Emitidas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PO",
                "PRAS",
                "PECFF",
                "Solventadas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Pendientes",
                "Monto Daño ($)",
            ]
        )
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            sheet.merge_cells(f"{col}1:{col}2")
        sheet.merge_cells("H1:N1")
        sheet.merge_cells("O1:U1")
        sheet.merge_cells("V1:AB1")
        for row_idx in (1, 2):
            for cell in sheet[row_idx]:
                cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = header_primary_fill if row_idx == 1 else header_secondary_fill
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 24
        sheet.freeze_panes = "A3"

        ente_key_counts: dict[str, int] = {}
        for group in groups:
            ente_key = "|".join(
                [
                    str(group.get("ente_id") or "").strip(),
                    str(group.get("ente_nombre") or "").strip(),
                ]
            )
            ente_key_counts[ente_key] = ente_key_counts.get(ente_key, 0) + 1
        general_totals = {
            "emitidas": build_status_metrics(),
            "solventadas": build_status_metrics(),
            "pendientes": build_status_metrics(),
        }

        def add_metric_totals(target: dict, source: dict) -> None:
            for anexo in anexos_orden:
                target[anexo] += int((source or {}).get(anexo) or 0)
            target["total"] += int((source or {}).get("total") or 0)
            target["monto_dano"] += float((source or {}).get("monto_dano") or 0)

        def totals_export_values(label: str, totals: dict) -> list:
            return [
                label,
                "",
                "",
                "",
                "",
                "",
                "",
                int((totals.get("emitidas") or {}).get("R") or 0),
                int((totals.get("emitidas") or {}).get("SA") or 0),
                int((totals.get("emitidas") or {}).get("PDP") or 0),
                int((totals.get("emitidas") or {}).get("PRAS") or 0),
                int((totals.get("emitidas") or {}).get("PEFCF") or 0),
                int((totals.get("emitidas") or {}).get("total") or 0),
                (
                    float((totals.get("emitidas") or {}).get("monto_dano") or 0)
                    if float((totals.get("emitidas") or {}).get("monto_dano") or 0) > 0
                    else "-"
                ),
                int((totals.get("solventadas") or {}).get("R") or 0),
                int((totals.get("solventadas") or {}).get("SA") or 0),
                int((totals.get("solventadas") or {}).get("PDP") or 0),
                int((totals.get("solventadas") or {}).get("PRAS") or 0),
                int((totals.get("solventadas") or {}).get("PEFCF") or 0),
                int((totals.get("solventadas") or {}).get("total") or 0),
                (
                    float((totals.get("solventadas") or {}).get("monto_dano") or 0)
                    if float((totals.get("solventadas") or {}).get("monto_dano") or 0) > 0
                    else "-"
                ),
                int((totals.get("pendientes") or {}).get("R") or 0),
                int((totals.get("pendientes") or {}).get("SA") or 0),
                int((totals.get("pendientes") or {}).get("PDP") or 0),
                int((totals.get("pendientes") or {}).get("PRAS") or 0),
                int((totals.get("pendientes") or {}).get("PEFCF") or 0),
                int((totals.get("pendientes") or {}).get("total") or 0),
                (
                    float((totals.get("pendientes") or {}).get("monto_dano") or 0)
                    if float((totals.get("pendientes") or {}).get("monto_dano") or 0) > 0
                    else "-"
                ),
            ]

        def style_totals_export_row(row_idx: int, *, fill: PatternFill, label_alignment) -> None:
            sheet.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=7,
            )
            for col_idx in range(1, 29):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = fill
                cell.border = thin_border
                if col_idx == 1:
                    cell.alignment = label_alignment
                elif col_idx in (14, 21, 28):
                    cell.alignment = right_alignment
                else:
                    cell.alignment = center_alignment
            for amount_col in (14, 21, 28):
                amount_cell = sheet.cell(row=row_idx, column=amount_col)
                if isinstance(amount_cell.value, (int, float)):
                    amount_cell.number_format = "#,##0.00"

        if not groups:
            sheet.append(["Sin resultados para los filtros seleccionados."])
            sheet.merge_cells("A3:AB3")
            empty_cell = sheet["A3"]
            empty_cell.alignment = center_alignment
            empty_cell.font = Font(italic=True)
            empty_cell.fill = zebra_fill
            for col_idx in range(1, 29):
                sheet.cell(row=3, column=col_idx).border = thin_border
        else:
            for group in groups:
                visual_rows = build_observaciones_responsables_visual_rows(group)
                group_no = str(group.get("ente_numero") or group.get("ente_id") or "—")
                ente_key = "|".join(
                    [
                        str(group.get("ente_id") or "").strip(),
                        str(group.get("ente_nombre") or "").strip(),
                    ]
                )
                ente_label = str(group.get("ente_nombre") or "—")
                if ente_key_counts.get(ente_key, 0) > 1 and group.get("tipo_auditoria"):
                    ente_label = f"{ente_label}\n{group.get('tipo_auditoria')}"

                if not visual_rows:
                    sheet.append([group_no, ente_label, "Sin periodos disponibles para este ente."] + [""] * 25)
                    row_idx = sheet.max_row
                    sheet.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=28)
                    for col_idx in range(1, 29):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = left_alignment if col_idx in (2, 3) else center_alignment
                        if row_idx % 2 == 0:
                            cell.fill = zebra_fill
                    continue

                titular_periodo_cells = build_collapsed_cells(
                    [row["titular_periodo"] for row in visual_rows]
                )
                titular_nombre_cells = build_collapsed_cells(
                    [row["titular_nombre"] for row in visual_rows]
                )
                administrativo_periodo_cells = build_collapsed_cells(
                    [row["administrativo_periodo"] for row in visual_rows]
                )
                administrativo_nombre_cells = build_collapsed_cells(
                    [row["administrativo_nombre"] for row in visual_rows]
                )
                cedula_cells = build_collapsed_cells(
                    [row["periodo_cedula"] for row in visual_rows]
                )

                group_start_row = sheet.max_row + 1
                for index, visual_row in enumerate(visual_rows):
                    emitidas = visual_row["emitidas"]
                    solventadas = visual_row["solventadas"]
                    pendientes = visual_row["pendientes"]
                    show_metrics = bool(visual_row["show_metrics"])
                    sheet.append(
                        [
                            group_no if index == 0 else "",
                            ente_label if index == 0 else "",
                            titular_periodo_cells[index]["value"] if titular_periodo_cells[index] else "",
                            titular_nombre_cells[index]["value"] if titular_nombre_cells[index] else "",
                            administrativo_periodo_cells[index]["value"] if administrativo_periodo_cells[index] else "",
                            administrativo_nombre_cells[index]["value"] if administrativo_nombre_cells[index] else "",
                            cedula_cells[index]["value"] if cedula_cells[index] else "",
                            int(emitidas.get("R") or 0) if show_metrics else "",
                            int(emitidas.get("SA") or 0) if show_metrics else "",
                            int(emitidas.get("PDP") or 0) if show_metrics else "",
                            int(emitidas.get("PRAS") or 0) if show_metrics else "",
                            int(emitidas.get("PEFCF") or 0) if show_metrics else "",
                            int(emitidas.get("total") or 0) if show_metrics else "",
                            float(emitidas.get("monto_dano") or 0) if show_metrics and float(emitidas.get("monto_dano") or 0) > 0 else ("-" if show_metrics else ""),
                            int(solventadas.get("R") or 0) if show_metrics else "",
                            int(solventadas.get("SA") or 0) if show_metrics else "",
                            int(solventadas.get("PDP") or 0) if show_metrics else "",
                            int(solventadas.get("PRAS") or 0) if show_metrics else "",
                            int(solventadas.get("PEFCF") or 0) if show_metrics else "",
                            int(solventadas.get("total") or 0) if show_metrics else "",
                            float(solventadas.get("monto_dano") or 0) if show_metrics and float(solventadas.get("monto_dano") or 0) > 0 else ("-" if show_metrics else ""),
                            int(pendientes.get("R") or 0) if show_metrics else "",
                            int(pendientes.get("SA") or 0) if show_metrics else "",
                            int(pendientes.get("PDP") or 0) if show_metrics else "",
                            int(pendientes.get("PRAS") or 0) if show_metrics else "",
                            int(pendientes.get("PEFCF") or 0) if show_metrics else "",
                            int(pendientes.get("total") or 0) if show_metrics else "",
                            float(pendientes.get("monto_dano") or 0) if show_metrics and float(pendientes.get("monto_dano") or 0) > 0 else ("-" if show_metrics else ""),
                        ]
                    )
                    row_idx = sheet.max_row
                    for col_idx in range(1, 29):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if col_idx in (2, 3, 4, 5, 6, 7):
                            cell.alignment = left_alignment
                        elif col_idx in (14, 21, 28):
                            cell.alignment = right_alignment
                        else:
                            cell.alignment = center_alignment
                        if row_idx % 2 == 0:
                            cell.fill = zebra_fill
                    for amount_col in (14, 21, 28):
                        amount_cell = sheet.cell(row=row_idx, column=amount_col)
                        if isinstance(amount_cell.value, (int, float)):
                            amount_cell.number_format = "#,##0.00"

                group_row_count = len(visual_rows)
                if group_row_count > 1:
                    sheet.merge_cells(
                        start_row=group_start_row,
                        start_column=1,
                        end_row=group_start_row + group_row_count - 1,
                        end_column=1,
                    )
                    sheet.merge_cells(
                        start_row=group_start_row,
                        start_column=2,
                        end_row=group_start_row + group_row_count - 1,
                        end_column=2,
                    )
                for col_idx, cells in (
                    (3, titular_periodo_cells),
                    (4, titular_nombre_cells),
                    (5, administrativo_periodo_cells),
                    (6, administrativo_nombre_cells),
                    (7, cedula_cells),
                ):
                    for offset, cell_info in enumerate(cells):
                        if not cell_info or int(cell_info.get("rowspan") or 0) <= 1:
                            continue
                        start_row = group_start_row + offset
                        end_row = start_row + int(cell_info["rowspan"]) - 1
                        sheet.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )
                for offset, visual_row in enumerate(visual_rows):
                    if not visual_row["show_metrics"] or int(visual_row["metric_rowspan"] or 0) <= 1:
                        continue
                    start_row = group_start_row + offset
                    end_row = start_row + int(visual_row["metric_rowspan"]) - 1
                    for col_idx in range(8, 29):
                        sheet.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )

                totals = group.get("totals", {}) or {}
                for status_key in ("emitidas", "solventadas", "pendientes"):
                    add_metric_totals(general_totals[status_key], (totals.get(status_key) or {}))
                sheet.append(totals_export_values("SUBTOTAL", totals))
                subtotal_row_idx = sheet.max_row
                style_totals_export_row(subtotal_row_idx, fill=total_fill, label_alignment=right_alignment)

            if len(groups) > 1:
                grand_total_fill = PatternFill("solid", fgColor="DFECE5")
                sheet.append(totals_export_values("TOTAL GENERAL", general_totals))
                grand_total_row_idx = sheet.max_row
                style_totals_export_row(
                    grand_total_row_idx,
                    fill=grand_total_fill,
                    label_alignment=right_alignment,
                )

        base_widths = {
            1: 10,
            2: 34,
            3: 22,
            4: 28,
            5: 24,
            6: 28,
            7: 20,
            8: 8,
            9: 8,
            10: 9,
            11: 9,
            12: 10,
            13: 11,
            14: 16,
            15: 8,
            16: 8,
            17: 8,
            18: 9,
            19: 10,
            20: 12,
            21: 16,
            22: 8,
            23: 8,
            24: 9,
            25: 9,
            26: 10,
            27: 11,
            28: 16,
        }
        for col_idx in range(1, 29):
            max_len = 0
            for row_idx in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                string_value = "" if value is None else str(value)
                if len(string_value) > max_len:
                    max_len = len(string_value)
            width = max(base_widths.get(col_idx, 10), min(max_len + 2, 42))
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"responsables_periodo_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/observaciones-desglose-exportar")
    @luis_required
    def observaciones_desglose_exportar():
        ejercicio = request.args.get("ejercicio", "").strip()
        group_by = request.args.get("group_by", "tipo_auditoria").strip()
        selected_filters = parse_selected_filters()
        if not ejercicio:
            return jsonify({"error": "ejercicio requerido"}), 400

        payload = build_observaciones_breakdown(get_db(), ejercicio, selected_filters, group_by)
        rows = payload.get("rows", []) or []
        label = payload.get("label") or "Grupo"
        totals = payload.get("totals", {}) or {}
        is_tipo_auditoria = payload.get("group_by") == "tipo_auditoria"
        is_fuente_financiamiento = payload.get("group_by") == "fuente_financiamiento"
        left_headers = ["No", "Nombre del Ente", "Tipo de Auditoría"] if is_tipo_auditoria else ["No", label]
        left_count = len(left_headers)
        max_col = left_count + 21
        amount_cols = (left_count + 7, left_count + 14, left_count + 21)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resumen"
        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        header_primary_fill = PatternFill("solid", fgColor="1F3B2C")
        header_secondary_fill = PatternFill("solid", fgColor="2A503D")
        total_fill = PatternFill("solid", fgColor="EDF4EF")
        zebra_fill = PatternFill("solid", fgColor="F8FAF7")
        white_bold_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_alignment = Alignment(horizontal="right", vertical="center")

        def metric_export_values(metric: dict) -> list:
            monto = float((metric or {}).get("monto_dano") or 0)
            return [
                int((metric or {}).get("R") or 0),
                int((metric or {}).get("SA") or 0),
                int((metric or {}).get("PDP") or 0),
                int((metric or {}).get("PRAS") or 0),
                int((metric or {}).get("PEFCF") or 0),
                int((metric or {}).get("total") or 0),
                monto if monto > 0 else "-",
            ]

        if is_fuente_financiamiento:
            left_headers = [
                "No",
                "Nombre del Ente",
                "Tipo de Auditoría",
                "Fuente de Financiamiento",
                "Periodo Cédula",
                "Periodo Titular",
            ]
            left_count = len(left_headers)
            max_col = left_count + 21
            amount_cols = (left_count + 7, left_count + 14, left_count + 21)
            sheet.append([
                *left_headers,
                "EMITIDAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "SOLVENTADAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "PENDIENTES",
                "",
                "",
                "",
                "",
                "",
                "",
            ])
            sheet.append([
                *([""] * left_count),
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Emitidas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Solventadas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Pendientes",
                "Monto Daño ($)",
            ])
            for col_idx in range(1, left_count + 1):
                sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            sheet.merge_cells(start_row=1, start_column=left_count + 1, end_row=1, end_column=left_count + 7)
            sheet.merge_cells(start_row=1, start_column=left_count + 8, end_row=1, end_column=left_count + 14)
            sheet.merge_cells(start_row=1, start_column=left_count + 15, end_row=1, end_column=left_count + 21)
            for row_idx in (1, 2):
                for cell in sheet[row_idx]:
                    cell.font = white_bold_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.fill = header_primary_fill if row_idx == 1 else header_secondary_fill
            sheet.row_dimensions[1].height = 28
            sheet.row_dimensions[2].height = 24
            sheet.freeze_panes = "A3"

            merge_ranges: list[tuple[int, int, int]] = []
            subtotal_rows: set[int] = set()
            if not rows:
                sheet.append(["Sin resultados para los filtros seleccionados."])
                sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
                empty_cell = sheet["A3"]
                empty_cell.alignment = center_alignment
                empty_cell.font = Font(italic=True)
                empty_cell.fill = zebra_fill
            else:
                for row in rows:
                    if row.get("row_type") == "subtotal":
                        sheet.append([
                            "SUBTOTAL",
                            "",
                            "",
                            "",
                            "",
                            "",
                            *metric_export_values((row.get("totals") or {}).get("emitidas") or {}),
                            *metric_export_values((row.get("totals") or {}).get("solventadas") or {}),
                            *metric_export_values((row.get("totals") or {}).get("pendientes") or {}),
                        ])
                        row_idx = sheet.max_row
                        subtotal_rows.add(row_idx)
                        sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=left_count)
                        continue

                    row_idx = sheet.max_row + 1
                    if row.get("show_entity") and int(row.get("entity_rowspan") or 1) > 1:
                        merge_ranges.append((row_idx, row_idx + int(row.get("entity_rowspan") or 1) - 1, 1))
                        merge_ranges.append((row_idx, row_idx + int(row.get("entity_rowspan") or 1) - 1, 2))
                    if row.get("show_tipo_auditoria") and int(row.get("tipo_auditoria_rowspan") or 1) > 1:
                        merge_ranges.append((row_idx, row_idx + int(row.get("tipo_auditoria_rowspan") or 1) - 1, 3))
                    if row.get("show_fuente_financiamiento") and int(row.get("fuente_financiamiento_rowspan") or 1) > 1:
                        merge_ranges.append((row_idx, row_idx + int(row.get("fuente_financiamiento_rowspan") or 1) - 1, 4))
                    sheet.append([
                        row.get("no") if row.get("show_entity") else "",
                        row.get("label") if row.get("show_entity") else "",
                        row.get("tipo_auditoria") if row.get("show_tipo_auditoria") else "",
                        row.get("fuente_financiamiento") if row.get("show_fuente_financiamiento") else "",
                        row.get("periodo_cedula") or "—",
                        row.get("periodo_titular") or "—",
                        *metric_export_values((row.get("totals") or {}).get("emitidas") or {}),
                        *metric_export_values((row.get("totals") or {}).get("solventadas") or {}),
                        *metric_export_values((row.get("totals") or {}).get("pendientes") or {}),
                    ])

                sheet.append([
                    "TOTAL GENERAL",
                    "",
                    "",
                    "",
                    "",
                    "",
                    *metric_export_values((totals or {}).get("emitidas") or {}),
                    *metric_export_values((totals or {}).get("solventadas") or {}),
                    *metric_export_values((totals or {}).get("pendientes") or {}),
                ])
                total_row_idx = sheet.max_row
                sheet.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=left_count)
                subtotal_rows.add(total_row_idx)

            for start_row, end_row, col_idx in merge_ranges:
                sheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

            for row_idx in range(3, sheet.max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = right_alignment if col_idx in amount_cols else (left_alignment if col_idx in range(2, left_count + 1) else center_alignment)
                    if row_idx in subtotal_rows:
                        cell.font = Font(bold=True)
                        cell.fill = total_fill
                    elif row_idx % 2 == 0:
                        cell.fill = zebra_fill
                for amount_col in amount_cols:
                    amount_cell = sheet.cell(row=row_idx, column=amount_col)
                    if isinstance(amount_cell.value, (int, float)):
                        amount_cell.number_format = "#,##0.00"

            base_widths = {
                1: 8,
                2: 42,
                3: 22,
                4: 38,
                5: 28,
                6: 28,
                7: 8,
                8: 8,
                9: 9,
                10: 9,
                11: 10,
                12: 12,
                13: 16,
                14: 8,
                15: 8,
                16: 9,
                17: 9,
                18: 10,
                19: 13,
                20: 16,
                21: 8,
                22: 8,
                23: 9,
                24: 9,
                25: 10,
                26: 12,
                27: 16,
            }
            for col_idx in range(1, max_col + 1):
                max_len = 0
                for row_idx in range(1, sheet.max_row + 1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    string_value = "" if value is None else str(value)
                    if len(string_value) > max_len:
                        max_len = len(string_value)
                sheet.column_dimensions[get_column_letter(col_idx)].width = max(
                    base_widths.get(col_idx, 10),
                    min(max_len + 2, 54),
                )

            stream = BytesIO()
            workbook.save(stream)
            stream.seek(0)
            filename = f"resumen_fuente_financiamiento_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                stream,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        sheet.append(
            [
                *left_headers,
                "EMITIDAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "SOLVENTADAS",
                "",
                "",
                "",
                "",
                "",
                "",
                "PENDIENTES",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.append(
            [
                *([""] * left_count),
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Emitidas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Solventadas",
                "Monto Daño ($)",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PECFF",
                "Pendientes",
                "Monto Daño ($)",
            ]
        )
        for col_idx in range(1, left_count + 1):
            sheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        sheet.merge_cells(start_row=1, start_column=left_count + 1, end_row=1, end_column=left_count + 7)
        sheet.merge_cells(start_row=1, start_column=left_count + 8, end_row=1, end_column=left_count + 14)
        sheet.merge_cells(start_row=1, start_column=left_count + 15, end_row=1, end_column=left_count + 21)
        for row_idx in (1, 2):
            for cell in sheet[row_idx]:
                cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = header_primary_fill if row_idx == 1 else header_secondary_fill
        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 24
        sheet.freeze_panes = "A3"

        def append_export_row(row_payload: dict | None, row_totals: dict, index: int | str = "") -> None:
            row_payload = row_payload or {}
            if is_tipo_auditoria:
                if row_payload.get("row_type") == "subtotal":
                    fixed_values = ["", "SUBTOTAL", ""]
                elif row_payload.get("row_type") == "grand_total":
                    fixed_values = ["", "TOTAL GENERAL", ""]
                else:
                    fixed_values = [
                        row_payload.get("no", "") if row_payload.get("show_entity") else "",
                        row_payload.get("label", "") if row_payload.get("show_entity") else "",
                        row_payload.get("detail", ""),
                    ]
            else:
                fixed_values = [
                    row_payload.get("no") or index,
                    row_payload.get("label") or "—",
                ]
            sheet.append(
                [
                    *fixed_values,
                    *metric_export_values((row_totals or {}).get("emitidas") or {}),
                    *metric_export_values((row_totals or {}).get("solventadas") or {}),
                    *metric_export_values((row_totals or {}).get("pendientes") or {}),
                ]
            )

        merge_ranges: list[tuple[int, int, int]] = []
        if not rows:
            sheet.append(["Sin resultados para los filtros seleccionados."])
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
            empty_cell = sheet["A3"]
            empty_cell.alignment = center_alignment
            empty_cell.font = Font(italic=True)
            empty_cell.fill = zebra_fill
            for col_idx in range(1, max_col + 1):
                sheet.cell(row=3, column=col_idx).border = thin_border
        else:
            for index, row in enumerate(rows, start=1):
                next_row_idx = sheet.max_row + 1
                if is_tipo_auditoria and row.get("show_entity") and int(row.get("entity_rowspan") or 1) > 1:
                    merge_ranges.append((next_row_idx, next_row_idx + int(row.get("entity_rowspan") or 1) - 1, 1))
                    merge_ranges.append((next_row_idx, next_row_idx + int(row.get("entity_rowspan") or 1) - 1, 2))
                append_export_row(row, row.get("totals") or {}, index)
                row_idx = sheet.max_row
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = left_alignment if col_idx in range(2, left_count + 1) else center_alignment
                    if col_idx in amount_cols:
                        cell.alignment = right_alignment
                    if row.get("row_type") == "subtotal":
                        cell.font = Font(bold=True)
                        cell.fill = total_fill
                    elif row_idx % 2 == 0:
                        cell.fill = zebra_fill

            append_export_row({"row_type": "grand_total", "no": "", "label": "TOTAL GENERAL", "detail": ""}, totals, "")
            total_row_idx = sheet.max_row
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=total_row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx in range(2, left_count + 1) else center_alignment
                if col_idx in amount_cols:
                    cell.alignment = right_alignment

            for start_row, end_row, col_idx in merge_ranges:
                sheet.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

        for row_idx in range(3, sheet.max_row + 1):
            for amount_col in amount_cols:
                amount_cell = sheet.cell(row=row_idx, column=amount_col)
                if isinstance(amount_cell.value, (int, float)):
                    amount_cell.number_format = "#,##0.00"

        base_widths = {
            1: 8,
            2: 42,
            3: 8,
            4: 8,
            5: 9,
            6: 9,
            7: 10,
            8: 12,
            9: 16,
            10: 8,
            11: 8,
            12: 9,
            13: 9,
            14: 10,
            15: 13,
            16: 16,
            17: 8,
            18: 8,
            19: 9,
            20: 9,
            21: 10,
            22: 12,
            23: 16,
        }
        for col_idx in range(1, max_col + 1):
            max_len = 0
            for row_idx in range(1, sheet.max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                string_value = "" if value is None else str(value)
                if len(string_value) > max_len:
                    max_len = len(string_value)
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(
                base_widths.get(col_idx, 10),
                min(max_len + 2, 54),
            )

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        safe_group_map = {
            "general": "general",
            "tipo_auditoria": "tipo_auditoria",
            "fuente_financiamiento": "fuente_financiamiento",
        }
        safe_group = safe_group_map.get(payload.get("group_by"), "tipo_auditoria")
        filename = f"resumen_{safe_group}_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/observaciones-exportar")
    def observaciones_exportar():
        user = get_current_user()
        if user is None:
            return redirect(url_for("login", next=request.path))
        if not (is_luis_user(user) or is_gabo_user(user)):
            return redirect(url_for(home_endpoint_for_user(user), notice="no_permission"))

        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        if not ejercicio:
            return jsonify({"error": "ejercicio requerido"}), 400
    
        db = get_db()
        params = [ejercicio]
        filter_clauses = []
        for key in filter_order:
            apply_filter_clause(
                filter_clauses,
                params,
                key,
                selected_filters.get(key, []),
                alias="observaciones",
            )
    
        filter_sql = ""
        if filter_clauses:
            filter_sql = " AND " + " AND ".join(filter_clauses)
    
        rows = db.execute(
            f"""
            SELECT
                observaciones.ente_nombre,
                observaciones.tipo_anexo,
                observaciones.numero_observacion,
                observaciones.estado,
                observaciones.fecha_notificacion,
                observaciones.fuente_financiamiento,
                observaciones.modalidad,
                observaciones.convenio_nombre,
                observaciones.convenio_ente_nombre,
                observaciones.pdp_concepto_irregularidad,
                observaciones.monto_pdp_emitido,
                observaciones.monto_pdp_solventado,
                observaciones.monto_pdp_pendiente,
                observaciones.tipo_auditoria
            FROM observaciones
            LEFT JOIN entes_detalle
                ON {normalize_ente_id_sql("observaciones.ente_id")} = {normalize_ente_id_sql("entes_detalle.ente_id")}
                AND observaciones.ejercicio = entes_detalle.ejercicio
            WHERE observaciones.ejercicio = ?
            {filter_sql}
            ORDER BY
                {ente_numero_sort_sql('entes_detalle.ente_numero')} ASC,
                entes_detalle.ente_numero ASC,
                observaciones.ente_id ASC,
                observaciones.tipo_anexo ASC,
                observaciones.numero_observacion ASC
            """,
            params,
        ).fetchall()
    
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Observaciones"
        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        header_primary_fill = PatternFill("solid", fgColor="1F3B2C")
        total_fill = PatternFill("solid", fgColor="EDF4EF")
        zebra_fill = PatternFill("solid", fgColor="F8FAF7")
        white_bold_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        headers = [
            "Ente",
            "Tipo auditoria",
            "Anexo",
            "No. Obs",
            "Estado",
            "Fecha",
            "Fuente de Financiamiento",
            "Convenio",
            "Concepto de Irregularidad",
            "Monto emitido",
            "Monto solventado",
            "Monto pendiente",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = white_bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_primary_fill
        sheet.row_dimensions[1].height = 22
        sheet.freeze_panes = "A2"

        total_observaciones = 0
        total_emitido = 0.0
        total_solventado = 0.0
        total_pendiente = 0.0
        conteo_pdp = 0
        conteo_pendiente = 0
        conteo_solventado = 0
    
        for row in rows:
            monto_emitido = float(row["monto_pdp_emitido"] or 0)
            monto_solventado = float(row["monto_pdp_solventado"] or 0)
            monto_pendiente = float(row["monto_pdp_pendiente"] or 0)
            total_observaciones += 1
            total_emitido += monto_emitido
            total_solventado += monto_solventado
            total_pendiente += monto_pendiente
            if (row["tipo_anexo"] or "") == "PDP":
                conteo_pdp += 1
            if (row["estado"] or "").strip().lower() == "pendiente":
                conteo_pendiente += 1
            if (row["estado"] or "").strip().lower() == "solventado":
                conteo_solventado += 1
    
            sheet.append(
                [
                    row["ente_nombre"] or "—",
                    row["tipo_auditoria"] or "—",
                    row["tipo_anexo"] or "—",
                    row["numero_observacion"] if row["numero_observacion"] is not None else "—",
                    row["estado"] or "—",
                    row["fecha_notificacion"] or "—",
                    row["fuente_financiamiento"] or "—",
                    (
                        row["convenio_ente_nombre"]
                        or row["convenio_nombre"]
                        or "—"
                    ) if (row["modalidad"] or "") == "Convenio" else "—",
                    row["pdp_concepto_irregularidad"] or "—",
                    monto_emitido if (row["tipo_anexo"] or "") == "PDP" else 0,
                    monto_solventado if (row["tipo_anexo"] or "") == "PDP" else 0,
                    monto_pendiente if (row["tipo_anexo"] or "") == "PDP" else 0,
                ]
            )

        last_data_row = sheet.max_row
        for row_idx in range(2, last_data_row + 1):
            for col_idx in range(1, 13):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                if col_idx in (1, 2, 6, 8, 9):
                    cell.alignment = left_alignment
                elif col_idx in (10, 11, 12):
                    cell.alignment = right_alignment
                else:
                    cell.alignment = center_alignment
            for col in (10, 11, 12):
                sheet.cell(row=row_idx, column=col).number_format = "#,##0.00"
        sheet.auto_filter.ref = f"A1:L{max(last_data_row, 1)}"

        summary_start = sheet.max_row + 2
        sheet.merge_cells(start_row=summary_start, start_column=1, end_row=summary_start, end_column=12)
        summary_header = sheet.cell(row=summary_start, column=1, value="Subtotal / Resumen")
        summary_header.font = white_bold_font
        summary_header.alignment = left_alignment
        summary_header.fill = header_primary_fill
        for col_idx in range(1, 13):
            cell = sheet.cell(row=summary_start, column=col_idx)
            cell.border = thin_border
            cell.fill = header_primary_fill
        summary_rows = [
            ("Total observaciones", total_observaciones),
            ("Observaciones PDP", conteo_pdp),
            ("Observaciones pendientes", conteo_pendiente),
            ("Observaciones solventadas", conteo_solventado),
            ("Monto total emitido", total_emitido),
            ("Monto total solventado", total_solventado),
            ("Monto total pendiente", total_pendiente),
        ]
        for offset, (label, value) in enumerate(summary_rows, start=1):
            current_row = summary_start + offset
            label_cell = sheet.cell(row=current_row, column=1, value=label)
            value_cell = sheet.cell(row=current_row, column=2, value=value)
            label_cell.font = Font(bold=True)
            label_cell.alignment = left_alignment
            value_cell.alignment = right_alignment
            for col_idx in range(1, 13):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.fill = total_fill
            if "Monto" in label:
                value_cell.number_format = "#,##0.00"

        base_widths = {
            1: 40,
            2: 22,
            3: 10,
            4: 10,
            5: 14,
            6: 18,
            7: 22,
            8: 34,
            9: 30,
            10: 16,
            11: 16,
            12: 16,
        }
        for col_idx in range(1, 13):
            max_len = 0
            for row_idx in range(1, sheet.max_row + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                str_val = "" if val is None else str(val)
                if len(str_val) > max_len:
                    max_len = len(str_val)
            width = max(base_widths.get(col_idx, 12), min(max_len + 2, 58))
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"observaciones_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.get("/fuentes-financiamiento-exportar")
    def fuentes_financiamiento_exportar():
        user = get_current_user()
        if user is None:
            return redirect(url_for("login", next=request.path))
        if not (is_luis_user(user) or is_gabo_user(user)):
            return redirect(url_for(home_endpoint_for_user(user), notice="no_permission"))

        ejercicio = " ".join((request.args.get("ejercicio") or "").split())
        if not ejercicio:
            return jsonify({"error": "ejercicio requerido"}), 400

        db = get_db()
        raw_rows = db.execute(
            """
            SELECT ff.nombre AS fuente
            FROM entes_fuentes AS ef
            JOIN fuentes_financiamiento AS ff
              ON ff.id = ef.fuente_id
            WHERE TRIM(COALESCE(ef.ejercicio, '')) = ?
            UNION
            SELECT fuente_financiamiento AS fuente
            FROM observaciones
            WHERE TRIM(COALESCE(ejercicio, '')) = ?
            UNION
            SELECT COALESCE(NULLIF(TRIM(cm.fuente_nombre), ''), ff.nombre) AS fuente
            FROM cargas_manuales AS cm
            LEFT JOIN fuentes_financiamiento AS ff
              ON ff.id = cm.fuente_id
            WHERE TRIM(COALESCE(cm.ejercicio, '')) = ?
            """,
            (ejercicio, ejercicio, ejercicio),
        ).fetchall()

        seen = set()
        fuentes = []
        for row in raw_rows:
            fuente = normalize_fuente_financiamiento(" ".join((row["fuente"] or "").split()))
            if not fuente:
                continue
            key = fuente.casefold()
            if key in seen:
                continue
            seen.add(key)
            fuentes.append(fuente)
        fuentes.sort(key=lambda value: value.casefold())

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Fuentes"
        sheet.sheet_view.showGridLines = False
        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        title_fill = PatternFill("solid", fgColor="1F3B2C")
        subtitle_fill = PatternFill("solid", fgColor="EDF4EF")
        header_fill = PatternFill("solid", fgColor="2A503D")
        zebra_fill = PatternFill("solid", fgColor="F8FAF7")
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_font = Font(bold=True, color="FFFFFF")
        meta_font = Font(bold=True, color="1F3B2C")
        muted_font = Font(color="4B5F52")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right_alignment = Alignment(horizontal="right", vertical="center")

        sheet.merge_cells("A1:B1")
        sheet["A1"] = "Fuentes de Financiamiento"
        sheet["A1"].font = title_font
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = center_alignment

        metadata = [
            ("Ejercicio fiscal", ejercicio),
            ("Total de fuentes", len(fuentes)),
            ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        for row_idx, (label, value) in enumerate(metadata, start=2):
            label_cell = sheet.cell(row=row_idx, column=1, value=label)
            value_cell = sheet.cell(row=row_idx, column=2, value=value)
            label_cell.font = meta_font
            value_cell.font = muted_font
            label_cell.fill = subtitle_fill
            value_cell.fill = subtitle_fill
            label_cell.border = thin_border
            value_cell.border = thin_border
            label_cell.alignment = right_alignment
            value_cell.alignment = left_alignment

        header_row = 6
        sheet.cell(row=header_row, column=1, value="No.")
        sheet.cell(row=header_row, column=2, value="Fuente de Financiamiento")
        for col_idx in range(1, 3):
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_alignment if col_idx == 1 else left_alignment

        if fuentes:
            for index, fuente in enumerate(fuentes, start=1):
                row_idx = header_row + index
                sheet.cell(row=row_idx, column=1, value=index)
                sheet.cell(row=row_idx, column=2, value=fuente)
        else:
            row_idx = header_row + 1
            sheet.cell(row=row_idx, column=1, value="")
            sheet.cell(row=row_idx, column=2, value="Sin fuentes de financiamiento para el ejercicio seleccionado.")

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            for col_idx in range(1, 3):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_alignment if col_idx == 1 else left_alignment
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill

        sheet.freeze_panes = "A7"
        sheet.auto_filter.ref = f"A{header_row}:B{max(sheet.max_row, header_row)}"
        sheet.column_dimensions["A"].width = 9
        max_len = max((len(fuente) for fuente in fuentes), default=52)
        sheet.column_dimensions["B"].width = max(42, min(max_len + 4, 92))
        for row_idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 22
        sheet.page_setup.orientation = "portrait"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"fuentes_financiamiento_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.get("/pendientes-periodo-exportar")
    @luis_required
    def pendientes_periodo_exportar():
        ejercicio = request.args.get("ejercicio", "").strip()
        selected_filters = parse_selected_filters()
        if not ejercicio:
            return jsonify({"error": "ejercicio requerido"}), 400

        selected_entes = selected_values_for_key(selected_filters, "ente_id")
        if not selected_entes:
            return jsonify({"error": "selecciona al menos un ente"}), 400

        db = get_db()
        scope_sql, scope_params = build_observaciones_scope(
            ejercicio,
            selected_filters,
            include_ente=True,
        )
        pendientes_por_periodo = build_pendientes_por_periodo_summary(
            db,
            ejercicio,
            scope_sql,
            scope_params,
            selected_entes,
        )
        groups = pendientes_por_periodo.get("groups", []) or []

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pendientes por periodo"
        thin_border = Border(
            left=Side(style="thin", color="D7DFD9"),
            right=Side(style="thin", color="D7DFD9"),
            top=Side(style="thin", color="D7DFD9"),
            bottom=Side(style="thin", color="D7DFD9"),
        )
        header_primary_fill = PatternFill("solid", fgColor="1F3B2C")
        header_secondary_fill = PatternFill("solid", fgColor="2A503D")
        total_fill = PatternFill("solid", fgColor="EDF4EF")
        zebra_fill = PatternFill("solid", fgColor="F8FAF7")
        white_bold_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        sheet.append(
            [
                "No.",
                "Ente fiscalizable",
                "Tipo de Auditoria",
                "Período cédula",
                "Periodo Titular",
                "PENDIENTES",
                "",
                "",
                "",
                "",
                "",
                "",
                "Numerales de Observaciones No Solventadas",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.append(
            [
                "",
                "",
                "",
                "",
                "",
                "R",
                "SA",
                "PDP",
                "PRAS",
                "PEFCE",
                "Total",
                "Monto Daño ($)",
                "SA",
                "PDP",
                "PRAS",
                "PEFCF",
                "R",
            ]
        )
        sheet.merge_cells("A1:A2")
        sheet.merge_cells("B1:B2")
        sheet.merge_cells("C1:C2")
        sheet.merge_cells("D1:D2")
        sheet.merge_cells("E1:E2")
        sheet.merge_cells("F1:L1")
        sheet.merge_cells("M1:Q1")
        for row_idx in (1, 2):
            for cell in sheet[row_idx]:
                cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = header_primary_fill if row_idx == 1 else header_secondary_fill
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 22
        sheet.freeze_panes = "A3"
        total_rows: list[int] = []

        if not groups:
            sheet.append(["Sin resultados para los filtros seleccionados."])
            sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=17)
            empty_cell = sheet.cell(row=3, column=1)
            empty_cell.alignment = center_alignment
            empty_cell.font = Font(italic=True)
            empty_cell.fill = zebra_fill
            for col_idx in range(1, 18):
                sheet.cell(row=3, column=col_idx).border = thin_border
        else:
            for group in groups:
                periodos = group.get("periodos", []) or []
                if not periodos:
                    sheet.append(
                        [
                            group.get("ente_numero") or group.get("ente_id") or "—",
                            group.get("ente_nombre") or "—",
                            group.get("tipo_auditoria") or "—",
                            "Sin periodos disponibles para este ente.",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
                    row_idx = sheet.max_row
                    for col_idx in range(1, 18):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = left_alignment if col_idx in (2, 4, 5) else center_alignment
                        if col_idx == 12:
                            cell.alignment = right_alignment
                    continue

                for index, periodo in enumerate(periodos):
                    pendientes = periodo.get("pendientes", {}) or {}
                    numerales = periodo.get("numerales_no_solventadas", {}) or {}
                    monto_row = float(pendientes.get("monto_dano") or 0)
                    sheet.append(
                        [
                            (group.get("ente_numero") or group.get("ente_id") or "—") if index == 0 else "",
                            (group.get("ente_nombre") or "—") if index == 0 else "",
                            (group.get("tipo_auditoria") or "—") if index == 0 else "",
                            periodo.get("periodo_cedula") or "—",
                            periodo.get("periodo_titular") or "—",
                            int(pendientes.get("R") or 0),
                            int(pendientes.get("SA") or 0),
                            int(pendientes.get("PDP") or 0),
                            int(pendientes.get("PRAS") or 0),
                            int(pendientes.get("PEFCF") or 0),
                            int(pendientes.get("total") or 0),
                            monto_row if monto_row > 0 else "-",
                            numerales.get("SA") or "-",
                            numerales.get("PDP") or "-",
                            numerales.get("PRAS") or "-",
                            numerales.get("PEFCF") or "-",
                            numerales.get("R") or "-",
                        ]
                    )
                    row_idx = sheet.max_row
                    for col_idx in range(1, 18):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if col_idx in (2, 4, 5):
                            cell.alignment = left_alignment
                        elif col_idx == 12:
                            cell.alignment = right_alignment
                        else:
                            cell.alignment = center_alignment

                totales = group.get("totales", {}) or {}
                monto_total = float(totales.get("monto_dano") or 0)
                sheet.append(
                    [
                        "",
                        "TOTAL",
                        "",
                        "",
                        "",
                        int(totales.get("R") or 0),
                        int(totales.get("SA") or 0),
                        int(totales.get("PDP") or 0),
                        int(totales.get("PRAS") or 0),
                        int(totales.get("PEFCF") or 0),
                        int(totales.get("total") or 0),
                        monto_total if monto_total > 0 else "-",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
                total_row_idx = sheet.max_row
                total_rows.append(total_row_idx)
                for col_idx in range(1, 18):
                    cell = sheet.cell(row=total_row_idx, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                    cell.border = thin_border
                    if col_idx in (2, 4, 5):
                        cell.alignment = left_alignment
                    elif col_idx == 12:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = center_alignment

        for row_idx in range(3, sheet.max_row + 1):
            monto_cell = sheet.cell(row=row_idx, column=12)
            if isinstance(monto_cell.value, (int, float)):
                monto_cell.number_format = "#,##0.00"
            if groups and row_idx not in total_rows and row_idx % 2 == 0:
                for col_idx in range(1, 18):
                    sheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

        base_widths = {
            1: 8,
            2: 40,
            3: 22,
            4: 18,
            5: 18,
            6: 8,
            7: 8,
            8: 8,
            9: 8,
            10: 8,
            11: 10,
            12: 16,
            13: 16,
            14: 16,
            15: 16,
            16: 16,
            17: 14,
        }
        for col_idx in range(1, 18):
            max_len = 0
            for row_idx in range(1, sheet.max_row + 1):
                val = sheet.cell(row=row_idx, column=col_idx).value
                str_val = "" if val is None else str(val)
                if len(str_val) > max_len:
                    max_len = len(str_val)
            width = max(base_widths.get(col_idx, 11), min(max_len + 2, 58))
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.auto_filter.ref = f"A2:Q{sheet.max_row}"

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"pendientes_por_periodo_{ejercicio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.get("/comparativo-anual")
    @luis_required
    def comparativo_anual():
        user = get_current_user()
        can_edit = user["role"] == "editor" if user else False
        return render_template(
            "comparativo_anual.html",
            user=user,
            can_edit=can_edit,
        )

    def render_luis_operational_view(luis_view: str):
        user = get_current_user()
        can_edit = user["role"] == "editor" if user else False
        return render_template(
            "index.html",
            user=user,
            can_edit=can_edit,
            luis_view=luis_view,
        )


    @app.get("/comparativo-anual/stats")
    @luis_required
    def comparativo_anual_stats():
        db = get_db()
        available_years = get_available_comparison_years(db)
        selected_years = parse_comparison_years(available_years)
        selected_filters = parse_comparison_filters()
        selected_filters_payload = {
            "years": selected_years,
            "ente_uid": selected_filters.get("ente_uid", []),
            "tipo_auditoria": selected_filters.get("tipo_auditoria", []),
            "tipo_anexo": selected_filters.get("tipo_anexo", []),
            "estado": selected_filters.get("estado", []),
            "fuente_financiamiento": selected_filters.get("fuente_financiamiento", []),
            "origen_fuente": selected_filters.get("origen_fuente", []),
            "ramo_33": selected_filters.get("ramo_33", []),
            "ramo_28": selected_filters.get("ramo_28", []),
            "universo": selected_filters.get("universo", "all"),
        }

        empty_response = {
            "available_years": available_years,
            "selected_years": selected_years,
            "selected_filters": selected_filters_payload,
            "filter_options": {
                "entes": [],
                "tipo_auditoria": [],
                "tipo_anexo": [],
                "estado": [],
                "fuente_financiamiento": [],
                "origen_fuente": [],
                "ramo_33": [],
                "ramo_28": [],
            },
            "summary": {
                "universo": {
                    "mode": selected_filters.get("universo", "all"),
                    "mode_label": (
                        "Solo entes presentes en todos los años"
                        if selected_filters.get("universo") == "complete"
                        else "Universo completo"
                    ),
                    "entes_filtrados": 0,
                    "entes_en_todos_los_anios": 0,
                },
                "kpis_by_year": [],
                "totals_by_year": [],
                "status_by_year": [],
                "anexo_totals_by_year": [],
                "stacked_by_anexo": [],
                "pdp_amounts_by_year": [],
                "top_variations": {
                    "compare_from": selected_years[0] if selected_years else "",
                    "compare_to": selected_years[-1] if selected_years else "",
                    "top_changes": [],
                    "increases": [],
                    "decreases": [],
                },
                "comparison_table": [],
            },
        }

        cache_key = (
            tuple(selected_years),
            *build_comparison_scope_cache_key(selected_filters),
        )
        now = time.time()
        cached = comparison_cache.get(cache_key)
        if cached and now - cached[0] < comparison_cache_ttl_seconds:
            return jsonify(cached[1])

        if not selected_years:
            comparison_cache[cache_key] = (now, empty_response)
            return jsonify(empty_response)

        base_filters = dict(selected_filters)
        base_filters["universo"] = "all"
        base_rows = fetch_comparison_base_rows(db, selected_years)
        scope_cache: dict = {}
        base_scope = build_comparison_scope_from_rows(
            base_rows,
            selected_years,
            base_filters,
            scope_cache=scope_cache,
        )
        entes_filtrados = len(
            {
                (row.get("ente_uid") or "").strip()
                for row in base_scope["rows"]
                if (row.get("ente_uid") or "").strip()
            }
        )
        entes_en_todos_los_anios = len(
            compute_common_comparison_entity_uids(base_scope["rows"], selected_years)
        )

        scope_payload = build_comparison_scope_from_rows(
            base_rows,
            selected_years,
            selected_filters,
            scope_cache=scope_cache,
        )
        scope_rows = scope_payload["rows"]

        def option_rows_for(exclude_key: str) -> list[dict]:
            return build_comparison_scope_from_rows(
                base_rows,
                selected_years,
                selected_filters,
                exclude_key=exclude_key,
                scope_cache=scope_cache,
            )["rows"]

        filter_options = {
            "entes": build_comparison_ente_options(
                option_rows_for("ente_uid"),
                selected_years,
            ),
            "tipo_auditoria": collect_comparison_distinct_values(
                option_rows_for("tipo_auditoria"),
                "tipo_auditoria",
            ),
            "tipo_anexo": collect_comparison_distinct_values(
                option_rows_for("tipo_anexo"),
                "tipo_anexo",
            ),
            "estado": collect_comparison_distinct_values(
                option_rows_for("estado"),
                "estado",
            ),
            "fuente_financiamiento": collect_comparison_distinct_values(
                option_rows_for("fuente_financiamiento"),
                "fuente_financiamiento",
            ),
            "origen_fuente": collect_comparison_distinct_values(
                option_rows_for("origen_fuente"),
                "origen_fuente",
            ),
            "ramo_33": collect_comparison_distinct_values(
                option_rows_for("ramo_33"),
                "ramo_33",
            ),
            "ramo_28": collect_comparison_distinct_values(
                option_rows_for("ramo_28"),
                "ramo_28",
            ),
        }
        scope_summary = summarize_comparison_scope(scope_rows, selected_years)
        comparison_rows = aggregate_comparison_rows_by_entity(scope_rows)
        comparison_table = build_comparison_table_rows(comparison_rows, selected_years)

        first_year = selected_years[0] if selected_years else ""
        last_year = selected_years[-1] if selected_years else ""
        top_changes = []
        top_increases = []
        top_decreases = []

        def serialize_variation_row(item: dict):
            return {
                "ente_uid": item["ente_uid"],
                "ente_numero": item["ente_numero"],
                "ente_nombre": item["ente_nombre"],
                "label": item["label"],
                "counts_by_year": item["counts_by_year"],
                "delta_abs": item["delta_abs"],
                "delta_pct": item["delta_pct"],
                "change_label": item.get("change_label", ""),
                "aliases": item["aliases"],
                "has_historical_names": item["has_historical_names"],
            }

        if len(selected_years) >= 2:
            top_changes = [
                serialize_variation_row(item)
                for item in comparison_table[:5]
            ]
            sorted_increases = sorted(
                [row for row in comparison_table if row["delta_abs"] > 0],
                key=lambda item: (
                    -int(item["delta_abs"] or 0),
                    item.get("ente_nombre") or item.get("label") or "",
                ),
            )
            sorted_decreases = sorted(
                [row for row in comparison_table if row["delta_abs"] < 0],
                key=lambda item: (
                    int(item["delta_abs"] or 0),
                    item.get("ente_nombre") or item.get("label") or "",
                ),
            )
            top_increases = [serialize_variation_row(item) for item in sorted_increases[:5]]
            top_decreases = [serialize_variation_row(item) for item in sorted_decreases[:5]]

        payload = {
            "available_years": available_years,
            "selected_years": selected_years,
            "selected_filters": selected_filters_payload,
            "filter_options": filter_options,
            "summary": {
                "universo": {
                    "mode": selected_filters.get("universo", "all"),
                    "mode_label": (
                        "Solo entes presentes en todos los años"
                        if selected_filters.get("universo") == "complete"
                        else "Universo completo"
                    ),
                    "entes_filtrados": entes_filtrados,
                    "entes_en_todos_los_anios": entes_en_todos_los_anios,
                },
                "kpis_by_year": scope_summary["kpis_by_year"],
                "totals_by_year": scope_summary["totals_by_year"],
                "status_by_year": scope_summary["status_by_year"],
                "anexo_totals_by_year": scope_summary["anexo_totals_by_year"],
                "stacked_by_anexo": scope_summary["stacked_by_anexo"],
                "pdp_amounts_by_year": scope_summary["pdp_amounts_by_year"],
                "top_variations": {
                    "compare_from": first_year,
                    "compare_to": last_year,
                    "top_changes": top_changes,
                    "increases": top_increases,
                    "decreases": top_decreases,
                },
                "comparison_table": comparison_table,
            },
        }
        comparison_cache[cache_key] = (now, payload)
        if len(comparison_cache) > 250:
            oldest_key = min(comparison_cache.items(), key=lambda item: item[1][0])[0]
            comparison_cache.pop(oldest_key, None)
        return jsonify(payload)


    @app.get("/observaciones-stats")
    @luis_required
    def observaciones_stats():
        ente_id = normalize_ente_id(request.args.get("ente_id", ""))
        tipo_auditoria = normalize_tipo_auditoria(request.args.get("tipo_auditoria", ""))
        where_clauses = []
        params = []
        if ente_id:
            where_clauses.append(f"{normalize_ente_id_sql('ente_id')} = ?")
            params.append(ente_id)
        if tipo_auditoria:
            where_clauses.append("tipo_auditoria = ?")
            params.append(tipo_auditoria)
    
        where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    
        db = get_db()
        totals = db.execute(
            f"""
            SELECT ejercicio, COUNT(*) as total
            FROM observaciones
            {where_clause}
            GROUP BY ejercicio
            ORDER BY ejercicio
            """,
            params,
        ).fetchall()
    
        estados = db.execute(
            f"""
            SELECT ejercicio, estado, COUNT(*) as total
            FROM observaciones
            {where_clause}
            GROUP BY ejercicio, estado
            ORDER BY ejercicio
            """,
            params,
        ).fetchall()
    
        tipos = db.execute(
            f"""
            SELECT ejercicio, tipo_anexo, COUNT(*) as total
            FROM observaciones
            {where_clause}
            GROUP BY ejercicio, tipo_anexo
            ORDER BY ejercicio
            """,
            params,
        ).fetchall()
    
        tipos_estados = db.execute(
            f"""
            SELECT ejercicio, tipo_anexo, estado, COUNT(*) as total
            FROM observaciones
            {where_clause}
            GROUP BY ejercicio, tipo_anexo, estado
            ORDER BY ejercicio
            """,
            params,
        ).fetchall()
    
        fuentes = db.execute(
            f"""
            SELECT ejercicio, fuente_financiamiento, COUNT(*) as total
            FROM observaciones
            {where_clause}
            GROUP BY ejercicio, fuente_financiamiento
            ORDER BY ejercicio
            """,
            params,
        ).fetchall()
    
        pdp_where = "WHERE tipo_anexo = 'PDP'"
        pdp_clauses = []
        pdp_params = []
        if ente_id:
            pdp_clauses.append(f"{normalize_ente_id_sql('ente_id')} = ?")
            pdp_params.append(ente_id)
        if tipo_auditoria:
            pdp_clauses.append("tipo_auditoria = ?")
            pdp_params.append(tipo_auditoria)
        if pdp_clauses:
            pdp_where = f"WHERE {' AND '.join(pdp_clauses)} AND tipo_anexo = 'PDP'"
    
        pdp = db.execute(
            f"""
            SELECT
                ejercicio,
                SUM(COALESCE(monto_pdp_emitido, 0)) as emitido,
                SUM(COALESCE(monto_pdp_solventado, 0)) as solventado,
                SUM(COALESCE(monto_pdp_pendiente, 0)) as pendiente
            FROM observaciones
            {pdp_where}
            GROUP BY ejercicio
            ORDER BY ejercicio
            """,
            pdp_params,
        ).fetchall()
    
        return jsonify(
            {
                "totals": [dict(row) for row in totals],
                "estados": [dict(row) for row in estados],
                "tipos": [dict(row) for row in tipos],
                "tipos_estados": [dict(row) for row in tipos_estados],
                "fuentes": [dict(row) for row in fuentes],
                "pdp": [dict(row) for row in pdp],
            }
        )
    
    
    @app.get("/catalogo-entes")
    @luis_required
    def catalogo_entes():
        ejercicio = request.args.get("ejercicio", "").strip()
        if not ejercicio:
            return jsonify([])
    
        db = get_db()
        rows = db.execute(
            f"""
            SELECT
                entes_detalle.ente_id,
                entes_detalle.ente_uid,
                entes_detalle.ente_numero,
                entes_detalle.ente_nombre,
                entes_detalle.ejercicio,
                entes_detalle.clasificacion,
                entes_detalle.ramo33,
                entes_detalle.ramo28,
                entes_detalle.responsable,
                (
                    SELECT ed.ente_nombre
                    FROM entes_detalle AS ed
                    WHERE COALESCE(ed.ente_uid, ed.ente_id)
                      = COALESCE(entes_detalle.ente_uid, entes_detalle.ente_id)
                      AND ed.ejercicio < entes_detalle.ejercicio
                    ORDER BY ed.ejercicio DESC
                    LIMIT 1
                ) AS nombre_anterior,
                (
                    SELECT COUNT(DISTINCT ed.ente_nombre)
                    FROM entes_detalle AS ed
                    WHERE COALESCE(ed.ente_uid, ed.ente_id)
                      = COALESCE(entes_detalle.ente_uid, entes_detalle.ente_id)
                ) AS nombres_distintos
            FROM entes_detalle
            WHERE entes_detalle.ejercicio = ?
            ORDER BY {ente_numero_sort_sql('entes_detalle.ente_numero')} ASC, entes_detalle.ente_numero ASC
            """,
            (ejercicio,),
        ).fetchall()
    
        return jsonify([dict(row) for row in rows])
    
    
    @app.post("/fuentes")
    @luis_required
    @role_required("editor")
    def fuentes():
        nombre = normalize_fuente_financiamiento(request.form.get("fuente_nombre", "").strip())
        if not nombre:
            return redirect(url_for("index", notice="fuente_error"))
    
        db = get_db()
        db.execute(
            """
            INSERT OR IGNORE INTO fuentes_financiamiento (nombre, created_at)
            VALUES (?, ?)
            """,
            (nombre, datetime.now().strftime("%Y-%m-%d %H:%M")),
        )
        db.commit()
        return redirect(url_for("index", notice="fuente_saved"))
    
    
    @app.post("/irregularidades")
    @luis_required
    @role_required("editor")
    def irregularidades():
        try:
            concepto = normalize_irregularidad_concepto(
                request.form.get("irregularidad_concepto", "").strip(),
                strict=True,
            )
        except ValueError:
            concepto = ""
        if not concepto:
            return redirect(url_for("index", notice="irregularidad_error"))
    
        db = get_db()
        db.execute(
            """
            INSERT OR IGNORE INTO catalogo_irregularidades (concepto, created_at)
            VALUES (?, ?)
            """,
            (concepto, datetime.now().strftime("%Y-%m-%d %H:%M")),
        )
        db.commit()
        return redirect(url_for("index", notice="irregularidad_saved"))
    
    
    @app.route("/stats")
    @luis_required
    def stats():
        db = get_db()
        totals = db.execute(
            """
            SELECT ejercicio, COUNT(*) as total
            FROM registros
            GROUP BY ejercicio
            ORDER BY ejercicio
            """
        ).fetchall()
    
        estados = db.execute(
            """
            SELECT ejercicio, estado, COUNT(*) as total
            FROM registros
            GROUP BY ejercicio, estado
            ORDER BY ejercicio
            """
        ).fetchall()
    
        tipos = db.execute(
            """
            SELECT ejercicio, tipo_anexo, COUNT(*) as total
            FROM registros
            GROUP BY ejercicio, tipo_anexo
            ORDER BY ejercicio
            """
        ).fetchall()
    
        return jsonify(
            {
                "totals": [dict(row) for row in totals],
                "estados": [dict(row) for row in estados],
                "tipos": [dict(row) for row in tipos],
            }
        )
    
    
    @app.post("/reclasificar/<int:registro_id>")
    @luis_required
    @role_required("editor")
    def reclasificar(registro_id: int):
        db = get_db()
        row = db.execute(
            """
            SELECT id, tipo_anexo, tipo_anexo_origen
            FROM registros
            WHERE id = ?
            """,
            (registro_id,),
        ).fetchone()
    
        if row is None:
            return redirect(url_for("index", saved="0"))
    
        if row["tipo_anexo"] not in {"PDP", "PRAS"}:
            return redirect(url_for("index"))
    
        nuevo_tipo = "PRAS" if row["tipo_anexo"] == "PDP" else "PDP"
        origen = row["tipo_anexo_origen"] or row["tipo_anexo"]
        db.execute(
            """
            UPDATE registros
            SET tipo_anexo = ?, tipo_anexo_origen = ?
            WHERE id = ?
            """,
            (nuevo_tipo, origen, registro_id),
        )
        db.commit()
        return redirect(url_for("index", saved="1"))
    
    
    @app.get("/")
    @luis_required
    def index():
        return render_luis_operational_view("consulta")

    @app.get("/resumen-general")
    @luis_required
    def luis_resumen_general():
        return render_luis_operational_view("resumen_general")

    @app.get("/tipo-auditoria")
    @luis_required
    def luis_tipo_auditoria():
        return render_luis_operational_view("tipo_auditoria")

    @app.get("/fuente-financiamiento")
    @luis_required
    def luis_fuente_financiamiento():
        return render_luis_operational_view("fuente_financiamiento")

    @app.get("/graficas")
    @luis_required
    def luis_graficas():
        return render_luis_operational_view("graficas")

    @app.get("/pendientes-periodo")
    @luis_required
    def luis_pendientes_periodo():
        return render_luis_operational_view("pendientes")

    @app.get("/titulares-administrativos")
    @luis_required
    def luis_titulares_administrativos():
        return render_luis_operational_view("titulares")

    @app.get("/catalogo")
    @luis_required
    def luis_catalogo():
        return render_luis_operational_view("catalogo")
    
    
