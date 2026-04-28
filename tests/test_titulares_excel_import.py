from io import BytesIO
import sqlite3

import pytest
from openpyxl import Workbook, load_workbook


@pytest.fixture
def isolated_client(monkeypatch, tmp_path):
    import app as app_module
    import scripts.gabo_routes as gabo_routes

    test_db_path = tmp_path / "titulares_excel_test.db"
    monkeypatch.setattr(app_module, "DB_PATH", str(test_db_path))
    monkeypatch.setattr(gabo_routes, "DB_PATH", str(test_db_path))
    monkeypatch.setattr(gabo_routes, "BASE_DIR", str(tmp_path))
    app_module.init_db()

    with sqlite3.connect(test_db_path) as conn:
        conn.execute(
            """
            INSERT INTO entes_detalle (
                ente_uid,
                ente_id,
                ejercicio,
                ente_numero,
                ente_nombre,
                responsable,
                clasificacion,
                ramo33,
                ramo28,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "ENTE-1-16",
                "1.16",
                "2025",
                "1.16",
                "Secretaría de Infraestructura (SI)",
                "",
                "",
                "No",
                "No",
                "2026-04-28 00:00",
            ),
        )

    app_module.app.config["TESTING"] = True
    with app_module.app.test_client() as client:
        with client.session_transaction() as session_data:
            session_data["user"] = "gabo"
            session_data["role"] = "loader"
        yield client, test_db_path


def _build_periodos_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SI"
    ws["A1"] = "EJERCICIO 2025"
    ws.merge_cells("A1:E1")
    ws["A2"] = "Nombre del ente:"
    ws["B2"] = "SECRETARIA DE INFRAESTRUCTURA"
    ws.append([])
    ws.append(
        [
            "Periodos Informe",
            "Titular",
            "Administrativo",
            "Director Administrativo a cargo",
            "Cédulas de resultados",
        ]
    )
    ws.append(
        [
            "01 de Enero al 15 de Mayo",
            "Diego Corona Cremean",
            "01 de enero al 31 de diciembre",
            "Julio Cesar Meneses Guerrero",
            "01 de enero al 15 de mayo",
        ]
    )
    ws.append(
        [
            "16 de Mayo al 31 de Diciembre",
            "Eduardo Ruben Hernandez Tapia",
            "",
            "",
            "16 de mayo al 30 de junio",
        ]
    )
    ws.append(["", "", "", "", "01 de julio al 31 de diciembre"])
    ws.merge_cells("A6:A7")
    ws.merge_cells("B6:B7")
    ws.merge_cells("C5:C7")
    ws.merge_cells("D5:D7")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def test_titulares_excel_preview_and_apply(isolated_client):
    client, db_path = isolated_client
    workbook_bytes = _build_periodos_workbook()

    preview_response = client.post(
        "/carga/titulares/excel/preview",
        data={
            "ejercicio": "2025",
            "ente_id": "1.16",
            "tipo_auditoria_destino": "Financiera",
            "titulares_file": (BytesIO(workbook_bytes), "1.16 Periodos_Cedulas 2025_SI_ok.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert preview_response.status_code == 200
    preview = preview_response.get_json()
    assert preview["summary"]["files_total"] == 1
    assert preview["summary"]["errors"] == 0
    assert preview["summary"]["titulares"] == 2
    assert preview["summary"]["administrativos"] == 1
    assert preview["summary"]["capturas"] == 2
    assert preview["summary"]["cedulas"] == 3

    apply_response = client.post(
        "/carga/titulares/excel/aplicar",
        json={
            "tipo_auditoria_destino": preview["tipo_auditoria_destino"],
            "history_rows": preview["history_rows"],
            "capture_rows": preview["capture_rows"],
        },
    )

    assert apply_response.status_code == 200
    applied = apply_response.get_json()
    assert applied["history"]["inserted"] == 3
    assert applied["captures"]["inserted"] == 2

    with sqlite3.connect(db_path) as conn:
        historial_count = conn.execute("SELECT COUNT(*) FROM historial_titulares").fetchone()[0]
        captura_count = conn.execute("SELECT COUNT(*) FROM cargas_titulares").fetchone()[0]
        second_cedulas = conn.execute(
            """
            SELECT cedula_resultados
            FROM cargas_titulares
            WHERE titular = 'Eduardo Ruben Hernandez Tapia'
            ORDER BY id
            LIMIT 1
            """
        ).fetchone()[0]

    assert historial_count == 3
    assert captura_count == 2
    assert "16 de mayo al 30 de junio" in second_cedulas
    assert "01 de julio al 31 de diciembre" in second_cedulas

    history_response = client.get(
        "/carga/titulares/historial?ejercicio=2025&ente_id=1.16&tipo_auditoria=Financiera"
    )
    assert history_response.status_code == 200
    history_payload = history_response.get_json()
    assert len(history_payload["rows"]) == 3
    assert len(history_payload["capture_rows"]) == 3
    assert history_payload["capture_rows"][0]["periodo_informe"] == "01 de enero al 15 de mayo"
    assert all("|" not in row["cedula_resultados"] for row in history_payload["capture_rows"])
    eduardo_rows = [
        row for row in history_payload["capture_rows"]
        if row["titular"] == "Eduardo Ruben Hernandez Tapia"
    ]
    assert [row["cedula_resultados"] for row in eduardo_rows] == [
        "16 de mayo al 30 de junio",
        "01 de julio al 31 de diciembre",
    ]

    export_response = client.get(
        "/carga/titulares/exportar?ejercicio=2025&ente_id=1.16&tipo_auditoria=Financiera"
    )
    assert export_response.status_code == 200
    assert export_response.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    export_workbook = load_workbook(BytesIO(export_response.data))
    worksheet = export_workbook.active
    merged_ranges = {str(merged_range) for merged_range in worksheet.merged_cells.ranges}
    assert "A2:A4" in merged_ranges
    assert "B2:B4" in merged_ranges
    assert "C3:C4" in merged_ranges
    assert "D3:D4" in merged_ranges


def test_titulares_manual_save_is_visible_in_consulta(isolated_client):
    client, db_path = isolated_client

    response = client.post(
        "/carga/titulares",
        data={
            "titular_ejercicio": "2025",
            "titular_ente_id": "1.16",
            "titular_tipo_auditoria": "Financiera",
            "titular_nombre": "Titular Manual",
            "titular_fecha_inicio": "2025-01-01",
            "titular_fecha_fin": "2025-07-31",
            "titular_administrativo": "Administrativo Manual",
            "titular_admin_fecha_inicio": "2025-01-01",
            "titular_admin_fecha_fin": "2025-04-24",
        },
    )

    assert response.status_code == 200

    with sqlite3.connect(db_path) as conn:
        historial_count = conn.execute("SELECT COUNT(*) FROM historial_titulares").fetchone()[0]
        captura_count = conn.execute("SELECT COUNT(*) FROM cargas_titulares").fetchone()[0]

    assert historial_count == 2
    assert captura_count == 1

    history_response = client.get(
        "/carga/titulares/historial?ejercicio=2025&ente_id=1.16&tipo_auditoria=Financiera"
    )
    assert history_response.status_code == 200
    history_payload = history_response.get_json()
    assert len(history_payload["rows"]) == 2
    assert len(history_payload["capture_rows"]) == 1
    assert history_payload["capture_rows"][0]["titular"] == "Titular Manual"
    assert history_payload["capture_rows"][0]["administrativo"] == "Administrativo Manual"

    general_history_response = client.get("/carga/titulares/historial?ejercicio=2025")
    assert general_history_response.status_code == 200
    general_history_payload = general_history_response.get_json()
    assert len(general_history_payload["rows"]) == 2
    assert len(general_history_payload["capture_rows"]) == 1


def test_titulares_page_uses_general_table_and_hidden_capture(isolated_client):
    client, _db_path = isolated_client

    response = client.get("/carga/titulares")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert 'id="titularHistoryTableBody"' in html
    assert "Agregar Titular" in html
    assert "Exportar Excel" in html
    assert 'id="titularCaptureWorkspace"' in html
    assert "Todas las Auditorías" in html
    assert "rowspan" in html
